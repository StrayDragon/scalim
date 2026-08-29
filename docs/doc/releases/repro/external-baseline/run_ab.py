#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""外部基线对比矩阵（scalim vs pandas/polars 惯用法）· 版本锚定证据脚本.

本页数字的唯一生成入口；数据资产: docs/doc/assets/data/external-baseline-0.10.json。
官方测量: scalim 0.10.3 / pandas 2.3.3 / polars 1.42.1 / openpyxl 3.1.5 / Python 3.10.18。

与 releases 下其它 repro 的差异:
  - 对照物是外部库惯用法（pandas/polars：DataFrame 全量物化 + 向量化派生 + 库内写出），
    不是 scalim 自身新旧路径；不含手写 openpyxl write_only 流式（那等于用户自己手写流式管道）
  - 测量环境 MUST 为 Python 3.10+(pandas/polars 2.x/1.x 依赖)；不做 Python 3.6 兼容
  - polars 侧需 --with polars（xlsx 形状另需 --with xlsxwriter）；未安装时该侧记录 skipped 不失败
  - 子进程隔离 + VmHWM 真实峰值；golden = run0 全量读回校验 + 其余 run 行数/首50行精确校验
  - 单 run 峰值内存预算 = MemAvailable - 10%*MemTotal(超预算时脚本告警)
  - scalim 侧测 seq 与 adaptive(max_workers=4) 两种模式；polars 自身默认多线程(页面注明)

Shapes(7 个典型,只含计数、无业务字段名):
  S1 report_wide_xlsx   报表宽表 xlsx      30k rows, 3 base + 100 flat,           sink=xlsx
  S2 wide_export_csv    大宽表导出 csv     150k rows, 3 base + 100 flat,          sink=csv
  S3 chain_boundary_xls 链式边界 xlsx      20k rows, 3 base + 4 flat + 30 chain,  sink=xlsx
  S4 long_rows_csv      大长表 csv        2M rows, 3 base + 4 flat,              sink=csv
  S5 wide_cols_csv      大宽表·超多列 csv  20k rows, 3 base + 600 flat,           sink=csv
  S6 wide_cols_xlsx     大宽表·超多列 xlsx  12k rows, 3 base + 400 flat,          sink=xlsx
  S7 relation_sqlite_csv 多源关联 csv     30k main(fk->side 5k, 真实IO SQLite),
                                           输出 3 base + s0/s1(lookup) + d0/d1(派生)

两侧实现同一用户任务（端到端：消费源行 -> 算派生 -> 写真实文件）：
  pandas : 生成器/SQL -> DataFrame -> 向量化派生 -> to_excel/to_csv
  polars : 生成器/SQL -> DataFrame -> with_columns 向量化 -> write_excel/write_csv
  scalim : 生成器/SQL 惰性供给 -> DemandIr/Engine(call_by 每行每字段一次) -> ExcelSink/CSVSink

运行（仓库根目录；推荐经 just target 固定依赖）：
    just bench-external --runs 3          # 等价下方命令
    uv run --with polars --with xlsxwriter python docs/doc/releases/repro/external-baseline/run_ab.py --runs 3
    uv run --with polars python docs/doc/releases/repro/external-baseline/run_ab.py --runs 1 --rows-scale 0.05  # 冒烟

输出: <repo>/.tmp/evidence/external-baseline/<timestamp>/result.json (rebuildable; not committed)
诚实边界: 薄算术 calculator;单机单环境;S7 为本机 SQLite 真实 IO(确定性 fixture);数字不可跨机迁移。
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import time
from typing import Any, Dict, List


REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
_cur = REPO_ROOT
while _cur != os.path.dirname(_cur):
    if os.path.isfile(os.path.join(_cur, "pyproject.toml")) and os.path.isdir(os.path.join(_cur, "src")):
        REPO_ROOT = _cur
        break
    _cur = os.path.dirname(_cur)

# ---------------------------------------------------------------------------
# Shapes（只含计数参数；rows 按 rows-scale 缩放用于冒烟）
# ---------------------------------------------------------------------------
SHAPES = [
    {
        "id": "S1_report_wide_xlsx",
        "label": "报表宽表 xlsx",
        "rows": 30000,
        "n_flat": 100,
        "chain_depth": 0,
        "sink": "xlsx",
        "batch_size": 500,
        "short": False,
        "why": "报表宽表典型：3M cells；pandas 对照 = to_excel(openpyxl)",
    },
    {
        "id": "S2_wide_export_csv",
        "label": "大宽表导出 csv",
        "rows": 150000,
        "n_flat": 100,
        "chain_depth": 0,
        "sink": "csv",
        "batch_size": 1000,
        "short": False,
        "why": "大宽表 csv 导出：15M cells；pandas to_csv / polars write_csv 均为 C/多线程强项",
    },
    {
        "id": "S3_chain_boundary_xlsx",
        "label": "边界区·链式派生 xlsx",
        "rows": 20000,
        "n_flat": 4,
        "chain_depth": 30,
        "sink": "xlsx",
        "batch_size": 500,
        "short": True,
        "why": "对齐 c10 诚实边界场景（chain 前缀驻留，收益最小叙事）",
    },
    {
        "id": "S4_long_rows_csv",
        "label": "大长表·多行少列 csv",
        "rows": 2000000,
        "n_flat": 4,
        "chain_depth": 0,
        "sink": "csv",
        "batch_size": 1000,
        "short": True,
        "why": "大长表典型（行数主导 2M rows）：流式行吞吐 vs 全量 df",
    },
    {
        "id": "S5_wide_cols_csv",
        "label": "大宽表·超多列 csv",
        "rows": 20000,
        "n_flat": 600,
        "chain_depth": 0,
        "sink": "csv",
        "batch_size": 500,
        "short": False,
        "why": "大宽表典型（列数主导 600 派生列）：宽表 csv 下内存/时间取舍",
    },
    {
        "id": "S6_wide_cols_xlsx",
        "label": "大宽表·超多列 xlsx",
        "rows": 12000,
        "n_flat": 400,
        "chain_depth": 0,
        "sink": "xlsx",
        "batch_size": 500,
        "short": False,
        "why": "大宽表 Excel 典型（scalim 写出布局主场景）：与 S1 构成宽度扩展轴",
    },
    {
        "id": "S7_relation_sqlite_csv",
        "label": "多源关联·真实IO csv",
        "rows": 30000,
        "n_flat": 0,
        "chain_depth": 0,
        "sink": "csv",
        "batch_size": 500,
        "short": True,
        "relation": True,
        "side_rows": 5000,
        "why": "多源关联叙事：main(fk)->side 1:1 lookup 经真实 SQLite IO；pandas 对照 = read_sql+merge",
    },
]

BASE_FIELDS = ("id", "v0", "v1")
SAMPLE_ROWS = 50
CHECK_FIELDS_COUNT = 2  # 校验和抽检的派生列数
RELATION_SIDE_ROWS = 5000  # S7 侧表行数（不随 rows-scale 缩放，保证 fk 取模口径一致）


def scaled_rows(rows: int, scale: float) -> int:
    return max(1000, int(rows * scale))


# ---------------------------------------------------------------------------
# 共享数据语义（各侧严格同值；确定性生成，无随机）
# ---------------------------------------------------------------------------
def source_row(i: int) -> Dict[str, Any]:
    return {"id": i, "v0": float(i % 97), "v1": float(i % 13)}


def expected_flat(v0: float, v1: float, j: int) -> float:
    if j % 3 == 0:
        return v0 + v1
    if j % 3 == 1:
        return v0 - v1
    return v0 * v1


def expected_chain(v0: float, k: int) -> float:
    return v0 + float(k + 1)


def side_row(sid: int) -> Dict[str, Any]:
    return {"sid": sid, "s0": (sid % 17) * 1.5, "s1": (sid % 29) * 0.5}


def out_fields(n_flat: int, chain_depth: int, relation: bool = False) -> List[str]:
    fields = list(BASE_FIELDS) + ["d{}".format(j) for j in range(n_flat)]
    fields += ["c{}".format(k) for k in range(chain_depth)]
    if relation:
        fields += ["s0", "s1", "r0", "r1"]  # lookup 字段 + 关联派生
    return fields


def checksum_targets(n_flat: int, chain_depth: int, relation: bool) -> List[str]:
    targets: List[str] = []
    if n_flat:
        targets.append("d0")
    if chain_depth:
        targets.append("c0")
    elif n_flat > 1:
        targets.append("d1")
    if relation:
        targets += ["r0", "r1"]
    return targets


def expected_checksums(rows: int, n_flat: int, chain_depth: int, relation: bool = False,
                       only_first_n: int = -1) -> Dict[str, float]:
    """解析式期望校验和；only_first_n>=0 时只累计前 n 行（sample 校验用）。"""
    targets = checksum_targets(n_flat, chain_depth, relation)
    sums: Dict[str, float] = {name: 0.0 for name in targets}
    limit = rows if only_first_n < 0 else min(rows, only_first_n)
    for i in range(limit):
        v0, v1 = float(i % 97), float(i % 13)
        if "d0" in sums:
            sums["d0"] += expected_flat(v0, v1, 0)
        if "c0" in sums:
            sums["c0"] += expected_chain(v0, 0)
        if "d1" in sums and chain_depth == 0:
            sums["d1"] += expected_flat(v0, v1, 1)
        if relation:
            s = side_row(i % RELATION_SIDE_ROWS)
            if "r0" in sums:
                sums["r0"] += v0 + s["s0"]
            if "r1" in sums:
                sums["r1"] += v1 * s["s1"]
    return sums


def checksum_close(a: float, b: float) -> bool:
    return abs(a - b) <= 1e-9 * max(1.0, abs(a), abs(b))


# ---------------------------------------------------------------------------
# S7 SQLite fixture（真实 IO；确定性内容；双侧共读同一 db 文件）
# ---------------------------------------------------------------------------
def ensure_sqlite_fixture(db_path: str, main_rows: int, side_rows: int) -> None:
    import sqlite3

    if os.path.exists(db_path):
        return
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        cur.execute("CREATE TABLE main_rows (id INTEGER PRIMARY KEY, fk INTEGER, v0 REAL, v1 REAL)")
        cur.execute("CREATE TABLE side_rows (sid INTEGER PRIMARY KEY, s0 REAL, s1 REAL)")
        cur.executemany(
            "INSERT INTO main_rows VALUES (?,?,?,?)",
            [(i, i % side_rows, float(i % 97), float(i % 13)) for i in range(main_rows)],
        )
        cur.executemany(
            "INSERT INTO side_rows VALUES (?,?,?)",
            [(sid, (sid % 17) * 1.5, (sid % 29) * 0.5) for sid in range(side_rows)],
        )
        conn.commit()
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# pandas 侧（用户诚实写法：DataFrame + 向量化 + to_excel/to_csv / read_sql+merge）
# ---------------------------------------------------------------------------
def run_pandas(shape: Dict[str, Any], out_path: str) -> Dict[str, Any]:
    import pandas as pd

    rows = shape["_rows"]
    n_flat, chain = shape["n_flat"], shape["chain_depth"]

    t0 = time.perf_counter()
    if shape.get("relation"):
        import sqlite3

        db = shape["_db_path"]
        conn = sqlite3.connect(db)
        try:
            dfm = pd.read_sql_query("SELECT id, fk, v0, v1 FROM main_rows", conn)
            dfs = pd.read_sql_query("SELECT sid, s0, s1 FROM side_rows", conn)
        finally:
            conn.close()
        df = dfm.merge(dfs, left_on="fk", right_on="sid", how="left")
        df["r0"] = df["v0"] + df["s0"]
        df["r1"] = df["v1"] * df["s1"]
        df = df[out_fields(n_flat, chain, relation=True)]
        df.to_csv(out_path, header=True, index=False)
        elapsed = time.perf_counter() - t0
        return {"total_s": elapsed}

    df = pd.DataFrame(source_row(i) for i in range(rows))
    v0, v1 = df["v0"].to_numpy(), df["v1"].to_numpy()
    for j in range(n_flat):
        if j % 3 == 0:
            df["d{}".format(j)] = v0 + v1
        elif j % 3 == 1:
            df["d{}".format(j)] = v0 - v1
        else:
            df["d{}".format(j)] = v0 * v1
    for k in range(chain):
        # 语义等价 c_k = c_{k-1} + 1 = v0 + k + 1；向量化为 pandas 诚实写法
        df["c{}".format(k)] = v0 + float(k + 1)
    df = df[out_fields(n_flat, chain)]
    if shape["sink"] == "xlsx":
        df.to_excel(out_path, engine="openpyxl", header=True, index=False)
    else:
        df.to_csv(out_path, header=True, index=False)
    elapsed = time.perf_counter() - t0
    return {"total_s": elapsed}


# ---------------------------------------------------------------------------
# polars 侧（惯用法：DataFrame + with_columns 向量化 + write_excel/write_csv）
# ---------------------------------------------------------------------------
def run_polars(shape: Dict[str, Any], out_path: str) -> Dict[str, Any]:
    import polars as pl

    rows = shape["_rows"]
    n_flat, chain = shape["n_flat"], shape["chain_depth"]

    t0 = time.perf_counter()
    if shape.get("relation"):
        import sqlite3

        db = shape["_db_path"]
        conn = sqlite3.connect(db)
        try:
            dfm = pl.DataFrame(
                conn.execute("SELECT id, fk, v0, v1 FROM main_rows").fetchall(),
                schema={"id": pl.Int64, "fk": pl.Int64, "v0": pl.Float64, "v1": pl.Float64},
                orient="row",
            )
            dfs = pl.DataFrame(
                conn.execute("SELECT sid, s0, s1 FROM side_rows").fetchall(),
                schema={"sid": pl.Int64, "s0": pl.Float64, "s1": pl.Float64},
                orient="row",
            )
        finally:
            conn.close()
        df = dfm.join(dfs, left_on="fk", right_on="sid", how="left")
        df = df.with_columns([
            (pl.col("v0") + pl.col("s0")).alias("r0"),
            (pl.col("v1") * pl.col("s1")).alias("r1"),
        ])
        df = df.select(out_fields(n_flat, chain, relation=True))
        df.write_csv(out_path, include_header=True)
        elapsed = time.perf_counter() - t0
        return {"total_s": elapsed}

    df = pl.from_dicts([source_row(i) for i in range(rows)],
                       schema={"id": pl.Int64, "v0": pl.Float64, "v1": pl.Float64})
    exprs = []
    for j in range(n_flat):
        if j % 3 == 0:
            exprs.append((pl.col("v0") + pl.col("v1")).alias("d{}".format(j)))
        elif j % 3 == 1:
            exprs.append((pl.col("v0") - pl.col("v1")).alias("d{}".format(j)))
        else:
            exprs.append((pl.col("v0") * pl.col("v1")).alias("d{}".format(j)))
    for k in range(chain):
        exprs.append((pl.col("v0") + float(k + 1)).alias("c{}".format(k)))
    df = df.with_columns(exprs).select(out_fields(n_flat, chain))
    if shape["sink"] == "xlsx":
        df.write_excel(out_path, include_header=True)
    else:
        df.write_csv(out_path, include_header=True)
    elapsed = time.perf_counter() - t0
    return {"total_s": elapsed}


# ---------------------------------------------------------------------------
# scalim 侧（DemandIr + Engine，call_by 每行每字段一次，流式分批）
# ---------------------------------------------------------------------------
def _build_scalim_demand(shape: Dict[str, Any]):
    from scalim.planning import PlanBuilder
    from scalim.spec.ir import (
        BindingIr,
        CallBySpecIr,
        CallByValueIr,
        DemandIr,
        DerivedFieldIr,
        FieldIr,
        LoaderIr,
        KeyIr,
        MainSourceIr,
        RuntimeHandleIdIr,
        SourceIr,
    )

    n_flat, chain = shape["n_flat"], shape["chain_depth"]
    main = MainSourceIr(source_id="main", loader_ref=RuntimeHandleIdIr(handle_id="main.loader"))
    field_irs: List[Any] = [
        FieldIr(field_id="id", name="id", source_id=main.source_id),
        FieldIr(field_id="v0", name="v0", source_id=main.source_id),
        FieldIr(field_id="v1", name="v1", source_id=main.source_id),
    ]
    calculators: Dict[str, Any] = {}
    sources: List[Any] = []

    if shape.get("relation"):
        side_loader_ir = LoaderIr(
            callable_ref=RuntimeHandleIdIr(handle_id="side.loader"),
            bindings={
                # 绑定身份按侧源 key 字段（"sid"），与 join 的 to_field 对齐（参照 test_execution relation_model）
                "sid": BindingIr(
                    key_field="sid",
                    params_builder_ref=RuntimeHandleIdIr(handle_id="side.sid.params_builder"),
                ),
            },
        )
        side_source = SourceIr(source_id="side", key=KeyIr(key="sid"), loader_spec=side_loader_ir)
        sources = [side_source]
        # lookup 的 from_field 必须是已声明字段，否则上下文无值、关联全部 miss
        field_irs.append(FieldIr(field_id="fk", name="fk", source_id=main.source_id))
        relation = main["fk"].join(side_source["sid"])
        for fid, dkey in (("s0", "s0"), ("s1", "s1")):
            field_irs.append(
                FieldIr(field_id=fid, name=fid, source_id=side_source.source_id, data_key=dkey, relation=relation)
            )

        def _side_loader(sids=None):  # type: ignore[no-unused-argument]
            import sqlite3

            keys = sorted(set(sids or ()))
            if not keys:
                return {}
            conn = sqlite3.connect(shape["_db_path"])
            try:
                q = "SELECT sid, s0, s1 FROM side_rows WHERE sid IN ({})".format(",".join("?" * len(keys)))
                return {r[0]: {"sid": r[0], "s0": r[1], "s1": r[2]} for r in conn.execute(q, keys)}
            finally:
                conn.close()

        def _side_params(ctx):  # type: ignore[no-untyped-def]
            return (), {"sids": set(ctx.lookup_keys or set())}

        shape["_runtime_extras"] = {
            "source_loaders": {"side": _side_loader},
            "params_builders": {("side", "sid"): _side_params},
        }
        calculators["r0"] = lambda v0, s0: float(v0) + float(s0)  # noqa: E731
        calculators["r1"] = lambda v1, s1: float(v1) * float(s1)  # noqa: E731
        field_irs.append(
            DerivedFieldIr(
                field_id="r0", name="r0", dependencies=("v0", "s0"),
                call_by=CallBySpecIr(
                    reference=RuntimeHandleIdIr(handle_id="r0.calc"),
                    kwargs=(
                        ("v0", CallByValueIr(kind="field", value="v0")),
                        ("s0", CallByValueIr(kind="field", value="s0")),
                    ),
                    field_names=("v0", "s0"),
                ),
            )
        )
        field_irs.append(
            DerivedFieldIr(
                field_id="r1", name="r1", dependencies=("v1", "s1"),
                call_by=CallBySpecIr(
                    reference=RuntimeHandleIdIr(handle_id="r1.calc"),
                    kwargs=(
                        ("v1", CallByValueIr(kind="field", value="v1")),
                        ("s1", CallByValueIr(kind="field", value="s1")),
                    ),
                    field_names=("v1", "s1"),
                ),
            )
        )

    for j in range(n_flat):
        fid = "d{}".format(j)

        def _make_flat(calc_id: str):
            jj = int(calc_id[1:])

            def _calc(v0: float, v1: float) -> float:
                return expected_flat(v0, v1, jj)

            return _calc

        calculators[fid] = _make_flat(fid)
        field_irs.append(
            DerivedFieldIr(
                field_id=fid,
                name=fid,
                dependencies=("v0", "v1"),
                call_by=CallBySpecIr(
                    reference=RuntimeHandleIdIr(handle_id=fid + ".calc"),
                    kwargs=(
                        ("v0", CallByValueIr(kind="field", value="v0")),
                        ("v1", CallByValueIr(kind="field", value="v1")),
                    ),
                    field_names=("v0", "v1"),
                ),
            )
        )

    for k in range(chain):
        cid = "c{}".format(k)
        dep = "v0" if k == 0 else "c{}".format(k - 1)

        def _make_chain(kk: int):
            def _calc(prev: float) -> float:
                return prev + 1.0

            return _calc

        if k == 0:
            calculators[cid] = lambda v0: expected_chain(v0, 0)  # noqa: E731
        else:
            calculators[cid] = _make_chain(k)
        field_irs.append(
            DerivedFieldIr(
                field_id=cid,
                name=cid,
                dependencies=(dep,),
                call_by=CallBySpecIr(
                    reference=RuntimeHandleIdIr(handle_id=cid + ".calc"),
                    kwargs=((dep, CallByValueIr(kind="field", value=dep)),),
                    field_names=(dep,),
                ),
            )
        )

    demand = DemandIr.from_irs(sources=sources, fields=tuple(field_irs), main_source=main, name="compare_matrix")
    plan = PlanBuilder(demand).build()
    return demand, plan, calculators


def run_scalim(shape: Dict[str, Any], out_path: str, mode: str) -> Dict[str, Any]:
    from scalim.execution.engine import ScalimEngine
    from scalim.execution.runtime_bindings import RuntimeBindings
    from scalim.sinks import CSVSink, ExcelSink

    rows = shape["_rows"]
    n_flat, chain = shape["n_flat"], shape["chain_depth"]
    batch_size = shape["batch_size"]

    demand, plan, calculators = _build_scalim_demand(shape)
    extras = shape.get("_runtime_extras", {})
    fields = out_fields(n_flat, chain, relation=bool(shape.get("relation")))
    sink: Any
    if shape["sink"] == "xlsx":
        sink = ExcelSink(out_path, field_names=fields, header_names=fields, sheet_name="Sheet1")
    else:
        sink = CSVSink(out_path, field_names=fields, header_names=fields)

    t0 = time.perf_counter()
    engine = ScalimEngine(
        demand=demand,
        plan=plan,
        runtime_bindings=RuntimeBindings(
            main_source_loaders={},
            derived_calculators=calculators,
            **extras,
        ),
        batch_size=batch_size,
        parallel_mode=mode,
        max_workers=4,
    )

    if shape.get("relation"):
        import sqlite3

        def _iter_main():
            conn = sqlite3.connect(shape["_db_path"])
            try:
                cur = conn.execute("SELECT id, fk, v0, v1 FROM main_rows")
                for r in cur:
                    yield {"id": r[0], "fk": r[1], "v0": r[2], "v1": r[3]}
            finally:
                conn.close()

        engine.run(main_rows=_iter_main(), sink=sink)
    else:
        # 注意: pipeline.run() 结束时会 sink.close()，因此整表必须单次 run()，
        # 以生成器惰性供给、由引擎按 batch_size 流式分批（与对照侧消费同一生成器对齐）。
        engine.run(main_rows=(source_row(i) for i in range(rows)), sink=sink)
    elapsed = time.perf_counter() - t0
    return {"total_s": elapsed}


# ---------------------------------------------------------------------------
# golden：从写出文件流式读回校验（csv 模块 / openpyxl read_only）
# full=True：全表行数 + 派生列校验和；full=False：行数 + 首 50 行精确校验
# （确定性生成器 + 同一代码路径，run0 全量锚定即可覆盖其余 run）
# ---------------------------------------------------------------------------
def verify(shape: Dict[str, Any], out_path: str, full: bool = True) -> Dict[str, Any]:
    import csv

    rows = shape["_rows"]
    n_flat, chain = shape["n_flat"], shape["chain_depth"]
    relation = bool(shape.get("relation"))
    targets = checksum_targets(n_flat, chain, relation)
    sums = {name: 0.0 for name in targets}
    count = 0
    if shape["sink"] == "csv":
        with open(out_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                count += 1
                if full or count <= SAMPLE_ROWS:
                    for name in targets:
                        sums[name] += float(row[name])
    else:
        from openpyxl import load_workbook

        wb = load_workbook(out_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(c) for c in next(it)]
        col_idx = {name: header.index(name) for name in targets}
        for values in it:
            count += 1
            if full or count <= SAMPLE_ROWS:
                for name, idx in col_idx.items():
                    v = values[idx]
                    sums[name] += float(v) if v is not None else 0.0
        wb.close()
    expected = expected_checksums(rows, n_flat, chain, relation) if full else expected_checksums(
        rows, n_flat, chain, relation, only_first_n=SAMPLE_ROWS)
    ok = count == rows and all(checksum_close(sums[k], expected[k]) for k in targets)
    return {"golden_ok": ok, "rows_read": count, "checksums": sums, "expected": expected, "full": full}


# ---------------------------------------------------------------------------
# 子进程 worker：隔离 RSS；结束后读 VmHWM 作为真实峰值
# ---------------------------------------------------------------------------
def _rss_now_kb() -> int:
    try:
        with open("/proc/self/statm") as f:
            parts = f.read().split()
        return int(int(parts[1]) * os.sysconf("SC_PAGE_SIZE") / 1024)
    except Exception:
        return 0


def _hwm_kb() -> int:
    try:
        with open("/proc/self/status") as f:
            for line in f:
                if line.startswith("VmHWM:"):
                    return int(line.split()[1])
    except Exception:
        pass
    return 0


def worker(task_json: str) -> int:
    task = json.loads(task_json)
    shape, side, run_i, out_path = task["shape"], task["side"], task["run"], task["out_path"]
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    before = _rss_now_kb()
    try:
        if side == "pandas":
            body = run_pandas(shape, out_path)
        elif side == "polars":
            body = run_polars(shape, out_path)
        else:
            body = run_scalim(shape, out_path, task.get("mode", "seq"))
    except ImportError as exc:
        result = {
            "shape": shape["id"], "side": side, "mode": task.get("mode", "seq"), "run": run_i,
            "rows": shape["_rows"], "skipped": "import: {}".format(exc), "golden_ok": None,
            "rss_begin_kb": before, "rss_hwm_kb": 0, "rss_end_kb": 0, "total_s": 0.0,
            "out_path": "",
        }
        print(json.dumps(result))
        return 0
    golden = verify(shape, out_path, full=(run_i == 0))
    result = {
        "shape": shape["id"],
        "side": side,
        "mode": task.get("mode", "-"),
        "run": run_i,
        "rows": shape["_rows"],
        "total_s": body["total_s"],
        "rss_begin_kb": before,
        "rss_hwm_kb": _hwm_kb(),
        "rss_end_kb": _rss_now_kb(),
        "golden_ok": golden["golden_ok"],
        "rows_read": golden["rows_read"],
        "out_path": out_path,
    }
    os.remove(out_path)  # 证据只留数字不留产物
    print(json.dumps(result))
    return 0


# ---------------------------------------------------------------------------
# 主进程：编排 runs（短 shape 默认 5），汇总 median/min/max
# ---------------------------------------------------------------------------
def median(xs: List[float]) -> float:
    s = sorted(xs)
    n = len(s)
    return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2.0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--runs", type=int, default=3)
    ap.add_argument("--runs-short", type=int, default=5, help="短耗时 shape 的 runs（降波动）")
    ap.add_argument("--rows-scale", type=float, default=1.0)
    ap.add_argument("--shapes", type=str, default="")
    ap.add_argument("--sides", type=str, default="pandas,polars,scalim")
    ap.add_argument("--outdir", type=str, default="")
    args = ap.parse_args()
    smoke = args.rows_scale < 1.0
    assert args.runs >= 3 or smoke, "正式证据要求 --runs >= 3（冒烟除外）"

    shapes = [dict(s) for s in SHAPES if not args.shapes or s["id"] in args.shapes.split(",")]
    for s in shapes:
        s["_rows"] = scaled_rows(s["rows"], args.rows_scale)
        s["_scale"] = args.rows_scale

    sides = [x.strip() for x in args.sides.split(",") if x.strip()]
    ts = time.strftime("%Y%m%d-%H%M%S")
    outdir = args.outdir or os.path.join(REPO_ROOT, ".tmp", "evidence", "external-baseline", ts)
    os.makedirs(outdir, exist_ok=True)
    py = sys.executable
    this = os.path.abspath(__file__)

    def _pkg_version(name: str) -> str:
        try:
            mod = __import__(name)
            return str(getattr(mod, "__version__", "unknown"))
        except Exception:
            return "not-installed"

    versions = {
        "python": sys.version.split()[0],
        "scalim": _pkg_version("scalim") or "local-src",
        "pandas": _pkg_version("pandas"),
        "polars": _pkg_version("polars"),
        "xlsxwriter": _pkg_version("xlsxwriter"),
        "openpyxl": _pkg_version("openpyxl"),
        "numpy": _pkg_version("numpy"),
    }

    def _meminfo_kb() -> Dict[str, int]:
        info: Dict[str, int] = {}
        with open("/proc/meminfo") as f:
            for line in f:
                key = line.split(":", 1)[0]
                if key in ("MemTotal", "MemAvailable"):
                    info[key] = int(line.split()[1])
        return info

    mem = _meminfo_kb()
    mem_budget_mib = max(0, mem["MemAvailable"] - int(mem["MemTotal"] * 0.10)) / 1024.0

    results: List[Dict[str, Any]] = []
    planned = 0
    for shape in shapes:
        run_n = args.runs_short if (shape.get("short") and not smoke) else args.runs
        configs: List[tuple] = []
        for side in sides:
            if side == "scalim":
                configs += [("scalim", "seq"), ("scalim", "adaptive")]
            else:
                configs += [(side, "-")]
        planned += len(configs) * run_n

    for shape in shapes:
        run_n = args.runs_short if (shape.get("short") and not smoke) else args.runs
        if shape.get("relation"):
            db = os.path.join(outdir, "fixture_{}.sqlite".format(shape["id"]))
            ensure_sqlite_fixture(db, shape["_rows"], RELATION_SIDE_ROWS)
            shape["_db_path"] = db
        configs: List[tuple] = []
        for side in sides:
            if side == "scalim":
                configs += [("scalim", "seq"), ("scalim", "adaptive")]
            else:
                configs += [(side, "-")]
        ext = "xlsx" if shape["sink"] == "xlsx" else "csv"
        for side, mode in configs:
            for run_i in range(run_n):
                task = {
                    "shape": shape,
                    "side": side,
                    "mode": mode,
                    "run": run_i,
                    "out_path": os.path.join(outdir, "_tmp_{}_{}_{}_{}.{}".format(shape["id"], side, mode, run_i, ext)),
                }
                t0 = time.perf_counter()
                proc = subprocess.run(
                    [py, this, "--worker", json.dumps(task)],
                    capture_output=True,
                    text=True,
                    cwd=REPO_ROOT,
                )
                wall = time.perf_counter() - t0
                if proc.returncode != 0:
                    print(proc.stderr[-4000:], file=sys.stderr)
                    print("FAILED: {} {} {} run{}".format(shape["id"], side, mode, run_i), file=sys.stderr)
                    return 1
                res = json.loads(proc.stdout.strip().splitlines()[-1])
                res["parent_wall_s"] = wall
                results.append(res)
                tag = "{}:{}{}".format(side, mode, "/skip" if res.get("skipped") else "")
                print(
                    "[{}/{}] {} {:>18} run{}: {:8.3f}s  HWM={:8.1f}MiB  golden={}".format(
                        len(results), planned, shape["id"], tag, run_i,
                        res["total_s"], res["rss_hwm_kb"] / 1024.0, res.get("golden_ok"),
                    ),
                    flush=True,
                )

    summary: List[Dict[str, Any]] = []
    for shape in shapes:
        run_n = args.runs_short if (shape.get("short") and not smoke) else args.runs
        configs: List[tuple] = []
        for side in sides:
            if side == "scalim":
                configs += [("scalim", "seq"), ("scalim", "adaptive")]
            else:
                configs += [(side, "-")]
        for side, mode in configs:
            rs = [r for r in results if r["shape"] == shape["id"] and r["side"] == side and r.get("mode") == mode]
            valid = [r for r in rs if not r.get("skipped")]
            entry: Dict[str, Any] = {
                "shape": shape["id"],
                "label": shape["label"],
                "side": side,
                "mode": mode,
                "runs": len(valid),
                "rows": shape["_rows"],
                "sink": shape["sink"],
                "n_flat": shape["n_flat"],
                "chain_depth": shape["chain_depth"],
                "relation": bool(shape.get("relation")),
                "golden_all_ok": bool(valid) and all(r["golden_ok"] for r in valid),
                "skipped": bool(rs) and not valid,
            }
            if valid:
                times = [r["total_s"] for r in valid]
                hwms = [r["rss_hwm_kb"] for r in valid]
                entry.update({
                    "total_s_median": median(times),
                    "total_s_min": min(times),
                    "total_s_max": max(times),
                    "spread_pct": (max(times) - min(times)) / median(times) * 100.0 if median(times) else 0.0,
                    "rss_hwm_mib_median": median(hwms) / 1024.0,
                })
            summary.append(entry)

    # 相对基线：每 shape 以 pandas(-) 为 baseline
    for shape in shapes:
        base = next((x for x in summary if x["shape"] == shape["id"] and x["side"] == "pandas" and not x.get("skipped")), None)
        if not base or "total_s_median" not in base:
            continue
        for x in summary:
            if x["shape"] == shape["id"] and x is not base and "total_s_median" in x:
                x["time_ratio_vs_pandas"] = x["total_s_median"] / base["total_s_median"]
                x["rss_ratio_vs_pandas"] = x["rss_hwm_mib_median"] / base["rss_hwm_mib_median"]

    out = {
        "meta": {
            "timestamp": ts,
            "runs": args.runs,
            "runs_short": args.runs_short,
            "rows_scale": args.rows_scale,
            "versions": versions,
            "mem_total_mib": mem["MemTotal"] / 1024.0,
            "mem_available_mib": mem["MemAvailable"] / 1024.0,
            "mem_budget_mib": mem_budget_mib,
            "host_note": "本机单次调研；跨机器/负载下数字不可直接迁移，仅作量级与取舍叙事依据",
            "method": "subprocess-isolated wall clock + VmHWM peak RSS; deterministic synthetic source; "
                      "S7=本机 SQLite 真实 IO(确定性 fixture); golden=run0 全量读回 + 其余 run 行数/首50行; "
                      "scalim seq+adaptive(W=4); polars 自身多线程; 薄算术 calculator",
        },
        "summary": summary,
        "runs": results,
    }
    result_path = os.path.join(outdir, "result.json")
    with open(result_path, "w") as f:
        json.dump(out, f, indent=2)

    print("\n=== summary (median; short shapes {} runs, others {}) ===".format(args.runs_short, args.runs))
    print("{:<26}{:>8}{:>10}{:>10}{:>10}{:>11}{:>11}{:>9}".format(
        "shape", "side", "mode", "rows", "time_s", "HWM_MiB", "t_vs_pd", "golden"))
    for x in summary:
        t = x.get("total_s_median")
        h = x.get("rss_hwm_mib_median")
        tr = x.get("time_ratio_vs_pandas")
        print("{:<26}{:>8}{:>10}{:>10}{:>10}{:>11}{:>11}{:>9}".format(
            x["shape"], x["side"], x["mode"], x["rows"],
            "{:.3f}".format(t) if t is not None else "skip",
            "{:.1f}".format(h) if h is not None else "-",
            "{:.2f}x".format(tr) if tr is not None else ("" if x.get("skipped") else "base"),
            "ok" if x["golden_all_ok"] else ("skip" if x.get("skipped") else "FAIL")))
    print("\nresult.json -> {}".format(result_path))
    over = [r for r in results if r.get("rss_hwm_kb") and r["rss_hwm_kb"] / 1024.0 > mem_budget_mib]
    if over:
        print("WARNING: {} 个 run 峰值超过内存预算 {:.0f} MiB（保留系统 10% 余量策略）".format(len(over), mem_budget_mib))
    else:
        print("内存预算检查: 全部 run 峰值 <= 预算 {:.0f} MiB（MemAvailable - 10% MemTotal）".format(mem_budget_mib))
    return 0


if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "--worker":
        sys.exit(worker(sys.argv[2]))
    sys.exit(main())
