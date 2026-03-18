import csv
import io
import os
import threading
from contextlib import suppress
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, FrozenSet, Iterator, List, Optional, Sequence, Tuple, Type, cast

from ....events.catalog import (
    EVENT_DIAGNOSTIC_WARNING,
    EVENT_WORKFLOW_RESOURCE_COMMIT,
    EVENT_WORKFLOW_RESOURCE_CREATE,
    EVENT_WORKFLOW_RESOURCE_DISCARD,
    EVENT_WORKFLOW_RESOURCE_WRITE,
)
from ....events.events import (
    DiagnosticWarningEvent,
    WorkflowResourceCommitEvent,
    WorkflowResourceCreateEvent,
    WorkflowResourceDiscardEvent,
    WorkflowResourceWriteEvent,
)
from ....sinks.sink_base import create_temp_path
from ....vendor.compact.importlibx import require_optional_dependency

if TYPE_CHECKING:
    from openpyxl import Workbook


_WRITE_LOCK_SUFFIX = ".scalim.lock"


def _get_openpyxl_workbook_class() -> "Type[Workbook]":
    openpyxl_mod = require_optional_dependency("openpyxl", context="scalim.dsl.by_yaml.runtime.workflow_resources")
    return cast("Any", openpyxl_mod).Workbook


class WorkflowWriteError(RuntimeError):
    diff: Optional[List[str]]

    def __init__(self, message: str, *, diff: Optional[List[str]] = None) -> None:
        super(WorkflowWriteError, self).__init__(message)
        self.diff = list(diff) if diff is not None else None


def _acquire_write_lock(output_path: str) -> Path:
    lock_path = Path(str(output_path) + _WRITE_LOCK_SUFFIX)
    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError:
        msg = "Output path is locked (possible concurrent writers): output_path={!r}, lock_path={!r}".format(
            str(output_path),
            str(lock_path),
        )
        raise WorkflowWriteError(msg) from None
    with os.fdopen(fd, "w") as f:
        _ = f.write("pid={}\n".format(os.getpid()))
    return lock_path


def _release_write_lock(lock_path: Path) -> None:
    try:
        lock_path.unlink()
    except FileNotFoundError:
        return
    except OSError:
        return


def _best_effort_close_write_only_workbook_worksheets(workbook: Any) -> None:
    worksheets = getattr(workbook, "worksheets", None)
    if worksheets is None:
        return
    try:
        worksheets = list(worksheets)
    except TypeError:
        return
    for ws in worksheets:
        try:
            is_closed = bool(ws.closed)
        except AttributeError:
            continue
        if is_closed:
            continue
        with suppress(Exception):
            ws.close()


def _read_csv_header(path: str) -> List[str]:
    p = Path(str(path))
    if not p.exists():
        msg = "Missing input CSV: {!r}".format(str(path))
        raise WorkflowWriteError(msg)
    with p.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        try:
            header = next(reader)
        except StopIteration:
            msg = "Input CSV is empty (missing header): {!r}".format(str(path))
            raise WorkflowWriteError(msg) from None
    header = [str(x or "").strip() for x in header]
    if not header or any(not x for x in header):
        msg = "Input CSV has invalid header (empty field): {!r}".format(str(path))
        raise WorkflowWriteError(msg)
    return header


def _iter_csv_rows(path: str) -> Iterator[List[str]]:
    p = Path(str(path))
    with p.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        _header = next(reader, None)
        for row in reader:
            yield [str(v) for v in row]


def _describe_header_diff(expected: Sequence[str], actual: Sequence[str]) -> List[str]:
    exp_set = {str(x) for x in expected}
    act_set = {str(x) for x in actual}
    missing = sorted(exp_set.difference(act_set))
    extra = sorted(act_set.difference(exp_set))
    return [
        "expected={}".format(",".join(str(x) for x in expected)),
        "actual={}".format(",".join(str(x) for x in actual)),
        "missing={}".format(",".join(missing) if missing else "(none)"),
        "extra={}".format(",".join(extra) if extra else "(none)"),
    ]


def _build_alignment_mapping(expected: Sequence[str], actual: Sequence[str]) -> List[int]:
    index_by_key: Dict[str, int] = {}
    for idx, key in enumerate(actual):
        k = str(key)
        if k not in index_by_key:
            index_by_key[k] = int(idx)
    mapping: List[int] = []
    for key in expected:
        mapping.append(int(index_by_key.get(str(key), -1)))
    return mapping


@dataclass
class _AppendSegment:
    input_csv_path: str
    header_policy: str
    mapping: List[int]
    on_mismatch: str
    align_by: str
    input_header: List[str]


@dataclass
class _SheetPlan:
    sheet: str
    baseline_header: List[str]
    segments: List[_AppendSegment]


@dataclass
class _WorkbookPlan:
    resource_id: str
    path: str
    lock_path: Optional[Path]
    sheet_order: List[str]
    sheets: Dict[str, _SheetPlan]
    last_workflow_node_id: Optional[str] = None


@dataclass
class _CsvPlan:
    resource_id: str
    path: str
    lock_path: Optional[Path]
    baseline_header: Optional[List[str]] = None
    segments: Optional[List[_AppendSegment]] = None
    last_workflow_node_id: Optional[str] = None


@dataclass(frozen=True)
class SheetBookDef:
    resource_id: str
    budget_max_sheets: int
    budget_max_total_cells: int
    export_path: Optional[str]
    export_write_lock: bool


@dataclass
class _SheetBookSegment:
    producer_node_id: str
    start_row: int
    end_row: int
    header_policy: str


@dataclass
class _SheetBookSheetPlan:
    sheet: str
    baseline_header: List[str]
    columns: Dict[str, List[str]]
    segments: List[_SheetBookSegment]
    row_count: int = 0


@dataclass
class _SheetBookPlan:
    resource_id: str
    budget_max_sheets: int
    budget_max_total_cells: int
    export_path: Optional[str]
    export_write_lock: bool
    sheet_order: List[str]
    sheets: Dict[str, _SheetBookSheetPlan]
    lock_path: Optional[Path] = None
    total_cells: int = 0
    last_workflow_node_id: Optional[str] = None


class WorkflowResourceManager:
    """工作流级共享输出资源管理器(延迟提交 + 原子落盘)."""

    _workflow_exec_id: str
    _instrumentation: Any
    _workbook_defs: Dict[str, str]
    _csv_defs: Dict[str, str]
    _sheetbook_defs: Dict[str, SheetBookDef]
    _workbooks: Dict[str, _WorkbookPlan]
    _csvs: Dict[str, _CsvPlan]
    _sheetbooks: Dict[str, _SheetBookPlan]
    _lock: threading.Lock

    def __init__(
        self,
        *,
        workflow_exec_id: str,
        instrumentation: Any,
        workbook_defs: Dict[str, str],
        csv_defs: Dict[str, str],
        sheetbook_defs: Dict[str, SheetBookDef],
    ) -> None:
        self._workflow_exec_id = str(workflow_exec_id)
        self._instrumentation = instrumentation
        self._workbook_defs = dict(workbook_defs)
        self._csv_defs = dict(csv_defs)
        self._sheetbook_defs = dict(sheetbook_defs)
        self._workbooks = {}
        self._csvs = {}
        self._sheetbooks = {}
        self._lock = threading.Lock()

    def _emit_resource_create(self, *, workflow_node_id: str, resource_type: str, resource_id: str, path: str) -> None:
        _ = self._instrumentation.emit(
            EVENT_WORKFLOW_RESOURCE_CREATE,
            WorkflowResourceCreateEvent(
                workflow_exec_id=self._workflow_exec_id,
                workflow_node_id=str(workflow_node_id),
                resource_type=str(resource_type),
                resource_id=str(resource_id),
                path=str(path),
            ),
            meta={
                "workflow_exec_id": self._workflow_exec_id,
                "workflow_node_id": str(workflow_node_id),
            },
        )

    def _emit_resource_write(
        self,
        *,
        workflow_node_id: str,
        resource_type: str,
        resource_id: str,
        path: str,
        write_kind: str,
        action: str,
        input_node_id: Optional[str] = None,
        input_output_id: Optional[str] = None,
        sheet: Optional[str] = None,
    ) -> None:
        _ = self._instrumentation.emit(
            EVENT_WORKFLOW_RESOURCE_WRITE,
            WorkflowResourceWriteEvent(
                workflow_exec_id=self._workflow_exec_id,
                workflow_node_id=str(workflow_node_id),
                resource_type=str(resource_type),
                resource_id=str(resource_id),
                path=str(path),
                write_kind=str(write_kind),
                action=str(action),
                input_node_id=str(input_node_id) if input_node_id is not None else None,
                input_output_id=str(input_output_id) if input_output_id is not None else None,
                sheet=str(sheet) if sheet is not None else None,
            ),
            meta={
                "workflow_exec_id": self._workflow_exec_id,
                "workflow_node_id": str(workflow_node_id),
            },
        )

    def _emit_resource_commit(self, *, workflow_node_id: str, resource_type: str, resource_id: str, path: str) -> None:
        _ = self._instrumentation.emit(
            EVENT_WORKFLOW_RESOURCE_COMMIT,
            WorkflowResourceCommitEvent(
                workflow_exec_id=self._workflow_exec_id,
                workflow_node_id=str(workflow_node_id),
                resource_type=str(resource_type),
                resource_id=str(resource_id),
                path=str(path),
            ),
            meta={
                "workflow_exec_id": self._workflow_exec_id,
                "workflow_node_id": str(workflow_node_id),
            },
        )

    def _emit_resource_discard(self, *, workflow_node_id: str, resource_type: str, resource_id: str, path: str, reason: str) -> None:
        _ = self._instrumentation.emit(
            EVENT_WORKFLOW_RESOURCE_DISCARD,
            WorkflowResourceDiscardEvent(
                workflow_exec_id=self._workflow_exec_id,
                workflow_node_id=str(workflow_node_id),
                resource_type=str(resource_type),
                resource_id=str(resource_id),
                path=str(path),
                reason=str(reason),
            ),
            meta={
                "workflow_exec_id": self._workflow_exec_id,
                "workflow_node_id": str(workflow_node_id),
            },
        )

    def _get_or_create_workbook(self, workbook_id: str, *, workflow_node_id: str) -> _WorkbookPlan:
        key = str(workbook_id)
        with self._lock:
            existing = self._workbooks.get(key)
            if existing is not None:
                return existing

        raw_path = self._workbook_defs.get(key)
        if raw_path is None:
            msg = "Unknown workbook resource id: {!r}".format(key)
            raise WorkflowWriteError(msg)
        lock_path = _acquire_write_lock(raw_path)
        plan = _WorkbookPlan(
            resource_id=key,
            path=str(raw_path),
            lock_path=lock_path,
            sheet_order=[],
            sheets={},
        )
        with self._lock:
            self._workbooks[key] = plan
        self._emit_resource_create(workflow_node_id=str(workflow_node_id), resource_type="workbook", resource_id=key, path=str(raw_path))
        return plan

    def _get_or_create_csv(self, csv_id: str, *, workflow_node_id: str) -> _CsvPlan:
        key = str(csv_id)
        with self._lock:
            existing = self._csvs.get(key)
            if existing is not None:
                return existing

        raw_path = self._csv_defs.get(key)
        if raw_path is None:
            msg = "Unknown csv resource id: {!r}".format(key)
            raise WorkflowWriteError(msg)
        lock_path = _acquire_write_lock(raw_path)
        plan = _CsvPlan(resource_id=key, path=str(raw_path), lock_path=lock_path)
        with self._lock:
            self._csvs[key] = plan
        self._emit_resource_create(workflow_node_id=str(workflow_node_id), resource_type="csv", resource_id=key, path=str(raw_path))
        return plan

    def _get_or_create_sheetbook(self, sheetbook_id: str, *, workflow_node_id: str) -> _SheetBookPlan:
        key = str(sheetbook_id)
        with self._lock:
            existing = self._sheetbooks.get(key)
            if existing is not None:
                return existing

        raw_def = self._sheetbook_defs.get(key)
        if raw_def is None:
            msg = "Unknown sheetbook resource id: {!r}".format(key)
            raise WorkflowWriteError(msg)

        plan = _SheetBookPlan(
            resource_id=str(raw_def.resource_id),
            budget_max_sheets=int(raw_def.budget_max_sheets),
            budget_max_total_cells=int(raw_def.budget_max_total_cells),
            export_path=str(raw_def.export_path) if raw_def.export_path is not None else None,
            export_write_lock=bool(raw_def.export_write_lock),
            sheet_order=[],
            sheets={},
        )
        with self._lock:
            self._sheetbooks[key] = plan

        display_path = plan.export_path if plan.export_path is not None else "<memory>"
        self._emit_resource_create(
            workflow_node_id=str(workflow_node_id),
            resource_type="sheetbook",
            resource_id=str(key),
            path=str(display_path),
        )
        return plan

    def _check_sheetbook_budget(
        self,
        plan: _SheetBookPlan,
        *,
        new_total_cells: int,
        allow_over_budget: bool = False,
    ) -> None:
        if allow_over_budget:  # pragma: no cover
            return  # pragma: no cover
        limit = int(plan.budget_max_total_cells)
        if int(new_total_cells) <= limit:
            return
        diff = [
            "budget.max_total_cells={}".format(limit),
            "new_total_cells={}".format(int(new_total_cells)),
        ]
        msg = "Sheetbook budget exceeded: sheetbook={!r}".format(str(plan.resource_id))
        raise WorkflowWriteError(msg, diff=diff)

    def apply_sheetbook_sheet(
        self,
        *,
        workflow_node_id: str,
        sheetbook_id: str,
        sheet: str,
        input_node_id: str,
        input_output_id: str,
        input_csv_path: str,
        on_conflict: str,
    ) -> None:
        plan = self._get_or_create_sheetbook(sheetbook_id, workflow_node_id=str(workflow_node_id))
        sheet_name = str(sheet)
        action = "write"

        input_header = _read_csv_header(input_csv_path)

        with self._lock:
            existing = plan.sheets.get(sheet_name)
            if existing is not None:
                if on_conflict == "skip":
                    action = "skip"
                    display_path = plan.export_path if plan.export_path is not None else "<memory>"
                    self._emit_resource_write(
                        workflow_node_id=str(workflow_node_id),
                        resource_type="sheetbook",
                        resource_id=str(sheetbook_id),
                        path=str(display_path),
                        write_kind="sheetbook_sheet",
                        action=action,
                        input_node_id=str(input_node_id),
                        input_output_id=str(input_output_id),
                        sheet=sheet_name,
                    )
                    plan.last_workflow_node_id = str(workflow_node_id)
                    return
                if on_conflict == "error":
                    msg = "Sheet conflict (sheetbook_sheet): sheetbook={!r}, sheet={!r}".format(str(sheetbook_id), sheet_name)
                    raise WorkflowWriteError(msg, diff=["on_conflict=error", "existing_sheet=present"])
                if on_conflict == "overwrite":
                    action = "overwrite"
            elif len(plan.sheets) >= int(plan.budget_max_sheets):
                msg = "Sheetbook budget exceeded: max_sheets (sheetbook={!r})".format(str(sheetbook_id))
                diff = [
                    "budget.max_sheets={}".format(int(plan.budget_max_sheets)),
                    "current_sheets={}".format(len(plan.sheets)),
                    "new_sheet={!r}".format(sheet_name),
                ]
                raise WorkflowWriteError(msg, diff=diff)

        expected = list(input_header)
        mapping = _build_alignment_mapping(expected, input_header)

        # 读取并物化到内存(列式),用于下游读取与导出.
        col_lists: List[List[str]] = [[] for _ in expected]
        for row in _iter_csv_rows(input_csv_path):
            for col_idx, src_idx in enumerate(mapping):
                col_lists[col_idx].append(row[src_idx] if src_idx >= 0 and src_idx < len(row) else "")

        columns: Dict[str, List[str]] = {str(key): col_lists[idx] for idx, key in enumerate(expected)}
        row_count = len(col_lists[0]) if col_lists else 0
        new_sheet_cells = int(row_count) * len(expected)

        with self._lock:
            existing = plan.sheets.get(sheet_name)
            old_cells = 0
            if existing is not None:
                old_cells = int(existing.row_count) * len(existing.baseline_header)

            new_total = int(plan.total_cells) - int(old_cells) + int(new_sheet_cells)
            self._check_sheetbook_budget(plan, new_total_cells=new_total)

            if existing is None:
                plan.sheet_order.append(sheet_name)

            sheet_plan = _SheetBookSheetPlan(
                sheet=sheet_name,
                baseline_header=list(expected),
                columns=columns,
                segments=[],
                row_count=int(row_count),
            )
            sheet_plan.segments.append(
                _SheetBookSegment(
                    producer_node_id=str(input_node_id),
                    start_row=0,
                    end_row=int(row_count),
                    header_policy="once",
                )
            )
            plan.sheets[sheet_name] = sheet_plan
            plan.total_cells = int(new_total)

        display_path = plan.export_path if plan.export_path is not None else "<memory>"
        self._emit_resource_write(
            workflow_node_id=str(workflow_node_id),
            resource_type="sheetbook",
            resource_id=str(sheetbook_id),
            path=str(display_path),
            write_kind="sheetbook_sheet",
            action=action,
            input_node_id=str(input_node_id),
            input_output_id=str(input_output_id),
            sheet=sheet_name,
        )
        plan.last_workflow_node_id = str(workflow_node_id)

    def apply_sheetbook_append(  # noqa: C901, PLR0912, PLR0915
        self,
        *,
        workflow_node_id: str,
        sheetbook_id: str,
        sheet: str,
        input_node_id: str,
        input_output_id: str,
        input_csv_path: str,
        align_by: str,
        header_policy: str,
        on_mismatch: str,
    ) -> None:
        plan = self._get_or_create_sheetbook(sheetbook_id, workflow_node_id=str(workflow_node_id))
        sheet_name = str(sheet)
        input_header = _read_csv_header(input_csv_path)

        with self._lock:
            sheet_plan = plan.sheets.get(sheet_name)
            if sheet_plan is None:
                if len(plan.sheets) >= int(plan.budget_max_sheets):
                    msg = "Sheetbook budget exceeded: max_sheets (sheetbook={!r})".format(str(sheetbook_id))
                    diff = [
                        "budget.max_sheets={}".format(int(plan.budget_max_sheets)),
                        "current_sheets={}".format(len(plan.sheets)),
                        "new_sheet={!r}".format(sheet_name),
                    ]
                    raise WorkflowWriteError(msg, diff=diff)
                plan.sheet_order.append(sheet_name)
                sheet_plan = _SheetBookSheetPlan(sheet=sheet_name, baseline_header=list(input_header), columns={}, segments=[], row_count=0)
                plan.sheets[sheet_name] = sheet_plan

            expected = list(sheet_plan.baseline_header)
            mapping = _build_alignment_mapping(expected, input_header)

            mismatch = False
            if str(align_by) == "field_id":
                exp_set = {str(x) for x in expected}
                act_set = {str(x) for x in input_header}
                missing = exp_set.difference(act_set)
                extra = act_set.difference(exp_set)
                mismatch = bool(missing or extra)
            else:
                mismatch = list(input_header) != expected

            if mismatch:
                diff = _describe_header_diff(expected, input_header)
                if on_mismatch == "error":
                    msg = "Field alignment mismatch (sheetbook_append): sheetbook={!r}, sheet={!r}".format(str(sheetbook_id), sheet_name)
                    raise WorkflowWriteError(msg, diff=diff)
                if on_mismatch == "warn":
                    _ = self._instrumentation.emit(
                        EVENT_DIAGNOSTIC_WARNING,
                        DiagnosticWarningEvent(
                            message="Field alignment mismatch (warn): sheetbook={!r}, sheet={!r}".format(str(sheetbook_id), sheet_name),
                            source_id=None,
                            field_id=None,
                            lookup_key={"expected": expected, "actual": list(input_header)},
                            row_id=None,
                        ),
                        meta={"workflow_exec_id": self._workflow_exec_id, "workflow_node_id": str(workflow_node_id)},
                    )
                if on_mismatch == "skip":
                    display_path = plan.export_path if plan.export_path is not None else "<memory>"
                    self._emit_resource_write(
                        workflow_node_id=str(workflow_node_id),
                        resource_type="sheetbook",
                        resource_id=str(sheetbook_id),
                        path=str(display_path),
                        write_kind="sheetbook_append",
                        action="skip",
                        input_node_id=str(input_node_id),
                        input_output_id=str(input_output_id),
                        sheet=sheet_name,
                    )
                    plan.last_workflow_node_id = str(workflow_node_id)
                    return

            # 重复写入检测: 同一生产者不应写入同一工作表多次.
            for seg in sheet_plan.segments:
                if str(seg.producer_node_id) == str(input_node_id):
                    msg = "Duplicate sheetbook write for the same producer: sheetbook={!r}, sheet={!r}, producer={!r}".format(
                        str(sheetbook_id),
                        sheet_name,
                        str(input_node_id),
                    )
                    raise WorkflowWriteError(msg, diff=["producer_node_id={!r}".format(str(input_node_id))])

            start_row = int(sheet_plan.row_count)

        # 读取并对齐到基线表头(列式追加).
        col_lists: List[List[str]] = [[] for _ in expected]
        for row in _iter_csv_rows(input_csv_path):
            for col_idx, src_idx in enumerate(mapping):
                col_lists[col_idx].append(row[src_idx] if src_idx >= 0 and src_idx < len(row) else "")

        append_rows = len(col_lists[0]) if col_lists else 0
        append_cells = int(append_rows) * len(expected)

        with self._lock:
            sheet_plan = plan.sheets.get(sheet_name)
            if sheet_plan is None:  # pragma: no cover
                msg = "Sheetbook sheet missing during append: sheetbook={!r}, sheet={!r}".format(str(sheetbook_id), sheet_name)
                raise WorkflowWriteError(msg)

            new_total = int(plan.total_cells) + int(append_cells)
            self._check_sheetbook_budget(plan, new_total_cells=new_total)

            for idx, field_key in enumerate(expected):
                col = sheet_plan.columns.get(field_key)
                if col is None:
                    col = []
                    sheet_plan.columns[str(field_key)] = col
                col.extend(col_lists[idx])

            sheet_plan.row_count = int(sheet_plan.row_count) + int(append_rows)
            sheet_plan.segments.append(
                _SheetBookSegment(
                    producer_node_id=str(input_node_id),
                    start_row=int(start_row),
                    end_row=int(start_row) + int(append_rows),
                    header_policy=str(header_policy),
                )
            )
            plan.total_cells = int(new_total)

        display_path = plan.export_path if plan.export_path is not None else "<memory>"
        self._emit_resource_write(
            workflow_node_id=str(workflow_node_id),
            resource_type="sheetbook",
            resource_id=str(sheetbook_id),
            path=str(display_path),
            write_kind="sheetbook_append",
            action="append",
            input_node_id=str(input_node_id),
            input_output_id=str(input_output_id),
            sheet=sheet_name,
        )
        plan.last_workflow_node_id = str(workflow_node_id)

    def iter_sheetbook_sheet_rows(  # noqa: C901
        self,
        *,
        consumer_node_id: str,
        visible_producer_node_ids: FrozenSet[str],
        producer_node_id: str,
        sheetbook_id: str,
        sheet: str,
    ) -> Iterator[Dict[str, object]]:
        """读取 `sheetbook` 的行快照(按 `ref.node` 截断;按依赖可见性过滤).

        返回: `Iterator[Dict[str, object]]`, 每行必须为 `JSON-like mapping`.
        """
        _ = str(consumer_node_id)
        producer = str(producer_node_id)
        sb_id = str(sheetbook_id)
        sheet_name = str(sheet)
        visible = frozenset(str(x) for x in visible_producer_node_ids)

        with self._lock:
            plan = self._sheetbooks.get(sb_id)
            if plan is None:
                msg = "Unknown sheetbook resource id: {!r}".format(sb_id)
                raise ValueError(msg)
            sheet_plan = plan.sheets.get(sheet_name)
            if sheet_plan is None:
                msg = "Unknown sheetbook sheet: sheetbook={!r}, sheet={!r}".format(sb_id, sheet_name)
                raise ValueError(msg)

            cutoff_idx = None
            for idx, seg in enumerate(sheet_plan.segments):
                if str(seg.producer_node_id) == producer:
                    cutoff_idx = int(idx)
                    break
            if cutoff_idx is None:
                msg = "Unknown sheetbook ref node for sheet: node={!r}, sheetbook={!r}, sheet={!r}".format(producer, sb_id, sheet_name)
                raise ValueError(msg)

            baseline_header = list(sheet_plan.baseline_header)
            columns = dict(sheet_plan.columns)
            segments: List[Tuple[str, int, int]] = []
            for seg in sheet_plan.segments[: cutoff_idx + 1]:
                seg_producer = str(seg.producer_node_id)
                if seg_producer != producer and seg_producer not in visible:
                    continue
                segments.append((seg_producer, int(seg.start_row), int(seg.end_row)))

        def _iter() -> Iterator[Dict[str, object]]:
            for _seg_producer, start, end in segments:
                for row_idx in range(int(start), int(end)):
                    row: Dict[str, object] = {}
                    for key in baseline_header:
                        col = columns.get(str(key))
                        if col is None:  # pragma: no cover
                            row[str(key)] = ""  # pragma: no cover
                            continue  # pragma: no cover
                        row[str(key)] = col[row_idx] if row_idx >= 0 and row_idx < len(col) else ""
                    yield row

        return _iter()

    def apply_workbook_sheet(
        self,
        *,
        workflow_node_id: str,
        workbook_id: str,
        sheet: str,
        input_node_id: str,
        input_output_id: str,
        input_csv_path: str,
        on_conflict: str,
    ) -> None:
        plan = self._get_or_create_workbook(workbook_id, workflow_node_id=str(workflow_node_id))
        sheet_name = str(sheet)
        action = "write"

        input_header = _read_csv_header(input_csv_path)

        with self._lock:
            existing = plan.sheets.get(sheet_name)
            if existing is not None:
                if on_conflict == "skip":
                    action = "skip"
                    self._emit_resource_write(
                        workflow_node_id=str(workflow_node_id),
                        resource_type="workbook",
                        resource_id=str(workbook_id),
                        path=str(plan.path),
                        write_kind="workbook_sheet",
                        action=action,
                        input_node_id=str(input_node_id),
                        input_output_id=str(input_output_id),
                        sheet=sheet_name,
                    )
                    plan.last_workflow_node_id = str(workflow_node_id)
                    return
                if on_conflict == "error":
                    msg = "Sheet conflict (workbook_sheet): workbook={!r}, sheet={!r}".format(str(workbook_id), sheet_name)
                    raise WorkflowWriteError(msg, diff=["on_conflict=error", "existing_sheet=present"])
                if on_conflict == "overwrite":
                    action = "overwrite"

            if existing is None:
                plan.sheet_order.append(sheet_name)

            mapping = _build_alignment_mapping(input_header, input_header)
            segment = _AppendSegment(
                input_csv_path=str(input_csv_path),
                header_policy="once",
                mapping=mapping,
                on_mismatch="error",
                align_by="header",
                input_header=input_header,
            )
            plan.sheets[sheet_name] = _SheetPlan(sheet=sheet_name, baseline_header=list(input_header), segments=[segment])

        self._emit_resource_write(
            workflow_node_id=str(workflow_node_id),
            resource_type="workbook",
            resource_id=str(workbook_id),
            path=str(plan.path),
            write_kind="workbook_sheet",
            action=action,
            input_node_id=str(input_node_id),
            input_output_id=str(input_output_id),
            sheet=sheet_name,
        )
        plan.last_workflow_node_id = str(workflow_node_id)

    def apply_workbook_append(
        self,
        *,
        workflow_node_id: str,
        workbook_id: str,
        sheet: str,
        input_node_id: str,
        input_output_id: str,
        input_csv_path: str,
        align_by: str,
        header_policy: str,
        on_mismatch: str,
    ) -> None:
        plan = self._get_or_create_workbook(workbook_id, workflow_node_id=str(workflow_node_id))
        sheet_name = str(sheet)
        input_header = _read_csv_header(input_csv_path)

        with self._lock:
            sheet_plan = plan.sheets.get(sheet_name)
            if sheet_plan is None:
                plan.sheet_order.append(sheet_name)
                sheet_plan = _SheetPlan(sheet=sheet_name, baseline_header=list(input_header), segments=[])
                plan.sheets[sheet_name] = sheet_plan

            expected = list(sheet_plan.baseline_header)
            mapping = _build_alignment_mapping(expected, input_header)

            if list(input_header) != expected:
                diff = _describe_header_diff(expected, input_header)
                if on_mismatch == "error":
                    msg = "Field alignment mismatch (workbook_append): workbook={!r}, sheet={!r}".format(str(workbook_id), sheet_name)
                    raise WorkflowWriteError(msg, diff=diff)
                if on_mismatch == "warn":
                    _ = self._instrumentation.emit(
                        EVENT_DIAGNOSTIC_WARNING,
                        DiagnosticWarningEvent(
                            message="Field alignment mismatch (warn): workbook={!r}, sheet={!r}".format(str(workbook_id), sheet_name),
                            source_id=None,
                            field_id=None,
                            lookup_key={"expected": expected, "actual": list(input_header)},
                            row_id=None,
                        ),
                        meta={"workflow_exec_id": self._workflow_exec_id, "workflow_node_id": str(workflow_node_id)},
                    )
                if on_mismatch == "skip":
                    self._emit_resource_write(
                        workflow_node_id=str(workflow_node_id),
                        resource_type="workbook",
                        resource_id=str(workbook_id),
                        path=str(plan.path),
                        write_kind="workbook_append",
                        action="skip",
                        input_node_id=str(input_node_id),
                        input_output_id=str(input_output_id),
                        sheet=sheet_name,
                    )
                    plan.last_workflow_node_id = str(workflow_node_id)
                    return

            sheet_plan.segments.append(
                _AppendSegment(
                    input_csv_path=str(input_csv_path),
                    header_policy=str(header_policy),
                    mapping=mapping,
                    on_mismatch=str(on_mismatch),
                    align_by=str(align_by),
                    input_header=list(input_header),
                )
            )
        self._emit_resource_write(
            workflow_node_id=str(workflow_node_id),
            resource_type="workbook",
            resource_id=str(workbook_id),
            path=str(plan.path),
            write_kind="workbook_append",
            action="append",
            input_node_id=str(input_node_id),
            input_output_id=str(input_output_id),
            sheet=sheet_name,
        )
        plan.last_workflow_node_id = str(workflow_node_id)

    def apply_csv_append(
        self,
        *,
        workflow_node_id: str,
        csv_id: str,
        input_node_id: str,
        input_output_id: str,
        input_csv_path: str,
        header_policy: str,
        on_mismatch: str,
    ) -> None:
        plan = self._get_or_create_csv(csv_id, workflow_node_id=str(workflow_node_id))
        input_header = _read_csv_header(input_csv_path)

        with self._lock:
            if plan.baseline_header is None:
                plan.baseline_header = list(input_header)
                plan.segments = []

            expected = list(plan.baseline_header or [])
            mapping = _build_alignment_mapping(expected, input_header)

            if list(input_header) != expected:
                diff = _describe_header_diff(expected, input_header)
                if on_mismatch == "error":
                    msg = "Field alignment mismatch (csv_append): csv={!r}".format(str(csv_id))
                    raise WorkflowWriteError(msg, diff=diff)
                if on_mismatch == "warn":
                    _ = self._instrumentation.emit(
                        EVENT_DIAGNOSTIC_WARNING,
                        DiagnosticWarningEvent(
                            message="Field alignment mismatch (warn): csv={!r}".format(str(csv_id)),
                            source_id=None,
                            field_id=None,
                            lookup_key={"expected": expected, "actual": list(input_header)},
                            row_id=None,
                        ),
                        meta={"workflow_exec_id": self._workflow_exec_id, "workflow_node_id": str(workflow_node_id)},
                    )
                if on_mismatch == "skip":
                    self._emit_resource_write(
                        workflow_node_id=str(workflow_node_id),
                        resource_type="csv",
                        resource_id=str(csv_id),
                        path=str(plan.path),
                        write_kind="csv_append",
                        action="skip",
                        input_node_id=str(input_node_id),
                        input_output_id=str(input_output_id),
                    )
                    plan.last_workflow_node_id = str(workflow_node_id)
                    return

            cast("List[_AppendSegment]", plan.segments).append(
                _AppendSegment(
                    input_csv_path=str(input_csv_path),
                    header_policy=str(header_policy),
                    mapping=mapping,
                    on_mismatch=str(on_mismatch),
                    align_by="header",
                    input_header=list(input_header),
                )
            )
        self._emit_resource_write(
            workflow_node_id=str(workflow_node_id),
            resource_type="csv",
            resource_id=str(csv_id),
            path=str(plan.path),
            write_kind="csv_append",
            action="append",
            input_node_id=str(input_node_id),
            input_output_id=str(input_output_id),
        )
        plan.last_workflow_node_id = str(workflow_node_id)

    def commit_all(self) -> None:
        for plan in list(self._workbooks.values()):
            self._commit_workbook(plan)
        for plan in list(self._csvs.values()):
            self._commit_csv(plan)
        for plan in list(self._sheetbooks.values()):
            self._commit_sheetbook(plan)

    def discard_all(self, *, workflow_node_id: str, reason: str) -> None:
        for plan in list(self._workbooks.values()):
            self._discard_workbook(plan, workflow_node_id=str(workflow_node_id), reason=str(reason))
        for plan in list(self._csvs.values()):
            self._discard_csv(plan, workflow_node_id=str(workflow_node_id), reason=str(reason))
        for plan in list(self._sheetbooks.values()):
            self._discard_sheetbook(plan, workflow_node_id=str(workflow_node_id), reason=str(reason))

    def _commit_workbook(self, plan: _WorkbookPlan) -> None:  # noqa: C901, PLR0912
        if not plan.sheets:
            if plan.lock_path is not None:
                _release_write_lock(plan.lock_path)
                plan.lock_path = None
            return

        output_path = str(plan.path)
        output_dir = Path(output_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)

        try:
            workbook_cls = _get_openpyxl_workbook_class()
        except ImportError as exc:
            if plan.lock_path is not None:
                _release_write_lock(plan.lock_path)
                plan.lock_path = None
            raise WorkflowWriteError(str(exc)) from exc

        wb = workbook_cls(write_only=True)
        try:
            for sheet_name in plan.sheet_order:
                sheet_plan = plan.sheets.get(sheet_name)
                if sheet_plan is None:  # pragma: no cover
                    continue  # pragma: no cover
                ws = wb.create_sheet(str(sheet_name))
                header_written = False
                for seg in sheet_plan.segments:
                    if seg.header_policy == "always" or (seg.header_policy == "once" and not header_written):
                        _ = ws.append(list(sheet_plan.baseline_header))
                        header_written = True
                    # `never`: 不输出 `header`

                    for row in _iter_csv_rows(seg.input_csv_path):
                        out_row: List[object] = []
                        for idx in seg.mapping:
                            out_row.append(row[idx] if idx >= 0 and idx < len(row) else "")
                        _ = ws.append(out_row)

            temp_path = create_temp_path(output_path, ".xlsx.tmp")
            temp_obj = Path(temp_path)
            try:
                wb.save(temp_obj)
                _ = temp_obj.replace(output_path)
            except Exception as exc:
                with suppress(Exception):
                    if temp_obj.exists():
                        temp_obj.unlink()
                msg = "Workbook commit failed: {}: {}".format(type(exc).__name__, exc)
                raise WorkflowWriteError(msg) from exc
        except Exception:
            _best_effort_close_write_only_workbook_worksheets(wb)
            raise
        finally:
            with suppress(Exception):
                wb.close()
            if plan.lock_path is not None:
                _release_write_lock(plan.lock_path)
                plan.lock_path = None

        node_id = plan.last_workflow_node_id or "__wf__commit"
        self._emit_resource_commit(workflow_node_id=node_id, resource_type="workbook", resource_id=plan.resource_id, path=str(plan.path))

    def _commit_csv(self, plan: _CsvPlan) -> None:
        if plan.segments is None or plan.baseline_header is None:
            if plan.lock_path is not None:
                _release_write_lock(plan.lock_path)
                plan.lock_path = None
            return

        output_path = str(plan.path)
        output_dir = Path(output_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)

        temp_path = create_temp_path(output_path, ".csv.tmp")
        temp_obj = Path(temp_path)

        try:
            with io.open(str(temp_obj), "w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                header_written = False
                for seg in plan.segments:
                    if seg.header_policy == "always" or (seg.header_policy == "once" and not header_written):
                        writer.writerow(list(plan.baseline_header))
                        header_written = True

                    for row in _iter_csv_rows(seg.input_csv_path):
                        out_row: List[str] = []
                        for idx in seg.mapping:
                            out_row.append(row[idx] if idx >= 0 and idx < len(row) else "")
                        writer.writerow(out_row)

            _ = temp_obj.replace(output_path)
        except Exception as exc:
            with suppress(Exception):
                if temp_obj.exists():
                    temp_obj.unlink()
            msg = "CSV commit failed: {}: {}".format(type(exc).__name__, exc)
            raise WorkflowWriteError(msg) from exc
        finally:
            if plan.lock_path is not None:
                _release_write_lock(plan.lock_path)
                plan.lock_path = None

        node_id = plan.last_workflow_node_id or "__wf__commit"
        self._emit_resource_commit(workflow_node_id=node_id, resource_type="csv", resource_id=plan.resource_id, path=str(plan.path))

    def _discard_workbook(self, plan: _WorkbookPlan, *, workflow_node_id: str, reason: str) -> None:
        if plan.lock_path is not None:
            _release_write_lock(plan.lock_path)
            plan.lock_path = None
        node_id = plan.last_workflow_node_id or str(workflow_node_id)
        self._emit_resource_discard(
            workflow_node_id=node_id,
            resource_type="workbook",
            resource_id=plan.resource_id,
            path=str(plan.path),
            reason=str(reason),
        )

    def _discard_csv(self, plan: _CsvPlan, *, workflow_node_id: str, reason: str) -> None:
        if plan.lock_path is not None:
            _release_write_lock(plan.lock_path)
            plan.lock_path = None
        node_id = plan.last_workflow_node_id or str(workflow_node_id)
        self._emit_resource_discard(
            workflow_node_id=node_id,
            resource_type="csv",
            resource_id=plan.resource_id,
            path=str(plan.path),
            reason=str(reason),
        )

    def _commit_sheetbook(self, plan: _SheetBookPlan) -> None:  # noqa: C901, PLR0912, PLR0915
        export_path = plan.export_path
        display_path = export_path if export_path is not None else "<memory>"
        if export_path is None:
            node_id = plan.last_workflow_node_id or "__wf__commit"
            self._emit_resource_commit(
                workflow_node_id=node_id, resource_type="sheetbook", resource_id=plan.resource_id, path=str(display_path)
            )
            return

        output_path = str(export_path)
        output_dir = Path(output_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)

        lock_path = None
        if bool(plan.export_write_lock):
            lock_path = _acquire_write_lock(output_path)
            plan.lock_path = lock_path

        try:
            workbook_cls = _get_openpyxl_workbook_class()
        except ImportError as exc:
            if lock_path is not None:
                _release_write_lock(lock_path)
                plan.lock_path = None
            raise WorkflowWriteError(str(exc)) from exc

        wb = workbook_cls(write_only=True)
        try:
            for sheet_name in plan.sheet_order:
                sheet_plan = plan.sheets.get(sheet_name)
                if sheet_plan is None:
                    continue  # pragma: no cover
                ws = wb.create_sheet(str(sheet_name))
                header_written = False
                fields = list(sheet_plan.baseline_header)
                for seg in sheet_plan.segments:
                    if seg.header_policy == "always" or (seg.header_policy == "once" and not header_written):
                        _ = ws.append(list(fields))
                        header_written = True
                    # `never`: 不输出表头

                    for row_idx in range(int(seg.start_row), int(seg.end_row)):
                        out_row: List[object] = []
                        for field_key in fields:
                            col = sheet_plan.columns.get(str(field_key))
                            out_row.append(col[row_idx] if col is not None and row_idx >= 0 and row_idx < len(col) else "")
                        _ = ws.append(out_row)

            temp_path = create_temp_path(output_path, ".xlsx.tmp")
            temp_obj = Path(temp_path)
            try:
                wb.save(temp_obj)
                _ = temp_obj.replace(output_path)
            except Exception as exc:
                with suppress(Exception):
                    if temp_obj.exists():
                        temp_obj.unlink()
                msg = "Sheetbook export failed: {}: {}".format(type(exc).__name__, exc)
                raise WorkflowWriteError(msg) from exc
        except Exception:
            _best_effort_close_write_only_workbook_worksheets(wb)
            raise
        finally:
            with suppress(Exception):
                wb.close()
            if plan.lock_path is not None:
                _release_write_lock(plan.lock_path)
                plan.lock_path = None

        node_id = plan.last_workflow_node_id or "__wf__commit"
        self._emit_resource_commit(
            workflow_node_id=node_id, resource_type="sheetbook", resource_id=plan.resource_id, path=str(display_path)
        )

    def _discard_sheetbook(self, plan: _SheetBookPlan, *, workflow_node_id: str, reason: str) -> None:
        if plan.lock_path is not None:
            _release_write_lock(plan.lock_path)
            plan.lock_path = None
        node_id = plan.last_workflow_node_id or str(workflow_node_id)
        display_path = plan.export_path if plan.export_path is not None else "<memory>"
        self._emit_resource_discard(
            workflow_node_id=node_id,
            resource_type="sheetbook",
            resource_id=plan.resource_id,
            path=str(display_path),
            reason=str(reason),
        )


__all__ = [
    "WorkflowResourceManager",
    "WorkflowWriteError",
]
