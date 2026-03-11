# region imports

import logging
from contextlib import suppress
from pathlib import Path
from typing import TYPE_CHECKING, Any, List, Optional, Sequence, Type

from ..vendor.compact.importlibx import require_optional_dependency

if TYPE_CHECKING:
    from openpyxl import Workbook
else:
    _openpyxl = require_optional_dependency("openpyxl", context="scalim.sinks.sink_excel")
    Workbook = _openpyxl.Workbook

from ..typedefs import RowData, SinkRowKeySeq
from ..vendor.compact.typing_extensionsx import Self, override
from .sink_base import (
    BaseRowSink,
    ColumnBatch,
    ColumnData,
    ColumnValues,
    IColumnSink,
    create_temp_path,
    iter_row_values,
    store_rows_as_columns,
    update_column,
    update_columns,
)

if TYPE_CHECKING:
    import types

# endregion

_LOGGER = logging.getLogger(__name__)

EXCEL_SINK_SAVE_FAILED = "ExcelSink 保存失败"
EXCEL_SINK_SAVE_FAILED_LOG = EXCEL_SINK_SAVE_FAILED + ": %s"

EXCEL_SINK_REMOVE_TEMP_FILE_FAILED = "ExcelSink 删除临时文件失败"
EXCEL_SINK_REMOVE_TEMP_FILE_FAILED_LOG = EXCEL_SINK_REMOVE_TEMP_FILE_FAILED + ": %s"

EXCEL_WORKBOOK_SINK_SAVE_FAILED = "ExcelWorkbookSink 保存失败"
EXCEL_WORKBOOK_SINK_SAVE_FAILED_LOG = EXCEL_WORKBOOK_SINK_SAVE_FAILED + ": %s"

EXCEL_WORKBOOK_SINK_REMOVE_TEMP_FILE_FAILED = "ExcelWorkbookSink 删除临时文件失败"
EXCEL_WORKBOOK_SINK_REMOVE_TEMP_FILE_FAILED_LOG = EXCEL_WORKBOOK_SINK_REMOVE_TEMP_FILE_FAILED + ": %s"

COLUMN_EXCEL_SINK_SAVE_FAILED = "ColumnExcelSink 保存失败"
COLUMN_EXCEL_SINK_SAVE_FAILED_LOG = COLUMN_EXCEL_SINK_SAVE_FAILED + ": %s"

COLUMN_EXCEL_SINK_REMOVE_TEMP_FILE_FAILED = "ColumnExcelSink 删除临时文件失败"
COLUMN_EXCEL_SINK_REMOVE_TEMP_FILE_FAILED_LOG = COLUMN_EXCEL_SINK_REMOVE_TEMP_FILE_FAILED + ": %s"


class ExcelSink(BaseRowSink):
    """`Excel` 行写入器:支持流式行写入.

    在内存中缓存数据,并在 `close()` 时一次性写入 `Excel` 文件.
    实现 `IRowSink` 接口,支持单行流式写入以优化内存使用(FR023).

    参数:
        `output_path`: 输出文件路径
        `field_names`: 字段标识列表,用于从行数据中取值
        `header_names`: 表头名称列表(可选),用于输出表头;默认等于 `field_names`
        `sheet_name`: 工作表名称,默认 `"Sheet1"`
        `include_header`: 是否包含表头,默认 `True`

    示例:
    ```python
    with ExcelSink("report.xlsx", field_names=["id", "name"]) as sink:
        sink.write_row({"id": 1, "name": "张三"})
        sink.write_batch([{"id": 2, "name": "李四"}])

    # 使用不同的表头名称
    with ExcelSink("report.xlsx", field_names=["id", "name"], header_names=["编号", "姓名"]) as sink:
        sink.write_row({"id": 1, "name": "张三"})
    ```
    """

    output_path: str
    sheet_name: str
    include_header: bool
    _closed: bool
    field_names: List[str]
    header_names: List[str]
    _workbook: Workbook
    _worksheet: Any

    def __init__(
        self,
        output_path: str,
        field_names: List[str],
        header_names: Optional[List[str]] = None,
        sheet_name: str = "Sheet1",
        include_header: bool = True,  # noqa: FBT001, FBT002
    ) -> None:
        self.output_path = output_path
        self.sheet_name = sheet_name
        self.include_header = include_header
        self._closed = False
        self.field_names = field_names
        self.header_names = header_names if header_names is not None else field_names
        self._workbook = Workbook(write_only=True)
        self._worksheet = self._workbook.create_sheet(self.sheet_name)
        if self.include_header:
            _ = self._worksheet.append(self.header_names)

    def _format_row(self, row: RowData) -> List[Any]:
        return [row.get(field_name) for field_name in self.field_names]

    @override
    def write_row(self, row: RowData) -> None:
        _ = self._worksheet.append(self._format_row(row))

    @override
    def write_batch(self, rows: Sequence[RowData]) -> None:
        for row in rows:
            _ = self._worksheet.append(self._format_row(row))

    @override
    def close(self) -> None:
        if self._closed:
            return

        # 使用临时文件,确保在同一目录以支持原子重命名
        temp_path = create_temp_path(self.output_path, ".xlsx.tmp")
        temp_path_obj = Path(temp_path)

        try:
            self._workbook.save(temp_path_obj)
            # 原子重命名临时文件到目标路径
            _ = temp_path_obj.replace(self.output_path)
        except Exception:
            _LOGGER.exception(EXCEL_SINK_SAVE_FAILED_LOG, self.output_path)
            # 清理临时文件
            if temp_path_obj.exists():
                try:
                    temp_path_obj.unlink()
                except OSError:
                    _LOGGER.warning(EXCEL_SINK_REMOVE_TEMP_FILE_FAILED_LOG, temp_path_obj, exc_info=True)
            # 尽力清理:在保存失败时避免出现只写生成器告警.
            try:
                is_closed = self._worksheet.closed
            except AttributeError:
                is_closed = True
            if not is_closed:
                with suppress(Exception):
                    self._worksheet.close()
            raise
        finally:
            with suppress(Exception):
                self._workbook.close()

        self._closed = True


class ExcelWorkbookSheetRowSink(BaseRowSink):
    """`Excel` `workbook` 中单个 `sheet` 的行写入器(共享同一 `workbook`).

    - 该 `sink` 仅负责向 `worksheet` 追加行.
    - `workbook` 的保存/原子替换由 `ExcelWorkbookSink.close()` 统一完成.
    """

    sheet_name: str
    include_header: bool
    field_names: List[str]
    header_names: List[str]
    _worksheet: Any
    _closed: bool

    def __init__(
        self,
        *,
        worksheet: Any,
        sheet_name: str,
        field_names: List[str],
        header_names: Optional[List[str]] = None,
        include_header: bool = True,
    ) -> None:
        self._worksheet = worksheet
        self.sheet_name = str(sheet_name)
        self.include_header = bool(include_header)
        self.field_names = list(field_names)
        self.header_names = list(header_names) if header_names is not None else list(self.field_names)
        self._closed = False
        if self.include_header:
            # 关键护栏: `header` 与 `rows` 分离,避免 `header` `list` 被复用污染.
            _ = self._worksheet.append(list(self.header_names))

    def _format_row(self, row: RowData) -> List[Any]:
        return [row.get(field_name) for field_name in self.field_names]

    @override
    def write_row(self, row: RowData) -> None:
        if self._closed:
            msg = "ExcelWorkbookSheetRowSink is closed: {}".format(self.sheet_name)
            raise RuntimeError(msg)
        _ = self._worksheet.append(self._format_row(row))

    @override
    def write_batch(self, rows: Sequence[RowData]) -> None:
        if self._closed:
            msg = "ExcelWorkbookSheetRowSink is closed: {}".format(self.sheet_name)
            raise RuntimeError(msg)
        for row in rows:
            _ = self._worksheet.append(self._format_row(row))

    @override
    def close(self) -> None:
        # `sheet` `sink` 仅标记关闭,不保存文件.
        self._closed = True


class ExcelWorkbookSink:
    """`Excel` `workbook` 容器: 单次保存,支持多 `sheet`.

    设计目标:
    - 多 `sheet` 共享同一 `workbook`,避免重复打开/保存文件
    - 支持 `write_only` 流式写入,避免“全量 `rows` 攒内存”
    - `sheet` 名冲突 `fail-fast`
    - `sheet` 顺序稳定: 由 `create_sheet()` 调用顺序决定
    """

    output_path: str
    _workbook: Workbook
    _sheet_names: List[str]
    _closed: bool

    def __init__(
        self,
        output_path: str,
    ) -> None:
        self.output_path = str(output_path)
        self._workbook = Workbook(write_only=True)
        self._sheet_names = []
        self._closed = False

    def create_sheet_row_sink(
        self,
        sheet_name: str,
        *,
        field_names: List[str],
        header_names: Optional[List[str]] = None,
        include_header: bool = True,
    ) -> ExcelWorkbookSheetRowSink:
        if self._closed:
            msg = "ExcelWorkbookSink is closed: {}".format(self.output_path)
            raise RuntimeError(msg)

        name = str(sheet_name)
        if name in self._sheet_names:
            msg = "Duplicate excel sheet name in workbook: {!r}".format(name)
            raise ValueError(msg)

        ws = self._workbook.create_sheet(name)
        self._sheet_names.append(name)
        return ExcelWorkbookSheetRowSink(
            worksheet=ws,
            sheet_name=name,
            field_names=field_names,
            header_names=header_names,
            include_header=include_header,
        )

    def close(self) -> None:
        if self._closed:
            return

        temp_path = create_temp_path(self.output_path, ".xlsx.tmp")
        temp_path_obj = Path(temp_path)

        try:
            self._workbook.save(temp_path_obj)
            _ = temp_path_obj.replace(self.output_path)
        except Exception:
            _LOGGER.exception(EXCEL_WORKBOOK_SINK_SAVE_FAILED_LOG, self.output_path)
            if temp_path_obj.exists():
                try:
                    temp_path_obj.unlink()
                except OSError:
                    _LOGGER.warning(EXCEL_WORKBOOK_SINK_REMOVE_TEMP_FILE_FAILED_LOG, temp_path_obj, exc_info=True)
            raise
        finally:
            with suppress(Exception):
                self._workbook.close()

        self._closed = True


class ColumnExcelSink(IColumnSink):
    """列式 `Excel` 写入器:用于生产环境的高性能列式写入(FR023).

    工作原理:
    1. 在内存中按列缓存数据
    2. 在 `close()` 时一次性将所有数据写入 `Excel` 文件

    优点:
    - 调用方可在 `write_column()` 后立即释放该列的源数据
    - 适合宽表场景(200+ 列)

    参数:
        `output_path`: 输出文件路径
        `field_names`: 字段标识列表,用于从列数据中取值
        `header_names`: 表头名称列表(可选),用于输出表头;默认等于 `field_names`
        `sheet_name`: 工作表名称,默认 `"Sheet1"`
        `include_header`: 是否包含表头,默认 `True`

    示例:
    ```python
    with ColumnExcelSink("/tmp/report.xlsx", ["id", "name"]) as sink:
        sink.set_row_ids([1, 2, 3])
        sink.write_column("id", {1: 1, 2: 2, 3: 3})
        sink.write_column("name", {1: "甲", 2: "乙", 3: "丙"})

    # 使用不同的表头名称
    with ColumnExcelSink("/tmp/report.xlsx", ["id", "name"], header_names=["编号", "姓名"]) as sink:
        sink.set_row_ids([1, 2, 3])
        sink.write_column("id", {1: 1, 2: 2, 3: 3})
    ```
    """

    output_path: str
    field_names: List[str]
    header_names: List[str]
    sheet_name: str
    include_header: bool
    _row_ids: List[Any]
    _columns: ColumnData
    _closed: bool

    def __init__(
        self,
        output_path: str,
        field_names: List[str],
        header_names: Optional[List[str]] = None,
        sheet_name: str = "Sheet1",
        include_header: bool = True,  # noqa: FBT001, FBT002
    ) -> None:
        self.output_path = output_path
        self.field_names = field_names
        self.header_names = header_names if header_names is not None else field_names
        self.sheet_name = sheet_name
        self.include_header = include_header
        self._row_ids = []
        self._columns = {}
        self._closed = False

    @override
    def set_row_ids(self, row_ids: "SinkRowKeySeq") -> None:
        self._row_ids.extend(row_ids)

    @override
    def write_column(self, field_key: str, values: ColumnValues) -> None:
        update_column(self._columns, field_key, values)

    @override
    def write_columns(self, columns: ColumnBatch) -> None:
        update_columns(self._columns, columns)

    @override
    def write_batch(self, rows: Sequence[RowData]) -> None:
        start_index = len(self._row_ids)

        def _pk_factory(row_idx: int) -> int:
            return start_index + row_idx

        store_rows_as_columns(rows, self._row_ids, self._columns, _pk_factory)

    @override
    def close(self) -> None:
        if self._closed:
            return

        # 使用临时文件,确保在同一目录以支持原子重命名
        temp_path = create_temp_path(self.output_path, ".xlsx.tmp")
        temp_path_obj = Path(temp_path)
        wb = None

        try:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet(self.sheet_name)
            else:
                ws.title = self.sheet_name

            if self.include_header:
                _ = ws.append(self.header_names)

            for row_values in iter_row_values(self._row_ids, self.field_names, self._columns):
                _ = ws.append(list(row_values))

            wb.save(temp_path_obj)
            # 原子重命名临时文件到目标路径
            _ = temp_path_obj.replace(self.output_path)
        except Exception:
            _LOGGER.exception(COLUMN_EXCEL_SINK_SAVE_FAILED_LOG, self.output_path)
            # 清理临时文件
            if temp_path_obj.exists():
                try:
                    temp_path_obj.unlink()
                except OSError:
                    _LOGGER.warning(COLUMN_EXCEL_SINK_REMOVE_TEMP_FILE_FAILED_LOG, temp_path_obj, exc_info=True)
            raise
        finally:
            if wb is not None:
                with suppress(Exception):
                    wb.close()

        self._closed = True

    def __enter__(self) -> Self:
        return self

    def __exit__(
        self,
        exc_type: Optional[Type[BaseException]],
        exc_val: Optional[BaseException],
        exc_tb: Optional["types.TracebackType"],  # noqa: PYI036
    ) -> None:
        self.close()


__all__ = [
    "ColumnExcelSink",
    "ExcelSink",
]
