"""OutputWriteLayout resolve + factory mapping tests."""

from pathlib import Path

import importlib

import pytest
from openpyxl import load_workbook

from scalim.dsl.yaml_dsl import DemandRunRuntimeOptions, OutputWriteLayout
from scalim.execution import ExcelColumnResidency, ExecutionRequest, ExportLayout, OutputSpec
from scalim.execution.output_composition import OutputCompositionSpec, OutputTargetSpec
from scalim.execution.output_write_layout import resolve_output_write_layout
from scalim.execution.runtime_bindings import RuntimeBindings
from scalim.sinks import CSVSink, ColumnCSVSink, ColumnExcelSink, ExcelSink, StreamingColumnExcelSink
from scalim.spec.ir import DemandIr, FieldIr, MainSourceIr, RuntimeHandleIdIr

run_ir_mod = importlib.import_module("scalim.execution.run_ir")


def test_layout_enum_closed_set_has_no_removed_names() -> None:
    assert not hasattr(OutputWriteLayout, "COLUMN_HOLD")
    assert not hasattr(OutputWriteLayout, "COLUMN_WINDOW")
    assert {member.value for member in OutputWriteLayout} == {"row_stream", "column_buffered", "column_chunked"}


@pytest.mark.parametrize(
    "streaming,fmt,residency,expected",
    [
        (True, "csv", ExcelColumnResidency.BUFFERED, OutputWriteLayout.ROW_STREAM),
        (True, "excel", ExcelColumnResidency.BUFFERED, OutputWriteLayout.ROW_STREAM),
        (False, "csv", ExcelColumnResidency.BUFFERED, OutputWriteLayout.COLUMN_BUFFERED),
        (False, "csv", ExcelColumnResidency.CHUNKED, OutputWriteLayout.COLUMN_BUFFERED),
        (False, "excel", ExcelColumnResidency.BUFFERED, OutputWriteLayout.COLUMN_BUFFERED),
        (False, "excel", ExcelColumnResidency.CHUNKED, OutputWriteLayout.COLUMN_CHUNKED),
    ],
)
def test_resolve_derives_legacy_table(streaming, fmt, residency, expected) -> None:
    got = resolve_output_write_layout(
        output_write_layout=None,
        streaming=streaming,
        output_format=fmt,
        excel_column_residency=residency,
        has_output_composition=False,
    )
    assert got is expected


def test_resolve_composition_forces_row_stream() -> None:
    got = resolve_output_write_layout(
        output_write_layout=None,
        streaming=False,
        output_format="excel",
        excel_column_residency=ExcelColumnResidency.BUFFERED,
        has_output_composition=True,
    )
    assert got is OutputWriteLayout.ROW_STREAM


def test_explicit_layout_wins() -> None:
    got = resolve_output_write_layout(
        output_write_layout=OutputWriteLayout.COLUMN_BUFFERED,
        streaming=True,
        output_format="excel",
        excel_column_residency=ExcelColumnResidency.CHUNKED,
        has_output_composition=False,
    )
    assert got is OutputWriteLayout.COLUMN_BUFFERED


def test_resolve_rejects_non_enum_layout() -> None:
    with pytest.raises(TypeError, match="output_write_layout must be an OutputWriteLayout"):
        resolve_output_write_layout(
            output_write_layout="column_buffered",  # type: ignore[arg-type]
            streaming=False,
            output_format="excel",
            excel_column_residency=ExcelColumnResidency.BUFFERED,
            has_output_composition=False,
        )


def test_options_reject_string_layout() -> None:
    with pytest.raises(TypeError, match="OutputWriteLayout"):
        _ = DemandRunRuntimeOptions(output_write_layout="row_stream")  # type: ignore[arg-type]
    with pytest.raises(TypeError, match="OutputWriteLayout"):
        _ = ExecutionRequest(
            export_layout=ExportLayout(field_ids=("a",)),
            output_write_layout="column_buffered",  # type: ignore[arg-type]
        )


@pytest.mark.parametrize(
    "streaming,fmt,residency,sink_type",
    [
        (True, "csv", ExcelColumnResidency.BUFFERED, CSVSink),
        (False, "csv", ExcelColumnResidency.BUFFERED, ColumnCSVSink),
        (False, "csv", ExcelColumnResidency.CHUNKED, ColumnCSVSink),
        (True, "excel", ExcelColumnResidency.BUFFERED, ExcelSink),
        (False, "excel", ExcelColumnResidency.BUFFERED, ColumnExcelSink),
        (False, "excel", ExcelColumnResidency.CHUNKED, StreamingColumnExcelSink),
    ],
)
def test_unset_layout_preserves_legacy_sink_types(tmp_path: Path, streaming, fmt, residency, sink_type) -> None:
    layout = ExportLayout(field_ids=("id",))
    ext = "csv" if fmt == "csv" else "xlsx"
    sink = run_ir_mod._create_file_sink(
        OutputSpec(format=fmt, path=str(tmp_path / ("out." + ext)), streaming=streaming),
        layout,
        excel_column_residency=residency,
        output_write_layout=None,
    )
    assert isinstance(sink, sink_type)
    if isinstance(sink, StreamingColumnExcelSink):
        with pytest.raises(RuntimeError, match="set_row_ids"):
            sink.close()
    else:
        sink.close()


def test_explicit_csv_column_chunked_fails_fast(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="column_chunked"):
        _ = run_ir_mod._create_file_sink(
            OutputSpec(format="csv", path=str(tmp_path / "x.csv"), streaming=False),
            ExportLayout(field_ids=("id",)),
            output_write_layout=OutputWriteLayout.COLUMN_CHUNKED,
        )


def test_composition_plus_explicit_column_layout_fails_fast(tmp_path: Path) -> None:
    main_source = MainSourceIr(source_id="orders", loader_ref=RuntimeHandleIdIr(handle_id="orders.loader"))
    runtime_bindings = RuntimeBindings(main_source_loaders={"orders": (lambda: [{"id": 1}])})
    demand_ir = DemandIr.from_irs(
        sources=[],
        fields=[FieldIr(field_id="id", name="ID", source=main_source)],
        main_source=main_source,
    )
    composition = OutputCompositionSpec(
        targets=(
            OutputTargetSpec(
                target_id="t1",
                layout=ExportLayout(field_ids=("id",)),
                output=OutputSpec(format="excel", path=str(tmp_path / "out.xlsx"), streaming=True),
            ),
        ),
    )
    req = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("id",)),
        output_composition=composition,
        runtime_bindings=runtime_bindings,
        output_write_layout=OutputWriteLayout.COLUMN_BUFFERED,
    )
    with pytest.raises(ValueError, match="output_composition"):
        _ = run_ir_mod.run_ir(demand_ir, req)


def test_explicit_column_chunked_selects_streaming_column_excel(tmp_path: Path) -> None:
    layout = ExportLayout(field_ids=("id", "name"))
    path = tmp_path / "w.xlsx"
    sink = run_ir_mod._create_file_sink(
        OutputSpec(format="excel", path=str(path), streaming=False),
        layout,
        excel_column_residency=ExcelColumnResidency.BUFFERED,
        output_write_layout=OutputWriteLayout.COLUMN_CHUNKED,
    )
    assert isinstance(sink, StreamingColumnExcelSink)
    assert sink is not None
    row_ids = [1, 2, 3]
    sink.set_row_ids(row_ids)
    sink.write_column("id", {1: 1, 2: 2, 3: 3})
    sink.write_column("name", {1: "a", 2: "b", 3: "c"})
    sink.close()

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
    finally:
        wb.close()
    assert rows == [["id", "name"], [1, "a"], [2, "b"], [3, "c"]]
    assert sink._flushed_rows == 3
