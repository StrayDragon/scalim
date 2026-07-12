from pathlib import Path

import pytest
from openpyxl import load_workbook

from scalim.sinks import ColumnExcelSink, StreamingColumnExcelSink


def _read_matrix(path: Path):
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def test_streaming_column_excel_matches_column_excel_with_row_windows(tmp_path: Path) -> None:
    field_names = ["id", "name", "score"]
    row_ids = list(range(20))
    hold_path = tmp_path / "hold.xlsx"
    stream_path = tmp_path / "stream.xlsx"

    hold = ColumnExcelSink(str(hold_path), field_names=field_names)
    hold.set_row_ids(row_ids)
    for name in field_names:
        hold.write_column(name, {r: "{}-{}".format(name, r) for r in row_ids})
    hold.close()

    stream = StreamingColumnExcelSink(str(stream_path), field_names=field_names)
    stream.set_row_ids(row_ids)
    for start in range(0, 20, 5):
        end = min(20, start + 5)
        win = row_ids[start:end]
        for name in field_names:
            stream.write_column(name, {r: "{}-{}".format(name, r) for r in win})
    stream.close()

    assert _read_matrix(hold_path) == _read_matrix(stream_path)
    assert stream._flushed_rows == 20


def test_streaming_column_excel_multi_batch_set_row_ids(tmp_path: Path) -> None:
    """对齐 pipeline 列模式:每 batch set_row_ids(本批) → 写满全部列."""
    field_names = ["id", "name", "score"]
    row_ids = list(range(12))
    hold_path = tmp_path / "hold.xlsx"
    stream_path = tmp_path / "stream.xlsx"

    hold = ColumnExcelSink(str(hold_path), field_names=field_names)
    for start in range(0, 12, 4):
        win = row_ids[start : start + 4]
        hold.set_row_ids(win)
        for name in field_names:
            hold.write_column(name, {r: "{}-{}".format(name, r) for r in win})
    hold.close()

    stream = StreamingColumnExcelSink(str(stream_path), field_names=field_names)
    for start in range(0, 12, 4):
        win = row_ids[start : start + 4]
        stream.set_row_ids(win)
        for name in field_names:
            stream.write_column(name, {r: "{}-{}".format(name, r) for r in win})
        assert stream._flushed_rows == start + 4
    stream.close()

    assert _read_matrix(hold_path) == _read_matrix(stream_path)
    assert stream._flushed_rows == 12


@pytest.mark.filterwarnings("ignore::pytest.PytestUnraisableExceptionWarning")
def test_streaming_column_excel_incomplete_rows_fail_fast(tmp_path: Path) -> None:
    path = tmp_path / "incomplete.xlsx"
    sink = StreamingColumnExcelSink(str(path), field_names=["a", "b"])
    sink.set_row_ids([1, 2])
    sink.write_column("a", {1: 1, 2: 2})
    with pytest.raises(RuntimeError, match="未齐备"):
        sink.close()


def test_streaming_column_excel_duplicate_row_id_fail_fast(tmp_path: Path) -> None:
    sink = StreamingColumnExcelSink(str(tmp_path / "x.xlsx"), field_names=["a"])
    sink.set_row_ids([1])
    with pytest.raises(RuntimeError, match="重复"):
        sink.set_row_ids([1])


def test_streaming_column_excel_empty_set_row_ids_noop(tmp_path: Path) -> None:
    sink = StreamingColumnExcelSink(str(tmp_path / "x.xlsx"), field_names=["a"])
    sink.set_row_ids([])
    assert sink._row_ids == []
    sink.set_row_ids([1])
    sink.set_row_ids([])
    assert sink._row_ids == [1]
    sink.write_column("a", {1: 1})
    sink.close()


def test_streaming_column_excel_edge_guards(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    path = tmp_path / "edge.xlsx"
    with StreamingColumnExcelSink(str(path), field_names=["a", "b"], include_header=False) as sink:
        with pytest.raises(RuntimeError, match="set_row_ids"):
            sink.write_column("a", {1: 1})
        sink.set_row_ids([1])
        sink.write_columns({"a": {1: 10}, "b": {1: 20, 999: 0}})
        assert sink._flushed_rows == 1
        # 已刷行再写:跳过
        sink.write_column("a", {1: 99})
        with pytest.raises(KeyError):
            sink.write_column("missing", {1: 1})
        with pytest.raises(RuntimeError, match="write_column"):
            sink.write_batch([{"a": 1}])
    assert path.exists()
    sink.close()  # 幂等
    with pytest.raises(RuntimeError, match="已关闭"):
        sink.set_row_ids([2])
    with pytest.raises(RuntimeError, match="已关闭"):
        sink.write_column("a", {2: 1})


def test_streaming_column_excel_close_without_set_row_ids(tmp_path: Path) -> None:
    sink = StreamingColumnExcelSink(str(tmp_path / "x.xlsx"), field_names=["a"])
    with pytest.raises(RuntimeError, match="set_row_ids"):
        sink.close()


def test_streaming_column_excel_save_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    import scalim.sinks._internal.streaming_column_excel as mod

    def _boom(*_args, **_kwargs):
        raise OSError("save failed")

    monkeypatch.setattr(mod, "save_openpyxl_workbook_atomic", _boom)
    sink = StreamingColumnExcelSink(str(tmp_path / "x.xlsx"), field_names=["a"])
    sink.set_row_ids([1])
    sink.write_column("a", {1: 1})
    with pytest.raises(OSError, match="save failed"):
        sink.close()


def test_streaming_column_excel_flush_skips_null_values(tmp_path: Path) -> None:
    sink = StreamingColumnExcelSink(str(tmp_path / "x.xlsx"), field_names=["a"])
    sink.set_row_ids([1])
    sink._pending[0] = set()
    sink._values[0] = None
    sink.write_column("a", {1: 1})
    assert sink._flushed_rows == 0
    sink._abandon_open_workbook()
    sink._abandon_open_workbook()
    sink._closed = True
