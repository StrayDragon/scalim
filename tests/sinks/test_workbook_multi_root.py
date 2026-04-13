from pathlib import Path

import pytest

from scalim.execution.run_ir import ExecutionRequest, ExportLayout, OutputSpec
from scalim.execution import workbook_multi_root as wm_mod
from scalim.execution.workbook_multi_root import run_multi_root_workbook
from tests.cases.minimal_ir import build_minimal_ir_case


def _read_workbook_sheet_names(path: Path) -> "list[str]":
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def test_run_multi_root_workbook_writes_two_demands(tmp_path: Path) -> None:
    out = tmp_path / "multi.xlsx"
    case_a = build_minimal_ir_case()
    case_b = build_minimal_ir_case()

    req_a = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id", "amount"), header_names=None),
        output=OutputSpec(path=None),
        sink=None,
        runtime_bindings=case_a.runtime_bindings,
    )
    req_b = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id", "customer_name"), header_names=None),
        output=OutputSpec(path=None),
        sink=None,
        runtime_bindings=case_b.runtime_bindings,
    )

    results = run_multi_root_workbook(
        output_path=str(out),
        runs=(
            ("SheetA", case_a.demand, req_a),
            ("SheetB", case_b.demand, req_b),
        ),
    )

    assert out.exists()
    assert _read_workbook_sheet_names(out) == ["SheetA", "SheetB"]
    assert len(results) == 2


class _FakeWorkbookSink:
    instances = []

    def __init__(self, _output_path: str) -> None:
        self.closed = False
        self.__class__.instances.append(self)

    def create_sheet_row_sink(  # type: ignore[no-untyped-def]
        self,
        _sheet_name: str,
        *,
        field_names,
        header_names,
        include_header: bool = True,  # noqa: FBT001, FBT002
    ):
        _ = (field_names, header_names, include_header)
        return object()

    def close(self) -> None:
        self.closed = True


def test_run_multi_root_workbook_rejects_unknown_failure_policy(tmp_path: Path) -> None:
    case = build_minimal_ir_case()
    req = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(path=None),
        sink=None,
        runtime_bindings=case.runtime_bindings,
    )
    with pytest.raises(ValueError, match="Unsupported failure_policy"):
        _ = run_multi_root_workbook(output_path=str(tmp_path / "x.xlsx"), runs=(("S", case.demand, req),), failure_policy="bad")


def test_run_multi_root_workbook_all_fail_wraps_sheet_error(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(wm_mod, "ExcelWorkbookSink", _FakeWorkbookSink)

    def _boom(_demand_ir, _req):  # type: ignore[no-untyped-def]
        raise ValueError("boom")

    monkeypatch.setattr(wm_mod, "run_ir", _boom)

    case = build_minimal_ir_case()
    req = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(path=None),
        sink=None,
        runtime_bindings=case.runtime_bindings,
    )

    with pytest.raises(wm_mod.ScalimMultiRootWorkbookRunError, match="Workbook sheet run failed"):
        _ = run_multi_root_workbook(output_path=str(tmp_path / "x.xlsx"), runs=(("SheetA", case.demand, req),), failure_policy="all_fail")

    assert _FakeWorkbookSink.instances[-1].closed is True


def test_run_multi_root_workbook_primary_only_continues_and_returns_results(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(wm_mod, "ExcelWorkbookSink", _FakeWorkbookSink)

    calls = {"n": 0}

    def _maybe_boom(_demand_ir, _req):  # type: ignore[no-untyped-def]
        calls["n"] += 1
        if calls["n"] == 1:
            raise ValueError("boom")
        return object()

    monkeypatch.setattr(wm_mod, "run_ir", _maybe_boom)

    case_a = build_minimal_ir_case()
    case_b = build_minimal_ir_case()
    req = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(path=None),
        sink=None,
        runtime_bindings=case_a.runtime_bindings,
    )

    results = run_multi_root_workbook(
        output_path=str(tmp_path / "x.xlsx"),
        runs=(("SheetA", case_a.demand, req), ("SheetB", case_b.demand, req)),
        failure_policy="primary_only",
    )
    assert len(results) == 1
    assert _FakeWorkbookSink.instances[-1].closed is True


def test_run_multi_root_workbook_primary_only_keeps_first_error_and_skips_subsequent_failures(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(wm_mod, "ExcelWorkbookSink", _FakeWorkbookSink)

    calls = {"n": 0}

    def _maybe_boom(_demand_ir, _req):  # type: ignore[no-untyped-def]
        calls["n"] += 1
        if calls["n"] in (1, 2):
            raise ValueError("boom")
        return object()

    monkeypatch.setattr(wm_mod, "run_ir", _maybe_boom)

    case_a = build_minimal_ir_case()
    case_b = build_minimal_ir_case()
    case_c = build_minimal_ir_case()
    req = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(path=None),
        sink=None,
        runtime_bindings=case_a.runtime_bindings,
    )

    results = run_multi_root_workbook(
        output_path=str(tmp_path / "x.xlsx"),
        runs=(("SheetA", case_a.demand, req), ("SheetB", case_b.demand, req), ("SheetC", case_c.demand, req)),
        failure_policy="primary_only",
    )
    assert len(results) == 1
    assert _FakeWorkbookSink.instances[-1].closed is True
