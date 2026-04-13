from collections.abc import Mapping

import pytest

from scalim.sinks import BlockColumnCSVSink, CSVSink, ColumnCSVSink
from scalim.sinks import ColumnExcelSink, ExcelSink
from scalim.sinks import InMemoryColumnSink, InMemoryRowSink
from scalim.sinks import PandasColumnSink, PandasRowSink


def test_in_memory_row_sink_write_row_aligned_and_mismatch() -> None:
    sink = InMemoryRowSink()
    sink.write_row_aligned(["a", "b"], [1, 2])
    assert sink.get_data() == [{"a": 1, "b": 2}]

    with pytest.raises(ValueError, match="write_row_aligned"):
        sink.write_row_aligned(["a"], [1, 2])


def test_in_memory_column_sink_write_column_aligned_and_mismatch() -> None:
    sink = InMemoryColumnSink()
    sink.set_row_ids([0, 1])
    sink.write_column_aligned("a", [0, 1], [1, 2])
    assert sink.get_column("a") == {0: 1, 1: 2}

    with pytest.raises(ValueError, match="write_column_aligned"):
        sink.write_column_aligned("a", [0, 1], [1])


def test_csv_sink_write_row_aligned_caches_and_handles_missing_fields(tmp_path) -> None:
    output_path = tmp_path / "rows.csv"
    sink = CSVSink(str(output_path), field_names=["id", "name"], include_header=False, flush_policy="always")
    sink.write_row_aligned(["id"], [1])
    sink.write_row_aligned(["id", "name"], [2, "b"])
    sink.write_row_aligned(["id", "name"], [3, "c"])
    sink.close()

    assert output_path.read_text(encoding="utf-8").splitlines() == ["1,", "2,b", "3,c"]

    sink2 = CSVSink(str(tmp_path / "rows2.csv"), field_names=["id"], include_header=False)
    try:
        with pytest.raises(ValueError, match="write_row_aligned"):
            sink2.write_row_aligned(["id"], [1, 2])
    finally:
        sink2.close()


def test_column_csv_sink_write_column_aligned_and_mismatch(tmp_path) -> None:
    output_path = tmp_path / "cols.csv"
    sink = ColumnCSVSink(str(output_path), ["id"], include_header=False)
    sink.set_row_ids([0, 1])
    sink.write_column_aligned("id", [0, 1], [10, 20])
    sink.close()

    assert output_path.read_text(encoding="utf-8").splitlines() == ["10", "20"]

    with pytest.raises(ValueError, match="write_column_aligned"):
        sink.write_column_aligned("id", [0, 1], [10])


def test_block_column_csv_sink_write_column_aligned_and_requires_row_ids(tmp_path) -> None:
    output_path = tmp_path / "block.csv"
    sink = BlockColumnCSVSink(str(output_path), ["id", "name"], col_width=8, write_delay=0.0)

    with pytest.raises(RuntimeError, match="set_row_ids"):
        sink.write_column_aligned("id", [0], [1])

    sink.set_row_ids([0, 1])
    sink.write_column_aligned("id", [0, 1], [1, 2])
    sink.write_column_aligned("name", [0, 1], ["vvv", "www"])
    sink.close()

    content = output_path.read_text(encoding="utf-8")
    assert "vvv" in content
    assert "www" in content

    sink2 = BlockColumnCSVSink(str(tmp_path / "block2.csv"), ["id"], col_width=8, write_delay=0.0)
    try:
        sink2.set_row_ids([0])
        with pytest.raises(ValueError, match="write_column_aligned"):
            sink2.write_column_aligned("id", [0], [])
    finally:
        sink2.close()


def test_pandas_sinks_aligned_write_and_mismatch() -> None:
    row_sink = PandasRowSink(field_names=["id", "name"])
    row_sink.write_row_aligned(["id", "name"], [1, "a"])
    assert row_sink.get_rows() == [{"id": 1, "name": "a"}]

    with pytest.raises(ValueError, match="write_row_aligned"):
        row_sink.write_row_aligned(["id"], [1, 2])

    col_sink = PandasColumnSink()
    col_sink.set_row_ids([0, 1])
    col_sink.write_column_aligned("id", [0, 1], [10, 20])
    assert col_sink.get_columns()["id"] == {0: 10, 1: 20}
    assert col_sink.field_names == ["id"]

    with pytest.raises(ValueError, match="write_column_aligned"):
        col_sink.write_column_aligned("id", [0, 1], [10])


def test_pandas_sinks_cover_internal_branch_arcs() -> None:
    class _DuplicateKeyRow(Mapping):
        def __init__(self) -> None:
            self._data = {"id": 1}

        def __getitem__(self, key):  # type: ignore[no-untyped-def]
            return self._data[key]

        def __iter__(self):  # type: ignore[no-untyped-def]
            return iter(["id", "id"])

        def __len__(self) -> int:
            return len(self._data)

    row_sink = PandasRowSink()
    row_sink.write_row(_DuplicateKeyRow())
    assert row_sink.field_names == ["id"]

    col_sink = PandasColumnSink()
    col_sink.write_column("id", {0: 10})
    col_sink.write_column("id", {1: 20})
    col_sink.write_column_aligned("id", [0, 1], [10, 20])

    col_sink.set_row_ids([0])
    col_sink.write_batch([{"id": 1}])


def test_block_column_csv_sink_write_column_aligned_unknown_field_and_sleep(monkeypatch: pytest.MonkeyPatch, tmp_path) -> None:
    import scalim.sinks._internal.sink_csv as csv_mod

    output_path = tmp_path / "block_sleep.csv"
    sink = BlockColumnCSVSink(str(output_path), ["id"], col_width=8, write_delay=0.1)
    sink.set_row_ids([0])

    slept = []

    def _fake_sleep(seconds: float) -> None:
        slept.append(seconds)

    monkeypatch.setattr(csv_mod.time, "sleep", _fake_sleep)

    sink.write_column_aligned("unknown", [0], [1])
    sink.write_column_aligned("id", [0], [1])
    sink.close()

    assert slept == [0.1]


def _read_first_column_xlsx(path) -> list:  # type: ignore[no-untyped-def]
    import openpyxl

    wb = openpyxl.load_workbook(path)
    try:
        ws = wb.active
        return [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]
    finally:
        wb.close()


def test_excel_sinks_aligned_write_and_mismatch(tmp_path) -> None:
    row_path = tmp_path / "rows.xlsx"
    row_sink = ExcelSink(str(row_path), field_names=["id"], include_header=False)
    row_sink.write_row_aligned(["id"], [1])
    row_sink.close()
    assert _read_first_column_xlsx(row_path) == [1]

    row_sink2 = ExcelSink(str(tmp_path / "rows2.xlsx"), field_names=["id"], include_header=False)
    try:
        with pytest.raises(ValueError, match="write_row_aligned"):
            row_sink2.write_row_aligned(["id"], [1, 2])
    finally:
        row_sink2.close()

    col_path = tmp_path / "cols.xlsx"
    col_sink = ColumnExcelSink(str(col_path), field_names=["id"], include_header=False)
    col_sink.set_row_ids([0, 1])
    col_sink.write_column_aligned("id", [0, 1], [10, 20])
    col_sink.close()
    assert _read_first_column_xlsx(col_path) == [10, 20]

    with pytest.raises(ValueError, match="write_column_aligned"):
        col_sink.write_column_aligned("id", [0, 1], [10])


def test_column_excel_sink_close_skips_workbook_close_when_creation_fails(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path,
) -> None:
    import scalim.sinks._internal.excel as excel_mod

    output_path = tmp_path / "column.xlsx"
    sink = ColumnExcelSink(str(output_path), ["id"], include_header=False)

    sink.set_row_ids([0, 2])
    sink.write_batch([{"id": 1}])

    def _boom(*_args, **_kwargs):  # type: ignore[no-untyped-def]
        raise TypeError("boom")

    monkeypatch.setattr(excel_mod, "Workbook", _boom)

    with pytest.raises(TypeError, match="boom"):
        sink.close()
