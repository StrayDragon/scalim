from pathlib import Path
from typing import Dict, List

import pytest
from openpyxl import load_workbook

try:
    import pandas as pd
except Exception as exc:
    pytest.skip("pandas unavailable in this environment: {}".format(exc), allow_module_level=True)

from scalim.sinks import BlockColumnCSVSink
from scalim.sinks import ColumnExcelSink, ExcelSink
from scalim.sinks.memory import InMemoryColumnSink, InMemoryRowDataSink
from scalim.sinks.pandas import PandasColumnSink, PandasRowSink


@pytest.fixture
def sample_rows(example_report_ir_module) -> List[Dict[str, object]]:
    orders = example_report_ir_module.data_loader.get_orders("2024-01-01", "2024-01-07", page=0, page_size=3)
    rows: List[Dict[str, object]] = []
    for row in sorted(orders, key=lambda r: r["order_id"]):
        rows.append(
            {
                "order_id": row["order_id"],
                "amount": row["amount"],
            }
        )
    return rows


def test_inmemory_sinks_collect_data(sample_rows) -> None:
    rows = sample_rows

    row_sink = InMemoryRowDataSink()
    row_sink.write_batch(rows)
    row_sink.close()
    assert row_sink.get_data() == rows

    pks = [row["order_id"] for row in rows]
    columns = {"order_id": {pk: pk for pk in pks}, "amount": {row["order_id"]: row["amount"] for row in rows}}

    col_sink = InMemoryColumnSink(field_names=["order_id", "amount"])
    col_sink.set_row_ids(pks)
    col_sink.write_columns(columns)
    col_sink.close()

    assert col_sink.get_row_ids() == pks
    assert col_sink.get_columns()["order_id"][pks[0]] == pks[0]
    assert col_sink.get_rows()[0]["order_id"] == pks[0]
    grid = col_sink.get_2d_list(include_header=True)
    assert grid[0] == ["order_id", "amount"]

    auto_col_sink = InMemoryColumnSink()
    auto_col_sink.write_column("order_id", {pks[0]: pks[0]})
    auto_col_sink.set_row_ids([pks[0]])
    auto_col_sink.close()
    assert "order_id" in auto_col_sink.get_field_names()


def test_pandas_sinks_roundtrip(sample_rows) -> None:
    rows = sample_rows
    field_names = ["order_id", "amount"]

    row_sink = PandasRowSink(field_names=field_names)
    row_sink.write_batch(rows)
    row_sink.close()
    df_rows = row_sink.to_dataframe()
    assert list(df_rows.columns) == field_names

    pks = [row["order_id"] for row in rows]
    col_sink = PandasColumnSink(field_names=field_names)
    col_sink.set_row_ids(pks)
    col_sink.write_column("order_id", {pk: pk for pk in pks})
    col_sink.write_column("amount", {row["order_id"]: row["amount"] for row in rows})
    col_sink.close()
    df_cols = col_sink.to_dataframe()
    assert list(df_cols.columns) == field_names
    assert df_cols["order_id"].tolist() == pks

    auto_row_sink = PandasRowSink()
    auto_row_sink.write_row(rows[0])
    auto_row_sink.close()
    auto_df = auto_row_sink.to_dataframe()
    assert "order_id" in auto_df.columns

    auto_col_sink = PandasColumnSink()
    auto_col_sink.write_column("order_id", {pks[0]: pks[0]})
    auto_col_sink.write_column("amount", {pks[0]: rows[0]["amount"]})
    auto_col_sink.set_row_ids([pks[0]])
    auto_col_sink.close()
    auto_df_cols = auto_col_sink.to_dataframe()
    assert "order_id" in auto_df_cols.columns


def test_excel_sinks_write_files(sample_rows, tmp_path: Path) -> None:
    rows = sample_rows
    field_names = ["order_id", "amount"]

    row_path = tmp_path / "rows.xlsx"
    with ExcelSink(str(row_path), field_names) as row_sink:
        row_sink.write_batch(rows)

    wb = load_workbook(row_path)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    wb.close()
    assert values[0] == tuple(field_names)
    assert values[1][0] == rows[0]["order_id"]

    col_path = tmp_path / "cols.xlsx"
    pks = [row["order_id"] for row in rows]
    with ColumnExcelSink(str(col_path), field_names) as col_sink:
        col_sink.set_row_ids(pks)
        col_sink.write_column("order_id", {pk: pk for pk in pks})
        col_sink.write_column("amount", {row["order_id"]: row["amount"] for row in rows})

    wb = load_workbook(col_path)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    wb.close()
    assert values[0] == tuple(field_names)
    assert values[1][1] == rows[0]["amount"]


def test_block_column_csv_sink(sample_rows, tmp_path: Path) -> None:
    rows = sample_rows
    field_names = ["order_id", "amount"]
    pks = [row["order_id"] for row in rows]

    output_path = tmp_path / "block.csv"
    sink = BlockColumnCSVSink(str(output_path), field_names, col_width=12, write_delay=0.0)

    with pytest.raises(RuntimeError, match="必须先调用 set_row_ids"):
        sink.write_column("order_id", {pks[0]: pks[0]})

    sink.set_row_ids(pks)
    sink.write_column("order_id", {pk: pk for pk in pks})
    sink.write_column("amount", {row["order_id"]: row["amount"] for row in rows})

    extra_pk = 999
    sink.set_row_ids([extra_pk])
    sink.write_column("order_id", {extra_pk: extra_pk})
    sink.close()

    lines = output_path.read_text(encoding="utf-8").splitlines()
    header = [cell.strip() for cell in lines[0].split(",")]
    assert header == field_names
    assert len(lines) == len(pks) + 2
    values = [cell.strip() for cell in lines[1].split(",")]
    assert values[0] == str(rows[0]["order_id"])
