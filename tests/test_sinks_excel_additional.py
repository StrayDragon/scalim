import logging
import os
from pathlib import Path

import pytest

from scalim.sinks import sink_excel as excel_mod


def _read_excel_rows(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(str(path))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _read_cell_value_and_type(path: Path, cell_ref: str):
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=False)
    try:
        ws = wb.active
        cell = ws[cell_ref]
        return cell.value, cell.data_type
    finally:
        wb.close()


def _write_excel_rows(path: Path, sink_cls, rows, header_names=None) -> None:
    if sink_cls is excel_mod.ExcelSink:
        with sink_cls(str(path), field_names=["id", "name"], header_names=header_names) as sink:
            for row in rows:
                sink.write_row(row)
        return

    sink = sink_cls(str(path), field_names=["id", "name"], header_names=header_names)
    row_ids = [row["id"] for row in rows]
    sink.set_row_ids(row_ids)
    sink.write_column("id", {row["id"]: row["id"] for row in rows})
    sink.write_column("name", {row["id"]: row["name"] for row in rows})
    sink.close()


class _FakeWorksheet:
    def __init__(self) -> None:
        self.title = ""

    def append(self, _row) -> None:  # type: ignore[no-untyped-def]
        return None


class _CloseableWorksheet:
    def __init__(self) -> None:
        self.closed = False
        self.closed_called = False

    def append(self, _row) -> None:  # type: ignore[no-untyped-def]
        return None

    def close(self) -> None:
        self.closed_called = True
        self.closed = True


class _FakeWorkbook:
    def __init__(self, *args, **kwargs) -> None:  # type: ignore[no-untyped-def]
        self.active = None
        self._sheet = _FakeWorksheet()

    def create_sheet(self, name: str) -> _FakeWorksheet:
        self._sheet.title = name
        self.active = self._sheet
        return self._sheet

    def save(self, _path: str) -> None:
        return None

    def close(self) -> None:
        return None


def test_excel_sink_close_twice(tmp_path: Path) -> None:
    output_path = tmp_path / "rows.xlsx"
    sink = excel_mod.ExcelSink(str(output_path), ["id"])
    sink.write_row({"id": 1})
    sink.close()
    sink.close()

    assert output_path.exists()


def test_excel_workbook_sheet_row_sink_raises_when_closed() -> None:
    sheet_sink = excel_mod.ExcelWorkbookSheetRowSink(
        worksheet=_FakeWorksheet(),
        sheet_name="Sheet1",
        field_names=["id"],
        include_header=False,
    )
    sheet_sink.close()

    with pytest.raises(RuntimeError, match="ExcelWorkbookSheetRowSink is closed"):
        sheet_sink.write_row({"id": 1})
    with pytest.raises(RuntimeError, match="ExcelWorkbookSheetRowSink is closed"):
        sheet_sink.write_batch([{"id": 1}])


def test_excel_workbook_sink_close_twice_and_rejects_new_sheet(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(excel_mod, "Workbook", _FakeWorkbook)

    output_path = tmp_path / "wb.xlsx"
    wb = excel_mod.ExcelWorkbookSink(str(output_path))
    _ = wb.create_sheet_row_sink("Sheet1", field_names=["id"])

    wb.close()
    wb.close()

    with pytest.raises(RuntimeError, match="ExcelWorkbookSink is closed"):
        _ = wb.create_sheet_row_sink("Sheet2", field_names=["id"])


def test_column_excel_sink_write_columns_and_batch(tmp_path: Path) -> None:
    output_path = tmp_path / "cols.xlsx"
    sink = excel_mod.ColumnExcelSink(str(output_path), ["id", "name"])
    sink.write_columns({"id": {1: 1}})
    sink.write_batch([{"id": 2, "name": "Bob"}])
    sink.close()
    sink.close()

    assert output_path.exists()


def test_column_excel_sink_handles_missing_active(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(excel_mod, "Workbook", _FakeWorkbook)

    output_path = tmp_path / "cols_fake.xlsx"
    sink = excel_mod.ColumnExcelSink(str(output_path), ["id"])
    sink.set_row_ids([1])
    sink.write_column("id", {1: 1})
    sink.close()


@pytest.mark.parametrize(
    "sink_cls",
    [excel_mod.ExcelSink, excel_mod.ColumnExcelSink],
    ids=["row-sink", "column-sink"],
)
def test_excel_sink_with_header_names(tmp_path: Path, sink_cls) -> None:
    """Test that Excel sinks correctly use header_names for header and field_names for values."""
    output_path = tmp_path / "output.xlsx"
    rows = [{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}]

    _write_excel_rows(output_path, sink_cls, rows, header_names=["编号", "姓名"])

    loaded_rows = _read_excel_rows(output_path)
    assert loaded_rows[0] == ("编号", "姓名")  # Header uses header_names
    assert loaded_rows[1] == (1, "Alice")  # Values from field_names keys
    assert loaded_rows[2] == (2, "Bob")


class _FailingSaveWorkbook:
    def __init__(self, write_only: bool = False) -> None:
        self._sheet = _FakeWorksheet()

    def create_sheet(self, name: str) -> _FakeWorksheet:
        self._sheet.title = name
        return self._sheet

    def save(self, _path) -> None:  # type: ignore[no-untyped-def]
        raise OSError("simulated save failure")

    def close(self) -> None:
        return None


class _FailingSaveWorkbookNonWriteOnly:
    def __init__(self) -> None:
        self._sheet = _FakeWorksheet()
        self.active = self._sheet

    def save(self, _path) -> None:  # type: ignore[no-untyped-def]
        raise OSError("simulated save failure")

    def close(self) -> None:
        return None


def _init_sink_for_failure(tmp_path: Path, sink_cls, filename: str):
    output_path = tmp_path / filename
    sink = sink_cls(str(output_path), ["id"])
    if sink_cls is excel_mod.ExcelSink:
        sink.write_row({"id": 1})
    else:
        sink.set_row_ids([1])
        sink.write_column("id", {1: 1})
    return sink


@pytest.mark.parametrize(
    "sink_cls,workbook_cls,log_match,filename,set_worksheet",
    [
        (excel_mod.ExcelSink, _FailingSaveWorkbook, excel_mod.EXCEL_SINK_SAVE_FAILED, "exc_rows.xlsx", True),
        (excel_mod.ColumnExcelSink, _FailingSaveWorkbookNonWriteOnly, excel_mod.COLUMN_EXCEL_SINK_SAVE_FAILED, "exc_cols.xlsx", False),
    ],
    ids=["row-sink", "column-sink"],
)
def test_excel_sink_close_exception_cleans_temp_file(
    tmp_path: Path,
    monkeypatch,
    caplog,
    sink_cls,
    workbook_cls,
    log_match: str,
    filename: str,
    set_worksheet: bool,
) -> None:
    monkeypatch.setattr(excel_mod, "Workbook", workbook_cls)
    sink = _init_sink_for_failure(tmp_path, sink_cls, filename)
    if set_worksheet:
        sink._worksheet = _FakeWorksheet()

    caplog.set_level(logging.ERROR, logger=excel_mod.__name__)
    with pytest.raises(OSError, match="simulated save failure"):
        sink.close()
    assert list(tmp_path.glob("*.xlsx.tmp")) == []
    assert any(log_match in record.message for record in caplog.records)


def test_excel_sink_close_exception_closes_open_worksheet(tmp_path: Path, monkeypatch, caplog) -> None:
    monkeypatch.setattr(excel_mod, "Workbook", _FailingSaveWorkbook)
    output_path = tmp_path / "exc_close.xlsx"
    sink = excel_mod.ExcelSink(str(output_path), ["id"])
    sink.write_row({"id": 1})
    worksheet = _CloseableWorksheet()
    sink._worksheet = worksheet

    caplog.set_level(logging.ERROR, logger=excel_mod.__name__)

    with pytest.raises(OSError, match="simulated save failure"):
        sink.close()

    assert worksheet.closed_called is True


@pytest.mark.parametrize(
    "sink_cls",
    [excel_mod.ExcelSink, excel_mod.ColumnExcelSink],
    ids=["row-sink", "column-sink"],
)
def test_excel_formula_injection_escape_and_allow(tmp_path: Path, sink_cls) -> None:
    escape_path = tmp_path / "escape.xlsx"
    allow_path = tmp_path / "allow.xlsx"

    if sink_cls is excel_mod.ExcelSink:
        sink = sink_cls(str(escape_path), field_names=["id", "name"], allow_formulas=False)
        sink.write_row({"id": 1, "name": "=1+1"})
        sink.close()

        sink2 = sink_cls(str(allow_path), field_names=["id", "name"], allow_formulas=True)
        sink2.write_row({"id": 1, "name": "=1+1"})
        sink2.close()
    else:
        sink = sink_cls(str(escape_path), field_names=["id", "name"], allow_formulas=False)
        sink.set_row_ids([1])
        sink.write_column("id", {1: 1})
        sink.write_column("name", {1: "=1+1"})
        sink.close()

        sink2 = sink_cls(str(allow_path), field_names=["id", "name"], allow_formulas=True)
        sink2.set_row_ids([1])
        sink2.write_column("id", {1: 1})
        sink2.write_column("name", {1: "=1+1"})
        sink2.close()

    escaped_value, escaped_type = _read_cell_value_and_type(escape_path, "B2")
    assert escaped_type != "f"
    assert escaped_value == "'=1+1"

    allowed_value, allowed_type = _read_cell_value_and_type(allow_path, "B2")
    assert allowed_type == "f"
    assert allowed_value == "=1+1"


def test_excel_sink_write_lock_removes_lock_file(tmp_path: Path) -> None:
    output_path = tmp_path / "locked.xlsx"
    lock_path = Path(str(output_path) + ".scalim.lock")

    sink = excel_mod.ExcelSink(str(output_path), ["id"], write_lock=True)
    sink.write_row({"id": 1})
    sink.close()

    assert output_path.exists()
    assert lock_path.exists() is False


def test_excel_sink_write_lock_conflict_fails_fast(tmp_path: Path) -> None:
    output_path = tmp_path / "locked_conflict.xlsx"
    lock_path = Path(str(output_path) + ".scalim.lock")
    lock_path.write_text("held", encoding="utf-8")

    sink = excel_mod.ExcelSink(str(output_path), ["id"], write_lock=True)
    sink.write_row({"id": 1})
    with pytest.raises(RuntimeError, match="Output path is locked"):
        sink.close()
    # close 失败后也要尽力关闭 `write_only worksheet`,避免 openpyxl 生成器在 GC 时抛出 unraisable warning.
    assert getattr(sink._worksheet, "closed", False) is True
    # 再次 close: 仍然会因锁冲突失败,但应为幂等清理(避免二次 close 触发 `write_only worksheet` 的异常分支).
    with pytest.raises(RuntimeError, match="Output path is locked"):
        sink.close()

    assert output_path.exists() is False
    assert lock_path.exists() is True


def test_excel_workbook_sink_write_lock_removes_lock_file(tmp_path: Path) -> None:
    output_path = tmp_path / "wb_lock.xlsx"
    lock_path = Path(str(output_path) + ".scalim.lock")

    wb = excel_mod.ExcelWorkbookSink(str(output_path), write_lock=True)
    sheet = wb.create_sheet_row_sink("Sheet1", field_names=["id"], include_header=True)
    sheet.write_row({"id": 1})
    wb.close()

    assert output_path.exists()
    assert lock_path.exists() is False


def test_excel_workbook_sink_write_lock_conflict_closes_write_only_sheets(tmp_path: Path) -> None:
    output_path = tmp_path / "wb_locked_conflict.xlsx"
    lock_path = Path(str(output_path) + ".scalim.lock")
    lock_path.write_text("held", encoding="utf-8")

    wb = excel_mod.ExcelWorkbookSink(str(output_path), write_lock=True)
    sheet = wb.create_sheet_row_sink("Sheet1", field_names=["id"], include_header=True)
    sheet.write_row({"id": 1})
    with pytest.raises(RuntimeError, match="Output path is locked"):
        wb.close()

    # close 失败后也要尽力关闭 `write_only worksheet`,避免 openpyxl 生成器在 GC 时抛出 unraisable warning.
    assert all(getattr(ws, "closed", False) is True for ws in getattr(wb._workbook, "worksheets", []))


def test_column_excel_sink_write_lock_removes_lock_file(tmp_path: Path) -> None:
    output_path = tmp_path / "col_lock.xlsx"
    lock_path = Path(str(output_path) + ".scalim.lock")

    sink = excel_mod.ColumnExcelSink(str(output_path), ["id"], write_lock=True)
    sink.set_row_ids([1])
    sink.write_column("id", {1: 1})
    sink.close()

    assert output_path.exists()
    assert lock_path.exists() is False


def test_excel_workbook_sink_close_exception_logs_unlink_failure(tmp_path: Path, monkeypatch, caplog) -> None:
    caplog.set_level(logging.WARNING, logger=excel_mod.__name__)
    monkeypatch.setattr(excel_mod, "Workbook", _FailingSaveWorkbook)

    output_path = tmp_path / "exc_wb.xlsx"
    wb = excel_mod.ExcelWorkbookSink(str(output_path))
    _ = wb.create_sheet_row_sink("Sheet1", field_names=["id"])

    original_unlink = excel_mod.Path.unlink

    def _failing_unlink(path, *args, **kwargs):  # type: ignore[no-untyped-def]
        if str(path).endswith(".xlsx.tmp"):
            raise OSError("simulated unlink failure")
        return original_unlink(path, *args, **kwargs)

    monkeypatch.setattr(excel_mod.Path, "unlink", _failing_unlink)

    with pytest.raises(OSError, match="simulated save failure"):
        wb.close()

    assert any(excel_mod.EXCEL_WORKBOOK_SINK_SAVE_FAILED in record.getMessage() for record in caplog.records)
    assert any(excel_mod.EXCEL_WORKBOOK_SINK_REMOVE_TEMP_FILE_FAILED in record.getMessage() for record in caplog.records)
    for temp_file in tmp_path.glob("*.xlsx.tmp"):
        os.unlink(temp_file)


@pytest.mark.parametrize(
    "sink_cls,workbook_cls,filename",
    [
        (excel_mod.ExcelSink, _FailingSaveWorkbook, "exc_unlink.xlsx"),
        (excel_mod.ColumnExcelSink, _FailingSaveWorkbookNonWriteOnly, "exc_cols_unlink.xlsx"),
    ],
    ids=["row-sink", "column-sink"],
)
def test_excel_sink_close_exception_logs_unlink_failure(
    tmp_path: Path,
    monkeypatch,
    caplog,
    sink_cls,
    workbook_cls,
    filename: str,
) -> None:
    caplog.set_level(logging.WARNING, logger=excel_mod.__name__)
    monkeypatch.setattr(excel_mod, "Workbook", workbook_cls)
    sink = _init_sink_for_failure(tmp_path, sink_cls, filename)

    original_unlink = excel_mod.Path.unlink

    def _failing_unlink(path, *args, **kwargs):  # type: ignore[no-untyped-def]
        if str(path).endswith(".xlsx.tmp"):
            raise OSError("simulated unlink failure")
        return original_unlink(path, *args, **kwargs)

    monkeypatch.setattr(excel_mod.Path, "unlink", _failing_unlink)

    with pytest.raises(OSError, match="simulated save failure"):
        sink.close()

    expected = (
        excel_mod.EXCEL_SINK_REMOVE_TEMP_FILE_FAILED
        if sink_cls is excel_mod.ExcelSink
        else excel_mod.COLUMN_EXCEL_SINK_REMOVE_TEMP_FILE_FAILED
    )
    assert any(expected in record.getMessage() for record in caplog.records)
    for temp_file in tmp_path.glob("*.xlsx.tmp"):
        os.unlink(temp_file)


def test_excel_write_lock_release_ignores_missing_lock_file(tmp_path: Path) -> None:
    lock_path = tmp_path / "missing.scalim.lock"
    excel_mod._release_write_lock(lock_path)


def test_excel_write_lock_release_logs_warning_on_oserror(tmp_path: Path, monkeypatch, caplog) -> None:
    lock_path = tmp_path / "held.scalim.lock"
    lock_path.write_text("held", encoding="utf-8")

    caplog.set_level(logging.WARNING, logger=excel_mod.__name__)

    original_unlink = excel_mod.Path.unlink

    def _failing_unlink(path, *args, **kwargs):  # type: ignore[no-untyped-def]
        if str(path) == str(lock_path):
            raise OSError("simulated unlink failure")
        return original_unlink(path, *args, **kwargs)

    monkeypatch.setattr(excel_mod.Path, "unlink", _failing_unlink)

    excel_mod._release_write_lock(lock_path)
    assert any("删除输出锁文件失败" in record.getMessage() for record in caplog.records)


def test_excel_formula_escape_skips_already_escaped_value() -> None:
    assert excel_mod.escape_excel_formula("'=1+1", allow_formulas=False) == "'=1+1"


def test_excel_best_effort_close_write_only_workbook_worksheets_ignores_typeerror() -> None:
    workbook = type("_Workbook", (), {"worksheets": 1})()
    excel_mod._best_effort_close_write_only_workbook_worksheets(workbook)
