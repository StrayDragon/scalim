import logging
import hashlib
from pathlib import Path

import pytest

from scalim.execution import output_composition as output_comp_mod
from scalim.execution.output_composition import (
    AggMetricSpec,
    AuditSheetSpec,
    DerivedGroupBySpec,
    DerivedOutputTargetSpec,
    MetaSheetSpec,
    OutputCompositionSpec,
    OutputTargetSpec,
    RankFieldSpec,
    build_output_composition,
)
from scalim.execution.run_ir import ExecutionRequest, ExportLayout, OutputSpec, run_ir
from scalim.sinks import ExcelWorkbookSink
from scalim.sinks import InMemoryColumnSink, InMemoryRowSink
from tests.cases.minimal_ir import build_minimal_ir_case


def _read_workbook_sheet_names(path: Path) -> "list[str]":
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _read_sheet_rows(path: Path, sheet_name: str) -> "list[list[object]]":
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def test_excel_workbook_sink_multi_sheet_and_conflict(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    wb = ExcelWorkbookSink(str(out))
    s1 = wb.create_sheet_row_sink("Detail", field_names=["id"], header_names=["ID"], include_header=True)
    s1.write_row({"id": 1})

    s2 = wb.create_sheet_row_sink("Summary", field_names=["k", "v"], header_names=["K", "V"], include_header=True)
    s2.write_row({"k": "a", "v": 1})

    with pytest.raises(ValueError, match="Duplicate excel sheet name"):
        _ = wb.create_sheet_row_sink("Detail", field_names=["id"], header_names=["ID"], include_header=True)

    wb.close()
    assert out.exists()
    assert _read_workbook_sheet_names(out) == ["Detail", "Summary"]


def test_run_ir_output_composition_workbook_detail_and_summary(tmp_path: Path) -> None:
    case = build_minimal_ir_case()
    rows = case.main_rows()

    out = tmp_path / "report.xlsx"

    detail = OutputTargetSpec(
        target_id="detail",
        layout=ExportLayout(field_ids=("order_id", "order_source", "amount", "cost", "profit"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Detail"),
        is_primary=True,
    )

    derived = DerivedOutputTargetSpec(
        target_id="summary_by_source",
        derived=DerivedGroupBySpec(
            group_by=("order_source",),
            metrics=(
                AggMetricSpec(out_field_id="order_cnt", op="count", field_id="order_id"),
                AggMetricSpec(out_field_id="sum_amount", op="sum", field_id="amount"),
                AggMetricSpec(out_field_id="sum_profit", op="sum", field_id="profit"),
            ),
            max_groups=100,
        ),
        output_layout=ExportLayout(field_ids=("order_source", "order_cnt", "sum_amount", "sum_profit"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Summary"),
    )

    meta = MetaSheetSpec(
        target_id="meta",
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True),
        sheet_name="Meta",
    )
    audit = AuditSheetSpec(
        target_id="audit",
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True),
        sheet_name="Audit",
    )

    spec = OutputCompositionSpec(
        targets=(detail,),
        derived_targets=(derived,),
        meta_sheet=meta,
        audit_sheet=audit,
        failure_policy="all_fail",
    )

    request = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(path=None),
        sink=None,
        output_composition=spec,
        parallel_mode="seq",
    )

    result = run_ir(case.demand, request)
    assert out.exists()
    assert result.outputs is not None
    assert result.outputs["detail"] == str(out)
    assert result.outputs["summary_by_source"] == str(out)
    assert result.total_rows == len(rows)

    sheet_names = _read_workbook_sheet_names(out)
    assert sheet_names == ["Detail", "Summary", "Meta", "Audit"]

    detail_rows = _read_sheet_rows(out, "Detail")
    # header + rows
    assert detail_rows[0][:2] == ["order_id", "order_source"]
    assert len(detail_rows) == len(rows) + 1

    summary_rows = _read_sheet_rows(out, "Summary")
    assert summary_rows[0] == ["order_source", "order_cnt", "sum_amount", "sum_profit"]
    assert len(summary_rows) >= 2

    meta_rows = _read_sheet_rows(out, "Meta")
    meta_kv = {r[0]: r[1] for r in meta_rows[1:]}
    fp = meta_kv.get("derived.summary_by_source.fingerprint")
    assert isinstance(fp, str)
    assert len(fp) == 40


def test_run_ir_output_composition_can_tee_to_row_sink(tmp_path: Path) -> None:
    case = build_minimal_ir_case()
    rows = case.main_rows()

    out = tmp_path / "report.xlsx"

    detail = OutputTargetSpec(
        target_id="detail",
        layout=ExportLayout(field_ids=("order_id", "order_source", "amount", "cost", "profit"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Detail"),
        is_primary=True,
    )
    spec = OutputCompositionSpec(targets=(detail,), derived_targets=(), meta_sheet=None, audit_sheet=None, failure_policy="all_fail")

    sink = InMemoryRowSink()
    request = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(path=None),
        sink=sink,
        output_composition=spec,
        parallel_mode="seq",
    )

    result = run_ir(case.demand, request)
    captured = sink.get_data()
    assert captured
    assert len(captured) == result.total_rows == len(rows)
    assert "order_id" in captured[0]


def test_run_ir_output_composition_rejects_column_sink_for_tee(tmp_path: Path) -> None:
    case = build_minimal_ir_case()

    out = tmp_path / "report.xlsx"
    detail = OutputTargetSpec(
        target_id="detail",
        layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Detail"),
        is_primary=True,
    )
    spec = OutputCompositionSpec(targets=(detail,), derived_targets=(), meta_sheet=None, audit_sheet=None, failure_policy="all_fail")

    request = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(path=None),
        sink=InMemoryColumnSink(["order_id"]),
        output_composition=spec,
        parallel_mode="seq",
    )

    with pytest.raises(ValueError, match=r"output_composition only supports streaming row sinks"):
        _ = run_ir(case.demand, request)


def test_output_composition_primary_only_disables_failed_derived(tmp_path: Path) -> None:
    case = build_minimal_ir_case()
    rows = case.main_rows()
    out = tmp_path / "report.xlsx"

    detail = OutputTargetSpec(
        target_id="detail",
        layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Detail"),
        is_primary=True,
    )

    derived = DerivedOutputTargetSpec(
        target_id="summary_overflow",
        derived=DerivedGroupBySpec(
            group_by=("order_id",),
            metrics=(AggMetricSpec(out_field_id="cnt", op="count", field_id="order_id"),),
            max_groups=1,  # overflow when more than 1 distinct key
        ),
        output_layout=ExportLayout(field_ids=("order_id", "cnt"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Summary"),
    )

    spec = OutputCompositionSpec(
        targets=(detail,),
        derived_targets=(derived,),
        meta_sheet=MetaSheetSpec(target_id="meta", output=OutputSpec(format="excel", path=str(out)), sheet_name="Meta"),
        audit_sheet=AuditSheetSpec(target_id="audit", output=OutputSpec(format="excel", path=str(out)), sheet_name="Audit"),
        failure_policy="primary_only",
    )

    request = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(path=None),
        sink=None,
        output_composition=spec,
        parallel_mode="seq",
    )
    result = run_ir(case.demand, request)

    assert out.exists()
    assert result.output_target_stats is not None
    derived_stat = [s for s in result.output_target_stats if s.target_id == "summary_overflow"][0]
    assert derived_stat.disabled is True
    assert derived_stat.error_count >= 1


def test_ranked_summary_orders_and_adds_rank(tmp_path: Path) -> None:
    case = build_minimal_ir_case()
    out = tmp_path / "report.xlsx"

    detail = OutputTargetSpec(
        target_id="detail",
        layout=ExportLayout(field_ids=("order_id", "order_source", "amount"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Detail"),
        is_primary=True,
    )

    ranked = DerivedOutputTargetSpec(
        target_id="ranked_summary",
        derived=DerivedGroupBySpec(
            group_by=("order_source",),
            metrics=(AggMetricSpec(out_field_id="sum_amount", op="sum", field_id="amount"),),
            rank_fields=(RankFieldSpec(out_field_id="rank", kind="dense_rank", by="sum_amount", order="desc"),),
        ),
        output_layout=ExportLayout(field_ids=("order_source", "sum_amount", "rank"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Rank"),
    )

    spec = OutputCompositionSpec(targets=(detail,), derived_targets=(ranked,))
    request = ExecutionRequest(
        export_layout=ExportLayout(field_ids=("order_id",), header_names=None),
        output=OutputSpec(path=None),
        sink=None,
        output_composition=spec,
        parallel_mode="seq",
    )

    _ = run_ir(case.demand, request)

    rows = _read_sheet_rows(out, "Rank")
    assert rows[0] == ["order_source", "sum_amount", "rank"]
    assert rows[1][0] == "app"
    assert rows[1][2] == 1


def test_meta_and_audit_redact_error_message_by_default_and_allow_full(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"

    ok = OutputTargetSpec(
        target_id="ok",
        layout=ExportLayout(field_ids=("id",), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="OK"),
        is_primary=True,
    )
    fail = OutputTargetSpec(
        target_id="fail",
        layout=ExportLayout(field_ids=("id",), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="FAIL"),
        is_primary=False,
    )

    meta = MetaSheetSpec(target_id="meta", output=OutputSpec(format="excel", path=str(out)), sheet_name="Meta")
    audit = AuditSheetSpec(target_id="audit", output=OutputSpec(format="excel", path=str(out)), sheet_name="Audit")

    secret = "SECRET=token-123"
    secret_hash = hashlib.sha256(secret.encode("utf-8")).hexdigest()

    class FailSink:
        def write_row(self, row):  # type: ignore[no-untyped-def]
            raise RuntimeError(secret)

        def write_batch(self, rows):  # type: ignore[no-untyped-def]
            for row in rows:
                self.write_row(row)

        def close(self) -> None:
            return

    spec = OutputCompositionSpec(targets=(ok, fail), meta_sheet=meta, audit_sheet=audit, failure_policy="primary_only")
    plan = build_output_composition(
        spec=spec,
        demand_name="d",
        demand_main_source_id="s",
        demand_target_fields=["id"],
        demand_field_fingerprints=[],
    )
    router = plan.sink
    for route in router._routes:
        if route.target_id == "fail":
            route.sink = FailSink()  # type: ignore[assignment]

    router.write_row({"id": 1})
    router.close()

    meta_rows = _read_sheet_rows(out, "Meta")
    meta_kv = {r[0]: r[1] for r in meta_rows[1:]}
    assert meta_kv["output.fail.error_type"] == "RuntimeError"
    assert meta_kv["output.fail.error_message_hash"] == secret_hash
    assert secret not in str(meta_kv["output.fail.error_message"])
    assert secret_hash in str(meta_kv["output.fail.error_message"])

    audit_rows = _read_sheet_rows(out, "Audit")
    header = audit_rows[0]
    idx_target = header.index("target_id")
    idx_msg = header.index("error_message")
    fail_row = [r for r in audit_rows[1:] if r[idx_target] == "fail"][0]
    assert secret not in str(fail_row[idx_msg])
    assert secret_hash in str(fail_row[idx_msg])

    # allow full message
    out2 = tmp_path / "report_full.xlsx"
    spec2 = OutputCompositionSpec(
        targets=(
            OutputTargetSpec(
                target_id="ok",
                layout=ExportLayout(field_ids=("id",), header_names=None),
                output=OutputSpec(format="excel", path=str(out2), streaming=True, include_header=True, sheet_name="OK"),
                is_primary=True,
            ),
            OutputTargetSpec(
                target_id="fail",
                layout=ExportLayout(field_ids=("id",), header_names=None),
                output=OutputSpec(format="excel", path=str(out2), streaming=True, include_header=True, sheet_name="FAIL"),
                is_primary=False,
            ),
        ),
        meta_sheet=MetaSheetSpec(target_id="meta", output=OutputSpec(format="excel", path=str(out2)), sheet_name="Meta"),
        audit_sheet=AuditSheetSpec(target_id="audit", output=OutputSpec(format="excel", path=str(out2)), sheet_name="Audit"),
        failure_policy="primary_only",
        include_full_error_message=True,
    )
    plan2 = build_output_composition(
        spec=spec2,
        demand_name="d",
        demand_main_source_id="s",
        demand_target_fields=["id"],
        demand_field_fingerprints=[],
    )
    router2 = plan2.sink
    for route in router2._routes:
        if route.target_id == "fail":
            route.sink = FailSink()  # type: ignore[assignment]

    router2.write_row({"id": 1})
    router2.close()

    meta_rows2 = _read_sheet_rows(out2, "Meta")
    meta_kv2 = {r[0]: r[1] for r in meta_rows2[1:]}
    assert secret in str(meta_kv2["output.fail.error_message"])


def test_derived_output_max_groups_zero_emits_warning(tmp_path: Path, caplog) -> None:
    caplog.set_level(logging.WARNING, logger="scalim.execution.output_composition")
    out = tmp_path / "report.xlsx"

    detail = OutputTargetSpec(
        target_id="detail",
        layout=ExportLayout(field_ids=("order_id", "order_source"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Detail"),
        is_primary=True,
    )
    derived = DerivedOutputTargetSpec(
        target_id="summary",
        derived=DerivedGroupBySpec(
            group_by=("order_source",),
            metrics=(AggMetricSpec(out_field_id="cnt", op="count", field_id="order_id"),),
            max_groups=0,
        ),
        output_layout=ExportLayout(field_ids=("order_source", "cnt"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Summary"),
    )

    spec = OutputCompositionSpec(targets=(detail,), derived_targets=(derived,))
    plan = build_output_composition(
        spec=spec,
        demand_name="d",
        demand_main_source_id="s",
        demand_target_fields=["order_id"],
        demand_field_fingerprints=[],
    )
    plan.sink.close()

    assert any("max_groups=0" in record.getMessage() for record in caplog.records)


def test_audit_includes_derived_truncate_events_and_fingerprint(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"

    derived = DerivedOutputTargetSpec(
        target_id="summary",
        derived=DerivedGroupBySpec(
            group_by=("g",),
            metrics=(AggMetricSpec(out_field_id="u", op="count_distinct", field_id="user_id"),),
            max_groups=10,
            max_distinct=1,
            distinct_on_overflow="truncate",
        ),
        output_layout=ExportLayout(field_ids=("g", "u"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Summary"),
        is_primary=True,
    )

    spec = OutputCompositionSpec(
        targets=(),
        derived_targets=(derived,),
        meta_sheet=MetaSheetSpec(target_id="meta", output=OutputSpec(format="excel", path=str(out)), sheet_name="Meta"),
        audit_sheet=AuditSheetSpec(target_id="audit", output=OutputSpec(format="excel", path=str(out)), sheet_name="Audit"),
    )
    plan = build_output_composition(
        spec=spec,
        demand_name="d",
        demand_main_source_id="s",
        demand_target_fields=["g"],
        demand_field_fingerprints=[],
        run_parallel_mode="seq",
    )
    router = plan.sink
    router.write_row({"g": "x", "user_id": "u1"})
    router.write_row({"g": "x", "user_id": "u2"})
    router.close()

    meta_rows = _read_sheet_rows(out, "Meta")
    meta_kv = {r[0]: r[1] for r in meta_rows[1:]}
    fp = meta_kv.get("derived.summary.fingerprint")
    assert isinstance(fp, str)
    assert len(fp) == 40

    audit_rows = _read_sheet_rows(out, "Audit")
    header = audit_rows[0]
    idx_target = header.index("target_id")
    idx_event = header.index("event_type")
    idx_fp = header.index("fingerprint")
    idx_msg = header.index("error_message")
    idx_hash = header.index("error_message_hash")

    row = [r for r in audit_rows[1:] if r[idx_target] == "summary" and r[idx_event] == "distinct_truncated"][0]
    assert row[idx_fp] == fp
    assert "sha256=" in str(row[idx_msg])
    assert len(str(row[idx_hash])) == 64


def test_dedup_by_first_fail_fast_under_adaptive(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    derived = DerivedOutputTargetSpec(
        target_id="summary",
        derived=output_comp_mod.DerivedDedupByGroupBySpec(
            dedup_by=output_comp_mod.DedupBySpec(key_fields=("k",), on_conflict="first"),
            group_by=DerivedGroupBySpec(
                group_by=("g",),
                metrics=(AggMetricSpec(out_field_id="cnt", op="count"),),
            ),
        ),
        output_layout=ExportLayout(field_ids=("g", "cnt"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Summary"),
        is_primary=True,
    )
    spec = OutputCompositionSpec(derived_targets=(derived,))
    with pytest.raises(ValueError, match="parallel_mode='adaptive'"):
        _ = build_output_composition(
            spec=spec,
            demand_name="d",
            demand_main_source_id="s",
            demand_target_fields=["g"],
            demand_field_fingerprints=[],
            run_parallel_mode="adaptive",
        )


def test_two_stage_group_by_stage2_requires_fields_produced_by_stage1(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    derived = DerivedOutputTargetSpec(
        target_id="summary",
        derived=output_comp_mod.TwoStageGroupBySpec(
            stage1=DerivedGroupBySpec(
                group_by=("g",),
                metrics=(AggMetricSpec(out_field_id="cnt", op="count", field_id="id"),),
            ),
            stage2=DerivedGroupBySpec(
                group_by=("g",),
                metrics=(AggMetricSpec(out_field_id="sum_x", op="sum", field_id="missing_stage1_field"),),
            ),
        ),
        output_layout=ExportLayout(field_ids=("g", "sum_x"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Summary"),
        is_primary=True,
    )
    spec = OutputCompositionSpec(derived_targets=(derived,))
    with pytest.raises(ValueError, match="stage2 requires fields not produced by stage1"):
        _ = build_output_composition(
            spec=spec,
            demand_name="d",
            demand_main_source_id="s",
            demand_target_fields=["g"],
            demand_field_fingerprints=[],
            run_parallel_mode="seq",
        )


def test_truncate_text_returns_empty_when_max_chars_nonpositive() -> None:
    assert output_comp_mod._truncate_text("abc", max_chars=0) == ""
    assert output_comp_mod._truncate_text("abc", max_chars=-1) == ""


def test_truncate_text_truncates_long_text() -> None:
    assert output_comp_mod._truncate_text("abcd", max_chars=3) == "abc…(truncated)"


def test_i_derived_aggregation_spec_base_methods_raise() -> None:
    class _DummySpec(output_comp_mod.IDerivedAggregationSpec):
        def required_fields(self):  # type: ignore[no-untyped-def]
            return super(_DummySpec, self).required_fields()

        def fingerprint_parts(self):  # type: ignore[no-untyped-def]
            return super(_DummySpec, self).fingerprint_parts()

        def validate_parallel_mode(self, parallel_mode: str) -> None:
            return super(_DummySpec, self).validate_parallel_mode(parallel_mode)

        def build_aggregator(self):  # type: ignore[no-untyped-def]
            return super(_DummySpec, self).build_aggregator()

    dummy = _DummySpec()
    with pytest.raises(NotImplementedError):
        dummy.required_fields()
    with pytest.raises(NotImplementedError):
        dummy.fingerprint_parts()
    with pytest.raises(NotImplementedError):
        dummy.validate_parallel_mode("seq")
    with pytest.raises(NotImplementedError):
        dummy.build_aggregator()


def test_derived_group_by_spec_rejects_invalid_distinct_on_overflow() -> None:
    spec = DerivedGroupBySpec(
        group_by=("g",),
        metrics=(AggMetricSpec(out_field_id="cnt", op="count"),),
        distinct_on_overflow="oops",
    )
    with pytest.raises(ValueError, match="Unsupported distinct_on_overflow"):
        spec.validate_parallel_mode("seq")


def test_dedup_by_spec_fingerprint_parts_and_validate_errors() -> None:
    spec = output_comp_mod.DedupBySpec(
        key_fields=("k1", "k2"),
        on_conflict="FIRST",
        max_distinct=3,
        on_overflow="ERROR",
    )
    parts = spec.fingerprint_parts()
    assert parts[0] == "kind=dedup_by"
    assert "key_fields=k1,k2" in parts
    assert "on_conflict=first" in parts
    assert "max_distinct=3" in parts
    assert "on_overflow=error" in parts

    with pytest.raises(ValueError, match="Unsupported dedup_by.on_conflict"):
        output_comp_mod.DedupBySpec(key_fields=("k",), on_conflict="oops").validate_parallel_mode("seq")
    with pytest.raises(ValueError, match="Unsupported dedup_by.on_overflow"):
        output_comp_mod.DedupBySpec(key_fields=("k",), on_overflow="oops").validate_parallel_mode("seq")


def test_output_composition_warns_on_count_distinct_without_max_distinct(tmp_path: Path, caplog) -> None:
    caplog.set_level(logging.WARNING, logger="scalim.execution.output_composition")
    out = tmp_path / "report.xlsx"

    derived = DerivedOutputTargetSpec(
        target_id="summary",
        derived=DerivedGroupBySpec(
            group_by=("g",),
            metrics=(AggMetricSpec(out_field_id="u", op="count_distinct", field_id="user_id"),),
            max_groups=1,
            max_distinct=0,
        ),
        output_layout=ExportLayout(field_ids=("g", "u"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Summary"),
        is_primary=True,
    )
    spec = OutputCompositionSpec(derived_targets=(derived,))
    plan = build_output_composition(
        spec=spec,
        demand_name="d",
        demand_main_source_id="s",
        demand_target_fields=["g"],
        demand_field_fingerprints=[],
        run_parallel_mode="seq",
    )
    plan.sink.close()

    assert any("max_distinct=0" in record.getMessage() and "count_distinct" in record.getMessage() for record in caplog.records)


def test_output_composition_warns_on_dedup_without_max_distinct_and_spec_methods(tmp_path: Path, caplog) -> None:
    caplog.set_level(logging.WARNING, logger="scalim.execution.output_composition")
    out = tmp_path / "report.xlsx"

    derived_spec = output_comp_mod.DerivedDedupByGroupBySpec(
        dedup_by=output_comp_mod.DedupBySpec(
            key_fields=("k",),
            on_conflict="first",
            max_distinct=0,
            on_overflow="error",
        ),
        group_by=DerivedGroupBySpec(
            group_by=("g",),
            metrics=(AggMetricSpec(out_field_id="cnt", op="count"),),
            max_groups=1,
        ),
    )
    assert derived_spec.required_fields() == ("k", "g")

    fp_parts = derived_spec.fingerprint_parts()
    assert fp_parts[0] == "kind=dedup_by+group_by"

    derived_spec.validate_parallel_mode("seq")

    agg = derived_spec.build_aggregator()
    agg.accumulate({"k": "a", "g": "x"})
    agg.accumulate({"k": "a", "g": "x"})
    assert agg.finalize_rows() == [{"g": "x", "cnt": 1}]

    derived = DerivedOutputTargetSpec(
        target_id="summary",
        derived=derived_spec,
        output_layout=ExportLayout(field_ids=("g", "cnt"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Summary"),
        is_primary=True,
    )
    spec = OutputCompositionSpec(derived_targets=(derived,))
    plan = build_output_composition(
        spec=spec,
        demand_name="d",
        demand_main_source_id="s",
        demand_target_fields=["g"],
        demand_field_fingerprints=[],
        run_parallel_mode="seq",
    )
    plan.sink.close()

    assert any("dedup_by.max_distinct=0" in record.getMessage() for record in caplog.records)


def test_two_stage_group_by_spec_methods_and_build(tmp_path: Path) -> None:
    stage1 = DerivedGroupBySpec(
        group_by=("g", "u"),
        metrics=(AggMetricSpec(out_field_id="cnt", op="count", field_id="id"),),
        max_groups=10,
    )
    stage2 = DerivedGroupBySpec(
        group_by=("g",),
        metrics=(AggMetricSpec(out_field_id="sum_cnt", op="sum", field_id="cnt"),),
        max_groups=10,
    )

    derived = output_comp_mod.TwoStageGroupBySpec(stage1=stage1, stage2=stage2)
    assert derived.required_fields() == ("g", "u", "id")
    fp_parts = derived.fingerprint_parts()
    assert fp_parts[0] == "kind=two_stage_group_by"
    derived.validate_parallel_mode("seq")

    agg = derived.build_aggregator()
    agg.accumulate({"g": "x", "u": "u1", "id": 1})
    agg.accumulate({"g": "x", "u": "u1", "id": 2})
    agg.accumulate({"g": "x", "u": "u2", "id": 3})
    assert agg.finalize_rows() == [{"g": "x", "sum_cnt": 3}]

    bad = output_comp_mod.TwoStageGroupBySpec(
        stage1=DerivedGroupBySpec(
            group_by=("g",),
            metrics=(AggMetricSpec(out_field_id="cnt", op="count", field_id="id"),),
            rank_fields=(RankFieldSpec(out_field_id="rank", kind="dense_rank", by="cnt", order="desc"),),
        ),
        stage2=stage2,
    )
    with pytest.raises(ValueError, match="two_stage_group_by does not support rank/post fields"):
        bad.validate_parallel_mode("seq")

    out = tmp_path / "report.xlsx"
    derived_target = DerivedOutputTargetSpec(
        target_id="summary",
        derived=derived,
        output_layout=ExportLayout(field_ids=("g", "sum_cnt"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Summary"),
        is_primary=True,
    )
    spec = OutputCompositionSpec(derived_targets=(derived_target,))
    plan = build_output_composition(
        spec=spec,
        demand_name="d",
        demand_main_source_id="s",
        demand_target_fields=["g"],
        demand_field_fingerprints=[],
        run_parallel_mode="seq",
    )
    plan.sink.write_row({"g": "x", "u": "u1", "id": 1})
    plan.sink.close()


def test_audit_includes_full_derived_event_message_when_enabled(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    derived = DerivedOutputTargetSpec(
        target_id="summary",
        derived=DerivedGroupBySpec(
            group_by=("g",),
            metrics=(AggMetricSpec(out_field_id="u", op="count_distinct", field_id="user_id"),),
            max_groups=10,
            max_distinct=1,
            distinct_on_overflow="truncate",
        ),
        output_layout=ExportLayout(field_ids=("g", "u"), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Summary"),
        is_primary=True,
    )
    spec = OutputCompositionSpec(
        derived_targets=(derived,),
        meta_sheet=MetaSheetSpec(target_id="meta", output=OutputSpec(format="excel", path=str(out)), sheet_name="Meta"),
        audit_sheet=AuditSheetSpec(target_id="audit", output=OutputSpec(format="excel", path=str(out)), sheet_name="Audit"),
        include_full_error_message=True,
    )
    plan = build_output_composition(
        spec=spec,
        demand_name="d",
        demand_main_source_id="s",
        demand_target_fields=["g"],
        demand_field_fingerprints=[],
        run_parallel_mode="seq",
    )
    router = plan.sink
    router.write_row({"g": "x", "user_id": "u1"})
    router.write_row({"g": "x", "user_id": "u2"})
    router.close()

    audit_rows = _read_sheet_rows(out, "Audit")
    header = audit_rows[0]
    idx_target = header.index("target_id")
    idx_event = header.index("event_type")
    idx_msg = header.index("error_message")

    row = [r for r in audit_rows[1:] if r[idx_target] == "summary" and r[idx_event] == "distinct_truncated"][0]
    assert "count_distinct truncated" in str(row[idx_msg])
    assert "sha256=" not in str(row[idx_msg])


def test_derived_dedup_by_group_by_spec_required_fields_dedupes_overlap() -> None:
    spec = output_comp_mod.DerivedDedupByGroupBySpec(
        dedup_by=output_comp_mod.DedupBySpec(key_fields=("g",), on_conflict="first"),
        group_by=DerivedGroupBySpec(
            group_by=("g",),
            metrics=(AggMetricSpec(out_field_id="cnt", op="count", field_id="id"),),
            max_groups=1,
        ),
    )
    assert spec.required_fields() == ("g", "id")


def test_collect_specs_for_derived_warnings_unknown_spec_returns_empty(tmp_path: Path) -> None:
    from scalim.execution.derived_outputs import AggregatorDiagnostics, IRowAggregator

    class _PassThroughAgg(IRowAggregator):
        def __init__(self) -> None:
            self._rows = []

        def required_fields(self):  # type: ignore[no-untyped-def]
            return ("x",)

        def accumulate(self, row):  # type: ignore[no-untyped-def]
            self._rows.append({"x": row.get("x")})

        def finalize_rows(self):  # type: ignore[no-untyped-def]
            return list(self._rows)

        def diagnostics(self) -> AggregatorDiagnostics:
            return AggregatorDiagnostics(meta={}, audit_events=[])

    class _CustomSpec(output_comp_mod.IDerivedAggregationSpec):
        def required_fields(self):  # type: ignore[no-untyped-def]
            return ("x",)

        def fingerprint_parts(self):  # type: ignore[no-untyped-def]
            return ("kind=custom",)

        def validate_parallel_mode(self, parallel_mode: str) -> None:
            return

        def build_aggregator(self, *, key_normalization: str = "raw") -> IRowAggregator:
            _ = key_normalization
            return _PassThroughAgg()

    out = tmp_path / "report.xlsx"
    derived = DerivedOutputTargetSpec(
        target_id="custom",
        derived=_CustomSpec(),
        output_layout=ExportLayout(field_ids=("x",), header_names=None),
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name="Custom"),
        is_primary=True,
    )
    spec = OutputCompositionSpec(derived_targets=(derived,))
    plan = build_output_composition(
        spec=spec,
        demand_name="d",
        demand_main_source_id="s",
        demand_target_fields=["x"],
        demand_field_fingerprints=[],
        run_parallel_mode="seq",
    )
    plan.sink.write_row({"x": 1})
    plan.sink.close()
