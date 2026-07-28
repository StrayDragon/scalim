"""c20/c999: unified books.xlsx; removed xlsx_file/xlsx_memory YAML aliases fail-fast."""

from __future__ import annotations

import warnings
from pathlib import Path
from typing import Any, Dict

import pytest

from scalim.dsl.yaml_dsl import (
    DemandRunOptions,
    DemandRunSecurityOptions,
    WorkflowRunOptions,
    run_workflow,
)
from scalim.dsl.yaml_dsl._internal.config_parsing.book_branch_parse import (
    MIGRATE_TO_XLSX_EMPTY_HINT,
    MIGRATE_TO_XLSX_PATH_HINT,
    REMOVED_XLSX_FILE_HINT,
    REMOVED_XLSX_MEMORY_HINT,
)
from scalim.dsl.yaml_dsl._internal.config_parsing.loader import YamlDemandLoader
from scalim.dsl.yaml_dsl._internal.config_parsing.validator import ConfigValidator
from scalim.dsl.yaml_dsl._internal.config_parsing.validators.issues import VALIDATION_SEVERITY_ERROR
from scalim.dsl.yaml_dsl.workflow import ScalimWorkflowConfigError, load_workflow_config_from_mapping
from scalim.execution import versioned_outputs
from scalim.vendor.yamlx import yaml


_ALLOWED = frozenset(["tests.fixtures.workflow_loaders", "scalim.workflow.loaders"])


def _run_options() -> WorkflowRunOptions:
    return WorkflowRunOptions(
        demand=DemandRunOptions(security=DemandRunSecurityOptions(allowed_modules=_ALLOWED)),
    )


def _latest_book_path(out_root: Path, *, book_id: str) -> Path:
    latest = versioned_outputs.read_latest(out_root)
    version_id = latest.get("version_id")
    assert isinstance(version_id, str)
    return out_root / "versions" / version_id / "books" / "{}.xlsx".format(str(book_id))


def test_parse_xlsx_empty_and_path_emit_no_deprecation() -> None:
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        empty = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml"}],
                    "resources": {"books": {"scratch": {"xlsx": {}}}},
                }
            }
        )
        with_path = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml"}],
                    "resources": {"books": {"report": {"xlsx": {"path": "./out"}}}},
                }
            }
        )
    assert empty.resources.books["scratch"].path is None
    assert with_path.resources.books["report"].path is not None
    assert with_path.resources.books["report"].path == "./out"
    assert not [w for w in caught if issubclass(w.category, DeprecationWarning)]


def test_parse_removed_aliases_fail_fast() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match=REMOVED_XLSX_MEMORY_HINT) as mem_exc:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml"}],
                    "resources": {"books": {"scratch": {"xlsx_memory": {}}}},
                }
            }
        )
    assert MIGRATE_TO_XLSX_EMPTY_HINT in str(mem_exc.value)

    with pytest.raises(ScalimWorkflowConfigError, match=REMOVED_XLSX_FILE_HINT) as file_exc:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml"}],
                    "resources": {"books": {"report": {"xlsx_file": {"path": "./out"}}}},
                }
            }
        )
    assert MIGRATE_TO_XLSX_PATH_HINT in str(file_exc.value)


def test_parse_xlsx_memory_export_alias_fail_fast() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match=r"xlsx_memory with export_xlsx was removed") as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml"}],
                    "resources": {
                        "books": {
                            "report": {"xlsx_memory": {"export_xlsx": {"path": "./out"}}},
                        }
                    },
                }
            }
        )
    assert MIGRATE_TO_XLSX_PATH_HINT in str(excinfo.value)


def test_xlsx_export_xlsx_is_error_not_warning() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match=r"xlsx\.export_xlsx is not allowed"):
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml"}],
                    "resources": {
                        "books": {
                            "bad": {"xlsx": {"export_xlsx": {"path": "./out"}}},
                        }
                    },
                }
            }
        )


def test_xlsx_write_defaults_and_budget_fail_fast() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match=r"xlsx\.write_defaults was removed"):
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml"}],
                    "resources": {"books": {"report": {"xlsx": {"path": "./out", "write_defaults": {"mode": "sheet"}}}}},
                }
            }
        )
    with pytest.raises(ScalimWorkflowConfigError, match=r"xlsx\.budget was removed"):
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml"}],
                    "resources": {"books": {"scratch": {"xlsx": {"budget": {"max_sheets": 1, "max_total_cells": 1}}}}},
                }
            }
        )


def test_demand_validator_accepts_xlsx_and_rejects_aliases() -> None:
    validator = ConfigValidator()
    ok = validator.validate_report({"resources": {"books": {"scratch": {"xlsx": {}}}}, "name": "n", "main_source": {}})
    assert not any("must declare exactly one" in i.message for i in ok.issues)

    report = validator.validate_report(
        {"resources": {"books": {"report": {"xlsx_file": {"path": "./out"}}}}, "name": "n", "main_source": {}}
    )
    error_msgs = [i.message for i in report.issues if i.severity == VALIDATION_SEVERITY_ERROR and REMOVED_XLSX_FILE_HINT in i.message]
    assert error_msgs
    assert any(MIGRATE_TO_XLSX_PATH_HINT in m for m in error_msgs)

    mem_export = validator.validate_report(
        {
            "resources": {"books": {"report": {"xlsx_memory": {"export_xlsx": {"path": "./out"}}}}},
            "name": "n",
            "main_source": {},
        }
    )
    assert any(
        "xlsx_memory with export_xlsx was removed" in i.message for i in mem_export.issues if i.severity == VALIDATION_SEVERITY_ERROR
    )

    xlsx_bad = validator.validate_report(
        {
            "resources": {
                "books": {
                    "bad_type": {"xlsx": []},
                    "bad_export": {"xlsx": {"export_xlsx": {"path": "./out"}}},
                    "bad_write": {"xlsx": {"write_defaults": {"mode": "sheet"}}},
                    "bad_budget": {"xlsx": {"budget": {"max_sheets": 1, "max_total_cells": 1}}},
                    "bad_empty_path": {"xlsx": {"path": ""}},
                    "bad_file_path": {"xlsx": {"path": "out.xlsx"}},
                    "bad_formulas": {"xlsx": {"allow_formulas": "yes"}},
                    "bad_unknown": {"xlsx": {"nope": 1}},
                }
            },
            "name": "n",
            "main_source": {},
        }
    )
    msgs = " ".join(i.message for i in xlsx_bad.issues)
    assert "xlsx must be an object" in msgs
    assert "export_xlsx is not allowed" in msgs
    assert "write_defaults was removed" in msgs
    assert "budget was removed" in msgs
    assert "non-empty output root" in msgs
    assert "output root directory" in msgs

    loader = YamlDemandLoader()
    with pytest.raises(Exception, match=r"xlsx\.export_xlsx is not allowed"):
        _ = loader._parse_book_config(  # noqa: SLF001
            {"xlsx": {"export_xlsx": {"path": "./out"}}},
            base_path="resources.books.bad",
        )
    with pytest.raises(Exception, match=r"xlsx must be a mapping"):
        _ = loader._parse_book_config({"xlsx": []}, base_path="resources.books.bad")  # noqa: SLF001
    with pytest.raises(Exception, match=r"unknown keys"):
        _ = loader._parse_book_config({"xlsx": {"nope": 1}}, base_path="resources.books.bad")  # noqa: SLF001
    with pytest.raises(Exception, match=r"non-empty output root"):
        _ = loader._parse_book_config({"xlsx": {"path": ""}}, base_path="resources.books.bad")  # noqa: SLF001
    with pytest.raises(Exception, match=r"output root directory"):
        _ = loader._parse_book_config({"xlsx": {"path": "out.xlsx"}}, base_path="resources.books.bad")  # noqa: SLF001
    with pytest.raises(Exception, match=r"allow_formulas must be a bool"):
        _ = loader._parse_book_config({"xlsx": {"allow_formulas": "no"}}, base_path="resources.books.bad")  # noqa: SLF001
    with pytest.raises(Exception, match=r"xlsx_memory.*was removed"):
        _ = loader._parse_book_config(  # noqa: SLF001
            {"xlsx_memory": {"export_xlsx": []}},
            base_path="resources.books.bad",
        )


def test_demand_validator_strips_xlsx_budget_and_write_defaults() -> None:
    validator = ConfigValidator()
    issues: list = []
    next_config = validator._error_and_strip_removed_resources_write_budget_fields(  # noqa: SLF001
        {
            "resources": {
                "books": {
                    "a": {"xlsx": {"path": "./out", "budget": {"max_sheets": 1, "max_total_cells": 1}}},
                    "b": {"xlsx": {"write_defaults": {"mode": "sheet"}}},
                    "both": {
                        "xlsx": {
                            "path": "./out",
                            "budget": {"max_sheets": 1, "max_total_cells": 1},
                            "write_defaults": {"mode": "sheet"},
                        }
                    },
                    "top_then_xlsx_budget": {
                        "write_defaults": {"mode": "sheet"},
                        "xlsx": {"budget": {"max_sheets": 1, "max_total_cells": 1}},
                    },
                }
            }
        },
        issues,
    )
    assert any("xlsx.budget was removed" in i.message for i in issues)
    assert any("xlsx.write_defaults was removed" in i.message for i in issues)
    assert "budget" not in next_config["resources"]["books"]["a"]["xlsx"]
    assert "write_defaults" not in next_config["resources"]["books"]["b"]["xlsx"]
    assert "budget" not in next_config["resources"]["books"]["both"]["xlsx"]
    assert "write_defaults" not in next_config["resources"]["books"]["both"]["xlsx"]


def test_parse_xlsx_init_var_and_removed_alias_errors() -> None:
    loader = YamlDemandLoader()
    with pytest.raises(Exception, match=REMOVED_XLSX_FILE_HINT):
        _ = loader._parse_book_config(  # noqa: SLF001
            {"xlsx_file": {"path": "out.xlsx"}},
            base_path="resources.books.report",
        )
    with pytest.raises(Exception, match=REMOVED_XLSX_FILE_HINT):
        _ = loader._parse_book_config(  # noqa: SLF001
            {"xlsx_file": {"path": "./out", "allow_formulas": "no"}},
            base_path="resources.books.report",
        )

    validator = ConfigValidator()
    report = validator.validate_report(
        {
            "resources": {
                "books": {
                    "report": {"xlsx": {"path": {"$init_var": "out_root"}}},
                }
            },
            "name": "n",
            "main_source": {},
        }
    )
    assert not any(i.path.endswith("xlsx.path") and "error" in str(i.severity).lower() for i in report.issues)

    path_issues: list = []
    validator._validate_resource_output_paths(  # noqa: SLF001
        {
            "resources": {
                "books": {
                    "w": {"xlsx": {"write_defaults": {"mode": "sheet"}}},
                    "b": {"xlsx": {"budget": {"max_sheets": 1, "max_total_cells": 1}}},
                    "bad_init": {"xlsx": {"path": {"nope": 1}}},
                }
            }
        },
        path_issues,
    )
    assert any("write_defaults" in i.message for i in path_issues)
    assert any("budget" in i.message for i in path_issues)
    assert any("bad_init" in i.path for i in path_issues)


def _write_isomorphic_tree(root: Path, *, books: Dict[str, Any], out_rel: str) -> Path:
    (root / "stage_a.demand.yaml").write_text(
        """
name: stage_a
resources:
  books:
    scratch:
      xlsx: {}
main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {extract: id}
    value: {extract: value}
outputs:
  - name: metrics_a
    to: {book: scratch, sheet: metrics_a}
    fields: [id, value]
""".lstrip(),
        encoding="utf-8",
    )
    (root / "stage_b.demand.yaml").write_text(
        """
name: stage_b
resources:
  books:
    scratch:
      xlsx: {}
main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {extract: id}
    value: {extract: value}
outputs:
  - name: metrics_b
    to: {book: scratch, sheet: metrics_b}
    fields: [id, value]
""".lstrip(),
        encoding="utf-8",
    )
    (root / "summary.demand.yaml").write_text(
        """
name: summary
resources:
  books:
    scratch:
      xlsx: {}
    report:
      xlsx:
        path: __OUT__
main_source:
  source_id: main
  loader: "scalim.workflow.loaders:book_sheet_rows"
  params:
    ref: {node: stage_a, book: scratch, sheet: metrics_a}
  fields:
    id: {extract: id}
    value: {extract: value}
outputs:
  - name: final_sheet
    to: {book: report, sheet: final_sheet}
    fields: [id, value]
""".replace("__OUT__", out_rel).lstrip(),
        encoding="utf-8",
    )
    wf = {
        "workflow": {
            "resources": {"books": books},
            "runs": [
                {"id": "stage_a", "demand": "./stage_a.demand.yaml"},
                {"id": "stage_b", "demand": "./stage_b.demand.yaml", "depends_on": ["stage_a"]},
                {"id": "summary", "demand": "./summary.demand.yaml", "depends_on": ["stage_b"]},
            ],
        }
    }
    path = root / "workflow.yaml"
    path.write_text(yaml.safe_dump(wf, allow_unicode=True, sort_keys=False), encoding="utf-8")
    return path


def test_unified_xlsx_after_runs_without_deprecation(tmp_path: Path) -> None:
    out = tmp_path / "out"
    wf = _write_isomorphic_tree(
        tmp_path,
        books={"scratch": {"xlsx": {}}, "report": {"xlsx": {"path": str(out)}}},
        out_rel=str(out),
    )
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        result = run_workflow(str(wf), options=_run_options())
    assert not result.errors()
    deprec = [
        w
        for w in caught
        if issubclass(w.category, DeprecationWarning) and ("xlsx_file" in str(w.message) or "xlsx_memory" in str(w.message))
    ]
    assert not deprec
    assert _latest_book_path(out, book_id="report").exists()


def test_unified_xlsx_before_runs_aliases_fail_fast(tmp_path: Path) -> None:
    out = tmp_path / "out"
    wf = _write_isomorphic_tree(
        tmp_path,
        books={"scratch": {"xlsx_memory": {}}, "report": {"xlsx_file": {"path": str(out)}}},
        out_rel=str(out),
    )
    with pytest.raises(ScalimWorkflowConfigError, match=r"xlsx_(file|memory) was removed"):
        _ = run_workflow(str(wf), options=_run_options())


def test_unified_xlsx_path_allow_formulas_false_escapes_formula_like_strings(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    out = tmp_path / "out"
    (tmp_path / "demand.yaml").write_text(
        """
name: formulas
main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_formula_like_rows"
  fields:
    label: {extract: label}
    payload: {extract: payload}
outputs:
  - name: detail
    to: {book: report, sheet: S}
    fields: [label, payload]
""".lstrip(),
        encoding="utf-8",
    )
    wf = {
        "workflow": {
            "resources": {"books": {"report": {"xlsx": {"path": str(out), "allow_formulas": False}}}},
            "runs": [{"id": "a", "demand": "./demand.yaml"}],
        }
    }
    path = tmp_path / "workflow.yaml"
    path.write_text(yaml.safe_dump(wf, allow_unicode=True, sort_keys=False), encoding="utf-8")
    result = run_workflow(str(path), options=_run_options())
    assert not result.errors()
    book_path = _latest_book_path(out, book_id="report")
    wb = load_workbook(str(book_path), data_only=False)
    try:
        ws = wb["S"]
        assert ws["A1"].value == "label"
        assert ws["B1"].value == "payload"
        assert ws["A2"].value == "ok"
        assert ws["B2"].value == "'=1+1"
        assert ws["B2"].data_type != "f"
        assert ws["B3"].value == "'  +SUM(A1:A2)"
        assert ws["B4"].value == "'@X"
    finally:
        wb.close()


def test_resource_defs_book_xlsx_memory_with_path_still_builds_sheetbook_export(tmp_path: Path) -> None:
    """内部 pathless `BookConfig(export_xlsx=...)` 编译产物仍走 sheetbook export 布局."""

    from scalim.spec.ir._workflow import WorkflowArtifactsIr, WorkflowIr, WorkflowOptionsIr, WorkflowResourceIr
    from scalim.workflow import execute as workflow_execute_mod

    out = tmp_path / "out"
    workflow_ir = WorkflowIr(
        nodes=(),
        edges=(),
        options=WorkflowOptionsIr(max_concurrency=1, failure_policy="all_fail"),
        resources=(
            WorkflowResourceIr(
                resource_id="report",
                resource_type="book",
                path=str(out),
                options={"pathful": False, "export_xlsx": {"allow_formulas": True}},
            ),
        ),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={}),
    )
    _workbooks, _allow, _csv, sheetbooks = workflow_execute_mod._build_workflow_resource_defs(  # noqa: SLF001
        workflow_ir, workflow_exec_id="wf_c20"
    )
    assert "report" in sheetbooks
    assert sheetbooks["report"].export_path is not None

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        # Archive examples were frozen out of tree; keep the load contract inline.
        cfg = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "resources": {
                        "books": {
                            "scratch": {"xlsx": {}},
                            "report": {"xlsx": {"path": "./out"}},
                        }
                    },
                    "runs": [
                        {"id": "stage_a", "demand": "./stage_a.demand.yaml"},
                        {"id": "stage_b", "demand": "./stage_b.demand.yaml", "depends_on": ["stage_a"]},
                        {"id": "summary", "demand": "./summary.demand.yaml", "depends_on": ["stage_b"]},
                    ],
                }
            }
        )
    assert cfg.resources.books["scratch"].path is None
    assert cfg.resources.books["report"].path is not None
    assert not [w for w in caught if issubclass(w.category, DeprecationWarning)]
