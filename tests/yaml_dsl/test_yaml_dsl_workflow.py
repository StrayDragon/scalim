import json
import threading
from pathlib import Path
from typing import Any, Dict, List, Optional, cast

import pytest

from scalim.dsl.yaml_dsl import DemandDiagnosticsOverride
from scalim.dsl.yaml_dsl import DemandDiagnosticsPolicy
from scalim.dsl.yaml_dsl import OutputOverride, OutputToOverride, OutputWriteOverride
from scalim.dsl.yaml_dsl import RunOverrides
from scalim.dsl.yaml_dsl import (
    DemandRunOptions,
    DemandRunOutputOptions,
    DemandRunResult,
    DemandRunRuntimeOptions,
    DemandRunSecurityOptions,
    DemandRunTemplateOptions,
)
from scalim.dsl.yaml_dsl import FileResourceOverride
from scalim.dsl.yaml_dsl import OutputExtrasOverride
from scalim.dsl.yaml_dsl import ResourcesOverride
from scalim.dsl.yaml_dsl import run_workflow as run_workflow_public
from scalim.dsl.yaml_dsl._internal.workflow_injected_entrypoints import run_workflow_injected
from scalim.dsl.yaml_dsl.runtime import compiler as by_yaml_compiler_mod
from scalim.dsl.yaml_dsl.workflow_types import (
    ComponentsExtend,
    ComponentsReplace,
    StageBarrierSchedulerOptions,
    WorkflowCachePoolPreloadForeverShared,
    WorkflowExecutionOptions,
    WorkflowNodePatch,
    WorkflowRunOptions,
    WorkflowRuntimeOptions,
)
from scalim.dsl.yaml_dsl import workflow_compile as workflow_compile_mod
from scalim.exceptions import ScalimInternalError
from scalim.execution import versioned_outputs
from scalim.workflow import execute as workflow_execute_mod
from scalim.workflow import loaders as workflow_loaders_mod
from scalim.workflow.errors import ScalimWorkflowConfigError as WorkflowRuntimeConfigError
from scalim.events import (
    EVENT_DIAGNOSTIC_WARNING,
    EVENT_LOADER_CALL,
    EVENT_PIPELINE_START,
    EVENT_WORKFLOW_CACHE_ACQUIRE,
    EVENT_WORKFLOW_CACHE_EVICT,
    EVENT_WORKFLOW_CACHE_RELEASE,
    EVENT_WORKFLOW_NODE_CANCELLED,
    EVENT_WORKFLOW_NODE_END,
    EVENT_WORKFLOW_NODE_START,
    EVENT_WORKFLOW_RESOURCE_COMMIT,
    EVENT_WORKFLOW_RESOURCE_CREATE,
    EVENT_WORKFLOW_RESOURCE_DISCARD,
    EVENT_WORKFLOW_RESOURCE_WRITE,
)
from scalim.hooks import BaseHook
from scalim.ob.manager import ObserverManager
from scalim.ob.observer import Observer
from scalim.execution.guardrails import GuardrailsPolicy
from scalim.execution.loader_retry import LoaderRetryPoliciesSpec
from scalim.dsl.yaml_dsl.workflow import (
    ScalimWorkflowConfigError,
    load_workflow_config,
    load_workflow_config_from_mapping,
    resolve_workflow_demand_path,
    validate_workflow_yaml_text_json,
)
from tests.fixtures import workflow_loaders


_ALLOWED_MODULES = frozenset(["tests.fixtures.workflow_loaders"])
_ALLOWED_MODULES_WITH_SHEETBOOK = frozenset(["tests.fixtures.workflow_loaders", "scalim.workflow.loaders"])


def _run_options(*, allowed_modules=_ALLOWED_MODULES, **kwargs):  # type: ignore[no-untyped-def] test helper
    overrides = kwargs.pop("overrides", None)
    components = kwargs.pop("components", None)
    batch_size = kwargs.pop("batch_size", None)
    parallel_mode = kwargs.pop("parallel_mode", None)
    max_workers = kwargs.pop("max_workers", None)
    init_vars = kwargs.pop("init_vars", None)
    demand_failure_policy = kwargs.pop("demand_failure_policy", None)
    demand_diagnostics = kwargs.pop("demand_diagnostics", None)
    sink = kwargs.pop("sink", None)

    if kwargs:
        msg = "Unknown run options keys: {}".format(", ".join(sorted(str(k) for k in kwargs.keys())))
        raise TypeError(msg)
    if sink is not None:
        raise TypeError("sink is not supported in DemandRunOptions")

    template_kwargs: Dict[str, Any] = {}
    if init_vars is not None:
        template_kwargs["init_vars"] = init_vars

    runtime_kwargs: Dict[str, Any] = {}
    if components is not None:
        runtime_kwargs["components"] = components
    if batch_size is not None:
        runtime_kwargs["batch_size"] = batch_size
    if parallel_mode is not None:
        runtime_kwargs["parallel_mode"] = parallel_mode
    if max_workers is not None:
        runtime_kwargs["max_workers"] = max_workers
    if demand_failure_policy is not None:
        runtime_kwargs["demand_failure_policy"] = demand_failure_policy
    if demand_diagnostics is not None:
        runtime_kwargs["demand_diagnostics"] = demand_diagnostics

    outputs_kwargs: Dict[str, Any] = {}
    if overrides is not None:
        outputs_kwargs["overrides"] = overrides

    return DemandRunOptions(
        security=DemandRunSecurityOptions(allowed_modules=allowed_modules),
        template=DemandRunTemplateOptions(**template_kwargs),
        runtime=DemandRunRuntimeOptions(**runtime_kwargs),
        outputs=DemandRunOutputOptions(**outputs_kwargs),
    )


def _workflow_runtime_options(  # type: ignore[no-untyped-def] test helper
    *,
    max_concurrency: int = 1,
    failure_policy: str = "all_fail",
    cache_pool_max_entries: Optional[int] = None,
    schedule_mode: str = "pipeline",
):
    execution = WorkflowExecutionOptions(max_concurrency=int(max_concurrency), failure_policy=str(failure_policy))
    kwargs: Dict[str, Any] = {}
    if str(schedule_mode).strip() == "stage_barrier":
        kwargs["scheduler"] = StageBarrierSchedulerOptions()
    if cache_pool_max_entries is None:
        return WorkflowRuntimeOptions(execution=execution, **kwargs)
    return WorkflowRuntimeOptions(
        execution=execution, cache_pool=WorkflowCachePoolPreloadForeverShared(max_entries=int(cache_pool_max_entries)), **kwargs
    )


def run_workflow(  # type: ignore[no-untyped-def] test helper
    workflow_yaml_path: str,
    *,
    options: DemandRunOptions,
    workflow_runtime_options: Optional[WorkflowRuntimeOptions] = None,
    path_aliases: Optional[Dict[str, str]] = None,
    run_options_patches_by_run_id: Optional[Dict[str, WorkflowNodePatch]] = None,
    run_ir_fn=None,
    compile_demand_yaml_fn=None,
):
    if workflow_runtime_options is None:
        workflow_runtime_options = WorkflowRuntimeOptions.preset_default()

    workflow_components = None if options.runtime.components is None else tuple(options.runtime.components)
    wf_options = WorkflowRunOptions(
        demand=options,
        patches_by_run_id=run_options_patches_by_run_id,
        runtime=workflow_runtime_options,
        path_aliases=path_aliases,
        workflow_components=workflow_components,
    )
    if run_ir_fn is None and compile_demand_yaml_fn is None:
        return run_workflow_public(workflow_yaml_path, options=wf_options)
    return run_workflow_injected(
        workflow_yaml_path,
        options=wf_options,
        run_ir_fn=run_ir_fn,
        compile_demand_yaml_fn=compile_demand_yaml_fn,
    )


class _WorkflowEventRecorder(Observer):
    event_types = {
        EVENT_DIAGNOSTIC_WARNING,
        EVENT_PIPELINE_START,
        EVENT_WORKFLOW_CACHE_ACQUIRE,
        EVENT_WORKFLOW_CACHE_EVICT,
        EVENT_WORKFLOW_CACHE_RELEASE,
        EVENT_WORKFLOW_NODE_START,
        EVENT_WORKFLOW_NODE_END,
        EVENT_WORKFLOW_NODE_CANCELLED,
        EVENT_WORKFLOW_RESOURCE_CREATE,
        EVENT_WORKFLOW_RESOURCE_WRITE,
        EVENT_WORKFLOW_RESOURCE_COMMIT,
        EVENT_WORKFLOW_RESOURCE_DISCARD,
    }

    def __init__(self) -> None:
        self._lock = threading.Lock()
        self.events: List[Any] = []

    def on_event(self, event) -> None:  # type: ignore[override]
        with self._lock:
            self.events.append(event)


class _LoaderCallObserverRecorder(Observer):
    event_types = {
        EVENT_LOADER_CALL,
    }

    def __init__(self) -> None:
        self._lock = threading.Lock()
        self.events: List[Any] = []

    def on_event(self, event) -> None:  # type: ignore[override]
        with self._lock:
            self.events.append(event)


class _LoaderCallHookRecorder(BaseHook):
    event_types = {
        EVENT_LOADER_CALL,
    }

    def __init__(self) -> None:
        super().__init__()
        self._lock = threading.Lock()
        self.events: List[Any] = []

    def on_loader_call(self, event) -> None:  # type: ignore[override]
        with self._lock:
            self.events.append(event)


def _write_text(path: Path, text: str) -> Path:
    path.write_text(text, encoding="utf-8")
    return path


def _cache_pool_config(
    *,
    conflict_policy: str = "error",
    release_policy: str = "dag_refcount",
    max_entries: int = 10,
    over_budget_policy: str = "fail_fast",
    pin: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    cfg: Dict[str, Any] = {
        "conflict_policy": str(conflict_policy),
        "release_policy": str(release_policy),
        "budget": {
            "max_entries": int(max_entries),
            "over_budget_policy": str(over_budget_policy),
        },
    }
    if pin:
        cfg["pin"] = pin
    return cfg


def _write_demand_yaml(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    main_loader_ref: str,
    preload_loader_ref: str,
    cache_mode: str = "preload_forever",
    preload_source_id: str = "preload",
    preload_params_yaml: Optional[str] = None,
) -> Path:
    preload_params_block = ""
    if preload_params_yaml:
        snippet = str(preload_params_yaml).strip("\n")
        indented = "\n".join(("    " + line if line.strip() else "") for line in snippet.splitlines())
        preload_params_block = "{}\n".format(indented)

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "{main_loader_ref}"
  fields:
    ref_id:
      name: ref_id

sources:
  {preload_source_id}:
    loader: "{preload_loader_ref}"
    key: id
    cache_mode: {cache_mode}
{preload_params_block}
    fields:
      value:
        name: value
        relation: main_to_preload

relations:
  main_to_preload:
    steps:
      - from: main.ref_id
        to: {preload_source_id}.id
"""
        )
        .format(
            name=name,
            main_loader_ref=main_loader_ref,
            preload_loader_ref=preload_loader_ref,
            cache_mode=cache_mode,
            preload_source_id=preload_source_id,
            preload_params_block=preload_params_block,
        )
        .lstrip(),
    )


def _write_table_demand_yaml_with_csv_output(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    loader_ref: str,
    output_name: str,
    output_root: Path,
    field_ids: List[str],
) -> Path:
    file_id = "{}_{}_csv".format(str(name), str(output_name))
    field_lines = []
    for fid in field_ids:
        field_lines.append("    {}: {{extract: {}}}".format(str(fid), str(fid)))

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "{loader_ref}"
  fields:
{fields}

resources:
  files:
    {file_id}:
      kind: csv_file
      path: "{output_root}"

outputs:
  - name: {output_name}
    to:
      file: {file_id}
    fields: {fields_list}
"""
        )
        .format(
            name=str(name),
            loader_ref=str(loader_ref),
            fields="\n".join(field_lines),
            file_id=file_id,
            output_name=str(output_name),
            output_root=str(output_root),
            fields_list=json.dumps([str(x) for x in field_ids]),
        )
        .lstrip(),
    )


def _write_table_demand_yaml_with_two_csv_outputs(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    loader_ref: str,
    output1_name: str,
    output1_root: Path,
    output1_field_ids: List[str],
    output2_name: str,
    output2_root: Path,
    output2_field_ids: List[str],
    main_field_ids: List[str],
) -> Path:
    output1_file_id = "{}_{}_csv".format(str(name), str(output1_name))
    output2_file_id = "{}_{}_csv".format(str(name), str(output2_name))
    field_lines = []
    for fid in main_field_ids:
        field_lines.append("    {}: {{extract: {}}}".format(str(fid), str(fid)))

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "{loader_ref}"
  fields:
{fields}

resources:
  files:
    {output1_file_id}:
      kind: csv_file
      path: "{output1_root}"
    {output2_file_id}:
      kind: csv_file
      path: "{output2_root}"

outputs:
  - name: {output1_name}
    to:
      file: {output1_file_id}
    fields: {output1_fields_list}
  - name: {output2_name}
    to:
      file: {output2_file_id}
    fields: {output2_fields_list}
"""
        )
        .format(
            name=str(name),
            loader_ref=str(loader_ref),
            fields="\n".join(field_lines),
            output1_name=str(output1_name),
            output1_file_id=output1_file_id,
            output1_root=str(output1_root),
            output1_fields_list=json.dumps([str(x) for x in output1_field_ids]),
            output2_name=str(output2_name),
            output2_file_id=output2_file_id,
            output2_root=str(output2_root),
            output2_fields_list=json.dumps([str(x) for x in output2_field_ids]),
        )
        .lstrip(),
    )


def _write_table_demand_yaml_with_book_output(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    loader_ref: str,
    output_name: str,
    book_id: str,
    sheet: str,
    field_ids: List[str],
) -> Path:
    field_lines = []
    for fid in field_ids:
        field_lines.append("    {}: {{extract: {}}}".format(str(fid), str(fid)))

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "{loader_ref}"
  fields:
{fields}

outputs:
  - name: {output_name}
    to:
      book: {book_id}
      sheet: {sheet}
    fields: {fields_list}
"""
        )
        .format(
            name=str(name),
            loader_ref=str(loader_ref),
            fields="\n".join(field_lines),
            book_id=str(book_id),
            output_name=str(output_name),
            sheet=str(sheet),
            fields_list=json.dumps([str(x) for x in field_ids]),
        )
        .lstrip(),
    )


def _write_table_demand_yaml_with_workbook_output(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    loader_ref: str,
    output_name: str,
    output_path: Path,
    sheet: str,
    field_ids: List[str],
) -> Path:
    field_lines = []
    for fid in field_ids:
        field_lines.append("    {}: {{extract: {}}}".format(str(fid), str(fid)))

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "{loader_ref}"
  fields:
{fields}

resources:
  books:
    {output_name}_book:
      kind: xlsx_file
      path: "{output_path}"

outputs:
  - name: {output_name}
    to:
      book: {output_name}_book
      sheet: "{sheet}"
    fields: {fields_list}
"""
        )
        .format(
            name=str(name),
            loader_ref=str(loader_ref),
            fields="\n".join(field_lines),
            output_name=str(output_name),
            output_path=str(output_path),
            sheet=str(sheet),
            fields_list=json.dumps([str(x) for x in field_ids]),
        )
        .lstrip(),
    )


def _write_table_demand_yaml_from_sheetbook_loader(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    init_var_name: str,
    output_name: str,
    output_root: Path,
    field_ids: List[str],
) -> Path:
    file_id = "{}_{}_csv".format(str(name), str(output_name))
    field_lines = []
    for fid in field_ids:
        field_lines.append("    {}: {{extract: {}}}".format(str(fid), str(fid)))

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "scalim.workflow.loaders:book_sheet_rows"
  params:
    ref:
      $init_var: {init_var_name}
  fields:
{fields}

resources:
  files:
    {file_id}:
      kind: csv_file
      path: "{output_root}"

outputs:
  - name: {output_name}
    to:
      file: {file_id}
    fields: {fields_list}
"""
        )
        .format(
            name=str(name),
            init_var_name=str(init_var_name),
            fields="\n".join(field_lines),
            file_id=file_id,
            output_name=str(output_name),
            output_root=str(output_root),
            fields_list=json.dumps([str(x) for x in field_ids]),
        )
        .lstrip(),
    )


def _write_duplicate_header_demand_yaml(tmp_path: Path, *, file_name: str, output_root: Path) -> Path:
    return _write_text(
        tmp_path / file_name,
        (
            """
name: duplicate_headers

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id:
      extract: id
      name: Dup
    value:
      extract: value
      name: Dup

sources: {{}}

resources:
  files:
    detail_csv:
      kind: csv_file
      path: "{output_root}"

outputs:
  - name: detail
    to:
      file: detail_csv
    fields: [id, value]
"""
        )
        .format(output_root=str(output_root))
        .lstrip(),
    )


def _write_workflow_yaml(
    tmp_path: Path,
    *,
    runs: list,
    max_concurrency: int = 1,
    failure_policy: str = "all_fail",
    resources: Optional[Dict[str, Any]] = None,
    cache_pool: Optional[Dict[str, Any]] = None,
    ctx: Optional[Dict[str, Any]] = None,
) -> Path:
    run_lines = []
    for item in runs:
        depends_on = item.get("depends_on") or []
        depends_on_lines = ""
        if depends_on:
            depends_on_lines = "\n      depends_on:\n{}".format("\n".join(["        - {}".format(d) for d in depends_on]))

        main_rows_from_lines = ""
        if "main_rows_from" in item:
            main_rows_from = cast("Any", item.get("main_rows_from"))  # pragma: allow-cast test yaml builder typed narrowing
            if main_rows_from is None:
                main_rows_from_lines = "\n      main_rows_from: null"
            else:
                if not isinstance(main_rows_from, dict):
                    raise ValueError("main_rows_from must be a mapping")
                producer = main_rows_from.get("run")
                main_rows_from_lines = "\n      main_rows_from:\n        run: {}".format(json.dumps(producer))

        init_vars = cast("Optional[Dict[str, object]]", item.get("init_vars"))  # pragma: allow-cast test yaml builder typed narrowing
        init_vars_lines = ""
        if init_vars:
            rendered = []
            for key, value in init_vars.items():
                rendered.append("        {}: {}".format(str(key), json.dumps(value)))
            init_vars_lines = "\n      init_vars:\n{}".format("\n".join(rendered))

        writes = cast("Any", item.get("writes"))  # pragma: allow-cast test yaml builder typed narrowing
        writes_lines = ""
        if writes is not None:
            if not isinstance(writes, list):
                raise ValueError("writes must be a list")
            if not writes:
                writes_lines = "\n      writes: []"
            else:
                rendered: List[str] = []
                for write in writes:
                    if not isinstance(write, dict) or len(write) != 1:
                        raise ValueError("writes items must be a mapping with exactly one key")
                    kind = next(iter(write.keys()))
                    cfg = write.get(kind) or {}
                    if not isinstance(cfg, dict):
                        raise ValueError("writes[*].<kind> must be a mapping")
                    cfg_lines = []
                    for key, value in cfg.items():
                        cfg_lines.append("            {}: {}".format(str(key), json.dumps(value)))
                    rendered.append("        - {}:\n{}".format(str(kind), "\n".join(cfg_lines)))
                writes_lines = "\n      writes:\n{}".format("\n".join(rendered))

        write_to = cast("Any", item.get("write_to"))  # pragma: allow-cast test yaml builder typed narrowing
        write_to_lines = ""
        if write_to:
            if writes is not None:
                raise ValueError("cannot set both write_to and writes")
            if not isinstance(write_to, dict) or len(write_to) != 1:
                raise ValueError("write_to must be a mapping with exactly one key")
            kind = next(iter(write_to.keys()))
            cfg = write_to.get(kind) or {}
            if not isinstance(cfg, dict):
                raise ValueError("write_to.<kind> must be a mapping")
            cfg_lines = []
            for key, value in cfg.items():
                cfg_lines.append("          {}: {}".format(str(key), json.dumps(value)))
            write_to_lines = "\n      write_to:\n        {}:\n{}".format(str(kind), "\n".join(cfg_lines))

        run_lines.append(
            "    - id: {}\n      demand: {}{}{}{}{}{}".format(
                item["id"],
                item["demand"],
                depends_on_lines,
                main_rows_from_lines,
                init_vars_lines,
                writes_lines,
                write_to_lines,
            )
        )

    resources_lines = ""
    if resources:
        group_lines: List[str] = []
        for group_key, group_cfg in resources.items():
            group_lines.append("    {}:".format(str(group_key)))
            for res_id, cfg in cast("Dict[str, Any]", group_cfg).items():  # pragma: allow-cast test yaml builder typed narrowing
                group_lines.append("      {}:".format(str(res_id)))
                for key, value in cast("Dict[str, Any]", cfg).items():  # pragma: allow-cast test yaml builder typed narrowing
                    group_lines.append("        {}: {}".format(str(key), json.dumps(value)))
        resources_lines = "\n  resources:\n{}".format("\n".join(group_lines))

    # NOTE: workflow-level runtime policy was moved out of YAML into `workflow_runtime_options`.
    _ = (max_concurrency, failure_policy, cache_pool, ctx)
    return _write_text(
        tmp_path / "workflow.yaml",
        (
            """
workflow:
  runs:
{runs}
{resources_lines}
"""
        )
        .format(
            runs="\n".join(run_lines),
            resources_lines=resources_lines.rstrip(),
        )
        .lstrip(),
    )


def test_load_workflow_config_semantic_validation(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: a
      demand: b.yaml
"""
        ).lstrip(),
    )

    with pytest.raises(ScalimWorkflowConfigError):
        _ = load_workflow_config(str(workflow_path))


def test_load_workflow_config_rejects_unknown_depends_on(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      depends_on: [nope]
"""
        ).lstrip(),
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "Unknown run.depends_on id" in str(excinfo.value)


def test_load_workflow_config_rejects_depends_on_cycles(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
      depends_on: [b]
    - id: b
      demand: b.yaml
      depends_on: [a]
"""
        ).lstrip(),
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "cycle_path" in str(excinfo.value)
    assert '["a", "b", "a"]' in str(excinfo.value)


def test_load_workflow_config_rejects_unknown_main_rows_from_run(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      main_rows_from:
        run: nope
"""
        ).lstrip(),
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "Unknown run.main_rows_from.run id" in str(excinfo.value)
    assert "path=workflow.runs.1.main_rows_from.run" in str(excinfo.value)


def test_load_workflow_config_requires_depends_on_for_main_rows_from(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      main_rows_from:
        run: a
"""
        ).lstrip(),
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "run.main_rows_from requires explicit depends_on" in str(excinfo.value)
    assert "path=workflow.runs.1.depends_on" in str(excinfo.value)


def test_load_workflow_config_rejects_main_rows_from_when_not_mapping(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      main_rows_from: a
"""
        ).lstrip(),
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "run.main_rows_from must be a mapping" in str(excinfo.value)
    assert "path=workflow.runs.1.main_rows_from" in str(excinfo.value)


def test_load_workflow_config_rejects_main_rows_from_unknown_keys(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      main_rows_from:
        run: a
        extra: 1
"""
        ).lstrip(),
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "run.main_rows_from has unknown keys" in str(excinfo.value)
    assert "path=workflow.runs.1.main_rows_from" in str(excinfo.value)


def test_load_workflow_config_rejects_main_rows_from_run_when_not_string(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      main_rows_from:
        run: 123
"""
        ).lstrip(),
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "run.main_rows_from.run must be a non-empty string" in str(excinfo.value)
    assert "path=workflow.runs.1.main_rows_from.run" in str(excinfo.value)


def test_load_workflow_config_rejects_main_rows_from_self_reference(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
      main_rows_from:
        run: a
"""
        ).lstrip(),
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "run.main_rows_from.run must not reference self" in str(excinfo.value)
    assert "path=workflow.runs.0.main_rows_from.run" in str(excinfo.value)


def test_load_workflow_config_rejects_write_to_with_migration_hint(tmp_path: Path) -> None:
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml", "write_to": {}}],
                }
            }
        )
    assert "run.write_to was removed" in str(excinfo.value)
    assert excinfo.value.path == "workflow.runs.0.write_to"


def test_load_workflow_config_rejects_writes_with_migration_hint(tmp_path: Path) -> None:
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml", "writes": []}],
                }
            }
        )
    assert "run.writes was removed" in str(excinfo.value)
    assert excinfo.value.path == "workflow.runs.0.writes"


def test_run_workflow_primary_only_collects_errors(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "ok", "demand": "ok.yaml"}, {"id": "bad", "demand": "bad.yaml"}],
    )

    runtime = WorkflowRuntimeOptions(execution=WorkflowExecutionOptions(max_concurrency=2, failure_policy="primary_only"))
    result = run_workflow(str(wf), options=_run_options(), workflow_runtime_options=runtime)
    assert [o.run_id for o in result.outcomes] == ["ok", "bad"]
    assert result.outcomes[0].result is not None
    assert result.outcomes[0].error is None
    assert result.outcomes[1].result is None
    assert result.outcomes[1].error is not None
    assert result.outcomes[1].error.exc_type in {"ValueError", "ScalimWorkflowRunFailedError", "RuntimeError"}
    assert result.errors()


def test_run_workflow_accepts_overrides_default_unset(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    result = run_workflow(str(wf), options=_run_options(overrides=RunOverrides()))
    assert [o.run_id for o in result.outcomes] == ["ok"]


def test_run_workflow_rejects_sink_in_base_options(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    from scalim.sinks import InMemoryRowSink

    with pytest.raises(TypeError, match=r"sink"):
        _ = run_workflow(str(wf), options=_run_options(sink=InMemoryRowSink()))


def test_run_workflow_concurrency_capture_replay_summarizes_loader_call_payload_for_observer(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=2,
        failure_policy="all_fail",
    )

    observer = _LoaderCallObserverRecorder()
    result = run_workflow(
        str(wf),
        options=_run_options(components=[observer]),
        workflow_runtime_options=_workflow_runtime_options(max_concurrency=2, failure_policy="all_fail"),
    )
    assert not result.errors()

    preload_payloads: List[Any] = []
    for event in observer.events:
        if str(getattr(event, "event_type", "")) != EVENT_LOADER_CALL:
            continue
        payload = getattr(event, "payload", None)
        if payload is None:
            continue
        if str(getattr(payload, "loader_name", "")) != "preload":
            continue
        preload_payloads.append(payload)

    assert preload_payloads
    for payload in preload_payloads:
        assert payload.result == {"type": "dict", "size": 1}


def test_run_workflow_concurrency_capture_replay_summarizes_loader_call_payload_for_typed_hook(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=2,
        failure_policy="all_fail",
    )

    hook = _LoaderCallHookRecorder()
    result = run_workflow(
        str(wf),
        options=_run_options(components=[hook]),
        workflow_runtime_options=_workflow_runtime_options(max_concurrency=2, failure_policy="all_fail"),
    )
    assert not result.errors()

    preload_events: List[Any] = []
    for event in hook.events:
        if str(getattr(event, "loader_name", "")) != "preload":
            continue
        preload_events.append(event)

    assert preload_events
    for event in preload_events:
        assert event.result == {"type": "dict", "size": 1}


def test_run_workflow_run_options_patches_by_run_id_batch_size_overrides_global(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    seen_batch_size_by_name: Dict[str, object] = {}

    def _compile_with_capture(yaml_path: str, *, options):  # type: ignore[no-untyped-def] test hook
        seen_batch_size_by_name[Path(str(yaml_path)).name] = options.runtime.batch_size
        return by_yaml_compiler_mod.compile(yaml_path, options=options)

    result = run_workflow(
        str(wf),
        options=_run_options(batch_size=2000),
        run_options_patches_by_run_id={"a": WorkflowNodePatch(batch_size=5000)},
        compile_demand_yaml_fn=_compile_with_capture,
    )
    assert not result.errors()
    assert seen_batch_size_by_name["a.yaml"] == 5000
    assert seen_batch_size_by_name["b.yaml"] == 2000


def test_run_workflow_run_options_patches_by_run_id_parallel_mode_overrides_global(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    seen_parallel_mode_by_name: Dict[str, object] = {}

    def _compile_with_capture(yaml_path: str, *, options):  # type: ignore[no-untyped-def] test hook
        seen_parallel_mode_by_name[Path(str(yaml_path)).name] = options.runtime.parallel_mode
        return by_yaml_compiler_mod.compile(yaml_path, options=options)

    result = run_workflow(
        str(wf),
        options=_run_options(parallel_mode="seq"),
        run_options_patches_by_run_id={"a": WorkflowNodePatch(parallel_mode="adaptive")},
        compile_demand_yaml_fn=_compile_with_capture,
    )
    assert not result.errors()
    assert seen_parallel_mode_by_name["a.yaml"] == "adaptive"
    assert seen_parallel_mode_by_name["b.yaml"] == "seq"


def test_run_workflow_run_options_patches_by_run_id_max_workers_overrides_global(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    seen_max_workers_by_name: Dict[str, object] = {}

    def _compile_with_capture(yaml_path: str, *, options):  # type: ignore[no-untyped-def] test hook
        seen_max_workers_by_name[Path(str(yaml_path)).name] = options.runtime.max_workers
        return by_yaml_compiler_mod.compile(yaml_path, options=options)

    result = run_workflow(
        str(wf),
        options=_run_options(max_workers=0),
        run_options_patches_by_run_id={"a": WorkflowNodePatch(max_workers=4)},
        compile_demand_yaml_fn=_compile_with_capture,
    )
    assert not result.errors()
    assert seen_max_workers_by_name["a.yaml"] == 4
    assert seen_max_workers_by_name["b.yaml"] == 0


def test_run_workflow_run_options_patches_by_run_id_rejects_invalid_parallelism_knobs(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    with pytest.raises((TypeError, ValueError), match=r"parallel_mode=seq\\|adaptive"):
        _ = run_workflow(
            str(wf),
            options=_run_options(),
            run_options_patches_by_run_id={"ok": WorkflowNodePatch(parallel_mode="thread")},
        )

    with pytest.raises((TypeError, ValueError), match=r"parallel_mode=seq\\|adaptive"):
        _ = run_workflow(
            str(wf),
            options=_run_options(),
            run_options_patches_by_run_id={"ok": WorkflowNodePatch(parallel_mode=123)},  # type: ignore[arg-type] intentional runtime boundary test
        )

    with pytest.raises((TypeError, ValueError), match=r"max_workers>=0"):
        _ = run_workflow(
            str(wf),
            options=_run_options(),
            run_options_patches_by_run_id={"ok": WorkflowNodePatch(max_workers=-1)},
        )

    with pytest.raises((TypeError, ValueError), match=r"max_workers>=0"):
        _ = run_workflow(  # type: ignore[arg-type] intentional runtime boundary test
            str(wf),
            options=_run_options(),
            run_options_patches_by_run_id={"ok": WorkflowNodePatch(max_workers="4")},
        )


def test_run_workflow_run_options_patches_by_run_id_rejects_unknown_id(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = run_workflow(
            str(wf),
            options=_run_options(),
            run_options_patches_by_run_id={"nope": WorkflowNodePatch(batch_size=5000)},
        )
    assert "nope" in str(excinfo.value)
    assert "ok" in str(excinfo.value)


def test_run_workflow_run_options_patches_by_run_id_rejects_dict_patch_payload(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    with pytest.raises(TypeError, match=r"dict patches are not supported"):
        _ = run_workflow(  # type: ignore[arg-type] intentional runtime boundary test
            str(wf),
            options=_run_options(),
            run_options_patches_by_run_id={"ok": {"batch_size": 5000}},
        )


def test_run_workflow_run_options_patches_by_run_id_components_extend_and_replace(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    base = _WorkflowEventRecorder()
    extra = _WorkflowEventRecorder()

    seen_components_by_name: Dict[str, object] = {}

    def _compile_with_capture(yaml_path: str, *, options):  # type: ignore[no-untyped-def] test hook
        seen_components_by_name[Path(str(yaml_path)).name] = list(options.runtime.components or [])
        return by_yaml_compiler_mod.compile(yaml_path, options=options)

    result = run_workflow(
        str(wf),
        options=_run_options(components=[base]),
        run_options_patches_by_run_id={
            "a": WorkflowNodePatch(components=ComponentsExtend([extra])),
            "b": WorkflowNodePatch(components=ComponentsReplace(())),
        },
        compile_demand_yaml_fn=_compile_with_capture,
    )
    assert not result.errors()
    assert seen_components_by_name["a.yaml"] == [base, extra]
    assert seen_components_by_name["b.yaml"] == []


def test_components_replace_normalizes_iterable_items_to_tuple() -> None:
    r = _WorkflowEventRecorder()
    patch = ComponentsReplace([r])
    assert isinstance(patch.items, tuple)
    assert patch.items == (r,)


def test_components_extend_accepts_tuple_items_without_normalization() -> None:
    r = _WorkflowEventRecorder()
    patch = ComponentsExtend((r,))
    assert isinstance(patch.items, tuple)
    assert patch.items == (r,)


def test_validate_run_options_patches_by_run_id_rejects_non_str_key() -> None:
    from scalim.dsl.yaml_dsl import workflow_entrypoints as workflow_entrypoints_mod

    with pytest.raises(TypeError, match=r"keys must be workflow run ids"):
        _ = workflow_entrypoints_mod._validate_patches_by_run_id(  # type: ignore[arg-type] intentional runtime boundary test
            {1: WorkflowNodePatch(batch_size=5000)},
            known_run_ids=frozenset(["ok"]),
        )


def test_validate_run_options_patches_by_run_id_rejects_non_patch_payload() -> None:
    from scalim.dsl.yaml_dsl import workflow_entrypoints as workflow_entrypoints_mod

    with pytest.raises(TypeError, match=r"must be a WorkflowNodePatch"):
        _ = workflow_entrypoints_mod._validate_patches_by_run_id(
            {"ok": object()},  # type: ignore[arg-type] intentional runtime boundary test
            known_run_ids=frozenset(["ok"]),
        )


def test_apply_workflow_run_options_patch_applies_demand_failure_policy_guardrails_loader_retry() -> None:
    from scalim.dsl.yaml_dsl import workflow_entrypoints as workflow_entrypoints_mod

    base = _run_options(demand_failure_policy="global")
    guardrails = GuardrailsPolicy(enabled=True)
    loader_retry = LoaderRetryPoliciesSpec()
    patch = WorkflowNodePatch(
        demand_failure_policy="patch",
        guardrails=guardrails,
        loader_retry=loader_retry,
    )

    next_options = workflow_entrypoints_mod._apply_workflow_node_patch(base, patch)
    assert next_options.runtime.demand_failure_policy == "patch"
    assert next_options.runtime.guardrails == guardrails
    assert next_options.runtime.loader_retry == loader_retry


def test_apply_workflow_run_options_patch_rejects_unknown_components_patch() -> None:
    from scalim.dsl.yaml_dsl import workflow_entrypoints as workflow_entrypoints_mod

    base = _run_options()
    with pytest.raises(TypeError, match=r"WorkflowNodePatch\.components must be"):
        _ = workflow_entrypoints_mod._apply_workflow_node_patch(
            base,
            WorkflowNodePatch(components=object()),  # type: ignore[arg-type] intentional runtime boundary test
        )


def test_run_workflow_run_options_patches_by_run_id_overrides_precedence_over_workflow_resources_and_global_overrides(
    tmp_path: Path,
) -> None:
    _ = _write_text(
        tmp_path / "a.yaml",
        (
            """
name: a

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id:
      extract: id
    value:
      extract: value

outputs:
  - name: detail
    to:
      file: detail_a
    fields: ["id", "value"]
"""
        ).lstrip(),
    )
    _ = _write_text(
        tmp_path / "b.yaml",
        (
            """
name: b

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_b_fast"
  fields:
    id:
      extract: id
    value:
      extract: value

outputs:
  - name: detail
    to:
      file: detail_b
    fields: ["id", "value"]
"""
        ).lstrip(),
    )

    wf_detail_a_root = tmp_path / "wf_detail_a"
    wf_detail_b_root = tmp_path / "wf_detail_b"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "files": {
                "detail_a": {"kind": "csv_file", "path": str(wf_detail_a_root)},
                "detail_b": {"kind": "csv_file", "path": str(wf_detail_b_root)},
            }
        },
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    global_detail_a_root = tmp_path / "global_detail_a"
    global_detail_b_root = tmp_path / "global_detail_b"
    overrides = RunOverrides(
        resources=ResourcesOverride(
            files={
                "detail_a": FileResourceOverride(path=str(global_detail_a_root)),
                "detail_b": FileResourceOverride(path=str(global_detail_b_root)),
            }
        )
    )

    patch_detail_a_root = tmp_path / "patch_detail_a"
    run_options_patches_by_run_id = {
        "a": WorkflowNodePatch(
            overrides=RunOverrides(
                resources=ResourcesOverride(
                    files={
                        "detail_a": FileResourceOverride(path=str(patch_detail_a_root)),
                    }
                )
            )
        )
    }

    result = run_workflow(
        str(wf),
        options=_run_options(overrides=overrides),
        run_options_patches_by_run_id=run_options_patches_by_run_id,
    )
    assert not result.errors()

    from scalim.execution.versioned_outputs import parse_versioned_output_path  # noqa: PLC0415

    outcomes = {str(o.run_id): o for o in result.outcomes}
    out_a = outcomes["a"]
    out_b = outcomes["b"]
    assert isinstance(out_a.result, DemandRunResult)
    assert isinstance(out_b.result, DemandRunResult)

    assert out_a.result.core.outputs is not None
    assert out_b.result.core.outputs is not None
    out_a_path = Path(str(out_a.result.core.outputs["detail"]))
    out_b_path = Path(str(out_b.result.core.outputs["detail"]))
    assert out_a_path.exists()
    assert out_b_path.exists()

    parsed_a = parse_versioned_output_path(out_a_path)
    parsed_b = parse_versioned_output_path(out_b_path)
    assert parsed_a.root == wf_detail_a_root.resolve(strict=False) or parsed_a.root == patch_detail_a_root.resolve(strict=False)
    assert parsed_a.kind == "files"
    assert parsed_a.artifact_id == "detail_a"
    assert parsed_a.root == patch_detail_a_root.resolve(strict=False)

    assert parsed_b.kind == "files"
    assert parsed_b.artifact_id == "detail_b"
    assert parsed_b.root == global_detail_b_root.resolve(strict=False)

    assert parsed_a.root != global_detail_a_root.resolve(strict=False)
    assert parsed_b.root != wf_detail_b_root.resolve(strict=False)


def test_run_workflow_all_fail_raises(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    with pytest.raises(Exception) as excinfo:
        _ = run_workflow(str(wf), options=_run_options())

    assert "run_id=bad" in str(excinfo.value)


def test_workflow_outcomes_are_in_declared_order(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="slow.yaml",
        name="slow",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="fast.yaml",
        name="fast",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "slow", "demand": "slow.yaml"}, {"id": "fast", "demand": "fast.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), options=_run_options())
    assert [o.run_id for o in result.outcomes] == ["slow", "fast"]


def test_workflow_dag_respects_depends_on_under_concurrency(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml", "depends_on": ["a"]}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), options=_run_options(components=[recorder]))
    assert [o.run_id for o in result.outcomes] == ["a", "b"]

    start = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_START]
    end = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]

    by_start = {e.payload.workflow_node_id: e for e in start}
    by_end = {e.payload.workflow_node_id: e for e in end}

    assert by_start["b"].seq > by_end["a"].seq


def test_workflow_pipeline_allows_cross_stage_overlap_under_concurrency(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table_alt",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="x.yaml",
        name="x",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_very_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table_alt",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "x", "demand": "x.yaml"},
            {"id": "b", "demand": "b.yaml", "depends_on": ["a"]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    runtime = _workflow_runtime_options(max_concurrency=2, failure_policy="primary_only", schedule_mode="pipeline")
    result = run_workflow(str(wf), options=_run_options(components=[recorder]), workflow_runtime_options=runtime)
    assert not result.errors()

    start = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_START]
    end = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]

    by_start = {e.payload.workflow_node_id: e for e in start}
    by_end = {e.payload.workflow_node_id: e for e in end}

    assert by_start["b"].timestamp >= by_end["a"].timestamp
    assert by_start["b"].timestamp < by_end["x"].timestamp
    assert by_start["b"].payload.schedule_mode == "pipeline"
    assert by_start["b"].payload.stage == 1


def test_workflow_stage_barrier_blocks_next_stage_until_all_nodes_in_stage_terminal(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table_alt",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="x.yaml",
        name="x",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_very_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table_alt",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "x", "demand": "x.yaml"},
            {"id": "b", "demand": "b.yaml", "depends_on": ["a"]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    runtime = _workflow_runtime_options(max_concurrency=2, failure_policy="primary_only", schedule_mode="stage_barrier")
    result = run_workflow(str(wf), options=_run_options(components=[recorder]), workflow_runtime_options=runtime)
    assert not result.errors()

    start = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_START]
    end = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]

    by_start = {e.payload.workflow_node_id: e for e in start}
    by_end = {e.payload.workflow_node_id: e for e in end}

    assert by_start["b"].timestamp >= by_end["a"].timestamp
    assert by_start["b"].timestamp >= by_end["x"].timestamp
    assert by_start["b"].payload.schedule_mode == "stage_barrier"
    assert by_start["b"].payload.stage == 1


def test_workflow_concurrency_does_not_call_components_concurrently_by_default(tmp_path: Path) -> None:
    for idx in range(4):
        _ = _write_demand_yaml(
            tmp_path,
            file_name="n{}.yaml".format(int(idx)),
            name="n{}".format(int(idx)),
            main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
            preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
            cache_mode="none",
        )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "n0", "demand": "n0.yaml"},
            {"id": "n1", "demand": "n1.yaml"},
            {"id": "n2", "demand": "n2.yaml"},
            {"id": "n3", "demand": "n3.yaml"},
        ],
        max_concurrency=4,
        failure_policy="primary_only",
    )

    class _ConcurrentCallProbe(Observer):
        event_types = {EVENT_PIPELINE_START}

        def __init__(self) -> None:
            self._guard = threading.Lock()
            self._in_call = threading.Lock()
            self.seen = 0
            self.concurrent_calls = 0

        def on_event(self, event) -> None:  # type: ignore[override]
            _ = event
            with self._guard:
                self.seen += 1

            if not self._in_call.acquire(False):
                with self._guard:
                    self.concurrent_calls += 1
                return
            try:
                # Deterministic busy work to widen the overlap window without time-based sleeps.
                x = 0
                for _idx in range(10_000):
                    x = (x + 1) ^ 0x1234
                _ = x
            finally:
                self._in_call.release()

    probe = _ConcurrentCallProbe()
    result = run_workflow(str(wf), options=_run_options(components=[probe]))
    assert not result.errors()
    assert int(probe.seen) > 0
    assert int(probe.concurrent_calls) == 0


def test_workflow_ctx_init_vars_are_resolved_after_prereq_completion(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="up.yaml",
        name="up",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="down.yaml",
        name="down",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
        preload_params_yaml=(
            """
params:
  p:
    $init_var: token
"""
        ).strip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "up", "demand": "up.yaml"},
            {
                "id": "down",
                "demand": "down.yaml",
                "depends_on": ["up"],
                "init_vars": {
                    "token": {
                        "$ctx": {"node": "up", "key": "total_rows"},
                    }
                },
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), options=_run_options(components=[recorder]))
    assert not result.errors()

    start = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_START]
    end = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]
    by_start = {e.payload.workflow_node_id: e for e in start}
    by_end = {e.payload.workflow_node_id: e for e in end}
    assert by_start["down"].seq > by_end["up"].seq


def test_workflow_ctx_ref_outside_depends_on_closure_fails_fast(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="c.yaml",
        name="c",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {
                "id": "c",
                "demand": "c.yaml",
                "init_vars": {"token": {"$ctx": {"node": "a", "key": "total_rows"}}},
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    with pytest.raises(ScalimWorkflowConfigError, match="declare depends_on"):
        _ = run_workflow(str(wf), options=_run_options())


def test_workflow_ctx_ref_node_self_fails_fast(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {
                "id": "a",
                "demand": "a.yaml",
                "init_vars": {"token": {"$ctx": {"node": "a", "key": "total_rows"}}},
            }
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    with pytest.raises(ScalimWorkflowConfigError, match="node=self"):
        _ = run_workflow(str(wf), options=_run_options())


def test_workflow_ctx_ref_unknown_node_fails_fast(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {
                "id": "b",
                "demand": "b.yaml",
                "init_vars": {"token": {"$ctx": {"node": "nope", "key": "total_rows"}}},
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    with pytest.raises(ScalimWorkflowConfigError, match="Unknown ctx node"):
        _ = run_workflow(str(wf), options=_run_options())


def test_workflow_ctx_guardrails_fail_fast() -> None:
    root = {
        "workflow": {
            "runs": [{"id": "a", "demand": "a.yaml"}],
            "options": {"ctx": {"max_value_bytes": 1, "max_bytes": 10}},
        }
    }

    with pytest.raises(ScalimWorkflowConfigError, match="moved out of workflow YAML") as excinfo:
        _ = load_workflow_config_from_mapping(root)
    assert excinfo.value.path == "workflow.options"


def test_workflow_primary_only_cancels_downstream_when_dep_fails(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="down.yaml",
        name="down",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "bad", "demand": "bad.yaml"},
            {"id": "down", "demand": "down.yaml", "depends_on": ["bad"]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    runtime = WorkflowRuntimeOptions(execution=WorkflowExecutionOptions(max_concurrency=2, failure_policy="primary_only"))
    result = run_workflow(str(wf), options=_run_options(components=[recorder]), workflow_runtime_options=runtime)
    assert [o.run_id for o in result.outcomes] == ["bad", "down"]
    assert result.outcomes[0].error is not None
    assert result.outcomes[1].result is None
    assert result.outcomes[1].error is not None
    assert "dependency" in result.outcomes[1].error.message.lower()

    cancelled = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_CANCELLED]
    assert len(cancelled) == 1
    assert cancelled[0].payload.workflow_node_id == "down"
    assert cancelled[0].payload.reason == "dependency_failed"


def test_workflow_observability_bridge_injects_meta_and_emits_workflow_events(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="slow.yaml",
        name="slow",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="fast.yaml",
        name="fast",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "slow", "demand": "slow.yaml"}, {"id": "fast", "demand": "fast.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    _ = run_workflow(str(wf), options=_run_options(components=[recorder]))

    workflow_start = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_START]
    workflow_end = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]
    pipeline_start = [e for e in recorder.events if e.event_type == EVENT_PIPELINE_START]

    assert {e.payload.workflow_node_id for e in workflow_start} == {"slow", "fast"}
    assert {e.payload.workflow_node_id for e in workflow_end} == {"slow", "fast"}
    assert {e.meta.get("workflow_node_id") for e in pipeline_start} == {"slow", "fast"}

    workflow_exec_id = workflow_start[0].meta.get("workflow_exec_id")
    assert workflow_exec_id

    for event in workflow_start + workflow_end:
        assert event.run_id == workflow_exec_id
        assert event.meta.get("workflow_exec_id") == workflow_exec_id
        assert event.meta.get("workflow_node_id") in {"slow", "fast"}
        assert event.payload.workflow_exec_id == workflow_exec_id

    for event in pipeline_start:
        assert event.meta.get("workflow_exec_id") == workflow_exec_id
        assert event.meta.get("workflow_node_id") in {"slow", "fast"}
        assert event.run_id != workflow_exec_id
        assert str(event.run_id).startswith("run_")
        assert event.seq >= 1


def test_workflow_observability_bridge_emits_cancelled_reason_for_all_fail(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}, {"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    recorder = _WorkflowEventRecorder()
    runtime = _workflow_runtime_options(max_concurrency=1, failure_policy="all_fail")
    with pytest.raises(Exception):
        _ = run_workflow(str(wf), options=_run_options(components=[recorder]), workflow_runtime_options=runtime)

    cancelled = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_CANCELLED]
    assert len(cancelled) == 1
    assert cancelled[0].payload.workflow_node_id == "ok"
    assert cancelled[0].payload.reason == "policy_all_fail"
    assert "failure_policy=all_fail" in cancelled[0].payload.message


def test_workflow_all_fail_skips_already_cancelled_nodes_on_late_terminal(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="slow.yaml",
        name="slow",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="child.yaml",
        name="child",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "bad", "demand": "bad.yaml"},
            {"id": "slow", "demand": "slow.yaml"},
            {"id": "child", "demand": "child.yaml", "depends_on": ["bad", "slow"]},
        ],
        max_concurrency=2,
        failure_policy="all_fail",
    )

    recorder = _WorkflowEventRecorder()
    runtime = _workflow_runtime_options(max_concurrency=2, failure_policy="all_fail")
    with pytest.raises(Exception):
        _ = run_workflow(str(wf), options=_run_options(components=[recorder]), workflow_runtime_options=runtime)

    cancelled = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_CANCELLED]
    assert len(cancelled) == 1
    assert cancelled[0].payload.workflow_node_id == "child"
    assert cancelled[0].payload.reason == "policy_all_fail"

    ended = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]
    assert any(e.payload.workflow_node_id == "slow" for e in ended)


def test_workflow_all_fail_compile_error_marks_end_and_cancels_pending(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:missing_loader",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}, {"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=2,
        failure_policy="all_fail",
    )

    recorder = _WorkflowEventRecorder()
    with pytest.raises(Exception) as excinfo:
        _ = run_workflow(str(wf), options=_run_options(components=[recorder]))
    assert "run_id=bad" in str(excinfo.value)

    ended = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]
    by_node = {e.payload.workflow_node_id: e for e in ended}
    assert by_node["bad"].payload.status == "error"

    cancelled = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_CANCELLED]
    assert len(cancelled) == 1
    assert cancelled[0].payload.workflow_node_id == "ok"
    assert cancelled[0].payload.reason == "policy_all_fail"


def test_observer_manager_workflow_attribution_keys_fail_fast_on_override() -> None:
    manager = ObserverManager(
        run_id="x",
        event_meta_defaults={
            "workflow_exec_id": "wf_1",
            "workflow_node_id": "a",
        },
    )
    with pytest.raises(ValueError, match="workflow attribution"):
        manager.emit_event("x", 1, meta={"workflow_node_id": "b"})


def test_observer_manager_workflow_attribution_meta_merges_non_reserved_meta() -> None:
    manager = ObserverManager(
        run_id="x",
        event_meta_defaults={
            "workflow_exec_id": "wf_1",
            "workflow_node_id": "a",
        },
    )
    event = manager.emit_event("x", 1, meta={"worker_id": "w1"})
    assert event.meta == {
        "workflow_exec_id": "wf_1",
        "workflow_node_id": "a",
        "worker_id": "w1",
    }


def test_workflow_observability_bridge_registers_hooks(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="fast.yaml",
        name="fast",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "fast", "demand": "fast.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    class _HookRecorder(BaseHook):
        def __init__(self) -> None:
            self._lock = threading.Lock()
            self.events: List[Any] = []

        def on_event(self, event) -> None:  # type: ignore[override]
            with self._lock:
                self.events.append(event)

    hook = _HookRecorder()
    _ = run_workflow(str(wf), options=_run_options(components=[hook]))
    assert any(e.event_type == EVENT_WORKFLOW_NODE_START for e in hook.events)


def test_cache_pool_reuses_preload_forever_across_runs(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    runtime = _workflow_runtime_options(max_concurrency=1, failure_policy="primary_only", cache_pool_max_entries=10)
    result = run_workflow(str(wf), options=_run_options(components=[recorder]), workflow_runtime_options=runtime)
    assert not result.errors()
    assert workflow_loaders.preload_calls() == 1

    acquires = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_ACQUIRE]
    acquires.sort(key=lambda e: e.seq)
    assert [e.payload.cache_status for e in acquires] == ["miss", "hit"]
    assert {e.payload.workflow_node_id for e in acquires} == {"a", "b"}

    releases = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_RELEASE]
    assert {e.payload.workflow_node_id for e in releases} == {"a", "b"}

    evicts = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_EVICT]
    assert len(evicts) == 1
    assert evicts[0].payload.reason == "refcount_zero"
    assert evicts[0].payload.workflow_node_id == "b"


def test_cache_pool_conflict_policy_error_fails_fast(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table_alt",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        runtime = _workflow_runtime_options(max_concurrency=1, failure_policy="primary_only", cache_pool_max_entries=10)
        _ = run_workflow(str(wf), options=_run_options(), workflow_runtime_options=runtime)

    msg = str(excinfo.value)
    assert "cache_pool signature conflict" in msg and "diff=" in msg
    assert "loader_ref" in msg
    assert workflow_loaders.preload_calls() == 1


def test_cache_pool_signature_uses_rendered_params_for_reuse(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        preload_params_yaml=(
            """
params:
  p:
    $init_var: token
"""
        ).strip(),
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        preload_params_yaml=(
            """
params:
  p: 1
"""
        ).strip(),
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    runtime = _workflow_runtime_options(max_concurrency=1, failure_policy="primary_only", cache_pool_max_entries=10)
    result = run_workflow(
        str(wf),
        options=_run_options(components=[recorder], init_vars={"token": 1}),
        workflow_runtime_options=runtime,
    )
    assert not result.errors()
    assert workflow_loaders.preload_calls() == 1

    acquires = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_ACQUIRE]
    assert acquires
    assert all(e.payload.conflict_detected is False for e in acquires)


def test_cache_pool_signature_includes_lookup_cast_meta(tmp_path: Path) -> None:
    demand_template = (
        """
name: demo

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_main_fast"
  fields:
    ref_id:
      name: ref_id

sources:
  preload:
    loader: "tests.fixtures.workflow_loaders:load_preload_table"
    key: id
    cache_mode: preload_forever
    lookup_cast:
      name: sep_first
      sep: "{sep}"
    fields:
      value:
        name: value
        relation: main_to_preload

relations:
  main_to_preload:
    steps:
      - from: main.ref_id
        to: preload.id
"""
    ).lstrip()

    _ = _write_text(tmp_path / "a.yaml", demand_template.format(sep=","))
    _ = _write_text(tmp_path / "b.yaml", demand_template.format(sep=";"))

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        runtime = _workflow_runtime_options(max_concurrency=1, failure_policy="primary_only", cache_pool_max_entries=10)
        _ = run_workflow(str(wf), options=_run_options(), workflow_runtime_options=runtime)
    assert "lookup_cast" in str(excinfo.value)


def test_cache_pool_budget_fail_fast_raises(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        preload_source_id="preload_a",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        preload_source_id="preload_b",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        runtime = _workflow_runtime_options(max_concurrency=2, failure_policy="primary_only", cache_pool_max_entries=1)
        _ = run_workflow(str(wf), options=_run_options(), workflow_runtime_options=runtime)
    assert "over budget" in str(excinfo.value).lower()
    assert workflow_loaders.preload_calls() == 1


def test_cache_pool_calls_node_done_on_compile_error_and_cancelled_dependents(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:missing_loader",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}, {"id": "ok", "demand": "ok.yaml", "depends_on": ["bad"]}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    runtime = _workflow_runtime_options(max_concurrency=1, failure_policy="primary_only", cache_pool_max_entries=10)
    result = run_workflow(str(wf), options=_run_options(), workflow_runtime_options=runtime)
    assert [o.run_id for o in result.outcomes] == ["bad", "ok"]
    assert result.outcomes[0].error is not None
    assert result.outcomes[1].error is not None
    assert result.outcomes[1].error.exc_type == "WorkflowCancelled"


def test_cache_pool_calls_node_done_on_runtime_error(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    runtime = _workflow_runtime_options(max_concurrency=1, failure_policy="primary_only", cache_pool_max_entries=10)
    result = run_workflow(str(wf), options=_run_options(), workflow_runtime_options=runtime)
    assert [o.run_id for o in result.outcomes] == ["bad"]
    assert result.outcomes[0].error is not None
    assert result.outcomes[0].error.exc_type == "ValueError"


def test_workflow_schema_validation() -> None:
    jsonschema = pytest.importorskip("jsonschema")
    yaml = pytest.importorskip("yaml")
    repo_root = Path(__file__).resolve().parents[2]
    schema_path = repo_root / "src" / "scalim" / "dsl" / "yaml_dsl" / "schema" / "workflow.gen.json"
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    ok = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
"""
        ).lstrip()
    )
    jsonschema.validate(ok, schema)

    ok_ctx_and_depends_on = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      depends_on: [a]
      init_vars:
        token:
          $ctx:
            node: a
            key: total_rows
"""
        ).lstrip()
    )
    jsonschema.validate(ok_ctx_and_depends_on, schema)

    bad_cache_pool = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
  options:
    max_concurrency: 1
    failure_policy: all_fail
    cache_pool:
      conflict_policy: error
      release_policy: dag_refcount
      budget:
        max_entries: 1
        over_budget_policy: fail_fast
"""
        ).lstrip()
    )
    with pytest.raises(jsonschema.exceptions.ValidationError):
        jsonschema.validate(bad_cache_pool, schema)

    bad_ctx_guardrails = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
  options:
    ctx:
      max_value_bytes: 65536
      max_bytes: 1048576
"""
        ).lstrip()
    )
    with pytest.raises(jsonschema.exceptions.ValidationError):
        jsonschema.validate(bad_ctx_guardrails, schema)

    bad_legacy_deps = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
      deps: [b]
"""
        ).lstrip()
    )
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(bad_legacy_deps, schema)

    bad = yaml.safe_load(
        (
            """
workflow:
  runs: []
"""
        ).lstrip()
    )
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(bad, schema)

    bad_share_preload_cache = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
  options:
    share_preload_cache: true
"""
        ).lstrip()
    )
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(bad_share_preload_cache, schema)

    bad_cache_pool = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
  options:
    cache_pool:
      conflict_policy: nope
      release_policy: dag_refcount
      budget:
        max_entries: 1
        over_budget_policy: fail_fast
"""
        ).lstrip()
    )
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(bad_cache_pool, schema)


def test_workflow_config_error_formats_without_path() -> None:
    assert str(ScalimWorkflowConfigError("msg")) == "msg"


def test_load_workflow_config_wraps_read_errors(tmp_path: Path) -> None:
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(tmp_path))
    assert "Failed to read workflow YAML" in str(excinfo.value)


def test_load_workflow_config_wraps_yaml_parse_errors(tmp_path: Path) -> None:
    workflow_path = _write_text(tmp_path / "wf.yaml", "workflow: [\n")
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "YAML parse error" in str(excinfo.value)


def test_load_workflow_config_root_must_be_mapping(tmp_path: Path) -> None:
    workflow_path = _write_text(tmp_path / "wf.yaml", "- 1\n")
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "root must be a mapping" in str(excinfo.value)


def test_resolve_workflow_demand_path_requires_non_empty_string(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path("", workflow_yaml_path=str(wf))
    assert "run.demand must be a non-empty string" in str(excinfo.value)


def test_resolve_workflow_demand_path_supports_at_alias(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    demand = resolve_workflow_demand_path(
        "@/a.yaml",
        workflow_yaml_path=str(wf),
        path_aliases={"@": str(tmp_path)},
        run_id="r1",
    )
    assert demand == (tmp_path / "a.yaml").resolve(strict=False)


def test_resolve_workflow_demand_path_supports_named_alias(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    demand = resolve_workflow_demand_path(
        "DATA:/b.yaml",
        workflow_yaml_path=str(wf),
        path_aliases={"DATA": str(tmp_path)},
        run_id="r1",
    )
    assert demand == (tmp_path / "b.yaml").resolve(strict=False)


def test_resolve_workflow_demand_path_rejects_unknown_alias(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "DATA:/b.yaml",
            workflow_yaml_path=str(wf),
            path_aliases={},
            run_id="r1",
        )
    assert "Unknown path alias" in str(excinfo.value)
    assert "run_id=r1" in str(excinfo.value)


def test_resolve_workflow_demand_path_unknown_alias_error_omits_run_id_when_missing(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "DATA:/b.yaml",
            workflow_yaml_path=str(wf),
            path_aliases={},
        )
    assert "Unknown path alias" in str(excinfo.value)
    assert "run_id=" not in str(excinfo.value)


def test_resolve_workflow_demand_path_rejects_empty_at_alias_path(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "@/",
            workflow_yaml_path=str(wf),
            path_aliases={"@": str(tmp_path)},
            run_id="r1",
        )
    assert "Invalid demand alias path" in str(excinfo.value)
    assert "run_id=r1" in str(excinfo.value)


def test_resolve_workflow_demand_path_empty_alias_error_omits_run_id_when_missing(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "@/",
            workflow_yaml_path=str(wf),
            path_aliases={"@": str(tmp_path)},
        )
    assert "Invalid demand alias path" in str(excinfo.value)
    assert "run_id=" not in str(excinfo.value)


def test_resolve_workflow_demand_path_supports_absolute_paths(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    abs_path = (tmp_path / "abs.yaml").resolve(strict=False)
    demand = resolve_workflow_demand_path(str(abs_path), workflow_yaml_path=str(wf))
    assert demand == abs_path


def test_resolve_workflow_demand_path_supports_paths_relative_to_workflow(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    demand = resolve_workflow_demand_path("rel.yaml", workflow_yaml_path=str(wf))
    assert demand == (tmp_path / "rel.yaml").resolve(strict=False)


def test_resolve_workflow_demand_path_rejects_relative_escape_by_default(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path("../escape.yaml", workflow_yaml_path=str(wf))
    assert "YAML path escapes allowed roots" in str(excinfo.value)


def test_resolve_workflow_demand_path_escape_error_includes_run_id(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path("../escape.yaml", workflow_yaml_path=str(wf), run_id="r1")
    assert "YAML path escapes allowed roots" in str(excinfo.value)
    assert "run_id=r1" in str(excinfo.value)


def test_resolve_workflow_demand_path_rejects_absolute_escape_by_default(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    outside = (tmp_path.parent / "escape.yaml").resolve(strict=False)
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(str(outside), workflow_yaml_path=str(wf))
    assert "YAML path escapes allowed roots" in str(excinfo.value)


def test_resolve_workflow_demand_path_rejects_alias_escape_by_default(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "DATA:/escape.yaml",
            workflow_yaml_path=str(wf),
            path_aliases={"DATA": str(tmp_path.parent)},
            run_id="r1",
        )
    assert "YAML path escapes allowed roots" in str(excinfo.value)
    assert "alias=DATA" in str(excinfo.value)


def test_resolve_workflow_demand_path_alias_escape_error_omits_run_id_when_missing(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "DATA:/escape.yaml",
            workflow_yaml_path=str(wf),
            path_aliases={"DATA": str(tmp_path.parent)},
        )
    assert "YAML path escapes allowed roots" in str(excinfo.value)
    assert "alias=DATA" in str(excinfo.value)
    assert "run_id=" not in str(excinfo.value)


def test_resolve_workflow_demand_path_allows_escape_with_explicit_allowed_roots(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    allowed_yaml_roots = [tmp_path.parent]
    demand = resolve_workflow_demand_path(
        "../escape.yaml",
        workflow_yaml_path=str(wf),
        allowed_yaml_roots=allowed_yaml_roots,
    )
    assert demand == (tmp_path.parent / "escape.yaml").resolve(strict=False)

    outside = (tmp_path.parent / "abs.yaml").resolve(strict=False)
    demand = resolve_workflow_demand_path(
        str(outside),
        workflow_yaml_path=str(wf),
        allowed_yaml_roots=allowed_yaml_roots,
    )
    assert demand == outside

    demand = resolve_workflow_demand_path(
        "DATA:/alias.yaml",
        workflow_yaml_path=str(wf),
        path_aliases={"DATA": str(tmp_path.parent)},
        run_id="r1",
        allowed_yaml_roots=allowed_yaml_roots,
    )
    assert demand == (tmp_path.parent / "alias.yaml").resolve(strict=False)


def test_resolve_workflow_demand_path_invalid_allowed_yaml_roots_is_wrapped(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    missing_root = tmp_path / "missing_root"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "rel.yaml",
            workflow_yaml_path=str(wf),
            allowed_yaml_roots=[missing_root],
            run_id="r1",
        )
    assert "Invalid allowed_yaml_roots" in str(excinfo.value)
    assert "run_id=r1" in str(excinfo.value)


def test_resolve_workflow_demand_path_invalid_allowed_yaml_roots_wrap_omits_run_id_when_missing(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    missing_root = tmp_path / "missing_root"
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "rel.yaml",
            workflow_yaml_path=str(wf),
            allowed_yaml_roots=[missing_root],
        )
    assert "Invalid allowed_yaml_roots" in str(excinfo.value)
    assert "run_id=" not in str(excinfo.value)


def test_validate_workflow_yaml_text_json_variants() -> None:
    bad_parse = json.loads(validate_workflow_yaml_text_json("workflow: [\n"))
    assert bad_parse["ok"] is False

    empty = json.loads(validate_workflow_yaml_text_json(""))
    assert empty["ok"] is False
    assert "empty" in empty["errors"][0]["message"].lower()

    root_not_mapping = json.loads(validate_workflow_yaml_text_json("- 1\n"))
    assert root_not_mapping["ok"] is False

    semantic_error = json.loads(validate_workflow_yaml_text_json("workflow: {}\n"))
    assert semantic_error["ok"] is False
    assert semantic_error["errors"]

    ok = json.loads(
        validate_workflow_yaml_text_json(
            (
                """
workflow:
  runs:
    - id: a
      demand: a.yaml
"""
            ).lstrip()
        )
    )
    assert ok["ok"] is True


@pytest.mark.parametrize(
    "bad_mapping",
    [
        {},
        {"workflow": {}},
        {"workflow": {"runs": []}},
        {"workflow": {"runs": [1]}},
        {"workflow": {"runs": [{"id": "", "demand": "a.yaml"}]}},
        {"workflow": {"runs": [{"id": "a", "demand": ""}]}},
        {"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": []}},
        {"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": {"max_concurrency": True}}},
        {"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": {"max_concurrency": "x"}}},
        {"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": {"max_concurrency": 0}}},
        {"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": {"failure_policy": "nope"}}},
    ],
)
def test_load_workflow_config_from_mapping_rejects_invalid_structures(bad_mapping: dict) -> None:
    with pytest.raises(ScalimWorkflowConfigError):
        _ = load_workflow_config_from_mapping(bad_mapping)


@pytest.mark.parametrize(
    "options_value",
    [
        None,
        {},
        {"max_concurrency": 2},
        {"failure_policy": "all_fail"},
        {"share_preload_cache": True},
        {"cache_pool": {"budget": {"max_entries": 16}}},
        {"ctx": {"max_value_bytes": 2, "max_bytes": 3}},
        {"resources_wait": {"max_wait_s": 12.5}},
        {"output_staging": {"dir_name": ".staging"}},
    ],
)
def test_load_workflow_config_from_mapping_rejects_workflow_options(options_value: object) -> None:
    with pytest.raises(ScalimWorkflowConfigError) as excinfo:
        _ = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": options_value}})
    assert excinfo.value.path == "workflow.options"


def test_load_workflow_config_from_mapping_rejects_legacy_deps_field() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match="run\\.deps was removed; use run\\.depends_on") as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [
                        {"id": "a", "demand": "a.yaml", "deps": ["b"]},
                    ]
                }
            }
        )
    assert excinfo.value.path == "workflow.runs.0.deps"


def test_load_workflow_config_from_mapping_rejects_depends_on_not_list() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match="run\\.depends_on must be a list"):
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [
                        {"id": "a", "demand": "a.yaml", "depends_on": "nope"},
                    ]
                }
            }
        )


def test_load_workflow_config_from_mapping_rejects_depends_on_empty_item() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match="depends_on items must be non-empty"):
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [
                        {"id": "a", "demand": "a.yaml"},
                        {"id": "b", "demand": "b.yaml", "depends_on": [""]},
                    ]
                }
            }
        )


def test_load_workflow_config_from_mapping_dedups_depends_on_preserves_first_order() -> None:
    cfg = load_workflow_config_from_mapping(
        {
            "workflow": {
                "runs": [
                    {"id": "a", "demand": "a.yaml"},
                    {"id": "b", "demand": "b.yaml", "depends_on": ["a", "a"]},
                ]
            }
        }
    )
    assert cfg.runs[1].depends_on == ("a",)


def test_load_workflow_config_from_mapping_accepts_forward_depends_on() -> None:
    cfg = load_workflow_config_from_mapping(
        {
            "workflow": {
                "runs": [
                    {"id": "b", "demand": "b.yaml", "depends_on": ["a"]},
                    {"id": "a", "demand": "a.yaml"},
                ]
            }
        }
    )
    assert [r.id for r in cfg.runs] == ["b", "a"]


def test_load_workflow_config_from_mapping_rejects_init_vars_not_mapping() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match="run\\.init_vars must be a mapping") as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [
                        {"id": "a", "demand": "a.yaml", "init_vars": []},
                    ]
                }
            }
        )
    assert excinfo.value.path == "workflow.runs.0.init_vars"


def test_load_workflow_config_from_mapping_rejects_init_vars_key_invalid() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match="run\\.init_vars keys must be non-empty strings") as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [
                        {"id": "a", "demand": "a.yaml", "init_vars": {"": 1}},
                    ]
                }
            }
        )
    assert excinfo.value.path == "workflow.runs.0.init_vars"


def test_load_workflow_config_from_mapping_accepts_null_resources() -> None:
    cfg = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "resources": None}})
    assert cfg.resources.books == {}


def test_load_workflow_config_from_mapping_accepts_resources_mapping() -> None:
    cfg = load_workflow_config_from_mapping(
        {
            "workflow": {
                "runs": [{"id": "a", "demand": "a.yaml"}],
                "resources": {
                    "books": {"report": {"kind": "xlsx_file", "path": "./out"}},
                },
            }
        }
    )
    assert cfg.resources.books["report"].path == "./out"


def test_load_workflow_config_from_mapping_rejects_resources_not_mapping() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match="workflow.resources must be a mapping"):
        _ = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "resources": []}})


def test_load_workflow_config_from_mapping_rejects_resources_key_invalid() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match="workflow\\.resources keys must be non-empty"):
        _ = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "resources": {"": {"kind": "x"}}}})


def test_load_workflow_config_from_mapping_rejects_self_depends_on() -> None:
    with pytest.raises(ScalimWorkflowConfigError, match="self dependency"):
        _ = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml", "depends_on": ["a"]}]}})


def test_run_workflow_requires_workflow_path(tmp_path: Path) -> None:
    _ = tmp_path
    with pytest.raises(ScalimWorkflowConfigError):
        _ = run_workflow("", options=_run_options())


def test_run_workflow_primary_only_submits_pending_after_failure(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok1.yaml",
        name="ok1",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok2.yaml",
        name="ok2",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}, {"id": "ok1", "demand": "ok1.yaml"}, {"id": "ok2", "demand": "ok2.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    runtime = WorkflowRuntimeOptions(execution=WorkflowExecutionOptions(max_concurrency=1, failure_policy="primary_only"))
    result = run_workflow(str(wf), options=_run_options(), workflow_runtime_options=runtime)
    assert [o.run_id for o in result.outcomes] == ["bad", "ok1", "ok2"]
    assert result.outcomes[0].error is not None
    assert result.outcomes[1].result is not None
    assert result.outcomes[2].result is not None


def test_run_workflow_all_fail_cancels_pending_queue(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}, {"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    with pytest.raises(Exception) as excinfo:
        _ = run_workflow(str(wf), options=_run_options())
    assert "run_id=bad" in str(excinfo.value)


def test_workflow_entrypoints_artifacts_directory_enforces_visibility() -> None:
    from scalim.spec.ir._workflow import (
        WorkflowArtifactsIr,
        WorkflowEdgeIr,
        WorkflowIr,
        WorkflowNodeIr,
        WorkflowNodeType,
        WorkflowOptionsIr,
    )

    ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(node_id="a", node_type=WorkflowNodeType.DEMAND, decl_order=0, deps=(), demand_path="a.yaml"),
            WorkflowNodeIr(node_id="b", node_type=WorkflowNodeType.DEMAND, decl_order=1, deps=("a",), demand_path="b.yaml"),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(),
        resources=(),
        artifacts=WorkflowArtifactsIr(
            slots_by_node_id={
                "a": ("output_path", "outputs"),
                "b": ("output_path", "outputs"),
            }
        ),
    )
    artifacts_dir = workflow_execute_mod.WorkflowArtifactsDirectory(ir)
    artifacts_dir.publish("a", "x", 1)
    assert artifacts_dir.get("b", "a", "x") == 1

    artifacts_dir.publish("b", "y", 2)
    assert artifacts_dir.get("b", "b", "y") == 2

    with pytest.raises(ValueError, match="not visible"):
        _ = artifacts_dir.get("a", "b", "y")

    with pytest.raises(KeyError, match="Unknown artifact"):
        _ = artifacts_dir.get("b", "a", "missing")


def test_workflow_entrypoints_ensure_json_like_variants() -> None:
    assert workflow_execute_mod.ensure_json_like(None, path="x") is None
    assert workflow_execute_mod.ensure_json_like(True, path="x") is True
    assert workflow_execute_mod.ensure_json_like(1, path="x") == 1
    assert workflow_execute_mod.ensure_json_like("s", path="x") == "s"
    assert workflow_execute_mod.ensure_json_like(1.5, path="x") == 1.5

    assert workflow_execute_mod.ensure_json_like([1, {"k": 2}], path="x") == [1, {"k": 2}]
    assert workflow_execute_mod.ensure_json_like((1, 2), path="x") == [1, 2]

    with pytest.raises(WorkflowRuntimeConfigError, match="finite"):
        _ = workflow_execute_mod.ensure_json_like(float("inf"), path="x")

    with pytest.raises(WorkflowRuntimeConfigError, match="dict key"):
        _ = workflow_execute_mod.ensure_json_like({1: "x"}, path="x")

    with pytest.raises(WorkflowRuntimeConfigError, match="JSON-like"):
        _ = workflow_execute_mod.ensure_json_like(object(), path="x")


def test_workflow_entrypoints_ctx_store_does_not_enforce_size_guardrails() -> None:
    from scalim.spec.ir._workflow import (
        WorkflowArtifactsIr,
        WorkflowEdgeIr,
        WorkflowIr,
        WorkflowNodeIr,
        WorkflowNodeType,
        WorkflowOptionsIr,
    )

    ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(node_id="a", node_type=WorkflowNodeType.DEMAND, decl_order=0, deps=(), demand_path="a.yaml"),
            WorkflowNodeIr(node_id="b", node_type=WorkflowNodeType.DEMAND, decl_order=1, deps=("a",), demand_path="b.yaml"),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={"a": (), "b": ()}),
    )
    ctx_store = workflow_execute_mod.WorkflowCtxStore(ir)
    large_value = "x" * 10000
    ctx_store.publish("a", "k", large_value, path="x")
    assert ctx_store.resolve("b", node="a", key="k", path="x") == large_value


def test_workflow_entrypoints_ctx_store_resolve_errors() -> None:
    from scalim.spec.ir._workflow import (
        WorkflowArtifactsIr,
        WorkflowEdgeIr,
        WorkflowIr,
        WorkflowNodeIr,
        WorkflowNodeType,
        WorkflowOptionsIr,
    )

    ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(node_id="a", node_type=WorkflowNodeType.DEMAND, decl_order=0, deps=(), demand_path="a.yaml"),
            WorkflowNodeIr(node_id="b", node_type=WorkflowNodeType.DEMAND, decl_order=1, deps=("a",), demand_path="b.yaml"),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={"a": (), "b": ()}),
    )
    ctx_store = workflow_execute_mod.WorkflowCtxStore(ir)

    with pytest.raises(WorkflowRuntimeConfigError, match="node=self"):
        _ = ctx_store.resolve("a", node="a", key="k", path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="not visible"):
        _ = ctx_store.resolve("a", node="b", key="k", path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="Unknown ctx key"):
        _ = ctx_store.resolve("b", node="a", key="missing", path="p")


def test_workflow_entrypoints_iter_ctx_directives_variants() -> None:
    with pytest.raises(WorkflowRuntimeConfigError, match="directive must be a mapping"):
        _ = workflow_execute_mod.iter_ctx_directives({"$ctx": "nope"}, path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="\\$ctx\\.node"):
        _ = workflow_execute_mod.iter_ctx_directives({"$ctx": {"node": "", "key": "k"}}, path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="\\$ctx\\.key"):
        _ = workflow_execute_mod.iter_ctx_directives({"$ctx": {"node": "a", "key": ""}}, path="p")

    assert workflow_execute_mod.iter_ctx_directives({"outer": {"$ctx": {"node": "a", "key": "k"}}}, path="p") == [
        ("a", "k"),
    ]
    assert workflow_execute_mod.iter_ctx_directives([{"$ctx": {"node": "a", "key": "k"}}], path="p") == [
        ("a", "k"),
    ]


def test_workflow_entrypoints_render_ctx_directives_variants() -> None:
    from scalim.spec.ir._workflow import (
        WorkflowArtifactsIr,
        WorkflowEdgeIr,
        WorkflowIr,
        WorkflowNodeIr,
        WorkflowNodeType,
        WorkflowOptionsIr,
    )

    ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(node_id="a", node_type=WorkflowNodeType.DEMAND, decl_order=0, deps=(), demand_path="a.yaml"),
            WorkflowNodeIr(node_id="b", node_type=WorkflowNodeType.DEMAND, decl_order=1, deps=("a",), demand_path="b.yaml"),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={"a": (), "b": ()}),
    )
    ctx_store = workflow_execute_mod.WorkflowCtxStore(ir)
    ctx_store.publish("a", "k", 1, path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="directive must be a mapping"):
        _ = workflow_execute_mod.render_ctx_directives({"$ctx": "nope"}, consumer_node_id="b", ctx_store=ctx_store, path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="\\$ctx\\.node"):
        _ = workflow_execute_mod.render_ctx_directives(
            {"$ctx": {"node": "", "key": "k"}}, consumer_node_id="b", ctx_store=ctx_store, path="p"
        )

    with pytest.raises(WorkflowRuntimeConfigError, match="\\$ctx\\.key"):
        _ = workflow_execute_mod.render_ctx_directives(
            {"$ctx": {"node": "a", "key": ""}}, consumer_node_id="b", ctx_store=ctx_store, path="p"
        )

    assert workflow_execute_mod.render_ctx_directives(
        {"outer": {"$ctx": {"node": "a", "key": "k"}}}, consumer_node_id="b", ctx_store=ctx_store, path="p"
    ) == {"outer": 1}
    assert workflow_execute_mod.render_ctx_directives(
        [{"$ctx": {"node": "a", "key": "k"}}], consumer_node_id="b", ctx_store=ctx_store, path="p"
    ) == [1]
    assert workflow_execute_mod.render_ctx_directives(1, consumer_node_id="b", ctx_store=ctx_store, path="p") == 1


def test_cache_pool_single_run_accepts_lookup_cast_and_normalize(tmp_path: Path) -> None:
    demand_path = _write_text(
        tmp_path / "demand.yaml",
        (
            """
name: demo

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_main_fast"
  fields:
    ref_id:
      name: ref_id

sources:
  other:
    loader: "tests.fixtures.workflow_loaders:load_preload_table"
    key: id
    cache_mode: none
  preload:
    loader: "tests.fixtures.workflow_loaders:load_preload_table"
    key: id
    cache_mode: preload_forever
    lookup_cast:
      name: sep_first
      sep: ","
    normalize:
      kind: index_by_key
      key_field: id
      on_conflict: last
    fields:
      value:
        name: value
        relation: main_to_preload

relations:
  main_to_preload:
    steps:
      - from: main.ref_id
        to: preload.id
"""
        ).lstrip(),
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": str(demand_path)}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    runtime = _workflow_runtime_options(max_concurrency=1, failure_policy="primary_only", cache_pool_max_entries=10)
    result = run_workflow(str(wf), options=_run_options(components=[recorder]), workflow_runtime_options=runtime)
    assert not result.errors()
    assert workflow_loaders.preload_calls() == 1

    acquires = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_ACQUIRE]
    assert len(acquires) == 1
    assert acquires[0].payload.cache_status == "miss"

    evicts = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_EVICT]
    assert len(evicts) == 1
    assert evicts[0].payload.reason == "refcount_zero"


def _read_xlsx_rows(path: Path, sheet: str) -> List[List[object]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb[str(sheet)]
        rows: List[List[object]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else v for v in row])
        return rows
    finally:
        wb.close()


def _read_xlsx_sheetnames(path: Path) -> List[str]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _read_csv_rows(path: Path) -> List[List[str]]:
    import csv

    with path.open("r", encoding="utf-8", newline="") as handle:
        return [list(r) for r in csv.reader(handle)]


def _latest_version_id(out_root: Path) -> str:
    latest = versioned_outputs.read_latest(out_root)
    version_id = latest.get("version_id")
    assert isinstance(version_id, str)
    return str(version_id)


def _latest_book_path(out_root: Path, *, book_id: str) -> Path:
    version_id = _latest_version_id(out_root)
    return out_root / "versions" / version_id / "books" / "{}.xlsx".format(str(book_id))


def test_workflow_shared_workbook_sheet_writes_commit_and_emit_events(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="A",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        book_id="report",
        sheet="B",
        field_ids=["id", "value"],
    )

    out_root = tmp_path / "out"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={"books": {"report": {"kind": "xlsx_file", "path": str(out_root)}}},
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), options=_run_options(components=[recorder]))
    assert not result.errors()
    workbook_path = _latest_book_path(out_root, book_id="report")
    assert workbook_path.exists()
    assert list(out_root.rglob("*.scalim.lock")) == []

    assert _read_xlsx_sheetnames(workbook_path) == ["A", "B"]
    assert _read_xlsx_rows(workbook_path, "A") == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]
    assert _read_xlsx_rows(workbook_path, "B") == [
        ["id", "value"],
        ["b1", "B1"],
        ["b2", "B2"],
    ]

    creates = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_CREATE]
    writes = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_WRITE]
    commits = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_COMMIT]
    assert len(creates) == 1
    assert creates[0].payload.resource_id == "report"
    assert creates[0].payload.workflow_node_id == "__wf__write.a.0"
    assert len([e for e in writes if e.payload.resource_id == "report"]) == 2
    assert len([e for e in commits if e.payload.resource_id == "report"]) == 1
    assert commits[0].payload.workflow_node_id == "__wf__write.b.0"


def test_workflow_sheetbook_resources_export_xlsx_and_emit_events(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="A",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        book_id="report",
        sheet="B",
        field_ids=["id", "value"],
    )

    out_root = tmp_path / "out"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_memory",
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(out_root)},
                }
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), options=_run_options(components=[recorder]))
    assert not result.errors()
    export_path = _latest_book_path(out_root, book_id="report")
    assert export_path.exists()
    assert list(out_root.rglob("*.scalim.lock")) == []

    assert _read_xlsx_sheetnames(export_path) == ["A", "B"]
    assert _read_xlsx_rows(export_path, "A") == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]
    assert _read_xlsx_rows(export_path, "B") == [
        ["id", "value"],
        ["b1", "B1"],
        ["b2", "B2"],
    ]

    creates = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_CREATE]
    writes = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_WRITE]
    commits = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_COMMIT]
    assert len(creates) == 1
    assert creates[0].payload.resource_type == "sheetbook"
    assert creates[0].payload.resource_id == "report"
    assert creates[0].payload.workflow_node_id == "__wf__write.a.0"
    assert len([e for e in writes if e.payload.resource_type == "sheetbook" and e.payload.resource_id == "report"]) == 2
    assert len([e for e in commits if e.payload.resource_type == "sheetbook" and e.payload.resource_id == "report"]) == 1
    assert commits[0].payload.workflow_node_id == "__wf__write.b.0"


def test_workflow_run_writes_two_outputs_to_same_sheetbook_sheets(tmp_path: Path) -> None:
    _ = _write_text(
        tmp_path / "report.yaml",
        (
            """
name: report

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {extract: id}
    value: {extract: value}

outputs:
  - name: metrics
    to: {book: report, sheet: Metrics}
    fields: ["id"]
  - name: detail
    to: {book: report, sheet: Detail}
    fields: ["id", "value"]
"""
        ).lstrip(),
    )

    out_root = tmp_path / "out"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_memory",
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(out_root)},
                    "write_defaults": {"mode": "sheet"},
                }
            }
        },
        runs=[
            {"id": "report", "demand": "report.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), options=_run_options(components=[recorder]))
    assert not result.errors()
    export_path = _latest_book_path(out_root, book_id="report")
    assert export_path.exists()
    assert _read_xlsx_sheetnames(export_path) == ["Metrics", "Detail"]
    assert _read_xlsx_rows(export_path, "Metrics") == [
        ["id"],
        ["a1"],
        ["a2"],
    ]
    assert _read_xlsx_rows(export_path, "Detail") == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]

    write_events = [
        e
        for e in recorder.events
        if e.event_type == EVENT_WORKFLOW_RESOURCE_WRITE and e.payload.resource_type == "sheetbook" and e.payload.resource_id == "report"
    ]
    assert [e.payload.workflow_node_id for e in write_events] == ["__wf__write.report.0", "__wf__write.report.1"]


def test_workflow_sheetbook_loader_consumes_rows_and_enforces_visibility(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="Orders",
        field_ids=["id", "value"],
    )

    consume_root = tmp_path / "consume_out"
    _ = _write_table_demand_yaml_from_sheetbook_loader(
        tmp_path,
        file_name="consume.yaml",
        name="consume",
        init_var_name="orders_sheet_ref",
        output_name="detail",
        output_root=consume_root,
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        resources={"books": {"report": {"kind": "xlsx_memory", "budget": {"max_sheets": 8, "max_total_cells": 1000}}}},
        runs=[
            {
                "id": "a",
                "demand": "a.yaml",
            },
            {
                "id": "consume",
                "demand": "consume.yaml",
                "depends_on": ["a"],
                "init_vars": {"orders_sheet_ref": {"node": "a", "book": "report", "sheet": "Orders"}},
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    runtime = WorkflowRuntimeOptions(execution=WorkflowExecutionOptions(max_concurrency=2, failure_policy="primary_only"))
    result = run_workflow(
        str(wf),
        options=_run_options(allowed_modules=_ALLOWED_MODULES_WITH_SHEETBOOK),
        workflow_runtime_options=runtime,
    )
    assert not result.errors()
    outcomes = {str(o.run_id): o for o in result.outcomes}
    consume_outcome = outcomes["consume"]
    assert isinstance(consume_outcome.result, DemandRunResult)
    assert consume_outcome.result.core.outputs is not None
    consume_out_path = Path(str(consume_outcome.result.core.outputs["detail"]))
    assert consume_out_path.exists()
    assert _read_csv_rows(consume_out_path) == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]

    # Outside depends_on closure: MUST fail-fast.
    wf_bad = _write_workflow_yaml(
        tmp_path,
        resources={"books": {"report": {"kind": "xlsx_memory", "budget": {"max_sheets": 8, "max_total_cells": 1000}}}},
        runs=[
            {
                "id": "a",
                "demand": "a.yaml",
            },
            {
                "id": "c",
                "demand": "consume.yaml",
                "init_vars": {"orders_sheet_ref": {"node": "a", "book": "report", "sheet": "Orders"}},
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    bad = run_workflow(
        str(wf_bad),
        options=_run_options(allowed_modules=_ALLOWED_MODULES_WITH_SHEETBOOK),
        workflow_runtime_options=runtime,
    )
    errs = {e.run_id: e for e in bad.errors()}
    assert "c" in errs
    assert "declare depends_on" in errs["c"].message


def test_workflow_xlsx_memory_keeps_loader_keys_canonical_while_exporting_display_headers(tmp_path: Path) -> None:
    _ = _write_text(
        tmp_path / "producer.yaml",
        (
            """
name: producer

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id:
      extract: id
      name: Order ID
    value:
      extract: value
      name: Display Value

outputs:
  - name: detail
    to:
      book: report
      sheet: Orders
    write:
      header_fields_output_by: name
    fields: ["id", "value"]
"""
        ).lstrip(),
    )

    consume_root = tmp_path / "consume_out"
    _ = _write_table_demand_yaml_from_sheetbook_loader(
        tmp_path,
        file_name="consume_named.yaml",
        name="consume_named",
        init_var_name="orders_sheet_ref",
        output_name="detail",
        output_root=consume_root,
        field_ids=["id", "value"],
    )

    out_root = tmp_path / "out"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_memory",
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(out_root)},
                }
            }
        },
        runs=[
            {"id": "producer", "demand": "producer.yaml"},
            {
                "id": "consume_named",
                "demand": "consume_named.yaml",
                "depends_on": ["producer"],
                "init_vars": {"orders_sheet_ref": {"node": "producer", "book": "report", "sheet": "Orders"}},
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), options=_run_options(allowed_modules=_ALLOWED_MODULES_WITH_SHEETBOOK))
    assert not result.errors()
    outcomes = {str(o.run_id): o for o in result.outcomes}
    consume_outcome = outcomes["consume_named"]
    assert isinstance(consume_outcome.result, DemandRunResult)
    assert consume_outcome.result.core.outputs is not None
    consume_path = Path(str(consume_outcome.result.core.outputs["detail"]))
    assert consume_path.exists()
    assert _read_csv_rows(consume_path) == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]
    export_path = _latest_book_path(out_root, book_id="report")
    assert _read_xlsx_rows(export_path, "Orders") == [
        ["Order ID", "Display Value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]


def test_workflow_xlsx_memory_preserves_typed_values_and_export_boundary(tmp_path: Path) -> None:
    _ = _write_text(
        tmp_path / "producer_typed.yaml",
        (
            """
name: producer_typed

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_typed_values"
  fields:
    order_count: {extract: order_count}
    amount: {extract: amount}
    paid: {extract: paid}
    code: {extract: code}
    raw_text: {extract: raw_text}

outputs:
  - name: detail
    to:
      book: report
      sheet: Orders
    fields: ["order_count", "amount", "paid", "code", "raw_text"]
"""
        ).lstrip(),
    )

    consume_root = tmp_path / "typed_consume_out"
    _ = _write_text(
        tmp_path / "consume_typed.yaml",
        (
            """
name: consume_typed

main_source:
  source_id: main
  loader: "scalim.workflow.loaders:book_sheet_rows"
  params:
    ref:
      $init_var: orders_sheet_ref
  fields:
    order_count: {extract: order_count}
    amount: {extract: amount}
    paid: {extract: paid}
    code: {extract: code}
    raw_text: {extract: raw_text}

fields:
  order_count_type:
    call_by: "tests.fixtures.call_by_fns:type_name(order_count)"
  amount_type:
    call_by: "tests.fixtures.call_by_fns:type_name(amount)"
  paid_type:
    call_by: "tests.fixtures.call_by_fns:type_name(paid)"

resources:
  files:
    detail_csv:
      kind: csv_file
      path: "__OUT__"

outputs:
  - name: detail
    to:
      file: detail_csv
    fields: ["order_count_type", "amount_type", "paid_type", "code", "raw_text"]
"""
        )
        .replace("__OUT__", str(consume_root))
        .lstrip(),
    )

    out_root = tmp_path / "out"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_memory",
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(out_root)},
                }
            }
        },
        runs=[
            {"id": "producer_typed", "demand": "producer_typed.yaml"},
            {
                "id": "consume_typed",
                "demand": "consume_typed.yaml",
                "depends_on": ["producer_typed"],
                "init_vars": {"orders_sheet_ref": {"node": "producer_typed", "book": "report", "sheet": "Orders"}},
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    allowed_modules = frozenset(["tests.fixtures.workflow_loaders", "tests.fixtures.call_by_fns", "scalim.workflow.loaders"])
    result = run_workflow(str(wf), options=_run_options(allowed_modules=allowed_modules))
    assert not result.errors()
    outcomes = {str(o.run_id): o for o in result.outcomes}
    consume_outcome = outcomes["consume_typed"]
    assert isinstance(consume_outcome.result, DemandRunResult)
    assert consume_outcome.result.core.outputs is not None
    consume_out_path = Path(str(consume_outcome.result.core.outputs["detail"]))
    assert consume_out_path.exists()
    assert _read_csv_rows(consume_out_path) == [
        ["order_count_type", "amount_type", "paid_type", "code", "raw_text"],
        ["int", "Decimal", "bool", "007", ""],
    ]

    export_path = _latest_book_path(out_root, book_id="report")
    exported = _read_xlsx_rows(export_path, "Orders")
    assert exported[0] == ["order_count", "amount", "paid", "code", "raw_text"]
    assert exported[1][0] == 5
    assert exported[1][1] == 1.2
    assert exported[1][2] is True
    assert exported[1][3] == "007"
    assert exported[1][4] == ""


def test_workflow_sheetbook_budget_guards_and_discard_on_failure(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="A",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        book_id="report",
        sheet="B",
        field_ids=["id", "value"],
    )

    out_root = tmp_path / "out"
    wf_max_sheets = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_memory",
                    "budget": {"max_sheets": 1, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(out_root)},
                }
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    runtime = WorkflowRuntimeOptions(execution=WorkflowExecutionOptions(max_concurrency=2, failure_policy="primary_only"))
    result_sheets = run_workflow(str(wf_max_sheets), options=_run_options(), workflow_runtime_options=runtime)
    assert result_sheets.errors()
    assert not (out_root / "manifest" / "latest.json").exists()
    assert list(out_root.rglob("*.xlsx")) == []
    assert list(out_root.rglob("*.scalim.lock")) == []

    out_root2 = tmp_path / "out2"
    wf_max_cells = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_memory",
                    "budget": {"max_sheets": 8, "max_total_cells": 3},
                    "export_xlsx": {"path": str(out_root2)},
                }
            }
        },
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    result_cells = run_workflow(str(wf_max_cells), options=_run_options(), workflow_runtime_options=runtime)
    assert result_cells.errors()
    assert not (out_root2 / "manifest" / "latest.json").exists()
    assert list(out_root2.rglob("*.xlsx")) == []
    assert list(out_root2.rglob("*.scalim.lock")) == []

    out_root3 = tmp_path / "out3"
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf_failed = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_memory",
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(out_root3)},
                }
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "bad", "demand": "bad.yaml", "depends_on": ["a"]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result_failed = run_workflow(str(wf_failed), options=_run_options(), workflow_runtime_options=runtime)
    assert result_failed.errors()
    assert not (out_root3 / "manifest" / "latest.json").exists()
    assert list(out_root3.rglob("*.xlsx")) == []
    assert list(out_root3.rglob("*.scalim.lock")) == []


def test_workflow_excel_output_collision_precheck_and_reserved_paths(tmp_path: Path) -> None:
    out_root = tmp_path / "out"
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="a_book",
        sheet="A",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        book_id="b_book",
        sheet="B",
        field_ids=["id", "value"],
    )

    wf_collision = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "a_book": {"kind": "xlsx_file", "path": str(out_root)},
                "b_book": {"kind": "xlsx_file", "path": str(out_root)},
            }
        },
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    ok = run_workflow(str(wf_collision), options=_run_options())
    assert not ok.errors()

    a_book_path = _latest_book_path(out_root, book_id="a_book")
    b_book_path = _latest_book_path(out_root, book_id="b_book")
    assert a_book_path.exists()
    assert b_book_path.exists()
    assert list(out_root.rglob("*.scalim.lock")) == []

    assert _read_xlsx_sheetnames(a_book_path) == ["A"]
    assert _read_xlsx_rows(a_book_path, "A") == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]
    assert _read_xlsx_sheetnames(b_book_path) == ["B"]
    assert _read_xlsx_rows(b_book_path, "B") == [
        ["id", "value"],
        ["b1", "B1"],
        ["b2", "B2"],
    ]

    out_root2 = tmp_path / "out2"
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="file.yaml",
        name="file",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="file_book",
        sheet="S",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="mem.yaml",
        name="mem",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="mem_book",
        sheet="S",
        field_ids=["id", "value"],
    )
    wf_reserved = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "file_book": {"kind": "xlsx_file", "path": str(out_root2)},
                "mem_book": {
                    "kind": "xlsx_memory",
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(out_root2)},
                },
            }
        },
        runs=[{"id": "file", "demand": "file.yaml"}, {"id": "mem", "demand": "mem.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    ok2 = run_workflow(str(wf_reserved), options=_run_options())
    assert not ok2.errors()
    assert _latest_book_path(out_root2, book_id="file_book").exists()
    assert _latest_book_path(out_root2, book_id="mem_book").exists()


def test_workflow_excel_output_collision_precheck_allows_dynamic_init_var_paths(tmp_path: Path) -> None:
    a_root = tmp_path / "a_out"
    b_root = tmp_path / "b_out"

    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="a_book",
        sheet="S",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="b_book",
        sheet="S",
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "a_book": {"kind": "xlsx_file", "path": {"$init_var": "a_path"}},
                "b_book": {"kind": "xlsx_file", "path": {"$init_var": "b_path"}},
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result = run_workflow(str(wf), options=_run_options(init_vars={"a_path": str(a_root), "b_path": str(b_root)}))
    assert not result.errors()
    assert _latest_book_path(a_root, book_id="a_book").exists()
    assert _latest_book_path(b_root, book_id="b_book").exists()


def test_workflow_excel_output_allows_dynamic_init_var_shared_output_root(tmp_path: Path) -> None:
    out_root = tmp_path / "out"

    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="a_book",
        sheet="S",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="b_book",
        sheet="S",
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "a_book": {"kind": "xlsx_file", "path": {"$init_var": "a_path"}},
                "b_book": {"kind": "xlsx_file", "path": {"$init_var": "b_path"}},
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result = run_workflow(str(wf), options=_run_options(init_vars={"a_path": str(out_root), "b_path": str(out_root)}))
    assert not result.errors()
    assert _latest_book_path(out_root, book_id="a_book").exists()
    assert _latest_book_path(out_root, book_id="b_book").exists()


def test_workflow_resources_rejects_export_xlsx_write_lock(tmp_path: Path) -> None:
    reserved_export = tmp_path / "reserved.xlsx"

    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="file.yaml",
        name="file",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="file_book",
        sheet="S",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="mem.yaml",
        name="mem",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="mem_book",
        sheet="S",
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "file_book": {"kind": "xlsx_file", "path": {"$init_var": "reserved_path"}},
                "mem_book": {
                    "kind": "xlsx_memory",
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": {"$init_var": "reserved_path"}, "write_lock": True},
                },
            }
        },
        runs=[{"id": "file", "demand": "file.yaml"}, {"id": "mem", "demand": "mem.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(ScalimWorkflowConfigError, match="write_lock was removed") as excinfo:
        _ = run_workflow(str(wf), options=_run_options(init_vars={"reserved_path": str(reserved_export)}))
    assert excinfo.value.path == "workflow.resources.books.mem_book.export_xlsx.write_lock"


def test_workflow_excel_output_runtime_precheck_includes_meta_and_audit_paths(tmp_path: Path) -> None:
    out_root = tmp_path / "out"

    _ = _write_text(
        tmp_path / "demand.yaml",
        (
            f"""
name: demand

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {{extract: id}}
    value: {{extract: value}}

outputs:
  - name: detail
    to: {{book: report, sheet: S}}
    fields: [id, value]
"""
        ).lstrip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        resources={"books": {"report": {"kind": "xlsx_file", "path": str(out_root), "write_defaults": {"mode": "sheet"}}}},
        runs=[{"id": "a", "demand": "demand.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    result = run_workflow(
        str(wf),
        options=_run_options(overrides=RunOverrides(output_extras=OutputExtrasOverride(meta=True, audit=True))),
    )
    assert not result.errors()
    workbook_path = _latest_book_path(out_root, book_id="report")
    assert workbook_path.exists()
    assert _read_xlsx_sheetnames(workbook_path) == ["S", "__meta__", "__audit__"]


def test_workflow_excel_output_collision_precheck_reports_demand_yaml_load_failures(tmp_path: Path) -> None:
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "missing.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    cfg = load_workflow_config(str(wf))
    with pytest.raises(ScalimWorkflowConfigError, match="Failed to load demand YAML for workflow compile"):
        _ = workflow_compile_mod.compile_workflow_ir(cfg, workflow_yaml_path=str(wf), path_aliases=None)


def test_compile_workflow_ir_enforces_rendered_yaml_max_len_during_structural_preload(tmp_path: Path) -> None:
    _ = _write_text(
        tmp_path / "demand.yaml",
        (
            """
name: "{{ big }}"

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {extract: id}

sources: {}
"""
        ).lstrip(),
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "demand.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    cfg = load_workflow_config(str(wf))

    with pytest.raises(ScalimWorkflowConfigError, match="渲染后的 YAML 文本超出上限"):
        _ = workflow_compile_mod.compile_workflow_ir(
            cfg,
            workflow_yaml_path=str(wf),
            path_aliases=None,
            template_vars={"big": "x" * 300},
            rendered_yaml_max_len=200,
        )


def test_workflow_structural_preload_does_not_import_runtime_compiler() -> None:
    repo_root = Path(__file__).resolve().parents[2]
    workflow_compile_text = (repo_root / "src/scalim/dsl/yaml_dsl/workflow_compile.py").read_text(encoding="utf-8")
    loader_text = (repo_root / "src/scalim/dsl/yaml_dsl/_internal/config_parsing/loader.py").read_text(encoding="utf-8")

    for needle in (
        "from scalim.dsl.yaml_dsl.runtime import compiler",
        "from .runtime import compiler",
        "from ..runtime import compiler",
        "import scalim.dsl.yaml_dsl.runtime.compiler",
        "import scalim.dsl.yaml_dsl.runtime.compiler as",
    ):
        assert needle not in workflow_compile_text
        assert needle not in loader_text


def test_compile_workflow_ir_skips_runtime_duplicate_name_validation(tmp_path: Path) -> None:
    _ = _write_duplicate_header_demand_yaml(tmp_path, file_name="dup.yaml", output_root=tmp_path / "out")
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "dup", "demand": "dup.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    cfg = load_workflow_config(str(wf))
    compilation = workflow_compile_mod.compile_workflow_ir(cfg, workflow_yaml_path=str(wf), path_aliases=None)
    workflow_ir = compilation.workflow_ir

    assert any(node.node_id == "dup" for node in workflow_ir.nodes)


def test_derive_cache_pool_consumers_skips_runtime_duplicate_name_validation(tmp_path: Path) -> None:
    _ = _write_duplicate_header_demand_yaml(tmp_path, file_name="dup.yaml", output_root=tmp_path / "out")
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "dup", "demand": "dup.yaml"}, {"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(conflict_policy="error", release_policy="dag_refcount", max_entries=10),
    )

    cfg = load_workflow_config(str(wf))
    compilation = workflow_compile_mod.compile_workflow_ir(cfg, workflow_yaml_path=str(wf), path_aliases=None)
    workflow_ir = compilation.workflow_ir
    logical_keys_by_node_id, consumers_by_logical_key = workflow_compile_mod.derive_cache_pool_consumers(
        workflow_ir,
        demand_configs_by_run_id=compilation.demand_configs_by_run_id,
    )

    assert logical_keys_by_node_id["dup"] == frozenset()
    assert logical_keys_by_node_id["ok"] == frozenset({("preload_forever", "preload")})
    assert consumers_by_logical_key[("preload_forever", "preload")] == frozenset(["ok"])


def test_run_workflow_demand_diagnostics_can_disable_duplicate_name_validation(tmp_path: Path) -> None:
    out_root = tmp_path / "out"
    _ = _write_duplicate_header_demand_yaml(tmp_path, file_name="dup.yaml", output_root=out_root)
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "dup", "demand": "dup.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    result = run_workflow(
        str(wf),
        options=_run_options(demand_diagnostics=DemandDiagnosticsPolicy(validate_unique_field_names=False)),
    )

    assert not result.errors()
    outcome = next(o for o in result.outcomes if o.run_id == "dup")
    assert isinstance(outcome.result, DemandRunResult)
    assert outcome.result.core.outputs is not None
    out_path = Path(str(outcome.result.core.outputs["detail"]))
    assert out_path.exists()


def test_run_workflow_run_patch_demand_diagnostics_isolated_per_run_in_multi_demand_workflow(tmp_path: Path) -> None:
    a_out = tmp_path / "a_out"
    b_out = tmp_path / "b_out"
    for run_id, out_root in [("a", a_out), ("b", b_out)]:
        file_id = "{}_detail_csv".format(run_id)
        _ = _write_text(
            tmp_path / "{}.yaml".format(run_id),
            (
                """
name: duplicate_headers_{run_id}

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id:
      extract: id
      name: Dup
    value:
      extract: value
      name: Dup

sources: {{}}

resources:
  files:
    {file_id}:
      kind: csv_file
      path: "{output_root}"

outputs:
  - name: detail
    to:
      file: {file_id}
    fields: [id, value]
"""
            )
            .format(
                run_id=str(run_id),
                file_id=str(file_id),
                output_root=str(out_root),
            )
            .lstrip(),
        )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    with pytest.raises(ScalimWorkflowConfigError, match=r"Workflow preflight failed: run_id='b'"):
        _ = run_workflow(
            str(wf),
            options=_run_options(),
            run_options_patches_by_run_id={
                "a": WorkflowNodePatch(demand_diagnostics=DemandDiagnosticsOverride(validate_unique_field_names=False))
            },
        )

    assert not a_out.exists()
    assert not b_out.exists()


def test_run_workflow_global_demand_diagnostics_applies_to_all_runs_in_multi_demand_workflow(tmp_path: Path) -> None:
    a_out = tmp_path / "a_out"
    b_out = tmp_path / "b_out"
    for run_id, out_root in [("a", a_out), ("b", b_out)]:
        file_id = "{}_detail_csv".format(run_id)
        _ = _write_text(
            tmp_path / "{}.yaml".format(run_id),
            (
                """
name: duplicate_headers_{run_id}

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id:
      extract: id
      name: Dup
    value:
      extract: value
      name: Dup

sources: {{}}

resources:
  files:
    {file_id}:
      kind: csv_file
      path: "{output_root}"

outputs:
  - name: detail
    to:
      file: {file_id}
    fields: [id, value]
"""
            )
            .format(
                run_id=str(run_id),
                file_id=str(file_id),
                output_root=str(out_root),
            )
            .lstrip(),
        )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    result = run_workflow(
        str(wf),
        options=_run_options(demand_diagnostics=DemandDiagnosticsPolicy(validate_unique_field_names=False)),
    )

    assert not result.errors()

    a_outcome = next(o for o in result.outcomes if o.run_id == "a")
    b_outcome = next(o for o in result.outcomes if o.run_id == "b")

    assert a_outcome.error is None
    assert b_outcome.error is None
    assert a_outcome.result is not None
    assert b_outcome.result is not None

    a_result = a_outcome.result
    b_result = b_outcome.result
    assert isinstance(a_result, DemandRunResult)
    assert isinstance(b_result, DemandRunResult)
    assert a_result.config.validate_unique_field_names is False
    assert b_result.config.validate_unique_field_names is False
    assert a_result.core.outputs is not None
    assert b_result.core.outputs is not None
    assert Path(str(a_result.core.outputs["detail"])).exists()
    assert Path(str(b_result.core.outputs["detail"])).exists()


def test_run_workflow_run_patch_demand_diagnostics_none_disables_global_policy_for_one_run(tmp_path: Path) -> None:
    a_out = tmp_path / "a_out"
    b_out = tmp_path / "b_out"
    for run_id, out_root in [("a", a_out), ("b", b_out)]:
        file_id = "{}_detail_csv".format(run_id)
        _ = _write_text(
            tmp_path / "{}.yaml".format(run_id),
            (
                """
name: duplicate_headers_{run_id}

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id:
      extract: id
      name: Dup
    value:
      extract: value
      name: Dup

sources: {{}}

resources:
  files:
    {file_id}:
      kind: csv_file
      path: "{output_root}"

outputs:
  - name: detail
    to:
      file: {file_id}
    fields: [id, value]
"""
            )
            .format(
                run_id=str(run_id),
                file_id=str(file_id),
                output_root=str(out_root),
            )
            .lstrip(),
        )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    with pytest.raises(ScalimWorkflowConfigError, match=r"Workflow preflight failed: run_id='b'"):
        _ = run_workflow(
            str(wf),
            options=_run_options(demand_diagnostics=DemandDiagnosticsPolicy(validate_unique_field_names=False)),
            run_options_patches_by_run_id={"b": WorkflowNodePatch(demand_diagnostics=None)},
        )

    assert not a_out.exists()
    assert not b_out.exists()


def test_run_workflow_preflight_duplicate_names_fail_before_engine(tmp_path: Path) -> None:
    out_root = tmp_path / "out"
    _ = _write_duplicate_header_demand_yaml(tmp_path, file_name="dup.yaml", output_root=out_root)
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "dup", "demand": "dup.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    def _run_ir_fn(*args: Any, **kwargs: Any) -> object:
        _ = args, kwargs
        raise AssertionError("engine should not be called when preflight fails")

    with pytest.raises(ScalimWorkflowConfigError, match=r"Workflow preflight failed: run_id='dup'"):
        _ = run_workflow(
            str(wf),
            options=_run_options(),
            run_ir_fn=_run_ir_fn,
        )

    assert not out_root.exists()


def test_workflow_lifecycle_pipeline_harness_runs_to_preflight_without_engine_and_returns_effective_options(tmp_path: Path) -> None:
    a_root = tmp_path / "a_out"
    b_root = tmp_path / "b_out"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=a_root,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=b_root,
        field_ids=["id", "value"],
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    from scalim.dsl.yaml_dsl import workflow_entrypoints as workflow_entrypoints_mod

    wf_options = WorkflowRunOptions(
        demand=_run_options(batch_size=1000),
        patches_by_run_id={"b": WorkflowNodePatch(batch_size=123)},
        runtime=WorkflowRuntimeOptions.preset_default(),
        path_aliases=None,
    )
    lifecycle = workflow_entrypoints_mod.run_workflow_lifecycle_until_preflight(str(wf), options=wf_options)

    assert set(lifecycle.preload.demand_configs_by_run_id) == {"a", "b"}
    assert lifecycle.effective.options_by_run_id["a"].runtime.batch_size == 1000
    assert lifecycle.effective.options_by_run_id["b"].runtime.batch_size == 123

    assert not a_root.exists()
    assert not b_root.exists()


def test_workflow_lifecycle_pipeline_harness_merges_resources_overrides_into_compile_overrides_when_base_has_none(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    out_root = tmp_path / "out"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=out_root,
        field_ids=["id", "value"],
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    from scalim.dsl.yaml_dsl import workflow_entrypoints as workflow_entrypoints_mod

    seen = {"overrides": None}
    original_compile = workflow_entrypoints_mod.compile_workflow_ir

    def _compile_capture(*args: Any, **kwargs: Any) -> workflow_compile_mod.WorkflowCompileResult:  # noqa: ANN401
        seen["overrides"] = kwargs.get("overrides")
        return original_compile(*args, **kwargs)

    monkeypatch.setattr(workflow_entrypoints_mod, "compile_workflow_ir", _compile_capture)

    file_id = "a_detail_csv"
    patch_out_root = tmp_path / "patch_out"
    patches_by_run_id = {
        "a": WorkflowNodePatch(
            overrides=RunOverrides(resources=ResourcesOverride(files={file_id: FileResourceOverride(path=str(patch_out_root))}))
        )
    }

    wf_options = WorkflowRunOptions(
        demand=_run_options(),
        patches_by_run_id=patches_by_run_id,
        runtime=WorkflowRuntimeOptions.preset_default(),
        path_aliases=None,
    )
    _ = workflow_entrypoints_mod.run_workflow_lifecycle_until_preflight(str(wf), options=wf_options)

    assert seen["overrides"] is not None


def test_workflow_lifecycle_pipeline_rejects_missing_structural_preload_results(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=tmp_path / "a_out",
        field_ids=["id", "value"],
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    from scalim.dsl.yaml_dsl import workflow_entrypoints as workflow_entrypoints_mod

    original_compile = workflow_entrypoints_mod.compile_workflow_ir

    def _compile_drop_preload_results(*args: Any, **kwargs: Any) -> workflow_compile_mod.WorkflowCompileResult:
        compilation = original_compile(*args, **kwargs)
        return workflow_compile_mod.WorkflowCompileResult(workflow_ir=compilation.workflow_ir, demand_configs_by_run_id={})

    monkeypatch.setattr(workflow_entrypoints_mod, "compile_workflow_ir", _compile_drop_preload_results)

    wf_options = WorkflowRunOptions(
        demand=_run_options(),
        runtime=WorkflowRuntimeOptions.preset_default(),
        path_aliases=None,
    )
    with pytest.raises(ScalimWorkflowConfigError, match=r"Missing workflow structural preload result for run_id"):
        _ = workflow_entrypoints_mod.run_workflow_lifecycle_until_preflight(str(wf), options=wf_options)


def test_run_workflow_runtime_compile_rejects_unknown_run_id(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=tmp_path / "a_out",
        field_ids=["id", "value"],
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    from scalim.dsl.yaml_dsl import workflow_entrypoints as workflow_entrypoints_mod

    original = workflow_entrypoints_mod.run_workflow_lifecycle_until_preflight

    def _lifecycle_missing_options(*args: Any, **kwargs: Any) -> workflow_entrypoints_mod.WorkflowLifecyclePreflightResult:
        lifecycle = original(*args, **kwargs)
        bad_options = dict(lifecycle.effective.options_by_run_id)
        _ = bad_options.pop("a")
        bad_effective = workflow_entrypoints_mod.WorkflowLifecycleEffectiveMergeResult(
            workflow_yaml_path=lifecycle.effective.workflow_yaml_path,
            workflow_ir=lifecycle.effective.workflow_ir,
            runs=lifecycle.effective.runs,
            options_by_run_id=bad_options,
            patches_by_run_id=lifecycle.effective.patches_by_run_id,
            bundle_viz_base_config=lifecycle.effective.bundle_viz_base_config,
        )
        return workflow_entrypoints_mod.WorkflowLifecyclePreflightResult(
            parse=lifecycle.parse,
            preload=lifecycle.preload,
            effective=bad_effective,
        )

    monkeypatch.setattr(workflow_entrypoints_mod, "run_workflow_lifecycle_until_preflight", _lifecycle_missing_options)

    def _run_ir_fn(*args: Any, **kwargs: Any) -> object:
        _ = args, kwargs
        raise AssertionError("engine should not be called when runtime compile fails")

    with pytest.raises(ScalimWorkflowConfigError, match=r"Unknown workflow run id in runtime compile: 'a'"):
        _ = run_workflow(
            str(wf),
            options=_run_options(),
            run_ir_fn=_run_ir_fn,
        )


def test_run_workflow_bundle_viz_requires_overrides_in_runtime_compile(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=tmp_path / "a_out",
        field_ids=["id", "value"],
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    from scalim.dsl.yaml_dsl import workflow_entrypoints as workflow_entrypoints_mod
    from scalim.ob.presets.viz import VizObserverConfig

    original = workflow_entrypoints_mod.run_workflow_lifecycle_until_preflight

    def _lifecycle_drop_overrides(*args: Any, **kwargs: Any) -> workflow_entrypoints_mod.WorkflowLifecyclePreflightResult:
        lifecycle = original(*args, **kwargs)
        bad_options = dict(lifecycle.effective.options_by_run_id)
        bad_options["a"] = workflow_entrypoints_mod.replace(
            bad_options["a"],
            outputs=workflow_entrypoints_mod.replace(bad_options["a"].outputs, overrides=None),
        )
        bad_effective = workflow_entrypoints_mod.WorkflowLifecycleEffectiveMergeResult(
            workflow_yaml_path=lifecycle.effective.workflow_yaml_path,
            workflow_ir=lifecycle.effective.workflow_ir,
            runs=lifecycle.effective.runs,
            options_by_run_id=bad_options,
            patches_by_run_id=lifecycle.effective.patches_by_run_id,
            bundle_viz_base_config=lifecycle.effective.bundle_viz_base_config,
        )
        return workflow_entrypoints_mod.WorkflowLifecyclePreflightResult(
            parse=lifecycle.parse,
            preload=lifecycle.preload,
            effective=bad_effective,
        )

    monkeypatch.setattr(workflow_entrypoints_mod, "run_workflow_lifecycle_until_preflight", _lifecycle_drop_overrides)

    overrides = RunOverrides(viz_config=VizObserverConfig(output_dir=str(tmp_path / "viz")))
    with pytest.raises(ScalimWorkflowConfigError, match=r"workflow bundle viz requires"):
        _ = run_workflow(
            str(wf),
            options=_run_options(overrides=overrides),
        )


def test_run_workflow_preflight_uses_effective_overrides_outputs_header_policy(tmp_path: Path) -> None:
    out_root = tmp_path / "out"
    _ = _write_duplicate_header_demand_yaml(tmp_path, file_name="dup.yaml", output_root=out_root)
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "dup", "demand": "dup.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    # overrides.outputs replaces YAML outputs; force header_fields_output_by=field_id so duplicate display-name
    # diagnostics MUST NOT be triggered (effective outputs semantics).
    overrides = RunOverrides(
        outputs=(
            OutputOverride(
                name="detail",
                fields=("id", "value"),
                to=OutputToOverride(file="detail_csv"),
                write=OutputWriteOverride(header_fields_output_by="field_id"),
            ),
        )
    )

    result = run_workflow(
        str(wf),
        options=_run_options(overrides=overrides),
    )
    assert not result.errors()
    outcome = next(o for o in result.outcomes if o.run_id == "dup")
    assert isinstance(outcome.result, DemandRunResult)
    assert outcome.result.core.outputs is not None
    out_path = Path(str(outcome.result.core.outputs["detail"]))
    assert out_path.exists()


def test_run_workflow_preflight_outputs_override_can_enable_duplicate_name_trigger(tmp_path: Path) -> None:
    out_root = tmp_path / "out"
    _ = _write_text(
        tmp_path / "dup.yaml",
        (
            """
name: duplicate_headers

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id:
      extract: id
      name: Dup
    value:
      extract: value
      name: Dup

sources: {{}}

resources:
  files:
    detail_csv:
      kind: csv_file
      path: "{output_root}"

outputs:
  - name: detail
    to:
      file: detail_csv
    fields: [id, value]
    write:
      header_fields_output_by: field_id
"""
        )
        .format(output_root=str(out_root))
        .lstrip(),
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "dup", "demand": "dup.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    overrides = RunOverrides(
        outputs=(
            OutputOverride(
                name="detail",
                fields=("id", "value"),
                to=OutputToOverride(file="detail_csv"),
                write=OutputWriteOverride(header_fields_output_by="name"),
            ),
        )
    )

    with pytest.raises(ScalimWorkflowConfigError, match=r"Workflow preflight failed: run_id='dup'"):
        _ = run_workflow(
            str(wf),
            options=_run_options(overrides=overrides),
        )

    assert not out_root.exists()


def test_run_workflow_run_patch_demand_diagnostics_can_override_include_full_error_message(tmp_path: Path) -> None:
    a_out = tmp_path / "a_out"
    b_out = tmp_path / "b_out"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=b_out,
        field_ids=["id", "value"],
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    result = run_workflow(
        str(wf),
        options=_run_options(),
        run_options_patches_by_run_id={
            "a": WorkflowNodePatch(demand_diagnostics=DemandDiagnosticsOverride(include_full_error_message=True))
        },
    )
    assert not result.errors()

    a_outcome = next(o for o in result.outcomes if o.run_id == "a")
    b_outcome = next(o for o in result.outcomes if o.run_id == "b")

    assert a_outcome.result is not None
    assert b_outcome.result is not None

    a_result = a_outcome.result
    b_result = b_outcome.result
    assert isinstance(a_result, DemandRunResult)
    assert isinstance(b_result, DemandRunResult)
    assert a_result.config.include_full_error_message is True
    assert b_result.config.include_full_error_message is False


def test_run_workflow_run_patch_can_disable_duplicate_name_validation(tmp_path: Path) -> None:
    out_root = tmp_path / "out"
    _ = _write_duplicate_header_demand_yaml(tmp_path, file_name="dup.yaml", output_root=out_root)
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "dup", "demand": "dup.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    result = run_workflow(
        str(wf),
        options=_run_options(),
        run_options_patches_by_run_id={
            "dup": WorkflowNodePatch(demand_diagnostics=DemandDiagnosticsOverride(validate_unique_field_names=False))
        },
    )

    assert not result.errors()
    outcome = next(o for o in result.outcomes if o.run_id == "dup")
    assert isinstance(outcome.result, DemandRunResult)
    assert outcome.result.core.outputs is not None
    out_path = Path(str(outcome.result.core.outputs["detail"]))
    assert out_path.exists()


def test_workflow_sheetbook_append_export_xlsx_is_deterministic(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="S",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        book_id="report",
        sheet="S",
        field_ids=["id", "value"],
    )

    expected_rows = [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
        ["b1", "B1"],
        ["b2", "B2"],
    ]
    for idx in range(3):
        out_root = tmp_path / "out_{}".format(int(idx))
        wf = _write_workflow_yaml(
            tmp_path,
            resources={
                "books": {
                    "report": {
                        "kind": "xlsx_memory",
                        "budget": {"max_sheets": 8, "max_total_cells": 1000},
                        "export_xlsx": {"path": str(out_root)},
                    }
                }
            },
            runs=[
                {"id": "a", "demand": "a.yaml"},
                {"id": "b", "demand": "b.yaml"},
            ],
            max_concurrency=2,
            failure_policy="primary_only",
        )

        result = run_workflow(str(wf), options=_run_options())
        assert not result.errors()
        export_path = _latest_book_path(out_root, book_id="report")
        assert export_path.exists()
        assert _read_xlsx_sheetnames(export_path) == ["S"]
        assert _read_xlsx_rows(export_path, "S") == expected_rows


def test_book_sheet_rows_loader_requires_context() -> None:
    with pytest.raises(WorkflowRuntimeConfigError, match="requires workflow context"):
        _ = workflow_loaders_mod.book_sheet_rows(ref={"node": "a", "book": "sb", "sheet": "S"})


def test_book_sheet_rows_loader_validates_ref_and_context_cleanup() -> None:
    class _DummyManager:
        def iter_book_sheet_rows(self, **_kwargs: object) -> Any:  # noqa: ANN401
            return iter(())

    dummy = _DummyManager()

    with workflow_loaders_mod.workflow_loader_context(
        workflow_exec_id="wf",
        workflow_node_id="consumer",
        visible_producer_node_ids=frozenset(),
        resource_manager=dummy,  # type: ignore[arg-type]
    ):
        with pytest.raises(WorkflowRuntimeConfigError, match="params.ref"):
            _ = workflow_loaders_mod.book_sheet_rows(ref="nope")  # type: ignore[arg-type]
        with pytest.raises(WorkflowRuntimeConfigError, match="ref.node"):
            _ = workflow_loaders_mod.book_sheet_rows(ref={"book": "sb", "sheet": "S"})
        with pytest.raises(WorkflowRuntimeConfigError, match="ref.book"):
            _ = workflow_loaders_mod.book_sheet_rows(ref={"node": "a", "sheet": "S"})
        with pytest.raises(WorkflowRuntimeConfigError, match="ref.sheet"):
            _ = workflow_loaders_mod.book_sheet_rows(ref={"node": "a", "book": "sb"})

    with workflow_loaders_mod.workflow_loader_context(
        workflow_exec_id="wf",
        workflow_node_id="consumer",
        visible_producer_node_ids=frozenset(),
        resource_manager=dummy,  # type: ignore[arg-type]
    ):
        delattr(workflow_loaders_mod._TLS, "ctx")


def test_book_sheet_rows_loader_rejects_corrupted_context() -> None:
    workflow_loaders_mod._TLS.ctx = object()
    try:
        with pytest.raises(ScalimInternalError, match="context is corrupted"):
            _ = workflow_loaders_mod.book_sheet_rows(ref={"node": "a", "book": "sb", "sheet": "S"})
    finally:
        delattr(workflow_loaders_mod._TLS, "ctx")


def test_workflow_shared_workbook_append_is_deterministic_by_runs_order(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="slow.yaml",
        name="slow",
        loader_ref="tests.fixtures.workflow_loaders:load_table_c_slow",
        output_name="detail",
        book_id="report",
        sheet="All",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="fast.yaml",
        name="fast",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        book_id="report",
        sheet="All",
        field_ids=["id", "value"],
    )

    expected_rows = [
        ["id", "value"],
        ["c1", "C1"],
        ["b1", "B1"],
        ["b2", "B2"],
    ]
    expected_order_signature: Optional[List[Any]] = None
    for idx in range(3):
        out_root = tmp_path / "out_{}".format(int(idx))
        wf = _write_workflow_yaml(
            tmp_path,
            resources={"books": {"report": {"kind": "xlsx_file", "path": str(out_root)}}},
            runs=[
                {"id": "slow", "demand": "slow.yaml"},
                {"id": "fast", "demand": "fast.yaml"},
            ],
            max_concurrency=2,
            failure_policy="primary_only",
        )

        recorder = _WorkflowEventRecorder()
        result = run_workflow(str(wf), options=_run_options(components=[recorder]))
        assert not result.errors()
        workbook_path = _latest_book_path(out_root, book_id="report")
        assert workbook_path.exists()
        assert _read_xlsx_rows(workbook_path, "All") == expected_rows

        order_signature: List[Any] = []
        for event in recorder.events:
            if event.event_type == EVENT_PIPELINE_START:
                meta = event.meta if isinstance(event.meta, dict) else {}
                order_signature.append((event.event_type, str(meta.get("workflow_node_id") or ""), list(event.payload.targets)))
            if event.event_type in (EVENT_WORKFLOW_NODE_START, EVENT_WORKFLOW_NODE_END):
                order_signature.append((event.event_type, str(event.payload.workflow_node_id)))
            if event.event_type == EVENT_WORKFLOW_RESOURCE_COMMIT:
                order_signature.append(
                    (
                        event.event_type,
                        str(event.payload.workflow_node_id),
                        str(event.payload.resource_type),
                        str(event.payload.resource_id),
                    )
                )
        if expected_order_signature is None:
            expected_order_signature = list(order_signature)
        else:
            assert list(order_signature) == list(expected_order_signature)

        write_events = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_WRITE and e.payload.resource_id == "report"]
        assert [e.payload.workflow_node_id for e in write_events] == ["__wf__write.slow.0", "__wf__write.fast.0"]


def test_workflow_shared_append_header_policy_variants(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="All",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        book_id="report",
        sheet="All",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="c.yaml",
        name="c",
        loader_ref="tests.fixtures.workflow_loaders:load_table_c_slow",
        output_name="detail",
        book_id="report",
        sheet="All",
        field_ids=["id", "value"],
    )

    cases = {
        "once": [
            ["id", "value"],
            ["a1", "A1"],
            ["a2", "A2"],
            ["b1", "B1"],
            ["b2", "B2"],
            ["c1", "C1"],
        ],
        "always": [
            ["id", "value"],
            ["a1", "A1"],
            ["a2", "A2"],
            ["id", "value"],
            ["b1", "B1"],
            ["b2", "B2"],
            ["id", "value"],
            ["c1", "C1"],
        ],
        "never": [
            ["a1", "A1"],
            ["a2", "A2"],
            ["b1", "B1"],
            ["b2", "B2"],
            ["c1", "C1"],
        ],
    }

    for header_policy, expected_rows in cases.items():
        out_root = tmp_path / "out_{}".format(str(header_policy))
        wf = _write_workflow_yaml(
            tmp_path,
            resources={
                "books": {
                    "report": {
                        "kind": "xlsx_file",
                        "path": str(out_root),
                        "write_defaults": {"mode": "append", "header_policy": str(header_policy)},
                    }
                }
            },
            runs=[
                {"id": "a", "demand": "a.yaml"},
                {"id": "b", "demand": "b.yaml"},
                {"id": "c", "demand": "c.yaml"},
            ],
            max_concurrency=3,
            failure_policy="primary_only",
        )

        result = run_workflow(str(wf), options=_run_options())
        assert not result.errors()
        workbook_path = _latest_book_path(out_root, book_id="report")
        assert workbook_path.exists()
        assert _read_xlsx_rows(workbook_path, "All") == expected_rows


def test_workflow_shared_book_append_warn_skip_and_header_policies(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="All",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="m.yaml",
        name="m",
        loader_ref="tests.fixtures.workflow_loaders:load_table_mismatch",
        output_name="detail",
        book_id="report",
        sheet="All",
        field_ids=["id", "other"],
    )
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        book_id="report",
        sheet="All",
        field_ids=["id", "value"],
    )

    cases = (
        (
            "warn",
            [
                ["id", "value"],
                ["a1", "A1"],
                ["a2", "A2"],
                ["m1", ""],
                ["b1", "B1"],
                ["b2", "B2"],
            ],
            False,
        ),
        (
            "skip",
            [
                ["id", "value"],
                ["a1", "A1"],
                ["a2", "A2"],
                ["b1", "B1"],
                ["b2", "B2"],
            ],
            True,
        ),
    )

    for on_mismatch, expected_rows, expect_skips in cases:
        out_root = tmp_path / "out_{}".format(str(on_mismatch))
        wf = _write_workflow_yaml(
            tmp_path,
            resources={
                "books": {
                    "report": {
                        "kind": "xlsx_file",
                        "path": str(out_root),
                        "write_defaults": {"mode": "append", "header_policy": "once", "on_mismatch": str(on_mismatch)},
                    }
                }
            },
            runs=[
                {"id": "a", "demand": "a.yaml"},
                {"id": "m", "demand": "m.yaml"},
                {"id": "b", "demand": "b.yaml"},
            ],
            max_concurrency=3,
            failure_policy="primary_only",
        )

        recorder = _WorkflowEventRecorder()
        result = run_workflow(str(wf), options=_run_options(components=[recorder]))
        assert not result.errors()
        workbook_path = _latest_book_path(out_root, book_id="report")
        assert workbook_path.exists()
        assert _read_xlsx_rows(workbook_path, "All") == expected_rows

        skips = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_WRITE and e.payload.action == "skip"]
        assert bool(skips) is bool(expect_skips)


def test_workflow_shared_resources_discard_on_failure(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="OK",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        loader_ref="tests.fixtures.workflow_loaders:load_table_raises",
        output_name="detail",
        output_root=tmp_path / "bad_out",
        field_ids=["id", "value"],
    )

    out_root = tmp_path / "out"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={"books": {"report": {"kind": "xlsx_file", "path": str(out_root)}}},
        runs=[
            {"id": "ok", "demand": "ok.yaml"},
            {"id": "bad", "demand": "bad.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    runtime = WorkflowRuntimeOptions(execution=WorkflowExecutionOptions(max_concurrency=2, failure_policy="primary_only"))
    result = run_workflow(str(wf), options=_run_options(components=[recorder]), workflow_runtime_options=runtime)
    assert result.errors()
    assert not (out_root / "manifest" / "latest.json").exists()
    assert list(out_root.rglob("*.xlsx")) == []
    assert list(out_root.rglob("*.scalim.lock")) == []

    discards = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_DISCARD and e.payload.resource_id == "report"]
    assert discards


def test_workflow_shared_sheet_conflict_policies(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="S",
        field_ids=["id", "value"],
    )

    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        book_id="report",
        sheet="S",
        field_ids=["id", "value"],
    )

    runtime = WorkflowRuntimeOptions(execution=WorkflowExecutionOptions(max_concurrency=2, failure_policy="primary_only"))

    # on_conflict=error
    out_root_err = tmp_path / "out_err"
    wf_err = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_file",
                    "path": str(out_root_err),
                    "write_defaults": {"mode": "sheet", "on_conflict": "error"},
                }
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result_err = run_workflow(str(wf_err), options=_run_options(), workflow_runtime_options=runtime)
    assert result_err.errors()
    assert not (out_root_err / "manifest" / "latest.json").exists()
    assert list(out_root_err.rglob("*.xlsx")) == []

    # on_conflict=overwrite
    out_root_over = tmp_path / "out_over"
    wf_over = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_file",
                    "path": str(out_root_over),
                    "write_defaults": {"mode": "sheet", "on_conflict": "overwrite"},
                }
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result_over = run_workflow(str(wf_over), options=_run_options(), workflow_runtime_options=runtime)
    assert not result_over.errors()
    workbook_over = _latest_book_path(out_root_over, book_id="report")
    assert workbook_over.exists()
    assert _read_xlsx_rows(workbook_over, "S")[-1] == ["b2", "B2"]

    # on_conflict=skip
    out_root_skip = tmp_path / "out_skip"
    wf_skip = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_file",
                    "path": str(out_root_skip),
                    "write_defaults": {"mode": "sheet", "on_conflict": "skip"},
                }
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result_skip = run_workflow(str(wf_skip), options=_run_options(), workflow_runtime_options=runtime)
    assert not result_skip.errors()
    workbook_skip = _latest_book_path(out_root_skip, book_id="report")
    assert workbook_skip.exists()
    assert _read_xlsx_rows(workbook_skip, "S")[-1] == ["a2", "A2"]


def test_workflow_resources_rejects_write_lock_for_xlsx_file(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="S",
        field_ids=["id", "value"],
    )

    out_root = tmp_path / "out"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_file",
                    "path": str(out_root),
                    "write_lock": True,
                    "write_defaults": {"mode": "sheet"},
                }
            }
        },
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(ScalimWorkflowConfigError, match="write_lock was removed") as excinfo:
        _ = run_workflow(str(wf), options=_run_options())
    assert excinfo.value.path == "workflow.resources.books.report.write_lock"


def test_workflow_pathless_csv_output_without_writes_fails_fast(tmp_path: Path) -> None:
    _ = _write_text(
        tmp_path / "a.yaml",
        (
            """
name: a
main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {extract: id}
    value: {extract: value}
outputs:
  - name: detail
    to:
      file: detail_csv
    fields: [id, value]
resources:
  files:
    detail_csv:
      kind: csv_file
"""
        ).lstrip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    with pytest.raises(ScalimWorkflowConfigError, match="YAML DSL validation failed"):
        _ = run_workflow(str(wf), options=_run_options())


def test_workflow_managed_temp_outputs_does_not_require_pathless_csv_authoring_and_cleans_up(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="S",
        field_ids=["id", "value"],
    )

    out_root = tmp_path / "out"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={"books": {"report": {"kind": "xlsx_file", "path": str(out_root), "write_defaults": {"mode": "sheet"}}}},
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), options=_run_options())
    assert not result.errors()
    workbook_path = _latest_book_path(out_root, book_id="report")
    assert workbook_path.exists()
    assert _read_xlsx_rows(workbook_path, "S")[-1] == ["a2", "A2"]
    assert not (tmp_path / ".scalim").exists()


def test_workflow_main_rows_from_wires_upstream_typed_rows_into_downstream_main_rows(tmp_path: Path) -> None:
    a_root = tmp_path / "a_out"
    b_root = tmp_path / "b_out"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=a_root,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_raises",
        output_name="detail",
        output_root=b_root,
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml", "depends_on": ["a"], "main_rows_from": {"run": "a"}},
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), options=_run_options())
    assert not result.errors()

    b_outcome = next(o for o in result.outcomes if o.run_id == "b")
    assert isinstance(b_outcome.result, DemandRunResult)
    assert b_outcome.result.core.outputs is not None
    b_out_path = Path(str(b_outcome.result.core.outputs["detail"]))
    assert _read_csv_rows(b_out_path) == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]


def test_workflow_main_rows_from_releases_typed_rows_after_final_consumer(tmp_path: Path, monkeypatch) -> None:
    publish_calls = []
    discard_calls = []

    class _RecordingArtifactsDirectory(workflow_execute_mod.WorkflowArtifactsDirectory):  # type: ignore[misc]
        def publish(self, producer_node_id: str, artifact_id: str, value: object) -> None:
            publish_calls.append((str(producer_node_id), str(artifact_id)))
            super(_RecordingArtifactsDirectory, self).publish(producer_node_id, artifact_id, value)

        def discard(self, producer_node_id: str, artifact_id: str) -> None:
            discard_calls.append((str(producer_node_id), str(artifact_id)))
            super(_RecordingArtifactsDirectory, self).discard(producer_node_id, artifact_id)

    monkeypatch.setattr(workflow_execute_mod, "WorkflowArtifactsDirectory", _RecordingArtifactsDirectory)

    a_root = tmp_path / "a_out"
    b_root = tmp_path / "b_out"
    c_root = tmp_path / "c_out"
    d_root = tmp_path / "d_out"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=a_root,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_raises",
        output_name="detail",
        output_root=b_root,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="c.yaml",
        name="c",
        loader_ref="tests.fixtures.workflow_loaders:load_table_raises",
        output_name="detail",
        output_root=c_root,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="d.yaml",
        name="d",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_root=d_root,
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml", "depends_on": ["a"], "main_rows_from": {"run": "a"}},
            {"id": "c", "demand": "c.yaml", "depends_on": ["a"], "main_rows_from": {"run": "a"}},
            {"id": "d", "demand": "d.yaml", "depends_on": ["b", "c"]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), options=_run_options())
    assert not result.errors()
    outcomes = {str(o.run_id): o for o in result.outcomes}
    b_outcome = outcomes["b"]
    c_outcome = outcomes["c"]
    assert isinstance(b_outcome.result, DemandRunResult)
    assert isinstance(c_outcome.result, DemandRunResult)
    assert b_outcome.result.core.outputs is not None
    assert c_outcome.result.core.outputs is not None
    b_out_path = Path(str(b_outcome.result.core.outputs["detail"]))
    c_out_path = Path(str(c_outcome.result.core.outputs["detail"]))
    assert _read_csv_rows(b_out_path) == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]
    assert _read_csv_rows(c_out_path) == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]

    typed_rows_publishes = [c for c in publish_calls if c[1] == "in_memory_rows"]
    assert typed_rows_publishes == [("a", "in_memory_rows")]

    typed_rows_discards = [c for c in discard_calls if c[1] == "in_memory_rows"]
    assert typed_rows_discards == [("a", "in_memory_rows")]


def test_workflow_artifacts_directory_discard_all_in_memory_rows_removes_empty_producer_entry() -> None:
    from scalim.sinks._internal.rows import InMemoryRows
    from scalim.spec.ir._workflow import WorkflowArtifactsIr, WorkflowIr, WorkflowNodeIr, WorkflowNodeType, WorkflowOptionsIr

    workflow_ir = WorkflowIr(
        nodes=(WorkflowNodeIr(node_id="a", node_type=WorkflowNodeType.DEMAND, decl_order=0, deps=()),),
        edges=(),
        options=WorkflowOptionsIr(),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={}),
    )
    artifacts_dir = workflow_execute_mod.WorkflowArtifactsDirectory(workflow_ir)
    artifacts_dir.publish("a", "in_memory_rows", InMemoryRows(header=["id"], rows=[[1]]))
    artifacts_dir.discard_all_in_memory_rows()

    with pytest.raises(KeyError, match=r"Unknown artifact"):
        _ = artifacts_dir.get("a", "a", "in_memory_rows")


def test_workflow_main_rows_from_rejects_non_in_memory_rows_artifact(tmp_path: Path, monkeypatch) -> None:
    class _CorruptArtifactsDirectory(workflow_execute_mod.WorkflowArtifactsDirectory):  # type: ignore[misc]
        def publish(self, producer_node_id: str, artifact_id: str, value: object) -> None:
            if str(producer_node_id) == "a" and str(artifact_id) == "in_memory_rows":
                value = "not_in_memory_rows"
            super(_CorruptArtifactsDirectory, self).publish(producer_node_id, artifact_id, value)

    monkeypatch.setattr(workflow_execute_mod, "WorkflowArtifactsDirectory", _CorruptArtifactsDirectory)

    a_root = tmp_path / "a_out"
    b_root = tmp_path / "b_out"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_root=a_root,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_raises",
        output_name="detail",
        output_root=b_root,
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml", "depends_on": ["a"], "main_rows_from": {"run": "a"}},
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    runtime = WorkflowRuntimeOptions(execution=WorkflowExecutionOptions(max_concurrency=1, failure_policy="primary_only"))
    result = run_workflow(str(wf), options=_run_options(), workflow_runtime_options=runtime)
    assert result.errors()
    b_outcome = next(o for o in result.outcomes if o.run_id == "b")
    assert b_outcome.error is not None
    assert b_outcome.error.exc_type == "ScalimWorkflowWriteError"


def test_workflow_execute_release_main_rows_artifact_returns_when_missing_count_entry(tmp_path: Path) -> None:
    from typing import Any

    from scalim.execution.run_ir import ExecutionRequest, ExportLayout, OutputSpec
    from scalim.spec.ir._workflow import (
        WorkflowArtifactsIr,
        WorkflowEdgeIr,
        WorkflowIr,
        WorkflowNodeIr,
        WorkflowNodeType,
        WorkflowOptionsIr,
    )
    from scalim.vendor.dataclassesx import dataclass

    @dataclass(frozen=True)
    class _Compilation:
        demand_ir: object
        request: ExecutionRequest

    class _Core:
        output_path = None
        total_rows = 0
        duration = 0.0
        outputs = None
        in_memory_csv_outputs = {}
        in_memory_rows_outputs = {}
        workflow_managed_output_export_headers = None
        in_memory_rows = None

    def _compile_demand_node(demand_path: str, **kwargs: Any) -> object:
        _ = demand_path, kwargs
        request = ExecutionRequest(
            export_layout=ExportLayout(field_ids=(), header_names=None),
            output=OutputSpec(path=None),
            sink=None,
        )
        return _Compilation(demand_ir=object(), request=request)

    def _run_ir_fn(demand_ir: object, request: ExecutionRequest, **kwargs: Any) -> object:
        _ = demand_ir, request, kwargs
        return _Core()

    workflow_ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(
                node_id="a",
                node_type=WorkflowNodeType.DEMAND,
                decl_order=0,
                deps=(),
                demand_path="a.yaml",
            ),
            WorkflowNodeIr(
                node_id="b",
                node_type=WorkflowNodeType.DEMAND,
                decl_order=1,
                deps=("a",),
                demand_path="b.yaml",
                main_rows_from_run_id="a",
            ),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(max_concurrency=1, failure_policy="primary_only"),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={}),
    )

    prepared = workflow_execute_mod._prepare_workflow_run_ir(  # type: ignore[attr-defined]
        str(tmp_path / "workflow.yaml"),
        workflow_ir,
        components=None,
        bundle_viz_base_config=None,
        cache_pool_logical_keys_by_node_id=None,
        cache_pool_consumers_by_logical_key=None,
    )
    prepared.main_rows_consumers_remaining_by_run_id = {}
    try:
        outcomes, _failed, _exc = workflow_execute_mod._execute_workflow_run(  # type: ignore[attr-defined]
            prepared,
            compile_demand_fn=_compile_demand_node,
            build_demand_run_result_fn=None,
            run_ir_fn=_run_ir_fn,
        )
    finally:
        workflow_execute_mod._cleanup_workflow_finally(prepared, resources_finalized=False)  # type: ignore[attr-defined]

    b_outcome = next(o for o in outcomes if o.run_id == "b")
    assert b_outcome.error is not None
    assert b_outcome.error.exc_type == "KeyError"


def test_workflow_execute_release_main_rows_artifact_raises_on_negative_count(tmp_path: Path) -> None:
    from typing import Any

    from scalim.execution.run_ir import ExecutionRequest, ExportLayout, OutputSpec
    from scalim.spec.ir._workflow import (
        WorkflowArtifactsIr,
        WorkflowEdgeIr,
        WorkflowIr,
        WorkflowNodeIr,
        WorkflowNodeType,
        WorkflowOptionsIr,
    )
    from scalim.vendor.dataclassesx import dataclass

    @dataclass(frozen=True)
    class _Compilation:
        demand_ir: object
        request: ExecutionRequest

    class _Core:
        output_path = None
        total_rows = 0
        duration = 0.0
        outputs = None
        in_memory_csv_outputs = {}
        in_memory_rows_outputs = {}
        in_memory_rows = None

    def _compile_demand_node(demand_path: str, **kwargs: Any) -> object:
        _ = demand_path, kwargs
        request = ExecutionRequest(
            export_layout=ExportLayout(field_ids=(), header_names=None),
            output=OutputSpec(path=None),
            sink=None,
        )
        return _Compilation(demand_ir=object(), request=request)

    def _run_ir_fn(demand_ir: object, request: ExecutionRequest, **kwargs: Any) -> object:
        _ = demand_ir, request, kwargs
        return _Core()

    workflow_ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(
                node_id="a",
                node_type=WorkflowNodeType.DEMAND,
                decl_order=0,
                deps=(),
                demand_path="a.yaml",
            ),
            WorkflowNodeIr(
                node_id="b",
                node_type=WorkflowNodeType.DEMAND,
                decl_order=1,
                deps=("a",),
                demand_path="b.yaml",
                main_rows_from_run_id="a",
            ),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(max_concurrency=1, failure_policy="primary_only"),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={}),
    )

    prepared = workflow_execute_mod._prepare_workflow_run_ir(  # type: ignore[attr-defined]
        str(tmp_path / "workflow.yaml"),
        workflow_ir,
        components=None,
        bundle_viz_base_config=None,
        cache_pool_logical_keys_by_node_id=None,
        cache_pool_consumers_by_logical_key=None,
    )
    prepared.main_rows_consumers_remaining_by_run_id = {"a": 0}
    try:
        with pytest.raises(RuntimeError, match=r"negative main_rows consumer count"):
            _ = workflow_execute_mod._execute_workflow_run(  # type: ignore[attr-defined]
                prepared,
                compile_demand_fn=_compile_demand_node,
                build_demand_run_result_fn=None,
                run_ir_fn=_run_ir_fn,
            )
    finally:
        workflow_execute_mod._cleanup_workflow_finally(prepared, resources_finalized=False)  # type: ignore[attr-defined]


def test_workflow_derived_write_nodes_fail_fast_on_missing_to_book_binding(tmp_path: Path) -> None:
    _ = _write_text(
        tmp_path / "missing_binding.yaml",
        (
            """
name: demo
main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {extract: id}
    value: {extract: value}
outputs:
  - name: detail
    to:
      sheet: S
    fields: [id, value]
"""
        ).lstrip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "missing_binding.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(ScalimWorkflowConfigError, match=r"outputs\.0\.to must declare exactly one of to\.file or to\.book"):
        _ = run_workflow(str(wf), options=_run_options())


def test_workflow_derived_write_nodes_fail_fast_on_missing_book_resource(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="S",
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(ScalimWorkflowConfigError, match="Missing book resource id"):
        _ = run_workflow(str(wf), options=_run_options())


def test_workflow_shared_book_commit_failure_raises_workflow_config_error(tmp_path: Path) -> None:
    _ = _write_table_demand_yaml_with_book_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        book_id="report",
        sheet="S",
        field_ids=["id", "value"],
    )

    workbook_dir = tmp_path / "outdir"
    workbook_dir.mkdir()
    lock_path = Path(str(workbook_dir) + ".scalim.lock")

    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "books": {
                "report": {
                    "kind": "xlsx_file",
                    "path": str(workbook_dir),
                    "write_lock": True,
                    "write_defaults": {"mode": "sheet"},
                }
            }
        },
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(ScalimWorkflowConfigError, match="workflow.resources"):
        _ = run_workflow(str(wf), options=_run_options())
    assert workbook_dir.exists()
    assert not lock_path.exists()
