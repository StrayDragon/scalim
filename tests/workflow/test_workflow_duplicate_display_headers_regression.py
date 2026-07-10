"""回归复现: 重复展示列名(`header_fields_output_by=name`)导致 `xlsx` 数据错位.

背景:
- 底层输出字段 `field_id` 全局唯一(明确约束), 但用户希望导出给数据同事的“展示列名”
  可以重复(例如多个“人数”/“金额”指标块).
- 历史 bug: 若中间工件 `header` 用展示名(可重复)做列对齐, `_build_alignment_mapping`
  按“列名 -> 首次出现索引”建表, 重复列名坍缩到首次位置 -> 数据错位.
- 现行契约: `xlsx_file` / `xlsx_memory` 的 workflow-managed 中间态均为 `InMemoryRows`
  (`header`=`field_id`); `csv`/file 仍可为 `InMemoryCsv`。展示名经独立 `export_header`
  仅用于表头行。本文件部分用例仍用 `InMemoryCsv` 作为 tabular 适配输入,验证对齐契约。

说明: 本文件为可推送的脱敏 `MVP`, 所有字段名/取值均为占位符, 不含任何真实业务数据.
"""

from pathlib import Path
from typing import Any, Dict, List, Optional

from scalim.sinks.memory import InMemoryCsv
from scalim.sinks.rows import InMemoryRows
from scalim.workflow import resources as resources_mod
from scalim.workflow import resources_csv as resources_csv_mod
from scalim.workflow.resources_sheetbook import SheetBookDef


class _Instrumentation:
    def __init__(self) -> None:
        self.events: List[Dict[str, Any]] = []

    def emit(self, event_type: str, payload: Any, meta: Optional[Dict[str, Any]] = None) -> None:
        self.events.append({"event_type": str(event_type), "payload": payload, "meta": dict(meta or {})})


# 底层 field_id 唯一; 展示名(给数据同事看)允许重复.
_DUP_DISPLAY_HEADER = ["人数", "金额", "人数", "金额"]
_UNIQUE_FIELD_IDS = ["metric_count_a", "metric_amount_a", "metric_count_b", "metric_amount_b"]
# 占位假数据: 每个列各自不同, 便于一眼看出是否被首列值覆盖.
_INPUT_ROW = ["10", "20", "30", "40"]
_EXPECTED_DATA_ROW = ["10", "20", "30", "40"]


def test_build_alignment_mapping_dict_fallback() -> None:
    """`_build_alignment_mapping` 按 Dict 首现索引建映射;重复名坍缩到首现,缺失为 -1."""

    # 长度不同: actual 缺少 'c' -> -1; 仅记录首现索引。
    assert resources_csv_mod.build_alignment_mapping(["a", "b", "c"], ["a", "b"]) == [0, 1, -1]
    # 顺序不同: 按列名在 actual 中的首现索引对齐。
    assert resources_csv_mod.build_alignment_mapping(["b", "a"], ["a", "b"]) == [1, 0]
    # actual 含重复名但 expected 不同序 -> 后续同名取首现。
    assert resources_csv_mod.build_alignment_mapping(["y", "x"], ["x", "y", "x"]) == [1, 0]
    # expected==actual 且含重复名 -> Dict 坍缩为 [0,1,0,1](非恒等)。
    # 这已不是 bug: 修复后 `_build_alignment_mapping` 不再接收展示名,始终在唯一 field_id 上运行。
    assert resources_csv_mod.build_alignment_mapping(["a", "b", "a"], ["a", "b", "a"]) == [0, 1, 0]


def test_sheetbook_xlsx_memory_duplicate_display_headers_remains_correct(tmp_path: Path) -> None:
    """对照用例(`xlsx_memory`/`sheetbook`): 重复展示名下数据应保持正确.

    该路径中间工件为 `InMemoryRows`(header=field_id 唯一), 展示名经独立 `export_header`
    仅用于表头行, 因此不受 `_build_alignment_mapping` 坍缩影响. 此用例应始终通过,
    用以界定修复范围(不应破坏该路径的正确行为).
    """

    from openpyxl import load_workbook

    output_root = tmp_path / "out"
    export_path = output_root / "versions" / "wf" / "books" / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=_Instrumentation(),
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": SheetBookDef(
                resource_id="sb",
                budget_max_sheets=0,
                budget_max_total_cells=0,
                export_path=str(export_path),
                export_allow_formulas=True,
            )
        },
    )

    rows_artifact = InMemoryRows(header=list(_UNIQUE_FIELD_IDS), rows=[[10, 20, 30, 40]])
    manager.apply_sheetbook_sheet(
        workflow_node_id="n1",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S1",
        input_node_id="n1",
        input_output_id="detail",
        input_csv=rows_artifact,
        on_conflict="error",
        export_header=tuple(_DUP_DISPLAY_HEADER),
    )
    manager.commit_all()

    wb = load_workbook(str(export_path), read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb["S1"].iter_rows(values_only=True)]
    finally:
        wb.close()

    assert rows[0] == _DUP_DISPLAY_HEADER
    assert list(rows[1]) == [10, 20, 30, 40]


def test_workbook_sheet_writes_export_header_and_aligns_by_field_id(tmp_path: Path) -> None:
    """源头修复验证(`xlsx_file`/workbook): header=field_id + 独立 export_header.

    生产 managed 路径为 `InMemoryRows`;本用例用 `InMemoryCsv`(header=field_id) 作为
    tabular 适配输入。`apply_workbook_sheet` 存储 `export_header`;写出时表头用
    export_header(可重复),对齐用 baseline_header=field_id(identity, 不坍缩).
    """

    from openpyxl import load_workbook

    from scalim.execution import versioned_outputs

    out_root = tmp_path / "out"
    layout = versioned_outputs.ensure_output_root_layout(out_root)
    workbook_path = versioned_outputs.book_output_path(layout, version_id="wf", book_id="report")
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=_Instrumentation(),
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    # 修复后: 中间工件 header=field_id(唯一); 展示名经 export_header 独立透传(可重复).
    artifact = InMemoryCsv(header=list(_UNIQUE_FIELD_IDS), rows=[list(_INPUT_ROW)])
    manager.apply_workbook_sheet(
        workflow_node_id="n1",
        decl_order=0,
        workbook_id="report",
        sheet="S1",
        input_node_id="n1",
        input_output_id="detail",
        input_csv=artifact,
        on_conflict="error",
        export_header=tuple(_DUP_DISPLAY_HEADER),
    )
    manager.commit_all()

    wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb["S1"].iter_rows(values_only=True)]
    finally:
        wb.close()

    # 表头 = 展示名(可重复); 数据 = 按 field_id 位置(identity)正确, 不被首列值填充.
    assert rows[0] == _DUP_DISPLAY_HEADER, "header must be export_header (display names), got: {}".format(rows[0])
    assert list(rows[1]) == _EXPECTED_DATA_ROW, "data must align by field_id positionally, got: {}".format(rows[1])


def test_csv_append_writes_export_header_and_aligns_by_field_id(tmp_path: Path) -> None:
    """源头修复验证(`csv`/file): 中间工件 header=field_id + 独立 export_header.

    修复后 csv 路径 managed 工件 `InMemoryCsv.header` = field_id(唯一),
    `apply_csv_append` 接收+存储 `export_header`; `_commit_csv` 表头写 export_header,
    对齐用 baseline_header=field_id.
    """

    import csv as csv_mod

    from scalim.execution import versioned_outputs

    out_root = tmp_path / "out"
    layout = versioned_outputs.ensure_output_root_layout(out_root)
    csv_path = versioned_outputs.file_output_path(layout, version_id="wf", file_id="metrics")
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=_Instrumentation(),
        workbook_defs={},
        csv_defs={"metrics": str(csv_path)},
        sheetbook_defs={},
    )

    artifact = InMemoryCsv(header=list(_UNIQUE_FIELD_IDS), rows=[list(_INPUT_ROW)])
    manager.apply_csv_append(
        workflow_node_id="n1",
        decl_order=0,
        csv_id="metrics",
        input_node_id="n1",
        input_output_id="detail",
        input_csv=artifact,
        header_policy="once",
        on_mismatch="error",
        export_header=tuple(_DUP_DISPLAY_HEADER),
    )
    manager.commit_all()

    with csv_path.open(encoding="utf-8") as handle:
        rows = list(csv_mod.reader(handle))

    assert rows[0] == _DUP_DISPLAY_HEADER, "csv header must be export_header (display names), got: {}".format(rows[0])
    assert rows[1] == _EXPECTED_DATA_ROW, "csv data must align by field_id positionally, got: {}".format(rows[1])


def test_workbook_append_out_of_order_duplicate_display_headers_aligns_by_field_id(tmp_path: Path) -> None:
    """源头修复验证: append 乱序 + 重复展示名, 数据按 field_id 对齐(不坍缩).

    修复后 baseline_header=field_id(唯一). 第二次 append 的 input_header 为同一组
    field_id 的乱序排列; `_build_alignment_mapping(field_id, field_id_shuffled)` 按
    field_id 名匹配到正确列位置(而非按可重复展示名坍缩到首现索引). 表头仅写一次
    且为 export_header(展示名, 可重复).

    对照 bug(若工件 header=展示名): 映射会坍缩为 [1,0,1,0], 第二行被错位为
    ['30','40','30','40']; 修复后为 ['10','20','30','40'].

    注: 乱序 input_header != baseline 触发 mismatch, 这里用 `on_mismatch='warn'`
    容忍故意打乱以压测按 field_id 对齐(真实 workflow 同 sheet 的 append header 一致).
    """

    from openpyxl import load_workbook

    from scalim.execution import versioned_outputs

    out_root = tmp_path / "out"
    layout = versioned_outputs.ensure_output_root_layout(out_root)
    workbook_path = versioned_outputs.book_output_path(layout, version_id="wf", book_id="report")
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=_Instrumentation(),
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    # 首次 sheet 写入: field_id 顺序 [count_a, amount_a, count_b, amount_b], 行 [10,20,30,40].
    manager.apply_workbook_sheet(
        workflow_node_id="n1",
        decl_order=0,
        workbook_id="report",
        sheet="S1",
        input_node_id="n1",
        input_output_id="detail",
        input_csv=InMemoryCsv(header=list(_UNIQUE_FIELD_IDS), rows=[list(_INPUT_ROW)]),
        on_conflict="error",
        export_header=tuple(_DUP_DISPLAY_HEADER),
    )
    # 第二次 append: 同一组 field_id 但乱序输入, 行值按乱序 field_id 排列.
    shuffled_ids = list(reversed(_UNIQUE_FIELD_IDS))  # [amount_b, count_b, amount_a, count_a]
    # 乱序行: amount_b=40, count_b=30, amount_a=20, count_a=10
    manager.apply_workbook_append(
        workflow_node_id="n2",
        decl_order=1,
        workbook_id="report",
        sheet="S1",
        input_node_id="n2",
        input_output_id="detail",
        input_csv=InMemoryCsv(header=shuffled_ids, rows=[["40", "30", "20", "10"]]),
        align_by="field_id",
        header_policy="once",
        on_mismatch="warn",
        export_header=tuple(_DUP_DISPLAY_HEADER),
    )
    manager.commit_all()

    wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb["S1"].iter_rows(values_only=True)]
    finally:
        wb.close()

    # 表头仅写一次(展示名, 可重复); 两行数据均按 baseline field_id 顺序 = [10,20,30,40].
    assert rows[0] == _DUP_DISPLAY_HEADER, "header must be export_header, got: {}".format(rows[0])
    assert list(rows[1]) == _EXPECTED_DATA_ROW, "first row misaligned, got: {}".format(rows[1])
    assert list(rows[2]) == _EXPECTED_DATA_ROW, "appended row must align by field_id (not collapse), got: {}".format(rows[2])
