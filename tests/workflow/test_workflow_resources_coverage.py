import csv
import threading
from contextlib import suppress
from pathlib import Path
from typing import Any, Dict, List, Optional

import pytest

from scalim.workflow import resources as resources_mod
from scalim.workflow import resources_csv as resources_csv_mod
from scalim.workflow import resources_sheetbook as resources_sheetbook_mod
from scalim.workflow import resources_workbook as resources_workbook_mod
from scalim.events import (
    EVENT_DIAGNOSTIC_WARNING,
    EVENT_WORKFLOW_RESOURCE_DISCARD,
    EVENT_WORKFLOW_RESOURCE_WRITE,
)
from scalim.events._events import DiagnosticWarningEvent, WorkflowResourceWriteEvent

from tests.support.testing_utils import CI_TIMEOUT_S


class _Instrumentation:
    def __init__(self) -> None:
        self.events: List[Dict[str, Any]] = []
        self.warning_event = threading.Event()

    def emit(self, event_type: str, payload: Any, meta: Optional[Dict[str, Any]] = None) -> None:
        self.events.append({"event_type": str(event_type), "payload": payload, "meta": dict(meta or {})})
        if str(event_type) == EVENT_DIAGNOSTIC_WARNING:
            self.warning_event.set()


def _write_csv(path: Path, rows: List[List[str]]) -> Path:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)
    return path


def test_clone_exception_for_reraise_handles_fallbacks() -> None:
    from scalim._internal.utils.exceptions import clone_exception_for_reraise

    clone = clone_exception_for_reraise

    first = clone(resources_mod.ScalimWorkflowWriteError("boom"))
    assert isinstance(first, BaseException)

    try:
        raise ValueError("with traceback")
    except ValueError as exc:
        with_tb = exc

    cloned_with_tb = clone(with_tb)
    assert cloned_with_tb.__traceback__ is None

    class _CopyFails(Exception):
        def __reduce_ex__(self, _protocol: int) -> object:
            raise RuntimeError("no copy")

    exc2 = _CopyFails("x")
    second = clone(exc2)
    assert isinstance(second, _CopyFails)
    assert second is not exc2
    assert second.args == exc2.args

    class _CopyAndCtorFail(Exception):
        def __reduce_ex__(self, _protocol: int) -> object:
            raise RuntimeError("no copy")

        def __init__(self, code: int) -> None:
            if not isinstance(code, int):
                raise TypeError("code must be int")
            super(_CopyAndCtorFail, self).__init__(code)

    exc3 = _CopyAndCtorFail(1)
    exc3.args = ("bad",)  # type: ignore[assignment]
    third = clone(exc3)
    assert third is exc3


def test_best_effort_close_write_only_workbook_worksheets_handles_variants() -> None:
    class _NoWorksheets:
        pass

    class _BadWorksheets:
        worksheets = 123

    class _NoClosed:
        pass

    class _ClosedTrue:
        closed = True

    class _ClosedFalse:
        def __init__(self) -> None:
            self.closed = False

        def close(self) -> None:
            self.closed = True

    ws = _ClosedFalse()

    class _Workbook:
        worksheets = [_NoClosed(), _ClosedTrue(), ws]

    resources_workbook_mod._best_effort_close_write_only_workbook_worksheets(_NoWorksheets())
    resources_workbook_mod._best_effort_close_write_only_workbook_worksheets(_BadWorksheets())
    resources_workbook_mod._best_effort_close_write_only_workbook_worksheets(_Workbook())
    assert ws.closed is True


def test_read_csv_header_errors(tmp_path: Path) -> None:
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Missing input CSV"):
        _ = resources_csv_mod._read_csv_header(str(tmp_path / "nope.csv"))

    empty = _write_csv(tmp_path / "empty.csv", [])
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="empty"):
        _ = resources_csv_mod._read_csv_header(str(empty))

    invalid = _write_csv(tmp_path / "invalid.csv", [["", "ok"], ["x", "y"]])
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="invalid header"):
        _ = resources_csv_mod._read_csv_header(str(invalid))


def test_read_csv_header_in_memory_invalid_header_raises() -> None:
    from scalim.sinks import InMemoryCsv

    bad = InMemoryCsv(header=["", "ok"], rows=[])
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="<in_memory>"):
        _ = resources_csv_mod._read_csv_header(bad)


def test_resource_manager_unknown_resource_ids_raise(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a"]])
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Unknown workbook resource id"):
        manager.apply_workbook_sheet(
            workflow_node_id="n",
            decl_order=0,
            workbook_id="nope",
            sheet="S",
            input_node_id="a",
            input_output_id="out",
            input_csv=str(csv_path),
            on_conflict="error",
        )

    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Unknown csv resource id"):
        manager.apply_csv_append(
            workflow_node_id="n",
            decl_order=0,
            csv_id="nope",
            input_node_id="a",
            input_output_id="out",
            input_csv=str(csv_path),
            header_policy="once",
            on_mismatch="error",
        )

    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Unknown sheetbook resource id"):
        manager.apply_sheetbook_sheet(
            workflow_node_id="n",
            decl_order=0,
            sheetbook_id="nope",
            sheet="S",
            input_node_id="a",
            input_output_id="out",
            input_csv=str(csv_path),
            on_conflict="error",
        )


def test_resource_manager_book_unknown_resource_ids_raise(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a1"]])
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Unknown book resource id"):
        manager.apply_book_sheet(
            workflow_node_id="n",
            decl_order=0,
            book_id="nope",
            sheet="S",
            input_node_id="a",
            input_output_id="out",
            input_csv=str(csv_path),
            on_conflict="error",
        )

    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Unknown book resource id"):
        manager.apply_book_append(
            workflow_node_id="n",
            decl_order=0,
            book_id="nope",
            sheet="S",
            input_node_id="a",
            input_output_id="out",
            input_csv=str(csv_path),
            align_by="field_id",
            header_policy="once",
            on_mismatch="error",
        )


def test_resource_manager_iter_book_sheet_rows_requires_xlsx_memory(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(tmp_path / "report.xlsx")},
        csv_defs={},
        sheetbook_defs={},
    )

    with pytest.raises(ValueError, match="only supports xlsx_memory"):
        _ = manager.iter_book_sheet_rows(
            consumer_node_id="consumer",
            visible_producer_node_ids=frozenset(["producer"]),
            producer_node_id="producer",
            book_id="report",
            sheet="S",
        )


def test_resource_manager_workbook_append_mismatch_error_warn_skip(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    workbook_path = tmp_path / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    mismatch = _write_csv(tmp_path / "b.csv", [["id", "other"], ["a1", "A1"]])

    manager.apply_workbook_append(
        workflow_node_id="n0",
        decl_order=0,
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Field alignment mismatch"):
        manager.apply_workbook_append(
            workflow_node_id="n1",
            decl_order=1,
            workbook_id="report",
            sheet="S",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(mismatch),
            align_by="field_id",
            header_policy="once",
            on_mismatch="error",
        )
    manager.discard_all(workflow_node_id="n_discard", reason="test")

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(tmp_path / "warn.xlsx")},
        csv_defs={},
        sheetbook_defs={},
    )
    manager.apply_workbook_append(
        workflow_node_id="n0",
        decl_order=0,
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_workbook_append(
        workflow_node_id="n1",
        decl_order=1,
        workbook_id="report",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(mismatch),
        align_by="field_id",
        header_policy="once",
        on_mismatch="warn",
    )
    assert [e for e in instrumentation.events if e["event_type"] == EVENT_DIAGNOSTIC_WARNING]
    manager.discard_all(workflow_node_id="n_discard", reason="test")

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(tmp_path / "skip.xlsx")},
        csv_defs={},
        sheetbook_defs={},
    )
    manager.apply_workbook_append(
        workflow_node_id="n0",
        decl_order=0,
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_workbook_append(
        workflow_node_id="n1",
        decl_order=1,
        workbook_id="report",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(mismatch),
        align_by="field_id",
        header_policy="once",
        on_mismatch="skip",
    )
    skip_events = [e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_WRITE and e["payload"].action == "skip"]
    assert skip_events
    manager.discard_all(workflow_node_id="n_discard", reason="test")


def test_resource_manager_csv_append_mismatch_error_and_discard(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    csv_path = tmp_path / "merged.csv"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={"merged": str(csv_path)},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    mismatch = _write_csv(tmp_path / "b.csv", [["id", "other"], ["a1", "A1"]])

    manager.apply_csv_append(
        workflow_node_id="n0",
        decl_order=0,
        csv_id="merged",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Field alignment mismatch"):
        manager.apply_csv_append(
            workflow_node_id="n1",
            decl_order=1,
            csv_id="merged",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(mismatch),
            header_policy="once",
            on_mismatch="error",
        )

    manager.discard_all(workflow_node_id="n_discard", reason="test")


def test_resource_manager_csv_append_mismatch_warn_and_skip(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    csv_path = tmp_path / "merged.csv"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={"merged": str(csv_path)},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    manager.apply_csv_append(
        workflow_node_id="n0",
        decl_order=0,
        csv_id="merged",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        header_policy="once",
        on_mismatch="error",
    )

    mismatch = _write_csv(tmp_path / "b.csv", [["id", "other"], ["b1", "B1"]])
    manager.apply_csv_append(
        workflow_node_id="n1",
        decl_order=1,
        csv_id="merged",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(mismatch),
        header_policy="once",
        on_mismatch="warn",
    )
    assert [e for e in instrumentation.events if e["event_type"] == EVENT_DIAGNOSTIC_WARNING]

    mismatch2 = _write_csv(tmp_path / "c.csv", [["id", "another"], ["c1", "C1"]])
    manager.apply_csv_append(
        workflow_node_id="n2",
        decl_order=2,
        csv_id="merged",
        input_node_id="c",
        input_output_id="detail",
        input_csv=str(mismatch2),
        header_policy="once",
        on_mismatch="skip",
    )
    skip_events = [e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_WRITE and e["payload"].action == "skip"]
    assert skip_events

    manager.discard_all(workflow_node_id="n_discard", reason="test")


def test_resource_manager_sheetbook_sheet_conflict_skip_error_overwrite(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
            )
        },
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])

    manager.apply_sheetbook_sheet(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        on_conflict="error",
    )

    manager.apply_sheetbook_sheet(
        workflow_node_id="n1",
        decl_order=1,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(second),
        on_conflict="skip",
    )
    assert [e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_WRITE and e["payload"].action == "skip"]

    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Sheet conflict"):
        manager.apply_sheetbook_sheet(
            workflow_node_id="n2",
            decl_order=2,
            sheetbook_id="sb",
            sheet="S",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(second),
            on_conflict="error",
        )

    manager.apply_sheetbook_sheet(
        workflow_node_id="n3",
        decl_order=3,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="c",
        input_output_id="detail",
        input_csv=str(second),
        on_conflict="overwrite",
    )
    overwrite_events = [
        e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_WRITE and e["payload"].action == "overwrite"
    ]
    assert overwrite_events

    rows = list(
        manager.iter_sheetbook_sheet_rows(
            consumer_node_id="consumer",
            visible_producer_node_ids=frozenset(),
            producer_node_id="c",
            sheetbook_id="sb",
            sheet="S",
        )
    )
    assert rows == [{"id": "b1", "value": "B1"}]


def test_resource_manager_sheetbook_sheet_budget_max_sheets_is_enforced(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=1,
                budget_max_total_cells=1000,
                export_path=None,
            )
        },
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])

    manager.apply_sheetbook_sheet(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S1",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        on_conflict="error",
    )

    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="max_sheets"):
        manager.apply_sheetbook_sheet(
            workflow_node_id="n1",
            decl_order=1,
            sheetbook_id="sb",
            sheet="S2",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(first),
            on_conflict="error",
        )


def test_resource_manager_sheetbook_append_mismatch_error_warn_skip_and_budget(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
            )
        },
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    mismatch = _write_csv(tmp_path / "b.csv", [["id", "other"], ["b1", "B1"]])

    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Field alignment mismatch"):
        manager.apply_sheetbook_append(
            workflow_node_id="n1",
            decl_order=1,
            sheetbook_id="sb",
            sheet="S",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(mismatch),
            align_by="field_id",
            header_policy="once",
            on_mismatch="error",
        )
    manager.discard_all(workflow_node_id="n_discard", reason="test")

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
            )
        },
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n1",
        decl_order=1,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(mismatch),
        align_by="field_id",
        header_policy="once",
        on_mismatch="warn",
    )
    assert [e for e in instrumentation.events if e["event_type"] == EVENT_DIAGNOSTIC_WARNING]
    manager.discard_all(workflow_node_id="n_discard", reason="test")

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
            )
        },
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n1",
        decl_order=1,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(mismatch),
        align_by="field_id",
        header_policy="once",
        on_mismatch="skip",
    )
    skip_events = [e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_WRITE and e["payload"].action == "skip"]
    assert skip_events
    manager.discard_all(workflow_node_id="n_discard", reason="test")

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=1,
                budget_max_total_cells=1000,
                export_path=None,
            )
        },
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S1",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="max_sheets"):
        manager.apply_sheetbook_append(
            workflow_node_id="n1",
            decl_order=1,
            sheetbook_id="sb",
            sheet="S2",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(first),
            align_by="field_id",
            header_policy="once",
            on_mismatch="error",
        )
    manager.discard_all(workflow_node_id="n_discard", reason="test")


def test_resource_manager_sheetbook_append_duplicate_producer_is_rejected(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
            )
        },
    )
    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])

    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Duplicate sheetbook write"):
        manager.apply_sheetbook_append(
            workflow_node_id="n1",
            decl_order=1,
            sheetbook_id="sb",
            sheet="S",
            input_node_id="a",
            input_output_id="detail",
            input_csv=str(first),
            align_by="field_id",
            header_policy="once",
            on_mismatch="error",
        )


def test_resource_manager_sheetbook_export_header_metadata_controls_xlsx_header(tmp_path: Path) -> None:
    from scalim.execution import versioned_outputs

    openpyxl = pytest.importorskip("openpyxl")

    out_root = tmp_path / "out"
    export_path = versioned_outputs.book_output_path(
        versioned_outputs.ensure_output_root_layout(out_root),
        version_id="wf",
        book_id="sb",
    )
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(export_path),
            )
        },
    )
    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])

    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
        export_header=("Order ID", "Display Value"),
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n1",
        decl_order=1,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(second),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
        export_header=("Order ID", "Display Value"),
    )

    manager.commit_all()

    wb = openpyxl.load_workbook(str(export_path))
    ws = wb["S"]
    rows = list(ws.iter_rows(values_only=True))
    assert list(rows[0]) == ["Order ID", "Display Value"]
    assert list(rows[1]) == ["a1", "A1"]
    assert list(rows[2]) == ["b1", "B1"]

    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="export header baseline mismatch"):
        manager.apply_sheetbook_append(
            workflow_node_id="n2",
            decl_order=2,
            sheetbook_id="sb",
            sheet="S",
            input_node_id="c",
            input_output_id="detail",
            input_csv=str(second),
            align_by="field_id",
            header_policy="once",
            on_mismatch="error",
            export_header=("Order ID", "Other Display"),
        )


def test_resource_manager_sheetbook_rejects_header_alignment(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
            )
        },
    )
    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])

    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="align_by=header"):
        manager.apply_sheetbook_append(
            workflow_node_id="n0",
            decl_order=0,
            sheetbook_id="sb",
            sheet="S",
            input_node_id="a",
            input_output_id="detail",
            input_csv=str(first),
            align_by="header",
            header_policy="once",
            on_mismatch="error",
        )


def test_sheetbook_export_header_helper_rejects_width_mismatch() -> None:
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="width mismatch"):
        resources_mod.WorkflowSheetBookResourceMixin._normalize_sheetbook_export_header(  # noqa: SLF001
            ["id", "value"],
            ("仅一列",),
        )


def test_sheetbook_alignment_mismatch_helper_covers_header_mode() -> None:
    assert resources_mod.WorkflowSheetBookResourceMixin._sheetbook_has_alignment_mismatch(  # noqa: SLF001
        ["id", "value"],
        ["id", "other"],
        align_by="header",
    )


def test_workflow_resource_manager_emit_does_not_deadlock_on_reentry(tmp_path: Path) -> None:
    class _ReentrantInstrumentation:
        def __init__(self) -> None:
            self._reentered = False
            self.manager = None
            self.csv_input_path = None

        def emit(self, event_type: str, payload: Any, meta: Optional[Dict[str, Any]] = None) -> None:
            _ = event_type, meta
            if self._reentered:
                return
            if str(event_type) != EVENT_WORKFLOW_RESOURCE_WRITE:
                return
            write_payload = payload
            if not isinstance(write_payload, WorkflowResourceWriteEvent):
                return
            if write_payload.action != "skip":
                return
            self._reentered = True
            self.manager.apply_csv_append(  # type: ignore[union-attr]
                workflow_node_id="reentry",
                decl_order=0,
                csv_id="merged",
                input_node_id="r",
                input_output_id="out",
                input_csv=str(self.csv_input_path),  # type: ignore[arg-type]
                header_policy="once",
                on_mismatch="error",
            )

    instrumentation = _ReentrantInstrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(tmp_path / "report.xlsx")},
        csv_defs={"merged": str(tmp_path / "merged.csv")},
        sheetbook_defs={},
    )
    instrumentation.manager = manager

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    instrumentation.csv_input_path = str(first)

    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        decl_order=0,
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        on_conflict="error",
    )

    runner_done = threading.Event()
    runner_errors: List[BaseException] = []

    def _run() -> None:
        try:
            manager.apply_workbook_sheet(
                workflow_node_id="n1",
                decl_order=1,
                workbook_id="report",
                sheet="S",
                input_node_id="b",
                input_output_id="detail",
                input_csv=str(first),
                on_conflict="skip",
            )
        except BaseException as exc:
            runner_errors.append(exc)
        finally:
            runner_done.set()

    runner = threading.Thread(target=_run, daemon=True)
    runner.start()
    if not runner_done.wait(timeout=CI_TIMEOUT_S):
        pytest.fail("WorkflowResourceManager.emit appears to be called under internal locks (reentry deadlock)")
    runner.join(timeout=CI_TIMEOUT_S)
    assert not runner_errors
    assert instrumentation._reentered is True


def test_resource_manager_workbook_sheet_conflict_unknown_on_conflict_is_treated_as_write(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    workbook_path = output_root / "versions" / "wf" / "books" / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])

    manager.apply_workbook_sheet(
        workflow_node_id="n1",
        decl_order=0,
        workbook_id="report",
        sheet="S",
        input_node_id="n1",
        input_output_id="detail",
        input_csv=str(first),
        on_conflict="error",
    )
    manager.apply_workbook_sheet(
        workflow_node_id="n2",
        decl_order=1,
        workbook_id="report",
        sheet="S",
        input_node_id="n2",
        input_output_id="detail",
        input_csv=str(first),
        on_conflict="unknown",
    )

    write_events = [e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_WRITE]
    assert isinstance(write_events[-1]["payload"], WorkflowResourceWriteEvent)
    assert write_events[-1]["payload"].action == "write"


def test_resource_manager_commit_workbook_publishes_versioned_output_and_manifest(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from scalim.execution import versioned_outputs

    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    workbook_path = output_root / "versions" / "wf" / "books" / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])
    manager.apply_workbook_sheet(
        workflow_node_id="n1",
        decl_order=0,
        workbook_id="report",
        sheet="S1",
        input_node_id="n1",
        input_output_id="detail",
        input_csv=str(first),
        on_conflict="error",
    )
    manager.apply_workbook_sheet(
        workflow_node_id="n2",
        decl_order=1,
        workbook_id="report",
        sheet="S2",
        input_node_id="n2",
        input_output_id="detail",
        input_csv=str(second),
        on_conflict="error",
    )
    manager.commit_all()

    wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
    try:
        assert list(wb["S1"].iter_rows(values_only=True)) == [("id", "value"), ("a1", "A1")]
        assert list(wb["S2"].iter_rows(values_only=True)) == [("id", "value"), ("b1", "B1")]
    finally:
        with suppress(Exception):
            wb.close()

    latest = versioned_outputs.read_latest(output_root)
    assert latest.get("version_id") == "wf"
    assert latest.get("version_manifest_relpath") == versioned_outputs.version_manifest_relpath(version_id="wf")

    manifest = versioned_outputs.read_version_manifest(output_root, version_id="wf")
    assert manifest.get("books") == {"report": "books/report.xlsx"}
    assert manifest.get("files") == {}
    assert not list(output_root.rglob("*.scalim.lock"))


def test_resource_manager_commit_csv_publishes_versioned_output_and_manifest(tmp_path: Path) -> None:
    from scalim.execution import versioned_outputs

    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    output_path = output_root / "versions" / "wf" / "files" / "merged.csv"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={"merged": str(output_path)},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])
    manager.apply_csv_append(
        workflow_node_id="n1",
        decl_order=0,
        csv_id="merged",
        input_node_id="n1",
        input_output_id="detail",
        input_csv=str(first),
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_csv_append(
        workflow_node_id="n2",
        decl_order=1,
        csv_id="merged",
        input_node_id="n2",
        input_output_id="detail",
        input_csv=str(second),
        header_policy="once",
        on_mismatch="error",
    )
    manager.commit_all()

    with output_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    assert rows[0] == ["id", "value"]
    assert {tuple(row) for row in rows[1:]} == {("a1", "A1"), ("b1", "B1")}

    latest = versioned_outputs.read_latest(output_root)
    assert latest.get("version_id") == "wf"
    assert latest.get("version_manifest_relpath") == versioned_outputs.version_manifest_relpath(version_id="wf")

    manifest = versioned_outputs.read_version_manifest(output_root, version_id="wf")
    assert manifest.get("books") == {}
    assert manifest.get("files") == {"merged": "files/merged.csv"}
    assert not list(output_root.rglob("*.scalim.lock"))


def test_resource_manager_sheetbook_iter_rows_visibility_and_errors(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
            )
        },
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])

    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n1",
        decl_order=1,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(second),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )

    rows = list(
        manager.iter_sheetbook_sheet_rows(
            consumer_node_id="consumer",
            visible_producer_node_ids=frozenset(),
            producer_node_id="b",
            sheetbook_id="sb",
            sheet="S",
        )
    )
    assert rows == [{"id": "b1", "value": "B1"}]

    with pytest.raises(ValueError, match="Unknown sheetbook resource id"):
        _ = list(
            manager.iter_sheetbook_sheet_rows(
                consumer_node_id="consumer",
                visible_producer_node_ids=frozenset(),
                producer_node_id="a",
                sheetbook_id="nope",
                sheet="S",
            )
        )

    with pytest.raises(ValueError, match="Unknown sheetbook sheet"):
        _ = list(
            manager.iter_sheetbook_sheet_rows(
                consumer_node_id="consumer",
                visible_producer_node_ids=frozenset(),
                producer_node_id="a",
                sheetbook_id="sb",
                sheet="Other",
            )
        )

    with pytest.raises(ValueError, match="Unknown sheetbook ref node"):
        _ = list(
            manager.iter_sheetbook_sheet_rows(
                consumer_node_id="consumer",
                visible_producer_node_ids=frozenset(),
                producer_node_id="nope",
                sheetbook_id="sb",
                sheet="S",
            )
        )


def test_resource_manager_sheetbook_commit_import_error_and_save_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    output_root = tmp_path / "out"
    export_path = output_root / "versions" / "wf" / "books" / "sb.xlsx"
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(export_path),
            )
        },
    )
    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )

    def _raise_import(*_args: object, **_kwargs: object) -> Any:  # noqa: ANN401
        raise ImportError("no openpyxl")

    monkeypatch.setattr(resources_workbook_mod, "require_optional_dependency", _raise_import)
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="openpyxl"):
        manager.commit_all()
    assert not Path(str(export_path) + ".scalim.lock").exists()

    class _FakeWS:
        def append(self, row: List[object]) -> None:
            _ = row

    class _FakeWB:
        def __init__(self, write_only: bool = True) -> None:
            _ = write_only

        def create_sheet(self, name: str) -> _FakeWS:
            _ = name
            return _FakeWS()

        def save(self, path: Path) -> None:
            path.write_text("tmp", encoding="utf-8")
            raise RuntimeError("boom")

        def close(self) -> None:
            return

    output_root2 = tmp_path / "out2"
    export_path2 = output_root2 / "versions" / "wf" / "books" / "sb.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(export_path2),
            )
        },
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )

    temp_path = tmp_path / "temp.xlsx.tmp"
    monkeypatch.setattr(
        resources_workbook_mod,
        "require_optional_dependency",
        lambda *_args, **_kwargs: type("_M", (), {"Workbook": _FakeWB})(),
    )
    monkeypatch.setattr(resources_sheetbook_mod, "create_temp_path", lambda _path, _suffix: str(temp_path))
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Sheetbook export failed"):
        manager.commit_all()
    assert not temp_path.exists()


def test_resource_manager_sheetbook_discard_does_not_remove_external_files(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    export_path = output_root / "versions" / "wf" / "books" / "sb.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(export_path),
            )
        },
    )
    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )

    external_path = Path(str(export_path) + ".scalim.lock")
    external_path.parent.mkdir(parents=True, exist_ok=True)
    external_path.write_text("lock", encoding="utf-8")

    manager.discard_all(workflow_node_id="n_discard", reason="test")
    assert external_path.exists()
    assert [e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_DISCARD]


def test_resource_manager_commit_all_skips_empty_plans(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()

    output_root = tmp_path / "out"
    workbook_path = output_root / "versions" / "wf" / "books" / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )
    _ = manager._get_or_create_workbook("report", workflow_node_id="n")  # noqa: SLF001
    manager.commit_all()
    assert not workbook_path.exists()
    assert not output_root.exists()

    csv_output_path = output_root / "versions" / "wf" / "files" / "merged.csv"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=_Instrumentation(),
        workbook_defs={},
        csv_defs={"merged": str(csv_output_path)},
        sheetbook_defs={},
    )
    _ = manager._get_or_create_csv("merged", workflow_node_id="n")  # noqa: SLF001
    manager.commit_all()
    assert not csv_output_path.exists()
    assert not output_root.exists()


def test_resource_manager_commit_workbook_escapes_excel_formulas_by_default(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    workbook_path = output_root / "versions" / "wf" / "books" / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(
        tmp_path / "input.csv",
        [
            ["=h", "eq", "plus", "minus", "at"],
            ["ok", "=1+1", "  +SUM(A1:A2)", "-1+2", "@X"],
        ],
    )
    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        decl_order=0,
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )
    manager.commit_all()

    wb = load_workbook(str(workbook_path), data_only=False)
    try:
        ws = wb["S"]
        assert ws["A1"].value == "'=h"
        assert ws["A2"].value == "ok"

        eq_value = ws["B2"].value
        eq_type = ws["B2"].data_type
        assert eq_type != "f"
        assert eq_value == "'=1+1"

        assert ws["C2"].value == "'  +SUM(A1:A2)"
        assert ws["D2"].value == "'-1+2"
        assert ws["E2"].value == "'@X"
    finally:
        wb.close()
    assert not list(output_root.rglob("*.scalim.lock"))


def test_resource_manager_commit_workbook_allow_formulas_preserves_raw_strings(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    workbook_path = output_root / "versions" / "wf" / "books" / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        workbook_allow_formulas={"report": True},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(
        tmp_path / "input.csv",
        [
            ["=h", "eq", "plus", "minus", "at"],
            ["ok", "=1+1", "  +SUM(A1:A2)", "-1+2", "@X"],
        ],
    )
    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        decl_order=0,
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )
    manager.commit_all()

    wb = load_workbook(str(workbook_path), data_only=False)
    try:
        ws = wb["S"]
        assert ws["A1"].value == "=h"
        assert ws["A2"].value == "ok"

        eq_value = ws["B2"].value
        eq_type = ws["B2"].data_type
        assert eq_type == "f"
        assert eq_value == "=1+1"

        assert ws["C2"].value == "  +SUM(A1:A2)"
        assert ws["D2"].value == "-1+2"
        assert ws["E2"].value == "@X"
    finally:
        wb.close()
    assert not list(output_root.rglob("*.scalim.lock"))


def test_resource_manager_commit_sheetbook_escapes_excel_formulas_by_default(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    export_path = output_root / "versions" / "wf" / "books" / "sb.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(export_path),
            )
        },
    )

    csv_path = _write_csv(
        tmp_path / "input.csv",
        [
            ["=h", "eq", "plus", "minus", "at"],
            ["ok", "=1+1", "  +SUM(A1:A2)", "-1+2", "@X"],
        ],
    )
    manager.apply_sheetbook_sheet(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )
    manager.commit_all()

    wb = load_workbook(str(export_path), data_only=False)
    try:
        ws = wb["S"]
        assert ws["A1"].value == "'=h"
        assert ws["A2"].value == "ok"

        eq_value = ws["B2"].value
        eq_type = ws["B2"].data_type
        assert eq_type != "f"
        assert eq_value == "'=1+1"

        assert ws["C2"].value == "'  +SUM(A1:A2)"
        assert ws["D2"].value == "'-1+2"
        assert ws["E2"].value == "'@X"
    finally:
        wb.close()
    assert not list(output_root.rglob("*.scalim.lock"))


def test_resource_manager_commit_sheetbook_allow_formulas_preserves_raw_strings(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    export_path = output_root / "versions" / "wf" / "books" / "sb.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(export_path),
                export_allow_formulas=True,
            )
        },
    )

    csv_path = _write_csv(
        tmp_path / "input.csv",
        [
            ["=h", "eq", "plus", "minus", "at"],
            ["ok", "=1+1", "  +SUM(A1:A2)", "-1+2", "@X"],
        ],
    )
    manager.apply_sheetbook_sheet(
        workflow_node_id="n0",
        decl_order=0,
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )
    manager.commit_all()

    wb = load_workbook(str(export_path), data_only=False)
    try:
        ws = wb["S"]
        assert ws["A1"].value == "=h"
        assert ws["A2"].value == "ok"

        eq_value = ws["B2"].value
        eq_type = ws["B2"].data_type
        assert eq_type == "f"
        assert eq_value == "=1+1"

        assert ws["C2"].value == "  +SUM(A1:A2)"
        assert ws["D2"].value == "-1+2"
        assert ws["E2"].value == "@X"
    finally:
        wb.close()
    assert not list(output_root.rglob("*.scalim.lock"))


def test_resource_manager_commit_workbook_handles_missing_sheet_plan(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    workbook_path = output_root / "versions" / "wf" / "books" / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a1"]])
    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        decl_order=0,
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )

    plan = manager._workbooks["report"]  # noqa: SLF001
    plan.sheet_order.insert(0, "MISSING")

    manager.commit_all()
    assert workbook_path.exists()
    assert not list(output_root.rglob("*.scalim.lock"))


def test_resource_manager_commit_workbook_commit_failure_raises(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    output_path = output_root / "versions" / "wf" / "books" / "report.xlsx"
    output_path.mkdir(parents=True)

    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(output_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a1"]])
    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        decl_order=0,
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )

    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Publish staged output failed"):
        manager.commit_all()
    assert not Path(str(output_path) + ".scalim.lock").exists()
    assert not list(output_root.rglob("*.scalim.lock"))


def test_resource_manager_commit_csv_commit_failure_raises(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    output_path = output_root / "versions" / "wf" / "files" / "merged.csv"
    output_path.mkdir(parents=True)

    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={"merged": str(output_path)},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a1"]])
    manager.apply_csv_append(
        workflow_node_id="n0",
        decl_order=0,
        csv_id="merged",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="Publish staged output failed"):
        manager.commit_all()
    assert not Path(str(output_path) + ".scalim.lock").exists()
    assert not list(output_root.rglob("*.scalim.lock"))


def test_resource_manager_concurrent_workflows_publish_versions_and_latest_json_valid(tmp_path: Path) -> None:
    from scalim.execution import versioned_outputs

    output_root = tmp_path / "out"

    manager1 = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf1",
        instrumentation=_Instrumentation(),
        workbook_defs={},
        csv_defs={"merged": str(output_root / "versions" / "wf1" / "files" / "merged.csv")},
        sheetbook_defs={},
    )
    manager2 = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf2",
        instrumentation=_Instrumentation(),
        workbook_defs={},
        csv_defs={"merged": str(output_root / "versions" / "wf2" / "files" / "merged.csv")},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id"], ["a1"]])
    second = _write_csv(tmp_path / "b.csv", [["id"], ["b1"]])
    manager1.apply_csv_append(
        workflow_node_id="n1",
        decl_order=0,
        csv_id="merged",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        header_policy="once",
        on_mismatch="error",
    )
    manager2.apply_csv_append(
        workflow_node_id="n2",
        decl_order=0,
        csv_id="merged",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(second),
        header_policy="once",
        on_mismatch="error",
    )

    stop_reader = threading.Event()
    read_errors: List[BaseException] = []

    def _reader() -> None:
        while not stop_reader.is_set():
            try:
                latest = versioned_outputs.read_latest(output_root)
            except FileNotFoundError:
                stop_reader.wait(timeout=0.001)
                continue
            try:
                assert str(latest.get("version_id") or "") in {"wf1", "wf2"}
            except BaseException as exc:
                read_errors.append(exc)
                return

    reader = threading.Thread(target=_reader, daemon=True)
    reader.start()

    barrier = threading.Barrier(2)
    errors: List[BaseException] = []

    def _run(manager: resources_mod.WorkflowResourceManager) -> None:
        try:
            _ = barrier.wait(timeout=CI_TIMEOUT_S)
            manager.commit_all()
        except BaseException as exc:
            errors.append(exc)

    t1 = threading.Thread(target=lambda: _run(manager1), daemon=True)
    t2 = threading.Thread(target=lambda: _run(manager2), daemon=True)
    t1.start()
    t2.start()
    t1.join(timeout=CI_TIMEOUT_S)
    t2.join(timeout=CI_TIMEOUT_S)
    assert not t1.is_alive()
    assert not t2.is_alive()

    stop_reader.set()
    reader.join(timeout=CI_TIMEOUT_S)
    assert not reader.is_alive()

    assert errors == []
    assert read_errors == []
    assert (output_root / "versions" / "wf1" / "files" / "merged.csv").exists()
    assert (output_root / "versions" / "wf2" / "files" / "merged.csv").exists()

    latest = versioned_outputs.read_latest(output_root)
    assert str(latest.get("version_id") or "") in {"wf1", "wf2"}

    m1 = versioned_outputs.read_version_manifest(output_root, version_id="wf1")
    m2 = versioned_outputs.read_version_manifest(output_root, version_id="wf2")
    assert m1.get("files") == {"merged": "files/merged.csv"}
    assert m2.get("files") == {"merged": "files/merged.csv"}
    assert not list(output_root.rglob("*.scalim.lock"))


def test_resource_manager_commit_workbook_missing_openpyxl_does_not_write_output(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    instrumentation = _Instrumentation()
    output_root = tmp_path / "out"
    workbook_path = output_root / "versions" / "wf" / "books" / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a1"]])
    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        decl_order=0,
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )

    def _raise_import_error(*args: object, **kwargs: object) -> object:
        raise ImportError("missing openpyxl")

    monkeypatch.setattr(resources_workbook_mod, "require_optional_dependency", _raise_import_error)
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match="missing openpyxl"):
        manager.commit_all()
    assert not workbook_path.exists()


def test_csv_commit_replace_failure_raises_workflow_write_error(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from scalim.sinks import InMemoryCsv

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={},
    )

    final_path = str(tmp_path / "out.csv")
    staging_path = manager._staging_path_for_final_output(final_path)  # noqa: SLF001

    original_replace = Path.replace

    def _replace(self: Path, target: object) -> Path:  # noqa: ANN001
        if str(target) == str(staging_path):
            raise OSError("boom")
        return original_replace(self, target)

    monkeypatch.setattr(Path, "replace", _replace)

    seg = resources_csv_mod.AppendSegment(
        decl_order=0,
        input_csv=InMemoryCsv(header=["id"], rows=[["a1"]]),
        header_policy="once",
        mapping=[0],
        on_mismatch="error",
        align_by="header",
        input_header=["id"],
    )
    plan = resources_csv_mod.CsvPlan(resource_id="r", path=str(final_path), baseline_header=["id"], segments=[seg])
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match=r"CSV commit failed: OSError: boom"):
        manager._commit_csv(plan)  # noqa: SLF001


def test_workbook_commit_save_failure_raises_workflow_write_error(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from scalim.sinks import InMemoryCsv

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={},
    )

    workbook_cls = resources_workbook_mod.get_openpyxl_workbook_class()

    def _boom_save(self: object, *args: object, **kwargs: object) -> object:  # noqa: ARG001
        raise OSError("boom")

    monkeypatch.setattr(workbook_cls, "save", _boom_save)

    seg = resources_csv_mod.AppendSegment(
        decl_order=0,
        input_csv=InMemoryCsv(header=["id"], rows=[["a1"]]),
        header_policy="once",
        mapping=[0],
        on_mismatch="error",
        align_by="header",
        input_header=["id"],
    )
    sheet_plan = resources_workbook_mod.SheetPlan(sheet="S", baseline_header=["id"], segments=[seg])
    plan = resources_workbook_mod.WorkbookPlan(
        resource_id="report",
        path=str(tmp_path / "out.xlsx"),
        allow_formulas=False,
        sheet_decl_order={"S": 0},
        sheet_order=["S"],
        sheets={"S": sheet_plan},
    )
    with pytest.raises(resources_mod.ScalimWorkflowWriteError, match=r"Workbook commit failed: OSError: boom"):
        manager._commit_workbook(plan)  # noqa: SLF001
