from pathlib import Path

import pytest

from scalim.sinks.memory import InMemoryCsv
from scalim.sinks.rows import InMemoryRows
from scalim.spec.ir._workflow import WorkflowArtifactsIr, WorkflowIr, WorkflowOptionsIr
from scalim.workflow.artifacts import WorkflowArtifactsDirectory
from scalim.workflow.input_artifacts import resolve_workflow_input_tabular
from scalim.workflow.resources import WorkflowResourceManager
from scalim.workflow.resources_base import ScalimWorkflowWriteError
from scalim.workflow.tabular_artifacts import read_tabular_header


def _make_artifacts_dir() -> WorkflowArtifactsDirectory:
    workflow_ir = WorkflowIr(
        nodes=(),
        edges=(),
        options=WorkflowOptionsIr(max_concurrency=1, failure_policy="all_fail"),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={}),
    )
    return WorkflowArtifactsDirectory(workflow_ir)


def test_resolve_workflow_input_tabular_rejects_non_csv_output_path() -> None:
    artifacts_dir = _make_artifacts_dir()
    artifacts_dir.publish("a", "outputs", {"detail": "./out.xlsx"})

    with pytest.raises(ScalimWorkflowWriteError, match="only supports CSV outputs"):
        _ = resolve_workflow_input_tabular(
            artifacts_dir=artifacts_dir,
            consumer_node_id="a",
            consumer_decl_order=0,
            input_node_id="a",
            input_output_id="detail",
            error_prefix="write node",
        )


def test_resolve_workflow_input_tabular_missing_artifact_raises_tabular_error() -> None:
    artifacts_dir = _make_artifacts_dir()
    artifacts_dir.publish("a", "outputs", {"detail": ""})

    with pytest.raises(ScalimWorkflowWriteError, match="Missing workflow-managed tabular artifact"):
        _ = resolve_workflow_input_tabular(
            artifacts_dir=artifacts_dir,
            consumer_node_id="a",
            consumer_decl_order=0,
            input_node_id="a",
            input_output_id="detail",
            error_prefix="write node",
        )


def test_resolve_workflow_input_tabular_unknown_output_id_reraises_original_error() -> None:
    artifacts_dir = _make_artifacts_dir()
    artifacts_dir.publish("a", "outputs", {"other": ""})

    with pytest.raises(ScalimWorkflowWriteError, match="Unknown demand output id"):
        _ = resolve_workflow_input_tabular(
            artifacts_dir=artifacts_dir,
            consumer_node_id="a",
            consumer_decl_order=0,
            input_node_id="a",
            input_output_id="detail",
            error_prefix="write node",
        )


def test_read_tabular_header_rejects_invalid_in_memory_rows_header() -> None:
    artifact = object.__new__(InMemoryRows)
    object.__setattr__(artifact, "header", ["id", ""])
    object.__setattr__(artifact, "rows", [[1, 2]])
    with pytest.raises(ScalimWorkflowWriteError, match="in_memory_rows"):
        _ = read_tabular_header(artifact)


def test_read_tabular_header_rejects_invalid_in_memory_csv_header() -> None:
    with pytest.raises(ScalimWorkflowWriteError, match="in_memory_csv"):
        _ = read_tabular_header(InMemoryCsv(header=["id", ""], rows=[["1", "2"]]))


def test_read_tabular_header_rejects_missing_csv_path(tmp_path: Path) -> None:
    with pytest.raises(ScalimWorkflowWriteError, match="Missing input CSV"):
        _ = read_tabular_header(str(tmp_path / "missing.csv"))


def test_read_tabular_header_rejects_empty_csv(tmp_path: Path) -> None:
    path = tmp_path / "empty.csv"
    path.write_text("", encoding="utf-8")

    with pytest.raises(ScalimWorkflowWriteError, match="missing header"):
        _ = read_tabular_header(str(path))


def test_read_tabular_header_rejects_invalid_csv_header(tmp_path: Path) -> None:
    path = tmp_path / "bad.csv"
    path.write_text("id,,name\n1,2,3\n", encoding="utf-8")

    with pytest.raises(ScalimWorkflowWriteError, match="invalid header"):
        _ = read_tabular_header(str(path))


def test_workflow_resource_manager_accepts_rows_for_xlsx_file(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from scalim.execution import versioned_outputs
    from scalim.workflow.resources import WorkflowResourceManager as WRM

    class _Instrumentation:
        def emit(self, *_args: object, **_kwargs: object) -> None:
            return None

    layout = versioned_outputs.ensure_output_root_layout(tmp_path / "out")
    workbook_path = versioned_outputs.book_output_path(layout, version_id="wf", book_id="report")
    manager = WRM(
        workflow_exec_id="wf",
        instrumentation=_Instrumentation(),
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )
    manager.apply_book_sheet(
        workflow_node_id="n1",
        decl_order=0,
        book_id="report",
        sheet="S",
        input_node_id="n1",
        input_output_id="detail",
        input_csv=InMemoryRows(header=["id", "amount"], rows=[[7, 1.5]]),
        on_conflict="error",
    )
    manager.commit_all()

    wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
    try:
        values = [list(r) for r in wb["S"].iter_rows(values_only=True)]
    finally:
        wb.close()
    assert values[0] == ["id", "amount"]
    assert values[1] == [7, 1.5]


def test_workflow_artifacts_directory_discard_in_memory_rows_output_missing_producer_is_noop() -> None:
    artifacts_dir = _make_artifacts_dir()

    artifacts_dir.discard_in_memory_rows_output("missing", "detail")
    assert artifacts_dir.get_optional("missing", "missing", "in_memory_rows_outputs") is None
