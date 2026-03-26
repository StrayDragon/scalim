import json
import threading
from pathlib import Path
from typing import Any, Dict, List, Optional, cast

import pytest

from scalim.dsl.by_yaml import run_workflow
from scalim.dsl.by_yaml import workflow_compile as workflow_compile_mod
from scalim.workflow import execute as workflow_execute_mod
from scalim.workflow import loaders as workflow_loaders_mod
from scalim.workflow.errors import WorkflowConfigError as WorkflowRuntimeConfigError
from scalim.events.catalog import (
    EVENT_DIAGNOSTIC_WARNING,
    EVENT_PIPELINE_START,
    EVENT_WORKFLOW_CACHE_ACQUIRE,
    EVENT_WORKFLOW_CACHE_EVICT,
    EVENT_WORKFLOW_CACHE_RELEASE,
    EVENT_WORKFLOW_NODE_CANCELLED,
    EVENT_WORKFLOW_NODE_END,
    EVENT_WORKFLOW_NODE_START,
    EVENT_WORKFLOW_RESOURCE_COMMIT,
    EVENT_WORKFLOW_RESOURCE_CREATE,
    EVENT_WORKFLOW_RESOURCE_DISCARD,
    EVENT_WORKFLOW_RESOURCE_WRITE,
)
from scalim.hooks.base import BaseHook
from scalim.ob.manager import ObserverManager
from scalim.ob.observer import Observer
from scalim.dsl.by_yaml.workflow import (
    WorkflowConfigError,
    load_workflow_config,
    load_workflow_config_from_mapping,
    resolve_workflow_demand_path,
    validate_workflow_yaml_text_json,
)
from tests.fixtures import workflow_loaders


_ALLOWED_MODULES = frozenset(["tests.fixtures.workflow_loaders"])
_ALLOWED_MODULES_WITH_SHEETBOOK = frozenset(["tests.fixtures.workflow_loaders", "scalim.workflow.loaders"])


class _WorkflowEventRecorder(Observer):
    event_types = {
        EVENT_DIAGNOSTIC_WARNING,
        EVENT_PIPELINE_START,
        EVENT_WORKFLOW_CACHE_ACQUIRE,
        EVENT_WORKFLOW_CACHE_EVICT,
        EVENT_WORKFLOW_CACHE_RELEASE,
        EVENT_WORKFLOW_NODE_START,
        EVENT_WORKFLOW_NODE_END,
        EVENT_WORKFLOW_NODE_CANCELLED,
        EVENT_WORKFLOW_RESOURCE_CREATE,
        EVENT_WORKFLOW_RESOURCE_WRITE,
        EVENT_WORKFLOW_RESOURCE_COMMIT,
        EVENT_WORKFLOW_RESOURCE_DISCARD,
    }

    def __init__(self) -> None:
        self._lock = threading.Lock()
        self.events: List[Any] = []

    def on_event(self, event) -> None:  # type: ignore[override]
        with self._lock:
            self.events.append(event)


def _write_text(path: Path, text: str) -> Path:
    path.write_text(text, encoding="utf-8")
    return path


def _cache_pool_config(
    *,
    conflict_policy: str = "error",
    release_policy: str = "dag_refcount",
    max_entries: int = 10,
    over_budget_policy: str = "fail_fast",
    pin: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    cfg: Dict[str, Any] = {
        "conflict_policy": str(conflict_policy),
        "release_policy": str(release_policy),
        "budget": {
            "max_entries": int(max_entries),
            "over_budget_policy": str(over_budget_policy),
        },
    }
    if pin:
        cfg["pin"] = pin
    return cfg


def _write_demand_yaml(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    main_loader_ref: str,
    preload_loader_ref: str,
    cache_mode: str = "preload_forever",
    preload_source_id: str = "preload",
    preload_params_yaml: Optional[str] = None,
) -> Path:
    preload_params_block = ""
    if preload_params_yaml:
        snippet = str(preload_params_yaml).strip("\n")
        indented = "\n".join(("    " + line if line.strip() else "") for line in snippet.splitlines())
        preload_params_block = "{}\n".format(indented)

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "{main_loader_ref}"
  fields:
    ref_id:
      name: ref_id

sources:
  {preload_source_id}:
    loader: "{preload_loader_ref}"
    key: id
    cache_mode: {cache_mode}
{preload_params_block}
    fields:
      value:
        name: value
        relation: main_to_preload

relations:
  main_to_preload:
    steps:
      - from: main.ref_id
        to: {preload_source_id}.id
"""
        )
        .format(
            name=name,
            main_loader_ref=main_loader_ref,
            preload_loader_ref=preload_loader_ref,
            cache_mode=cache_mode,
            preload_source_id=preload_source_id,
            preload_params_block=preload_params_block,
        )
        .lstrip(),
    )


def _write_table_demand_yaml_with_csv_output(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    loader_ref: str,
    output_name: str,
    output_path: Path,
    field_ids: List[str],
) -> Path:
    field_lines = []
    for fid in field_ids:
        field_lines.append("    {}: {{extract: {}}}".format(str(fid), str(fid)))

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "{loader_ref}"
  fields:
{fields}

outputs:
  - name: {output_name}
    container:
      type: csv
      path: "{output_path}"
    fields: {fields_list}
"""
        )
        .format(
            name=str(name),
            loader_ref=str(loader_ref),
            fields="\n".join(field_lines),
            output_name=str(output_name),
            output_path=str(output_path),
            fields_list=json.dumps([str(x) for x in field_ids]),
        )
        .lstrip(),
    )


def _write_table_demand_yaml_with_two_csv_outputs(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    loader_ref: str,
    output1_name: str,
    output1_path: Path,
    output1_field_ids: List[str],
    output2_name: str,
    output2_path: Path,
    output2_field_ids: List[str],
    main_field_ids: List[str],
) -> Path:
    field_lines = []
    for fid in main_field_ids:
        field_lines.append("    {}: {{extract: {}}}".format(str(fid), str(fid)))

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "{loader_ref}"
  fields:
{fields}

outputs:
  - name: {output1_name}
    container:
      type: csv
      path: "{output1_path}"
    fields: {output1_fields_list}
  - name: {output2_name}
    container:
      type: csv
      path: "{output2_path}"
    fields: {output2_fields_list}
"""
        )
        .format(
            name=str(name),
            loader_ref=str(loader_ref),
            fields="\n".join(field_lines),
            output1_name=str(output1_name),
            output1_path=str(output1_path),
            output1_fields_list=json.dumps([str(x) for x in output1_field_ids]),
            output2_name=str(output2_name),
            output2_path=str(output2_path),
            output2_fields_list=json.dumps([str(x) for x in output2_field_ids]),
        )
        .lstrip(),
    )


def _write_table_demand_yaml_with_workbook_output(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    loader_ref: str,
    output_name: str,
    output_path: Path,
    sheet: str,
    field_ids: List[str],
) -> Path:
    field_lines = []
    for fid in field_ids:
        field_lines.append("    {}: {{extract: {}}}".format(str(fid), str(fid)))

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "{loader_ref}"
  fields:
{fields}

outputs:
  - name: {output_name}
    container:
      type: workbook
      path: "{output_path}"
      sheet: "{sheet}"
    fields: {fields_list}
"""
        )
        .format(
            name=str(name),
            loader_ref=str(loader_ref),
            fields="\n".join(field_lines),
            output_name=str(output_name),
            output_path=str(output_path),
            sheet=str(sheet),
            fields_list=json.dumps([str(x) for x in field_ids]),
        )
        .lstrip(),
    )


def _write_table_demand_yaml_from_sheetbook_loader(
    tmp_path: Path,
    *,
    file_name: str,
    name: str,
    init_var_name: str,
    output_name: str,
    output_path: Path,
    field_ids: List[str],
) -> Path:
    field_lines = []
    for fid in field_ids:
        field_lines.append("    {}: {{extract: {}}}".format(str(fid), str(fid)))

    return _write_text(
        tmp_path / file_name,
        (
            """
name: {name}

main_source:
  source_id: main
  loader: "scalim.workflow.loaders:sheetbook_sheet_rows"
  params:
    ref:
      $init_var: {init_var_name}
  fields:
{fields}

outputs:
  - name: {output_name}
    container:
      type: csv
      path: "{output_path}"
    fields: {fields_list}
"""
        )
        .format(
            name=str(name),
            init_var_name=str(init_var_name),
            fields="\n".join(field_lines),
            output_name=str(output_name),
            output_path=str(output_path),
            fields_list=json.dumps([str(x) for x in field_ids]),
        )
        .lstrip(),
    )


def _write_workflow_yaml(
    tmp_path: Path,
    *,
    runs: list,
    max_concurrency: int = 1,
    failure_policy: str = "all_fail",
    resources: Optional[Dict[str, Any]] = None,
    cache_pool: Optional[Dict[str, Any]] = None,
    ctx: Optional[Dict[str, Any]] = None,
) -> Path:
    run_lines = []
    for item in runs:
        depends_on = item.get("depends_on") or []
        depends_on_lines = ""
        if depends_on:
            depends_on_lines = "\n      depends_on:\n{}".format("\n".join(["        - {}".format(d) for d in depends_on]))

        main_rows_from_lines = ""
        if "main_rows_from" in item:
            main_rows_from = cast("Any", item.get("main_rows_from"))
            if main_rows_from is None:
                main_rows_from_lines = "\n      main_rows_from: null"
            else:
                if not isinstance(main_rows_from, dict):
                    raise ValueError("main_rows_from must be a mapping")
                producer = main_rows_from.get("run")
                main_rows_from_lines = "\n      main_rows_from:\n        run: {}".format(json.dumps(producer))

        init_vars = cast("Optional[Dict[str, object]]", item.get("init_vars"))
        init_vars_lines = ""
        if init_vars:
            rendered = []
            for key, value in init_vars.items():
                rendered.append("        {}: {}".format(str(key), json.dumps(value)))
            init_vars_lines = "\n      init_vars:\n{}".format("\n".join(rendered))

        writes = cast("Any", item.get("writes"))
        writes_lines = ""
        if writes is not None:
            if not isinstance(writes, list):
                raise ValueError("writes must be a list")
            if not writes:
                writes_lines = "\n      writes: []"
            else:
                rendered: List[str] = []
                for write in writes:
                    if not isinstance(write, dict) or len(write) != 1:
                        raise ValueError("writes items must be a mapping with exactly one key")
                    kind = next(iter(write.keys()))
                    cfg = write.get(kind) or {}
                    if not isinstance(cfg, dict):
                        raise ValueError("writes[*].<kind> must be a mapping")
                    cfg_lines = []
                    for key, value in cfg.items():
                        cfg_lines.append("            {}: {}".format(str(key), json.dumps(value)))
                    rendered.append("        - {}:\n{}".format(str(kind), "\n".join(cfg_lines)))
                writes_lines = "\n      writes:\n{}".format("\n".join(rendered))

        write_to = cast("Any", item.get("write_to"))
        write_to_lines = ""
        if write_to:
            if writes is not None:
                raise ValueError("cannot set both write_to and writes")
            if not isinstance(write_to, dict) or len(write_to) != 1:
                raise ValueError("write_to must be a mapping with exactly one key")
            kind = next(iter(write_to.keys()))
            cfg = write_to.get(kind) or {}
            if not isinstance(cfg, dict):
                raise ValueError("write_to.<kind> must be a mapping")
            cfg_lines = []
            for key, value in cfg.items():
                cfg_lines.append("          {}: {}".format(str(key), json.dumps(value)))
            write_to_lines = "\n      write_to:\n        {}:\n{}".format(str(kind), "\n".join(cfg_lines))

        run_lines.append(
            "    - id: {}\n      demand: {}{}{}{}{}{}".format(
                item["id"],
                item["demand"],
                depends_on_lines,
                main_rows_from_lines,
                init_vars_lines,
                writes_lines,
                write_to_lines,
            )
        )

    resources_lines = ""
    if resources:
        group_lines: List[str] = []
        for group_key, group_cfg in resources.items():
            group_lines.append("    {}:".format(str(group_key)))
            for res_id, cfg in cast("Dict[str, Any]", group_cfg).items():
                group_lines.append("      {}:".format(str(res_id)))
                for key, value in cast("Dict[str, Any]", cfg).items():
                    group_lines.append("        {}: {}".format(str(key), json.dumps(value)))
        resources_lines = "\n  resources:\n{}".format("\n".join(group_lines))

    cache_pool_lines = ""
    if cache_pool is not None:
        budget = cast("Dict[str, Any]", cache_pool.get("budget") or {})
        max_entries = int(budget.get("max_entries", 1))
        over_budget_policy = str(budget.get("over_budget_policy", "fail_fast"))
        pins = cast("List[Dict[str, Any]]", cache_pool.get("pin") or [])
        pin_lines = ""
        if pins:
            pin_item_lines = []
            for pin in pins:
                pin_item_lines.append(
                    "        - kind: {kind}\n          source_id: {source_id}".format(
                        kind=str(pin.get("kind", "")),
                        source_id=str(pin.get("source_id", "")),
                    )
                )
            pin_lines = "\n      pin:\n{}".format("\n".join(pin_item_lines))
        cache_pool_lines = (
            "\n    cache_pool:\n"
            "      conflict_policy: {conflict_policy}\n"
            "      release_policy: {release_policy}\n"
            "      budget:\n"
            "        max_entries: {max_entries}\n"
            "        over_budget_policy: {over_budget_policy}{pin_lines}"
        ).format(
            conflict_policy=str(cache_pool.get("conflict_policy", "")),
            release_policy=str(cache_pool.get("release_policy", "")),
            max_entries=max_entries,
            over_budget_policy=over_budget_policy,
            pin_lines=pin_lines,
        )

    ctx_lines = ""
    if ctx is not None:
        ctx_lines = ("\n    ctx:\n      max_value_bytes: {max_value_bytes}\n      max_bytes: {max_bytes}").format(
            max_value_bytes=int(ctx.get("max_value_bytes", 65536)),
            max_bytes=int(ctx.get("max_bytes", 1048576)),
        )
    return _write_text(
        tmp_path / "workflow.yaml",
        (
            """
workflow:
  runs:
{runs}
{resources_lines}
  options:
    max_concurrency: {max_concurrency}
    failure_policy: {failure_policy}
{cache_pool_lines}
{ctx_lines}
"""
        )
        .format(
            runs="\n".join(run_lines),
            resources_lines=resources_lines.rstrip(),
            max_concurrency=max_concurrency,
            failure_policy=failure_policy,
            cache_pool_lines=cache_pool_lines.rstrip(),
            ctx_lines=ctx_lines.rstrip(),
        )
        .lstrip(),
    )


def test_load_workflow_config_semantic_validation(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: a
      demand: b.yaml
"""
        ).lstrip(),
    )

    with pytest.raises(WorkflowConfigError):
        _ = load_workflow_config(str(workflow_path))


def test_load_workflow_config_rejects_unknown_depends_on(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      depends_on: [nope]
"""
        ).lstrip(),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "Unknown run.depends_on id" in str(excinfo.value)


def test_load_workflow_config_rejects_depends_on_cycles(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
      depends_on: [b]
    - id: b
      demand: b.yaml
      depends_on: [a]
"""
        ).lstrip(),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "cycle_path" in str(excinfo.value)
    assert '["a", "b", "a"]' in str(excinfo.value)


def test_load_workflow_config_rejects_unknown_main_rows_from_run(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      main_rows_from:
        run: nope
"""
        ).lstrip(),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "Unknown run.main_rows_from.run id" in str(excinfo.value)
    assert "path=workflow.runs.1.main_rows_from.run" in str(excinfo.value)


def test_load_workflow_config_requires_depends_on_for_main_rows_from(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      main_rows_from:
        run: a
"""
        ).lstrip(),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "run.main_rows_from requires explicit depends_on" in str(excinfo.value)
    assert "path=workflow.runs.1.depends_on" in str(excinfo.value)


def test_load_workflow_config_rejects_main_rows_from_when_not_mapping(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      main_rows_from: a
"""
        ).lstrip(),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "run.main_rows_from must be a mapping" in str(excinfo.value)
    assert "path=workflow.runs.1.main_rows_from" in str(excinfo.value)


def test_load_workflow_config_rejects_main_rows_from_unknown_keys(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      main_rows_from:
        run: a
        extra: 1
"""
        ).lstrip(),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "run.main_rows_from contains unknown keys" in str(excinfo.value)
    assert "path=workflow.runs.1.main_rows_from" in str(excinfo.value)


def test_load_workflow_config_rejects_main_rows_from_run_when_not_string(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      main_rows_from:
        run: 123
"""
        ).lstrip(),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "run.main_rows_from.run must be a non-empty string" in str(excinfo.value)
    assert "path=workflow.runs.1.main_rows_from.run" in str(excinfo.value)


def test_load_workflow_config_rejects_main_rows_from_self_reference(tmp_path: Path) -> None:
    workflow_path = _write_text(
        tmp_path / "wf.yaml",
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
      main_rows_from:
        run: a
"""
        ).lstrip(),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "run.main_rows_from.run must not reference self" in str(excinfo.value)
    assert "path=workflow.runs.0.main_rows_from.run" in str(excinfo.value)


def test_load_workflow_config_parses_multiple_writes(tmp_path: Path) -> None:
    cfg = load_workflow_config_from_mapping(
        {
            "workflow": {
                "resources": {
                    "workbooks": {"report": {"path": str(tmp_path / "out.xlsx")}},
                    "csvs": {"merged": {"path": str(tmp_path / "merged.csv")}},
                },
                "runs": [
                    {
                        "id": "a",
                        "demand": "a.yaml",
                        "writes": [
                            {"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}},
                            {"csv_append": {"csv": "merged", "output": "detail"}},
                        ],
                    }
                ],
            }
        }
    )
    assert cfg.runs[0].id == "a"
    assert len(cfg.runs[0].writes) == 2


def test_load_workflow_config_rejects_write_to_with_migration_hint(tmp_path: Path) -> None:
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "resources": {"csvs": {"merged": {"path": str(tmp_path / "merged.csv")}}},
                    "runs": [{"id": "a", "demand": "a.yaml", "write_to": {"csv_append": {"csv": "merged", "output": "detail"}}}],
                }
            }
        )
    assert "run.write_to was removed" in str(excinfo.value)
    assert "Migration: write_to:" in str(excinfo.value)


def test_load_workflow_config_rejects_writes_items_with_multiple_intent_keys(tmp_path: Path) -> None:
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "resources": {
                        "workbooks": {"report": {"path": str(tmp_path / "out.xlsx")}},
                        "csvs": {"merged": {"path": str(tmp_path / "merged.csv")}},
                    },
                    "runs": [
                        {
                            "id": "a",
                            "demand": "a.yaml",
                            "writes": [
                                {
                                    "workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"},
                                    "csv_append": {"csv": "merged", "output": "detail"},
                                }
                            ],
                        }
                    ],
                }
            }
        )
    assert "write intent must contain exactly one" in str(excinfo.value)


def test_run_workflow_primary_only_collects_errors(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "ok", "demand": "ok.yaml"}, {"id": "bad", "demand": "bad.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert [o.run_id for o in result.outcomes] == ["ok", "bad"]
    assert result.outcomes[0].result is not None
    assert result.outcomes[0].error is None
    assert result.outcomes[1].result is None
    assert result.outcomes[1].error is not None
    assert result.outcomes[1].error.exc_type in {"ValueError", "WorkflowRunFailedError", "RuntimeError"}
    assert result.errors()


def test_run_workflow_all_fail_raises(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    with pytest.raises(Exception) as excinfo:
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)

    assert "run_id=bad" in str(excinfo.value)


def test_workflow_outcomes_are_in_declared_order(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="slow.yaml",
        name="slow",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="fast.yaml",
        name="fast",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "slow", "demand": "slow.yaml"}, {"id": "fast", "demand": "fast.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert [o.run_id for o in result.outcomes] == ["slow", "fast"]


def test_workflow_dag_respects_depends_on_under_concurrency(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml", "depends_on": ["a"]}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert [o.run_id for o in result.outcomes] == ["a", "b"]

    start = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_START]
    end = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]

    by_start = {e.payload.workflow_node_id: e for e in start}
    by_end = {e.payload.workflow_node_id: e for e in end}

    assert by_start["b"].seq > by_end["a"].seq


def test_workflow_ctx_init_vars_are_resolved_after_prereq_completion(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="up.yaml",
        name="up",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="down.yaml",
        name="down",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
        preload_params_yaml=(
            """
params:
  p:
    $init_var: token
"""
        ).strip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "up", "demand": "up.yaml"},
            {
                "id": "down",
                "demand": "down.yaml",
                "depends_on": ["up"],
                "init_vars": {
                    "token": {
                        "$ctx": {"node": "up", "key": "total_rows"},
                    }
                },
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert not result.errors()

    start = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_START]
    end = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]
    by_start = {e.payload.workflow_node_id: e for e in start}
    by_end = {e.payload.workflow_node_id: e for e in end}
    assert by_start["down"].seq > by_end["up"].seq


def test_workflow_ctx_ref_outside_depends_on_closure_fails_fast(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="c.yaml",
        name="c",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {
                "id": "c",
                "demand": "c.yaml",
                "init_vars": {"token": {"$ctx": {"node": "a", "key": "total_rows"}}},
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    with pytest.raises(WorkflowConfigError, match="declare depends_on"):
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)


def test_workflow_ctx_ref_node_self_fails_fast(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {
                "id": "a",
                "demand": "a.yaml",
                "init_vars": {"token": {"$ctx": {"node": "a", "key": "total_rows"}}},
            }
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    with pytest.raises(WorkflowConfigError, match="node=self"):
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)


def test_workflow_ctx_ref_unknown_node_fails_fast(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {
                "id": "b",
                "demand": "b.yaml",
                "init_vars": {"token": {"$ctx": {"node": "nope", "key": "total_rows"}}},
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    with pytest.raises(WorkflowConfigError, match="Unknown ctx node"):
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)


def test_workflow_ctx_guardrails_fail_fast(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        ctx={"max_value_bytes": 1, "max_bytes": 10},
    )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert result.errors()
    assert result.outcomes[0].error is not None
    assert result.outcomes[0].error.exc_type == "WorkflowConfigError"
    assert "max_value_bytes" in result.outcomes[0].error.message


def test_workflow_primary_only_cancels_downstream_when_dep_fails(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="down.yaml",
        name="down",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "bad", "demand": "bad.yaml"},
            {"id": "down", "demand": "down.yaml", "depends_on": ["bad"]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert [o.run_id for o in result.outcomes] == ["bad", "down"]
    assert result.outcomes[0].error is not None
    assert result.outcomes[1].result is None
    assert result.outcomes[1].error is not None
    assert "dependency" in result.outcomes[1].error.message.lower()

    cancelled = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_CANCELLED]
    assert len(cancelled) == 1
    assert cancelled[0].payload.workflow_node_id == "down"
    assert cancelled[0].payload.reason == "dependency_failed"


def test_workflow_observability_bridge_injects_meta_and_emits_workflow_events(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="slow.yaml",
        name="slow",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="fast.yaml",
        name="fast",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "slow", "demand": "slow.yaml"}, {"id": "fast", "demand": "fast.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])

    workflow_start = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_START]
    workflow_end = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]
    pipeline_start = [e for e in recorder.events if e.event_type == EVENT_PIPELINE_START]

    assert {e.payload.workflow_node_id for e in workflow_start} == {"slow", "fast"}
    assert {e.payload.workflow_node_id for e in workflow_end} == {"slow", "fast"}
    assert {e.meta.get("workflow_node_id") for e in pipeline_start} == {"slow", "fast"}

    workflow_exec_id = workflow_start[0].meta.get("workflow_exec_id")
    assert workflow_exec_id

    for event in workflow_start + workflow_end:
        assert event.run_id == workflow_exec_id
        assert event.meta.get("workflow_exec_id") == workflow_exec_id
        assert event.meta.get("workflow_node_id") in {"slow", "fast"}
        assert event.payload.workflow_exec_id == workflow_exec_id

    for event in pipeline_start:
        assert event.meta.get("workflow_exec_id") == workflow_exec_id
        assert event.meta.get("workflow_node_id") in {"slow", "fast"}
        assert event.run_id != workflow_exec_id
        assert str(event.run_id).startswith("run_")
        assert event.seq >= 1


def test_workflow_observability_bridge_emits_cancelled_reason_for_all_fail(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}, {"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    recorder = _WorkflowEventRecorder()
    with pytest.raises(Exception):
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])

    cancelled = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_CANCELLED]
    assert len(cancelled) == 1
    assert cancelled[0].payload.workflow_node_id == "ok"
    assert cancelled[0].payload.reason == "policy_all_fail"
    assert "failure_policy=all_fail" in cancelled[0].payload.message


def test_workflow_all_fail_skips_already_cancelled_nodes_on_late_terminal(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="slow.yaml",
        name="slow",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_slow",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="child.yaml",
        name="child",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "bad", "demand": "bad.yaml"},
            {"id": "slow", "demand": "slow.yaml"},
            {"id": "child", "demand": "child.yaml", "depends_on": ["bad", "slow"]},
        ],
        max_concurrency=2,
        failure_policy="all_fail",
    )

    recorder = _WorkflowEventRecorder()
    with pytest.raises(Exception):
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])

    cancelled = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_CANCELLED]
    assert len(cancelled) == 1
    assert cancelled[0].payload.workflow_node_id == "child"
    assert cancelled[0].payload.reason == "policy_all_fail"

    ended = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]
    assert any(e.payload.workflow_node_id == "slow" for e in ended)


def test_workflow_all_fail_compile_error_marks_end_and_cancels_pending(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:missing_loader",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}, {"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=2,
        failure_policy="all_fail",
    )

    recorder = _WorkflowEventRecorder()
    with pytest.raises(Exception) as excinfo:
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert "run_id=bad" in str(excinfo.value)

    ended = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_END]
    by_node = {e.payload.workflow_node_id: e for e in ended}
    assert by_node["bad"].payload.status == "error"

    cancelled = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_NODE_CANCELLED]
    assert len(cancelled) == 1
    assert cancelled[0].payload.workflow_node_id == "ok"
    assert cancelled[0].payload.reason == "policy_all_fail"


def test_observer_manager_workflow_attribution_keys_fail_fast_on_override() -> None:
    manager = ObserverManager(
        run_id="x",
        event_meta_defaults={
            "workflow_exec_id": "wf_1",
            "workflow_node_id": "a",
        },
    )
    with pytest.raises(ValueError, match="workflow attribution"):
        manager.emit_event("x", 1, meta={"workflow_node_id": "b"})


def test_observer_manager_workflow_attribution_meta_merges_non_reserved_meta() -> None:
    manager = ObserverManager(
        run_id="x",
        event_meta_defaults={
            "workflow_exec_id": "wf_1",
            "workflow_node_id": "a",
        },
    )
    event = manager.emit_event("x", 1, meta={"worker_id": "w1"})
    assert event.meta == {
        "workflow_exec_id": "wf_1",
        "workflow_node_id": "a",
        "worker_id": "w1",
    }


def test_workflow_observability_bridge_registers_hooks(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="fast.yaml",
        name="fast",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "fast", "demand": "fast.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    class _HookRecorder(BaseHook):
        def __init__(self) -> None:
            self._lock = threading.Lock()
            self.events: List[Any] = []

        def on_event(self, event) -> None:  # type: ignore[override]
            with self._lock:
                self.events.append(event)

    hook = _HookRecorder()
    _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[hook])
    assert any(e.event_type == EVENT_WORKFLOW_NODE_START for e in hook.events)


def test_cache_pool_reuses_preload_forever_across_runs(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(conflict_policy="error", release_policy="dag_refcount", max_entries=10),
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert not result.errors()
    assert workflow_loaders.preload_calls() == 1

    acquires = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_ACQUIRE]
    acquires.sort(key=lambda e: e.seq)
    assert [e.payload.cache_status for e in acquires] == ["miss", "hit"]
    assert {e.payload.workflow_node_id for e in acquires} == {"a", "b"}

    releases = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_RELEASE]
    assert {e.payload.workflow_node_id for e in releases} == {"a", "b"}

    evicts = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_EVICT]
    assert len(evicts) == 1
    assert evicts[0].payload.reason == "refcount_zero"
    assert evicts[0].payload.workflow_node_id == "b"


def test_cache_pool_conflict_policy_error_fails_fast(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table_alt",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(conflict_policy="error", release_policy="dag_refcount", max_entries=10),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)

    msg = str(excinfo.value)
    assert "cache_pool signature conflict" in msg and "diff=" in msg
    assert "loader_ref" in msg
    assert workflow_loaders.preload_calls() == 1


def test_cache_pool_conflict_policy_separate_runs_and_warns(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table_alt",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(conflict_policy="separate", release_policy="dag_refcount", max_entries=10),
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert not result.errors()
    assert workflow_loaders.preload_calls() == 2

    warnings = [e for e in recorder.events if e.event_type == EVENT_DIAGNOSTIC_WARNING]
    assert warnings

    acquires = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_ACQUIRE]
    assert any(e.payload.conflict_detected is True for e in acquires)


def test_cache_pool_signature_uses_rendered_params_for_reuse(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        preload_params_yaml=(
            """
params:
  p:
    $init_var: token
"""
        ).strip(),
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        preload_params_yaml=(
            """
params:
  p: 1
"""
        ).strip(),
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(conflict_policy="error", release_policy="dag_refcount", max_entries=10),
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder], init_vars={"token": 1})
    assert not result.errors()
    assert workflow_loaders.preload_calls() == 1

    acquires = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_ACQUIRE]
    assert acquires
    assert all(e.payload.conflict_detected is False for e in acquires)


def test_cache_pool_signature_includes_lookup_cast_meta(tmp_path: Path) -> None:
    demand_template = (
        """
name: demo

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_main_fast"
  fields:
    ref_id:
      name: ref_id

sources:
  preload:
    loader: "tests.fixtures.workflow_loaders:load_preload_table"
    key: id
    cache_mode: preload_forever
    lookup_cast:
      name: sep_first
      sep: "{sep}"
    fields:
      value:
        name: value
        relation: main_to_preload

relations:
  main_to_preload:
    steps:
      - from: main.ref_id
        to: preload.id
"""
    ).lstrip()

    _ = _write_text(tmp_path / "a.yaml", demand_template.format(sep=","))
    _ = _write_text(tmp_path / "b.yaml", demand_template.format(sep=";"))

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(conflict_policy="error", release_policy="dag_refcount", max_entries=10),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert "lookup_cast" in str(excinfo.value)


def test_cache_pool_budget_evict_lru_evicts_idle_entry(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table_alt",
        preload_source_id="preload_a",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table_alt",
        preload_source_id="preload_b",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(
            conflict_policy="error",
            release_policy="workflow_end",
            max_entries=1,
            over_budget_policy="evict_lru",
        ),
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert not result.errors()
    assert workflow_loaders.preload_calls() == 2

    evicts = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_EVICT]
    assert any(e.payload.reason == "budget_lru" for e in evicts)
    assert any(e.payload.reason == "workflow_end" for e in evicts)


def test_cache_pool_budget_fail_fast_raises(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table_alt",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(
            conflict_policy="separate",
            release_policy="workflow_end",
            max_entries=1,
            over_budget_policy="fail_fast",
        ),
    )

    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert "over budget" in str(excinfo.value).lower()
    assert workflow_loaders.preload_calls() == 1


def test_cache_pool_pin_keeps_entry_until_workflow_end(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="a.yaml",
        name="a",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="b.yaml",
        name="b",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(
            conflict_policy="error",
            release_policy="dag_refcount",
            max_entries=10,
            pin=[{"kind": "preload_forever", "source_id": "preload"}],
        ),
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert not result.errors()
    assert workflow_loaders.preload_calls() == 1

    evicts = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_EVICT]
    assert evicts
    assert any(e.payload.reason == "workflow_end" for e in evicts)
    assert not any(e.payload.reason == "refcount_zero" for e in evicts)


def test_cache_pool_calls_node_done_on_compile_error_and_cancelled_dependents(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:missing_loader",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}, {"id": "ok", "demand": "ok.yaml", "depends_on": ["bad"]}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(conflict_policy="warn", release_policy="dag_refcount", max_entries=10),
    )
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert [o.run_id for o in result.outcomes] == ["bad", "ok"]
    assert result.outcomes[0].error is not None
    assert result.outcomes[1].error is not None
    assert result.outcomes[1].error.exc_type == "WorkflowCancelled"


def test_cache_pool_calls_node_done_on_runtime_error(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(conflict_policy="warn", release_policy="dag_refcount", max_entries=10),
    )
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert [o.run_id for o in result.outcomes] == ["bad"]
    assert result.outcomes[0].error is not None
    assert result.outcomes[0].error.exc_type == "ValueError"


def test_workflow_schema_validation() -> None:
    jsonschema = pytest.importorskip("jsonschema")
    yaml = pytest.importorskip("yaml")

    from scalim.dsl.by_yaml.schema_dsl.builder import build_workflow_schema

    schema = build_workflow_schema()
    ok = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
"""
        ).lstrip()
    )
    jsonschema.validate(ok, schema)

    ok_cache_pool = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
  options:
    max_concurrency: 1
    failure_policy: all_fail
    cache_pool:
      conflict_policy: error
      release_policy: dag_refcount
      budget:
        max_entries: 1
        over_budget_policy: fail_fast
"""
        ).lstrip()
    )
    jsonschema.validate(ok_cache_pool, schema)

    ok_ctx_and_depends_on = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
    - id: b
      demand: b.yaml
      depends_on: [a]
      init_vars:
        token:
          $ctx:
            node: a
            key: total_rows
  options:
    ctx:
      max_value_bytes: 65536
      max_bytes: 1048576
"""
        ).lstrip()
    )
    jsonschema.validate(ok_ctx_and_depends_on, schema)

    bad_legacy_deps = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
      deps: [b]
"""
        ).lstrip()
    )
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(bad_legacy_deps, schema)

    bad = yaml.safe_load(
        (
            """
workflow:
  runs: []
"""
        ).lstrip()
    )
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(bad, schema)

    bad_share_preload_cache = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
  options:
    share_preload_cache: true
"""
        ).lstrip()
    )
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(bad_share_preload_cache, schema)

    bad_cache_pool = yaml.safe_load(
        (
            """
workflow:
  runs:
    - id: a
      demand: a.yaml
  options:
    cache_pool:
      conflict_policy: nope
      release_policy: dag_refcount
      budget:
        max_entries: 1
        over_budget_policy: fail_fast
"""
        ).lstrip()
    )
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(bad_cache_pool, schema)


def test_workflow_config_error_formats_without_path() -> None:
    assert str(WorkflowConfigError("msg")) == "msg"


def test_load_workflow_config_wraps_read_errors(tmp_path: Path) -> None:
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(tmp_path))
    assert "Failed to read workflow YAML" in str(excinfo.value)


def test_load_workflow_config_wraps_yaml_parse_errors(tmp_path: Path) -> None:
    workflow_path = _write_text(tmp_path / "wf.yaml", "workflow: [\n")
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "YAML parse error" in str(excinfo.value)


def test_load_workflow_config_root_must_be_mapping(tmp_path: Path) -> None:
    workflow_path = _write_text(tmp_path / "wf.yaml", "- 1\n")
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config(str(workflow_path))
    assert "root must be a mapping" in str(excinfo.value)


def test_resolve_workflow_demand_path_requires_non_empty_string(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path("", workflow_yaml_path=str(wf))
    assert "run.demand must be a non-empty string" in str(excinfo.value)


def test_resolve_workflow_demand_path_supports_at_alias(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    demand = resolve_workflow_demand_path(
        "@/a.yaml",
        workflow_yaml_path=str(wf),
        path_aliases={"@": str(tmp_path)},
        run_id="r1",
    )
    assert demand == (tmp_path / "a.yaml").resolve(strict=False)


def test_resolve_workflow_demand_path_supports_named_alias(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    demand = resolve_workflow_demand_path(
        "DATA:/b.yaml",
        workflow_yaml_path=str(wf),
        path_aliases={"DATA": str(tmp_path)},
        run_id="r1",
    )
    assert demand == (tmp_path / "b.yaml").resolve(strict=False)


def test_resolve_workflow_demand_path_rejects_unknown_alias(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "DATA:/b.yaml",
            workflow_yaml_path=str(wf),
            path_aliases={},
            run_id="r1",
        )
    assert "Unknown path alias" in str(excinfo.value)
    assert "run_id=r1" in str(excinfo.value)


def test_resolve_workflow_demand_path_rejects_empty_at_alias_path(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "@/",
            workflow_yaml_path=str(wf),
            path_aliases={"@": str(tmp_path)},
            run_id="r1",
        )
    assert "Invalid demand alias path" in str(excinfo.value)
    assert "run_id=r1" in str(excinfo.value)


def test_resolve_workflow_demand_path_supports_absolute_paths(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    abs_path = (tmp_path / "abs.yaml").resolve(strict=False)
    demand = resolve_workflow_demand_path(str(abs_path), workflow_yaml_path=str(wf))
    assert demand == abs_path


def test_resolve_workflow_demand_path_supports_paths_relative_to_workflow(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    demand = resolve_workflow_demand_path("rel.yaml", workflow_yaml_path=str(wf))
    assert demand == (tmp_path / "rel.yaml").resolve(strict=False)


def test_resolve_workflow_demand_path_rejects_relative_escape_by_default(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path("../escape.yaml", workflow_yaml_path=str(wf))
    assert "YAML path escapes allowed roots" in str(excinfo.value)


def test_resolve_workflow_demand_path_escape_error_includes_run_id(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path("../escape.yaml", workflow_yaml_path=str(wf), run_id="r1")
    assert "YAML path escapes allowed roots" in str(excinfo.value)
    assert "run_id=r1" in str(excinfo.value)


def test_resolve_workflow_demand_path_rejects_absolute_escape_by_default(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    outside = (tmp_path.parent / "escape.yaml").resolve(strict=False)
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(str(outside), workflow_yaml_path=str(wf))
    assert "YAML path escapes allowed roots" in str(excinfo.value)


def test_resolve_workflow_demand_path_rejects_alias_escape_by_default(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "DATA:/escape.yaml",
            workflow_yaml_path=str(wf),
            path_aliases={"DATA": str(tmp_path.parent)},
            run_id="r1",
        )
    assert "YAML path escapes allowed roots" in str(excinfo.value)
    assert "alias=DATA" in str(excinfo.value)


def test_resolve_workflow_demand_path_allows_escape_with_explicit_allowed_roots(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    allowed_yaml_roots = [tmp_path.parent]
    demand = resolve_workflow_demand_path(
        "../escape.yaml",
        workflow_yaml_path=str(wf),
        allowed_yaml_roots=allowed_yaml_roots,
    )
    assert demand == (tmp_path.parent / "escape.yaml").resolve(strict=False)

    outside = (tmp_path.parent / "abs.yaml").resolve(strict=False)
    demand = resolve_workflow_demand_path(
        str(outside),
        workflow_yaml_path=str(wf),
        allowed_yaml_roots=allowed_yaml_roots,
    )
    assert demand == outside

    demand = resolve_workflow_demand_path(
        "DATA:/alias.yaml",
        workflow_yaml_path=str(wf),
        path_aliases={"DATA": str(tmp_path.parent)},
        run_id="r1",
        allowed_yaml_roots=allowed_yaml_roots,
    )
    assert demand == (tmp_path.parent / "alias.yaml").resolve(strict=False)


def test_resolve_workflow_demand_path_invalid_allowed_yaml_roots_is_wrapped(tmp_path: Path) -> None:
    wf = tmp_path / "workflow.yaml"
    missing_root = tmp_path / "missing_root"
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = resolve_workflow_demand_path(
            "rel.yaml",
            workflow_yaml_path=str(wf),
            allowed_yaml_roots=[missing_root],
            run_id="r1",
        )
    assert "Invalid allowed_yaml_roots" in str(excinfo.value)
    assert "run_id=r1" in str(excinfo.value)


def test_validate_workflow_yaml_text_json_variants() -> None:
    bad_parse = json.loads(validate_workflow_yaml_text_json("workflow: [\n"))
    assert bad_parse["ok"] is False

    empty = json.loads(validate_workflow_yaml_text_json(""))
    assert empty["ok"] is False
    assert "empty" in empty["errors"][0]["message"].lower()

    root_not_mapping = json.loads(validate_workflow_yaml_text_json("- 1\n"))
    assert root_not_mapping["ok"] is False

    semantic_error = json.loads(validate_workflow_yaml_text_json("workflow: {}\n"))
    assert semantic_error["ok"] is False
    assert semantic_error["errors"]

    ok = json.loads(
        validate_workflow_yaml_text_json(
            (
                """
workflow:
  runs:
    - id: a
      demand: a.yaml
"""
            ).lstrip()
        )
    )
    assert ok["ok"] is True


@pytest.mark.parametrize(
    "bad_mapping",
    [
        {},
        {"workflow": {}},
        {"workflow": {"runs": []}},
        {"workflow": {"runs": [1]}},
        {"workflow": {"runs": [{"id": "", "demand": "a.yaml"}]}},
        {"workflow": {"runs": [{"id": "a", "demand": ""}]}},
        {"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": []}},
        {"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": {"max_concurrency": True}}},
        {"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": {"max_concurrency": "x"}}},
        {"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": {"max_concurrency": 0}}},
        {"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": {"failure_policy": "nope"}}},
    ],
)
def test_load_workflow_config_from_mapping_rejects_invalid_structures(bad_mapping: dict) -> None:
    with pytest.raises(WorkflowConfigError):
        _ = load_workflow_config_from_mapping(bad_mapping)


@pytest.mark.parametrize(
    ("options", "path"),
    [
        ({"share_preload_cache": True}, "workflow.options.share_preload_cache"),
        ({"cache_pool": []}, "workflow.options.cache_pool"),
        (
            {
                "cache_pool": {
                    "conflict_policy": "nope",
                    "release_policy": "dag_refcount",
                    "budget": {"max_entries": 1, "over_budget_policy": "fail_fast"},
                }
            },
            "workflow.options.cache_pool.conflict_policy",
        ),
        (
            {
                "cache_pool": {
                    "conflict_policy": "warn",
                    "release_policy": "nope",
                    "budget": {"max_entries": 1, "over_budget_policy": "fail_fast"},
                }
            },
            "workflow.options.cache_pool.release_policy",
        ),
        (
            {
                "cache_pool": {
                    "conflict_policy": "warn",
                    "release_policy": "workflow_end",
                    "budget": [],
                }
            },
            "workflow.options.cache_pool.budget",
        ),
        (
            {
                "cache_pool": {
                    "conflict_policy": "warn",
                    "release_policy": "workflow_end",
                    "budget": {"max_entries": True, "over_budget_policy": "fail_fast"},
                }
            },
            "workflow.options.cache_pool.budget.max_entries",
        ),
        (
            {
                "cache_pool": {
                    "conflict_policy": "warn",
                    "release_policy": "workflow_end",
                    "budget": {"max_entries": "x", "over_budget_policy": "fail_fast"},
                }
            },
            "workflow.options.cache_pool.budget.max_entries",
        ),
        (
            {
                "cache_pool": {
                    "conflict_policy": "warn",
                    "release_policy": "workflow_end",
                    "budget": {"max_entries": 0, "over_budget_policy": "fail_fast"},
                }
            },
            "workflow.options.cache_pool.budget.max_entries",
        ),
        (
            {
                "cache_pool": {
                    "conflict_policy": "warn",
                    "release_policy": "workflow_end",
                    "budget": {"max_entries": 1, "over_budget_policy": "nope"},
                }
            },
            "workflow.options.cache_pool.budget.over_budget_policy",
        ),
        (
            {
                "cache_pool": {
                    "conflict_policy": "warn",
                    "release_policy": "workflow_end",
                    "budget": {"max_entries": 1, "over_budget_policy": "fail_fast"},
                    "pin": {},
                }
            },
            "workflow.options.cache_pool.pin",
        ),
        (
            {
                "cache_pool": {
                    "conflict_policy": "warn",
                    "release_policy": "workflow_end",
                    "budget": {"max_entries": 1, "over_budget_policy": "fail_fast"},
                    "pin": ["x"],
                }
            },
            "workflow.options.cache_pool.pin.0",
        ),
        (
            {
                "cache_pool": {
                    "conflict_policy": "warn",
                    "release_policy": "workflow_end",
                    "budget": {"max_entries": 1, "over_budget_policy": "fail_fast"},
                    "pin": [{"kind": "nope", "source_id": "s1"}],
                }
            },
            "workflow.options.cache_pool.pin.0.kind",
        ),
        (
            {
                "cache_pool": {
                    "conflict_policy": "warn",
                    "release_policy": "workflow_end",
                    "budget": {"max_entries": 1, "over_budget_policy": "fail_fast"},
                    "pin": [{"kind": "preload_forever", "source_id": ""}],
                }
            },
            "workflow.options.cache_pool.pin.0.source_id",
        ),
    ],
)
def test_load_workflow_config_from_mapping_rejects_invalid_cache_pool_options(options: Dict[str, Any], path: str) -> None:
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": options}})
    assert excinfo.value.path == path


def test_load_workflow_config_from_mapping_accepts_null_options() -> None:
    cfg = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "options": None}})
    assert cfg.options.max_concurrency == 1


def test_load_workflow_config_from_mapping_accepts_ctx_options() -> None:
    cfg = load_workflow_config_from_mapping(
        {
            "workflow": {
                "runs": [{"id": "a", "demand": "a.yaml"}],
                "options": {
                    "ctx": {
                        "max_value_bytes": 2,
                        "max_bytes": 3,
                    }
                },
            }
        }
    )
    assert cfg.options.ctx.max_value_bytes == 2
    assert cfg.options.ctx.max_bytes == 3


@pytest.mark.parametrize(
    ("ctx", "path"),
    [
        ([], "workflow.options.ctx"),
        ({"max_value_bytes": 0, "max_bytes": 1}, "workflow.options.ctx.max_value_bytes"),
        ({"max_value_bytes": "nope", "max_bytes": 1}, "workflow.options.ctx.max_value_bytes"),
        ({"max_value_bytes": 1, "max_bytes": 0}, "workflow.options.ctx.max_bytes"),
        ({"max_value_bytes": 1, "max_bytes": "nope"}, "workflow.options.ctx.max_bytes"),
        ({"max_value_bytes": True, "max_bytes": 1}, "workflow.options.ctx.max_value_bytes"),
        ({"max_value_bytes": 1, "max_bytes": True}, "workflow.options.ctx.max_bytes"),
    ],
)
def test_load_workflow_config_from_mapping_rejects_invalid_ctx_options(ctx: object, path: str) -> None:
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [{"id": "a", "demand": "a.yaml"}],
                    "options": {
                        "ctx": ctx,
                    },
                }
            }
        )
    assert excinfo.value.path == path


def test_load_workflow_config_from_mapping_rejects_legacy_deps_field() -> None:
    with pytest.raises(WorkflowConfigError, match="run\\.deps was removed; use run\\.depends_on") as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [
                        {"id": "a", "demand": "a.yaml", "deps": ["b"]},
                    ]
                }
            }
        )
    assert excinfo.value.path == "workflow.runs.0.deps"


def test_load_workflow_config_from_mapping_rejects_depends_on_not_list() -> None:
    with pytest.raises(WorkflowConfigError, match="run\\.depends_on must be a list"):
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [
                        {"id": "a", "demand": "a.yaml", "depends_on": "nope"},
                    ]
                }
            }
        )


def test_load_workflow_config_from_mapping_rejects_depends_on_empty_item() -> None:
    with pytest.raises(WorkflowConfigError, match="depends_on items must be non-empty"):
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [
                        {"id": "a", "demand": "a.yaml"},
                        {"id": "b", "demand": "b.yaml", "depends_on": [""]},
                    ]
                }
            }
        )


def test_load_workflow_config_from_mapping_dedups_depends_on_preserves_first_order() -> None:
    cfg = load_workflow_config_from_mapping(
        {
            "workflow": {
                "runs": [
                    {"id": "a", "demand": "a.yaml"},
                    {"id": "b", "demand": "b.yaml", "depends_on": ["a", "a"]},
                ]
            }
        }
    )
    assert cfg.runs[1].depends_on == ("a",)


def test_load_workflow_config_from_mapping_accepts_forward_depends_on() -> None:
    cfg = load_workflow_config_from_mapping(
        {
            "workflow": {
                "runs": [
                    {"id": "b", "demand": "b.yaml", "depends_on": ["a"]},
                    {"id": "a", "demand": "a.yaml"},
                ]
            }
        }
    )
    assert [r.id for r in cfg.runs] == ["b", "a"]


def test_load_workflow_config_from_mapping_rejects_init_vars_not_mapping() -> None:
    with pytest.raises(WorkflowConfigError, match="run\\.init_vars must be a mapping") as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [
                        {"id": "a", "demand": "a.yaml", "init_vars": []},
                    ]
                }
            }
        )
    assert excinfo.value.path == "workflow.runs.0.init_vars"


def test_load_workflow_config_from_mapping_rejects_init_vars_key_invalid() -> None:
    with pytest.raises(WorkflowConfigError, match="run\\.init_vars keys must be non-empty strings") as excinfo:
        _ = load_workflow_config_from_mapping(
            {
                "workflow": {
                    "runs": [
                        {"id": "a", "demand": "a.yaml", "init_vars": {"": 1}},
                    ]
                }
            }
        )
    assert excinfo.value.path == "workflow.runs.0.init_vars"


def test_load_workflow_config_from_mapping_accepts_null_resources() -> None:
    cfg = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "resources": None}})
    assert cfg.resources.workbooks == {}
    assert cfg.resources.csvs == {}


def test_load_workflow_config_from_mapping_accepts_resources_mapping() -> None:
    cfg = load_workflow_config_from_mapping(
        {
            "workflow": {
                "runs": [{"id": "a", "demand": "a.yaml"}],
                "resources": {
                    "workbooks": {"report": {"path": "./out/report.xlsx"}},
                    "csvs": {"merged": {"path": "./out/merged.csv"}},
                },
            }
        }
    )
    assert cfg.resources.workbooks["report"].path == "./out/report.xlsx"
    assert cfg.resources.csvs["merged"].path == "./out/merged.csv"


def test_load_workflow_config_from_mapping_rejects_resources_not_mapping() -> None:
    with pytest.raises(WorkflowConfigError, match="workflow.resources must be a mapping"):
        _ = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "resources": []}})


def test_load_workflow_config_from_mapping_rejects_resources_key_invalid() -> None:
    with pytest.raises(WorkflowConfigError, match="workflow\\.resources keys must be non-empty"):
        _ = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml"}], "resources": {"": {"kind": "x"}}}})


def test_load_workflow_config_from_mapping_rejects_self_depends_on() -> None:
    with pytest.raises(WorkflowConfigError, match="self dependency"):
        _ = load_workflow_config_from_mapping({"workflow": {"runs": [{"id": "a", "demand": "a.yaml", "depends_on": ["a"]}]}})


def test_run_workflow_requires_workflow_path(tmp_path: Path) -> None:
    _ = tmp_path
    with pytest.raises(WorkflowConfigError):
        _ = run_workflow("", allowed_modules=_ALLOWED_MODULES)


def test_run_workflow_primary_only_submits_pending_after_failure(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok1.yaml",
        name="ok1",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok2.yaml",
        name="ok2",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}, {"id": "ok1", "demand": "ok1.yaml"}, {"id": "ok2", "demand": "ok2.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert [o.run_id for o in result.outcomes] == ["bad", "ok1", "ok2"]
    assert result.outcomes[0].error is not None
    assert result.outcomes[1].result is not None
    assert result.outcomes[2].result is not None


def test_run_workflow_all_fail_cancels_pending_queue(tmp_path: Path) -> None:
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    _ = _write_demand_yaml(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_fast",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
        cache_mode="none",
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "bad", "demand": "bad.yaml"}, {"id": "ok", "demand": "ok.yaml"}],
        max_concurrency=1,
        failure_policy="all_fail",
    )

    with pytest.raises(Exception) as excinfo:
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert "run_id=bad" in str(excinfo.value)


def test_workflow_entrypoints_artifacts_directory_enforces_visibility() -> None:
    from scalim.spec.ir.workflow import WorkflowArtifactsIr, WorkflowEdgeIr, WorkflowIr, WorkflowNodeIr, WorkflowNodeType, WorkflowOptionsIr

    ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(node_id="a", node_type=WorkflowNodeType.DEMAND, decl_order=0, deps=(), demand_path="a.yaml"),
            WorkflowNodeIr(node_id="b", node_type=WorkflowNodeType.DEMAND, decl_order=1, deps=("a",), demand_path="b.yaml"),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(),
        resources=(),
        artifacts=WorkflowArtifactsIr(
            slots_by_node_id={
                "a": ("output_path", "outputs"),
                "b": ("output_path", "outputs"),
            }
        ),
    )
    artifacts_dir = workflow_execute_mod.WorkflowArtifactsDirectory(ir)
    artifacts_dir.publish("a", "x", 1)
    assert artifacts_dir.get("b", "a", "x") == 1

    artifacts_dir.publish("b", "y", 2)
    assert artifacts_dir.get("b", "b", "y") == 2

    with pytest.raises(ValueError, match="not visible"):
        _ = artifacts_dir.get("a", "b", "y")

    with pytest.raises(KeyError, match="Unknown artifact"):
        _ = artifacts_dir.get("b", "a", "missing")


def test_workflow_entrypoints_ensure_json_like_variants() -> None:
    assert workflow_execute_mod.ensure_json_like(None, path="x") is None
    assert workflow_execute_mod.ensure_json_like(True, path="x") is True
    assert workflow_execute_mod.ensure_json_like(1, path="x") == 1
    assert workflow_execute_mod.ensure_json_like("s", path="x") == "s"
    assert workflow_execute_mod.ensure_json_like(1.5, path="x") == 1.5

    assert workflow_execute_mod.ensure_json_like([1, {"k": 2}], path="x") == [1, {"k": 2}]
    assert workflow_execute_mod.ensure_json_like((1, 2), path="x") == [1, 2]

    with pytest.raises(WorkflowRuntimeConfigError, match="finite"):
        _ = workflow_execute_mod.ensure_json_like(float("inf"), path="x")

    with pytest.raises(WorkflowRuntimeConfigError, match="dict key"):
        _ = workflow_execute_mod.ensure_json_like({1: "x"}, path="x")

    with pytest.raises(WorkflowRuntimeConfigError, match="JSON-like"):
        _ = workflow_execute_mod.ensure_json_like(object(), path="x")


def test_workflow_entrypoints_ctx_store_total_bytes_guardrail() -> None:
    from scalim.spec.ir.workflow import (
        WorkflowArtifactsIr,
        WorkflowCtxOptionsIr,
        WorkflowEdgeIr,
        WorkflowIr,
        WorkflowNodeIr,
        WorkflowNodeType,
        WorkflowOptionsIr,
    )

    ir = WorkflowIr(
        nodes=(WorkflowNodeIr(node_id="a", node_type=WorkflowNodeType.DEMAND, decl_order=0, deps=(), demand_path="a.yaml"),),
        edges=(),
        options=WorkflowOptionsIr(ctx=WorkflowCtxOptionsIr(max_value_bytes=1000, max_bytes=5)),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={"a": ()}),
    )
    ctx_store = workflow_execute_mod.WorkflowCtxStore(ir)
    ctx_store.publish("a", "k1", "x", path="x")
    with pytest.raises(WorkflowRuntimeConfigError) as excinfo:
        ctx_store.publish("a", "k2", "y", path="x")
    assert excinfo.value.path == "workflow.options.ctx.max_bytes"


def test_workflow_entrypoints_ctx_store_resolve_errors() -> None:
    from scalim.spec.ir.workflow import WorkflowArtifactsIr, WorkflowEdgeIr, WorkflowIr, WorkflowNodeIr, WorkflowNodeType, WorkflowOptionsIr

    ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(node_id="a", node_type=WorkflowNodeType.DEMAND, decl_order=0, deps=(), demand_path="a.yaml"),
            WorkflowNodeIr(node_id="b", node_type=WorkflowNodeType.DEMAND, decl_order=1, deps=("a",), demand_path="b.yaml"),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={"a": (), "b": ()}),
    )
    ctx_store = workflow_execute_mod.WorkflowCtxStore(ir)

    with pytest.raises(WorkflowRuntimeConfigError, match="node=self"):
        _ = ctx_store.resolve("a", node="a", key="k", path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="not visible"):
        _ = ctx_store.resolve("a", node="b", key="k", path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="Unknown ctx key"):
        _ = ctx_store.resolve("b", node="a", key="missing", path="p")


def test_workflow_entrypoints_iter_ctx_directives_variants() -> None:
    with pytest.raises(WorkflowRuntimeConfigError, match="directive must be a mapping"):
        _ = workflow_execute_mod.iter_ctx_directives({"$ctx": "nope"}, path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="\\$ctx\\.node"):
        _ = workflow_execute_mod.iter_ctx_directives({"$ctx": {"node": "", "key": "k"}}, path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="\\$ctx\\.key"):
        _ = workflow_execute_mod.iter_ctx_directives({"$ctx": {"node": "a", "key": ""}}, path="p")

    assert workflow_execute_mod.iter_ctx_directives({"outer": {"$ctx": {"node": "a", "key": "k"}}}, path="p") == [
        ("a", "k"),
    ]
    assert workflow_execute_mod.iter_ctx_directives([{"$ctx": {"node": "a", "key": "k"}}], path="p") == [
        ("a", "k"),
    ]


def test_workflow_entrypoints_render_ctx_directives_variants() -> None:
    from scalim.spec.ir.workflow import WorkflowArtifactsIr, WorkflowEdgeIr, WorkflowIr, WorkflowNodeIr, WorkflowNodeType, WorkflowOptionsIr

    ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(node_id="a", node_type=WorkflowNodeType.DEMAND, decl_order=0, deps=(), demand_path="a.yaml"),
            WorkflowNodeIr(node_id="b", node_type=WorkflowNodeType.DEMAND, decl_order=1, deps=("a",), demand_path="b.yaml"),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={"a": (), "b": ()}),
    )
    ctx_store = workflow_execute_mod.WorkflowCtxStore(ir)
    ctx_store.publish("a", "k", 1, path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="directive must be a mapping"):
        _ = workflow_execute_mod.render_ctx_directives({"$ctx": "nope"}, consumer_node_id="b", ctx_store=ctx_store, path="p")

    with pytest.raises(WorkflowRuntimeConfigError, match="\\$ctx\\.node"):
        _ = workflow_execute_mod.render_ctx_directives(
            {"$ctx": {"node": "", "key": "k"}}, consumer_node_id="b", ctx_store=ctx_store, path="p"
        )

    with pytest.raises(WorkflowRuntimeConfigError, match="\\$ctx\\.key"):
        _ = workflow_execute_mod.render_ctx_directives(
            {"$ctx": {"node": "a", "key": ""}}, consumer_node_id="b", ctx_store=ctx_store, path="p"
        )

    assert workflow_execute_mod.render_ctx_directives(
        {"outer": {"$ctx": {"node": "a", "key": "k"}}}, consumer_node_id="b", ctx_store=ctx_store, path="p"
    ) == {"outer": 1}
    assert workflow_execute_mod.render_ctx_directives(
        [{"$ctx": {"node": "a", "key": "k"}}], consumer_node_id="b", ctx_store=ctx_store, path="p"
    ) == [1]
    assert workflow_execute_mod.render_ctx_directives(1, consumer_node_id="b", ctx_store=ctx_store, path="p") == 1


def test_cache_pool_single_run_accepts_lookup_cast_and_normalize(tmp_path: Path) -> None:
    demand_path = _write_text(
        tmp_path / "demand.yaml",
        (
            """
name: demo

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_main_fast"
  fields:
    ref_id:
      name: ref_id

sources:
  other:
    loader: "tests.fixtures.workflow_loaders:load_preload_table"
    key: id
    cache_mode: none
  preload:
    loader: "tests.fixtures.workflow_loaders:load_preload_table"
    key: id
    cache_mode: preload_forever
    lookup_cast:
      name: sep_first
      sep: ","
    normalize:
      kind: index_by_key
      key_field: id
      on_conflict: last
    fields:
      value:
        name: value
        relation: main_to_preload

relations:
  main_to_preload:
    steps:
      - from: main.ref_id
        to: preload.id
"""
        ).lstrip(),
    )

    workflow_loaders.reset_counters()
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": str(demand_path)}],
        max_concurrency=1,
        failure_policy="primary_only",
        cache_pool=_cache_pool_config(conflict_policy="error", release_policy="dag_refcount", max_entries=10),
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert not result.errors()
    assert workflow_loaders.preload_calls() == 1

    acquires = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_ACQUIRE]
    assert len(acquires) == 1
    assert acquires[0].payload.cache_status == "miss"

    evicts = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_CACHE_EVICT]
    assert len(evicts) == 1
    assert evicts[0].payload.reason == "refcount_zero"


def _read_xlsx_rows(path: Path, sheet: str) -> List[List[object]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb[str(sheet)]
        rows: List[List[object]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows
    finally:
        wb.close()


def _read_xlsx_sheetnames(path: Path) -> List[str]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _read_csv_rows(path: Path) -> List[List[str]]:
    import csv

    with path.open("r", encoding="utf-8", newline="") as handle:
        return [list(r) for r in csv.reader(handle)]


def test_workflow_shared_workbook_sheet_writes_commit_and_emit_events(tmp_path: Path) -> None:
    a_out = tmp_path / "a_detail.csv"
    b_out = tmp_path / "b_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_path=b_out,
        field_ids=["id", "value"],
    )

    workbook_path = tmp_path / "report.xlsx"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path)}}},
        runs=[
            {
                "id": "a",
                "demand": "a.yaml",
                "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "A", "output": "detail"}}],
            },
            {
                "id": "b",
                "demand": "b.yaml",
                "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "B", "output": "detail"}}],
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert not result.errors()
    assert workbook_path.exists()
    assert not Path(str(workbook_path) + ".scalim.lock").exists()

    assert _read_xlsx_rows(workbook_path, "A") == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]
    assert _read_xlsx_rows(workbook_path, "B") == [
        ["id", "value"],
        ["b1", "B1"],
        ["b2", "B2"],
    ]

    creates = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_CREATE]
    writes = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_WRITE]
    commits = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_COMMIT]
    assert len(creates) == 1
    assert creates[0].payload.resource_id == "report"
    assert creates[0].payload.workflow_node_id == "__wf__write.a.0"
    assert len([e for e in writes if e.payload.resource_id == "report"]) == 2
    assert len([e for e in commits if e.payload.resource_id == "report"]) == 1
    assert commits[0].payload.workflow_node_id == "__wf__write.b.0"


def test_workflow_sheetbook_resources_export_xlsx_and_emit_events(tmp_path: Path) -> None:
    a_out = tmp_path / "a_detail.csv"
    b_out = tmp_path / "b_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_path=b_out,
        field_ids=["id", "value"],
    )

    export_path = tmp_path / "report.xlsx"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "sheetbooks": {
                "report": {
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(export_path), "write_lock": True},
                }
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml", "writes": [{"sheetbook_sheet": {"sheetbook": "report", "sheet": "A", "output": "detail"}}]},
            {"id": "b", "demand": "b.yaml", "writes": [{"sheetbook_sheet": {"sheetbook": "report", "sheet": "B", "output": "detail"}}]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert not result.errors()
    assert export_path.exists()
    assert not Path(str(export_path) + ".scalim.lock").exists()

    assert _read_xlsx_sheetnames(export_path) == ["A", "B"]
    assert _read_xlsx_rows(export_path, "A") == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]
    assert _read_xlsx_rows(export_path, "B") == [
        ["id", "value"],
        ["b1", "B1"],
        ["b2", "B2"],
    ]

    creates = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_CREATE]
    writes = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_WRITE]
    commits = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_COMMIT]
    assert len(creates) == 1
    assert creates[0].payload.resource_type == "sheetbook"
    assert creates[0].payload.resource_id == "report"
    assert creates[0].payload.workflow_node_id == "__wf__write.a.0"
    assert len([e for e in writes if e.payload.resource_type == "sheetbook" and e.payload.resource_id == "report"]) == 2
    assert len([e for e in commits if e.payload.resource_type == "sheetbook" and e.payload.resource_id == "report"]) == 1
    assert commits[0].payload.workflow_node_id == "__wf__write.b.0"


def test_workflow_run_writes_two_outputs_to_same_sheetbook_sheets(tmp_path: Path) -> None:
    metrics_out = tmp_path / "metrics.csv"
    detail_out = tmp_path / "detail.csv"
    _ = _write_table_demand_yaml_with_two_csv_outputs(
        tmp_path,
        file_name="report.yaml",
        name="report",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output1_name="metrics",
        output1_path=metrics_out,
        output1_field_ids=["id"],
        output2_name="detail",
        output2_path=detail_out,
        output2_field_ids=["id", "value"],
        main_field_ids=["id", "value"],
    )

    export_path = tmp_path / "multi.xlsx"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "sheetbooks": {
                "report": {
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(export_path), "write_lock": True},
                }
            }
        },
        runs=[
            {
                "id": "report",
                "demand": "report.yaml",
                "writes": [
                    {"sheetbook_sheet": {"sheetbook": "report", "sheet": "Metrics", "output": "metrics"}},
                    {"sheetbook_sheet": {"sheetbook": "report", "sheet": "Detail", "output": "detail"}},
                ],
            }
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert not result.errors()
    assert export_path.exists()
    assert _read_xlsx_sheetnames(export_path) == ["Metrics", "Detail"]
    assert _read_xlsx_rows(export_path, "Metrics") == [
        ["id"],
        ["a1"],
        ["a2"],
    ]
    assert _read_xlsx_rows(export_path, "Detail") == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]

    write_events = [
        e
        for e in recorder.events
        if e.event_type == EVENT_WORKFLOW_RESOURCE_WRITE and e.payload.resource_type == "sheetbook" and e.payload.resource_id == "report"
    ]
    assert [e.payload.workflow_node_id for e in write_events] == ["__wf__write.report.0", "__wf__write.report.1"]


def test_workflow_sheetbook_loader_consumes_rows_and_enforces_visibility(tmp_path: Path) -> None:
    a_out = tmp_path / "a_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )

    consume_out = tmp_path / "consume.csv"
    _ = _write_table_demand_yaml_from_sheetbook_loader(
        tmp_path,
        file_name="consume.yaml",
        name="consume",
        init_var_name="orders_sheet_ref",
        output_name="detail",
        output_path=consume_out,
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        resources={"sheetbooks": {"report": {"budget": {"max_sheets": 8, "max_total_cells": 1000}}}},
        runs=[
            {
                "id": "a",
                "demand": "a.yaml",
                "writes": [{"sheetbook_sheet": {"sheetbook": "report", "sheet": "Orders", "output": "detail"}}],
            },
            {
                "id": "consume",
                "demand": "consume.yaml",
                "depends_on": ["a"],
                "init_vars": {"orders_sheet_ref": {"node": "a", "sheetbook": "report", "sheet": "Orders"}},
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES_WITH_SHEETBOOK)
    assert not result.errors()
    assert _read_csv_rows(consume_out) == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]

    # Outside depends_on closure: MUST fail-fast.
    wf_bad = _write_workflow_yaml(
        tmp_path,
        resources={"sheetbooks": {"report": {"budget": {"max_sheets": 8, "max_total_cells": 1000}}}},
        runs=[
            {
                "id": "a",
                "demand": "a.yaml",
                "writes": [{"sheetbook_sheet": {"sheetbook": "report", "sheet": "Orders", "output": "detail"}}],
            },
            {
                "id": "c",
                "demand": "consume.yaml",
                "init_vars": {"orders_sheet_ref": {"node": "a", "sheetbook": "report", "sheet": "Orders"}},
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    bad = run_workflow(str(wf_bad), allowed_modules=_ALLOWED_MODULES_WITH_SHEETBOOK)
    errs = {e.run_id: e for e in bad.errors()}
    assert "c" in errs
    assert "declare depends_on" in errs["c"].message


def test_workflow_sheetbook_budget_guards_and_discard_on_failure(tmp_path: Path) -> None:
    a_out = tmp_path / "a_detail.csv"
    b_out = tmp_path / "b_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_path=b_out,
        field_ids=["id", "value"],
    )

    export_path = tmp_path / "budget.xlsx"
    wf_max_sheets = _write_workflow_yaml(
        tmp_path,
        resources={
            "sheetbooks": {
                "report": {
                    "budget": {"max_sheets": 1, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(export_path), "write_lock": True},
                }
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml", "writes": [{"sheetbook_sheet": {"sheetbook": "report", "sheet": "A", "output": "detail"}}]},
            {"id": "b", "demand": "b.yaml", "writes": [{"sheetbook_sheet": {"sheetbook": "report", "sheet": "B", "output": "detail"}}]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result_sheets = run_workflow(str(wf_max_sheets), allowed_modules=_ALLOWED_MODULES)
    assert result_sheets.errors()
    assert not export_path.exists()
    assert not Path(str(export_path) + ".scalim.lock").exists()

    export_path2 = tmp_path / "budget_cells.xlsx"
    wf_max_cells = _write_workflow_yaml(
        tmp_path,
        resources={
            "sheetbooks": {
                "report": {
                    "budget": {"max_sheets": 8, "max_total_cells": 3},
                    "export_xlsx": {"path": str(export_path2), "write_lock": True},
                }
            }
        },
        runs=[{"id": "a", "demand": "a.yaml", "writes": [{"sheetbook_sheet": {"sheetbook": "report", "sheet": "A", "output": "detail"}}]}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    result_cells = run_workflow(str(wf_max_cells), allowed_modules=_ALLOWED_MODULES)
    assert result_cells.errors()
    assert not export_path2.exists()
    assert not Path(str(export_path2) + ".scalim.lock").exists()

    export_path3 = tmp_path / "failed.xlsx"
    _ = _write_demand_yaml(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        main_loader_ref="tests.fixtures.workflow_loaders:load_main_raises",
        preload_loader_ref="tests.fixtures.workflow_loaders:load_preload_table",
    )
    wf_failed = _write_workflow_yaml(
        tmp_path,
        resources={
            "sheetbooks": {
                "report": {
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(export_path3), "write_lock": True},
                }
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml", "writes": [{"sheetbook_sheet": {"sheetbook": "report", "sheet": "A", "output": "detail"}}]},
            {"id": "bad", "demand": "bad.yaml", "depends_on": ["a"]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result_failed = run_workflow(str(wf_failed), allowed_modules=_ALLOWED_MODULES)
    assert result_failed.errors()
    assert not export_path3.exists()
    assert not Path(str(export_path3) + ".scalim.lock").exists()


def test_workflow_excel_output_collision_precheck_and_reserved_paths(tmp_path: Path) -> None:
    out_path = tmp_path / "dup.xlsx"
    _ = _write_table_demand_yaml_with_workbook_output(
        tmp_path,
        file_name="a.xlsx.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=out_path,
        sheet="A",
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_workbook_output(
        tmp_path,
        file_name="b.xlsx.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_path=out_path,
        sheet="B",
        field_ids=["id", "value"],
    )

    wf_collision = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.xlsx.yaml"}, {"id": "b", "demand": "b.xlsx.yaml"}],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    with pytest.raises(WorkflowConfigError, match="collision"):
        _ = run_workflow(str(wf_collision), allowed_modules=_ALLOWED_MODULES)

    reserved_export = tmp_path / "reserved.xlsx"
    _ = _write_table_demand_yaml_with_workbook_output(
        tmp_path,
        file_name="reserved.xlsx.yaml",
        name="reserved",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=reserved_export,
        sheet="S",
        field_ids=["id", "value"],
    )
    wf_reserved = _write_workflow_yaml(
        tmp_path,
        resources={
            "sheetbooks": {
                "report": {
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(reserved_export), "write_lock": True},
                }
            }
        },
        runs=[{"id": "a", "demand": "reserved.xlsx.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(WorkflowConfigError, match="reserved"):
        _ = run_workflow(str(wf_reserved), allowed_modules=_ALLOWED_MODULES)


def test_workflow_excel_output_collision_precheck_allows_dynamic_init_var_paths(tmp_path: Path) -> None:
    a_path = tmp_path / "a.xlsx"
    b_path = tmp_path / "b.xlsx"

    _ = _write_text(
        tmp_path / "demand.yaml",
        (
            f"""
name: demand

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {{extract: id}}
    value: {{extract: value}}

outputs:
  - name: detail
    container:
      type: workbook
      path: {{$init_var: output_path}}
      sheet: "S"
    fields: [id, value]
"""
        ).lstrip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "demand.yaml", "init_vars": {"output_path": str(a_path)}},
            {"id": "b", "demand": "demand.yaml", "init_vars": {"output_path": str(b_path)}},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert not result.errors()
    assert a_path.exists()
    assert b_path.exists()


def test_workflow_excel_output_collision_precheck_rejects_dynamic_init_var_collisions(tmp_path: Path) -> None:
    out_path = tmp_path / "dup.xlsx"

    _ = _write_text(
        tmp_path / "demand.yaml",
        (
            f"""
name: demand

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {{extract: id}}
    value: {{extract: value}}

outputs:
  - name: detail
    container:
      type: workbook
      path: {{$init_var: output_path}}
      sheet: "S"
    fields: [id, value]
"""
        ).lstrip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "demand.yaml", "init_vars": {"output_path": str(out_path)}},
            {"id": "b", "demand": "demand.yaml", "init_vars": {"output_path": str(out_path)}},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    msg = str(excinfo.value)
    assert "collision" in msg
    assert "run_id='b'" in msg
    assert str(out_path.expanduser().resolve(strict=False)) in msg
    assert "nodes=a,b" in msg


def test_workflow_excel_output_reserved_paths_check_uses_resolved_dynamic_init_var_path(tmp_path: Path) -> None:
    reserved_export = tmp_path / "reserved.xlsx"

    _ = _write_text(
        tmp_path / "demand.yaml",
        (
            f"""
name: demand

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {{extract: id}}
    value: {{extract: value}}

outputs:
  - name: detail
    container:
      type: workbook
      path: {{$init_var: output_path}}
      sheet: "S"
    fields: [id, value]
"""
        ).lstrip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "sheetbooks": {
                "report": {
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(reserved_export), "write_lock": True},
                }
            }
        },
        runs=[{"id": "a", "demand": "demand.yaml", "init_vars": {"output_path": str(reserved_export)}}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    msg = str(excinfo.value)
    assert "reserved" in msg
    assert "run_id='a'" in msg
    assert str(reserved_export.expanduser().resolve(strict=False)) in msg


def test_workflow_excel_output_runtime_precheck_includes_meta_and_audit_paths(tmp_path: Path) -> None:
    workbook_path = tmp_path / "out.xlsx"
    audit_path = tmp_path / "audit.xlsx"

    _ = _write_text(
        tmp_path / "demand.yaml",
        (
            f"""
name: demand

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {{extract: id}}
    value: {{extract: value}}

outputs:
  - name: detail
    container:
      type: workbook
      path: "{workbook_path}"
      sheet: "S"
    fields: [id, value]

meta: true
audit:
  path: "{audit_path}"
  sheet: "__audit__"
"""
        ).lstrip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "demand.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert not result.errors()
    assert workbook_path.exists()
    assert audit_path.exists()


def test_workflow_excel_output_collision_precheck_reports_demand_yaml_load_failures(tmp_path: Path) -> None:
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "missing.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    cfg = load_workflow_config(str(wf))
    with pytest.raises(WorkflowConfigError, match="Failed to load demand YAML for workflow collision precheck"):
        _ = workflow_compile_mod.compile_workflow_ir(cfg, workflow_yaml_path=str(wf), path_aliases=None)


@pytest.mark.parametrize("extra_key", ("meta", "audit"))
def test_workflow_excel_output_collision_precheck_includes_meta_and_audit_paths(tmp_path: Path, extra_key: str) -> None:
    workbook_path = tmp_path / "out.xlsx"
    extra_path = tmp_path / "{}.xlsx".format(extra_key)

    _ = _write_text(
        tmp_path / "a.yaml",
        (
            f"""
name: a

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {{extract: id}}
    value: {{extract: value}}

outputs:
  - name: detail
    container:
      type: workbook
      path: "{workbook_path}"
      sheet: "S"
    fields: [id, value]

{extra_key}:
  path: "{extra_path}"
  sheet: "__{extra_key}__"
"""
        ).lstrip(),
    )
    _ = _write_text(
        tmp_path / "b.yaml",
        (
            f"""
name: b

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {{extract: id}}
    value: {{extract: value}}

outputs:
  - name: detail
    container:
      type: workbook
      path: "{extra_path}"
      sheet: "S"
    fields: [id, value]
"""
        ).lstrip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}, {"id": "b", "demand": "b.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    cfg = load_workflow_config(str(wf))
    with pytest.raises(WorkflowConfigError) as excinfo:
        _ = workflow_compile_mod.compile_workflow_ir(cfg, workflow_yaml_path=str(wf), path_aliases=None)
    msg = str(excinfo.value)
    assert "collision" in msg
    assert str(extra_path.expanduser().resolve(strict=False)) in msg


@pytest.mark.parametrize("extra_key", ("meta", "audit"))
def test_workflow_excel_output_collision_precheck_allows_meta_audit_true_without_explicit_path(tmp_path: Path, extra_key: str) -> None:
    workbook_path = tmp_path / "out.xlsx"
    _ = _write_text(
        tmp_path / "demand.yaml",
        (
            f"""
name: demand

main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {{extract: id}}
    value: {{extract: value}}

outputs:
  - name: detail
    container:
      type: workbook
      path: "{workbook_path}"
      sheet: "S"
    fields: [id, value]

{extra_key}: true
"""
        ).lstrip(),
    )
    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "demand.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    cfg = load_workflow_config(str(wf))
    ir = workflow_compile_mod.compile_workflow_ir(cfg, workflow_yaml_path=str(wf), path_aliases=None)
    assert ir is not None


def test_workflow_sheetbook_append_export_xlsx_is_deterministic(tmp_path: Path) -> None:
    a_out = tmp_path / "a_detail.csv"
    b_out = tmp_path / "b_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_path=b_out,
        field_ids=["id", "value"],
    )

    export_path = tmp_path / "append.xlsx"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "sheetbooks": {
                "report": {
                    "budget": {"max_sheets": 8, "max_total_cells": 1000},
                    "export_xlsx": {"path": str(export_path), "write_lock": True},
                }
            }
        },
        runs=[
            {"id": "a", "demand": "a.yaml", "writes": [{"sheetbook_append": {"sheetbook": "report", "sheet": "S", "output": "detail"}}]},
            {"id": "b", "demand": "b.yaml", "writes": [{"sheetbook_append": {"sheetbook": "report", "sheet": "S", "output": "detail"}}]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert not result.errors()
    assert export_path.exists()
    assert _read_xlsx_sheetnames(export_path) == ["S"]
    assert _read_xlsx_rows(export_path, "S") == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
        ["b1", "B1"],
        ["b2", "B2"],
    ]


def test_sheetbook_sheet_rows_loader_requires_context() -> None:
    with pytest.raises(ValueError, match="requires workflow context"):
        _ = workflow_loaders_mod.sheetbook_sheet_rows(ref={"node": "a", "sheetbook": "sb", "sheet": "S"})


def test_sheetbook_sheet_rows_loader_validates_ref_and_context_cleanup() -> None:
    class _DummyManager:
        def iter_sheetbook_sheet_rows(self, **_kwargs: object) -> Any:  # noqa: ANN401
            return iter(())

    dummy = _DummyManager()

    with workflow_loaders_mod.workflow_loader_context(
        workflow_exec_id="wf",
        workflow_node_id="consumer",
        visible_producer_node_ids=frozenset(),
        resource_manager=dummy,  # type: ignore[arg-type]
    ):
        with pytest.raises(TypeError, match="params.ref"):
            _ = workflow_loaders_mod.sheetbook_sheet_rows(ref="nope")  # type: ignore[arg-type]
        with pytest.raises(ValueError, match="ref.node"):
            _ = workflow_loaders_mod.sheetbook_sheet_rows(ref={"sheetbook": "sb", "sheet": "S"})
        with pytest.raises(ValueError, match="ref.sheetbook"):
            _ = workflow_loaders_mod.sheetbook_sheet_rows(ref={"node": "a", "sheet": "S"})
        with pytest.raises(ValueError, match="ref.sheet"):
            _ = workflow_loaders_mod.sheetbook_sheet_rows(ref={"node": "a", "sheetbook": "sb"})

    with workflow_loaders_mod.workflow_loader_context(
        workflow_exec_id="wf",
        workflow_node_id="consumer",
        visible_producer_node_ids=frozenset(),
        resource_manager=dummy,  # type: ignore[arg-type]
    ):
        delattr(workflow_loaders_mod._TLS, "ctx")


def test_sheetbook_sheet_rows_loader_rejects_corrupted_context() -> None:
    workflow_loaders_mod._TLS.ctx = object()
    try:
        with pytest.raises(TypeError, match="context is corrupted"):
            _ = workflow_loaders_mod.sheetbook_sheet_rows(ref={"node": "a", "sheetbook": "sb", "sheet": "S"})
    finally:
        delattr(workflow_loaders_mod._TLS, "ctx")


def test_workflow_shared_workbook_append_is_deterministic_by_runs_order(tmp_path: Path) -> None:
    slow_out = tmp_path / "slow_detail.csv"
    fast_out = tmp_path / "fast_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="slow.yaml",
        name="slow",
        loader_ref="tests.fixtures.workflow_loaders:load_table_c_slow",
        output_name="detail",
        output_path=slow_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="fast.yaml",
        name="fast",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_path=fast_out,
        field_ids=["id", "value"],
    )

    expected_rows = [
        ["id", "value"],
        ["c1", "C1"],
        ["b1", "B1"],
        ["b2", "B2"],
    ]
    for idx in range(3):
        workbook_path = tmp_path / "append_{}.xlsx".format(int(idx))
        wf = _write_workflow_yaml(
            tmp_path,
            resources={"workbooks": {"report": {"path": str(workbook_path)}}},
            runs=[
                {
                    "id": "slow",
                    "demand": "slow.yaml",
                    "writes": [
                        {
                            "workbook_append": {
                                "workbook": "report",
                                "sheet": "All",
                                "output": "detail",
                                "header_policy": "once",
                                "on_mismatch": "error",
                            }
                        }
                    ],
                },
                {
                    "id": "fast",
                    "demand": "fast.yaml",
                    "writes": [
                        {
                            "workbook_append": {
                                "workbook": "report",
                                "sheet": "All",
                                "output": "detail",
                                "header_policy": "once",
                                "on_mismatch": "error",
                            }
                        }
                    ],
                },
            ],
            max_concurrency=2,
            failure_policy="primary_only",
        )

        recorder = _WorkflowEventRecorder()
        result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
        assert not result.errors()
        assert workbook_path.exists()
        assert _read_xlsx_rows(workbook_path, "All") == expected_rows

        write_events = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_WRITE and e.payload.resource_id == "report"]
        assert [e.payload.workflow_node_id for e in write_events] == ["__wf__write.slow.0", "__wf__write.fast.0"]


def test_workflow_shared_append_header_policy_variants(tmp_path: Path) -> None:
    a_out = tmp_path / "a_detail.csv"
    b_out = tmp_path / "b_detail.csv"
    c_out = tmp_path / "c_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_path=b_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="c.yaml",
        name="c",
        loader_ref="tests.fixtures.workflow_loaders:load_table_c_slow",
        output_name="detail",
        output_path=c_out,
        field_ids=["id", "value"],
    )

    workbook_path = tmp_path / "policies.xlsx"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path)}}},
        runs=[
            {
                "id": "a",
                "demand": "a.yaml",
                "writes": [{"workbook_append": {"workbook": "report", "sheet": "All", "output": "detail", "header_policy": "once"}}],
            },
            {
                "id": "b",
                "demand": "b.yaml",
                "writes": [{"workbook_append": {"workbook": "report", "sheet": "All", "output": "detail", "header_policy": "always"}}],
            },
            {
                "id": "c",
                "demand": "c.yaml",
                "writes": [{"workbook_append": {"workbook": "report", "sheet": "All", "output": "detail", "header_policy": "never"}}],
            },
        ],
        max_concurrency=3,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert not result.errors()

    assert _read_xlsx_rows(workbook_path, "All") == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
        ["id", "value"],
        ["b1", "B1"],
        ["b2", "B2"],
        ["c1", "C1"],
    ]


def test_workflow_shared_csv_append_warn_skip_and_header_policies(tmp_path: Path) -> None:
    a_out = tmp_path / "a_detail.csv"
    b_out = tmp_path / "b_detail.csv"
    m_out = tmp_path / "m_detail.csv"
    m2_out = tmp_path / "m2_detail.csv"

    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_path=b_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="m.yaml",
        name="m",
        loader_ref="tests.fixtures.workflow_loaders:load_table_mismatch",
        output_name="detail",
        output_path=m_out,
        field_ids=["id", "other"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="m2.yaml",
        name="m2",
        loader_ref="tests.fixtures.workflow_loaders:load_table_mismatch",
        output_name="detail",
        output_path=m2_out,
        field_ids=["id", "other"],
    )

    merged_path = tmp_path / "merged.csv"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={"csvs": {"merged": {"path": str(merged_path)}}},
        runs=[
            {
                "id": "a",
                "demand": "a.yaml",
                "writes": [{"csv_append": {"csv": "merged", "output": "detail", "header_policy": "once", "on_mismatch": "error"}}],
            },
            {
                "id": "m",
                "demand": "m.yaml",
                "writes": [{"csv_append": {"csv": "merged", "output": "detail", "header_policy": "always", "on_mismatch": "warn"}}],
            },
            {
                "id": "b",
                "demand": "b.yaml",
                "writes": [{"csv_append": {"csv": "merged", "output": "detail", "header_policy": "never", "on_mismatch": "error"}}],
            },
            {
                "id": "m2",
                "demand": "m2.yaml",
                "writes": [{"csv_append": {"csv": "merged", "output": "detail", "header_policy": "once", "on_mismatch": "skip"}}],
            },
        ],
        max_concurrency=4,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert not result.errors()
    assert merged_path.exists()

    assert _read_csv_rows(merged_path) == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
        ["id", "value"],
        ["m1", ""],
        ["b1", "B1"],
        ["b2", "B2"],
    ]

    skips = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_WRITE and e.payload.action == "skip"]
    assert skips


def test_workflow_shared_resources_discard_on_failure(tmp_path: Path) -> None:
    ok_out = tmp_path / "ok_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=ok_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="bad.yaml",
        name="bad",
        loader_ref="tests.fixtures.workflow_loaders:load_table_raises",
        output_name="detail",
        output_path=tmp_path / "bad_detail.csv",
        field_ids=["id", "value"],
    )

    workbook_path = tmp_path / "discard.xlsx"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path)}}},
        runs=[
            {"id": "ok", "demand": "ok.yaml", "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "OK", "output": "detail"}}]},
            {"id": "bad", "demand": "bad.yaml"},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    recorder = _WorkflowEventRecorder()
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, components=[recorder])
    assert result.errors()
    assert not workbook_path.exists()
    assert not Path(str(workbook_path) + ".scalim.lock").exists()

    discards = [e for e in recorder.events if e.event_type == EVENT_WORKFLOW_RESOURCE_DISCARD and e.payload.resource_id == "report"]
    assert discards


def test_workflow_shared_sheet_conflict_policies(tmp_path: Path) -> None:
    a_out = tmp_path / "a_detail.csv"
    b_out = tmp_path / "b_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_path=b_out,
        field_ids=["id", "value"],
    )

    # on_conflict=error
    workbook_err = tmp_path / "conflict_err.xlsx"
    wf_err = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_err)}}},
        runs=[
            {"id": "a", "demand": "a.yaml", "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}}]},
            {
                "id": "b",
                "demand": "b.yaml",
                "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail", "on_conflict": "error"}}],
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result_err = run_workflow(str(wf_err), allowed_modules=_ALLOWED_MODULES)
    assert result_err.errors()
    assert not workbook_err.exists()

    # on_conflict=overwrite
    workbook_over = tmp_path / "conflict_over.xlsx"
    wf_over = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_over)}}},
        runs=[
            {"id": "a", "demand": "a.yaml", "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}}]},
            {
                "id": "b",
                "demand": "b.yaml",
                "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail", "on_conflict": "overwrite"}}],
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result_over = run_workflow(str(wf_over), allowed_modules=_ALLOWED_MODULES)
    assert not result_over.errors()
    assert workbook_over.exists()
    assert _read_xlsx_rows(workbook_over, "S")[-1] == ["b2", "B2"]

    # on_conflict=skip
    workbook_skip = tmp_path / "conflict_skip.xlsx"
    wf_skip = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_skip)}}},
        runs=[
            {"id": "a", "demand": "a.yaml", "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}}]},
            {
                "id": "b",
                "demand": "b.yaml",
                "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail", "on_conflict": "skip"}}],
            },
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )
    result_skip = run_workflow(str(wf_skip), allowed_modules=_ALLOWED_MODULES)
    assert not result_skip.errors()
    assert workbook_skip.exists()
    assert _read_xlsx_rows(workbook_skip, "S")[-1] == ["a2", "A2"]


def test_workflow_shared_output_lock_failure_fails_fast(tmp_path: Path) -> None:
    out = tmp_path / "a_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=out,
        field_ids=["id", "value"],
    )

    workbook_path = tmp_path / "locked.xlsx"
    lock_path = Path(str(workbook_path) + ".scalim.lock")
    lock_path.write_text("locked", encoding="utf-8")

    wf = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path)}}},
        runs=[{"id": "a", "demand": "a.yaml", "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}}]}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert result.errors()
    assert lock_path.exists()


def test_workflow_pathless_csv_output_without_writes_fails_fast(tmp_path: Path) -> None:
    _ = _write_text(
        tmp_path / "a.yaml",
        (
            """
name: a
main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {extract: id}
    value: {extract: value}
outputs:
  - name: detail
    container:
      type: csv
    fields: [id, value]
"""
        ).lstrip(),
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[{"id": "a", "demand": "a.yaml"}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    with pytest.raises(WorkflowConfigError, match="Pathless CSV output"):
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)


def test_workflow_managed_temp_outputs_injects_pathless_csv_and_cleans_up(tmp_path: Path) -> None:
    _ = _write_text(
        tmp_path / "a.yaml",
        (
            """
name: a
main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {extract: id}
    value: {extract: value}
outputs:
  - name: detail
    container:
      type: csv
    fields: [id, value]
"""
        ).lstrip(),
    )

    workbook_path = tmp_path / "managed.xlsx"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path)}}},
        runs=[{"id": "a", "demand": "a.yaml", "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}}]}],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert not result.errors()
    demand_outcome = next(o for o in result.outcomes if o.run_id == "a")
    assert demand_outcome.result is not None
    assert demand_outcome.result.core.in_memory_csv_outputs == {}
    assert workbook_path.exists()
    assert _read_xlsx_rows(workbook_path, "S")[-1] == ["a2", "A2"]
    assert not (tmp_path / ".scalim").exists()


def test_workflow_main_rows_from_wires_upstream_typed_rows_into_downstream_main_rows(tmp_path: Path) -> None:
    a_out = tmp_path / "a_detail.csv"
    b_out = tmp_path / "b_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_raises",
        output_name="detail",
        output_path=b_out,
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml", "depends_on": ["a"], "main_rows_from": {"run": "a"}},
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert not result.errors()
    assert _read_csv_rows(b_out) == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]


def test_workflow_main_rows_from_releases_typed_rows_after_final_consumer(tmp_path: Path, monkeypatch) -> None:
    publish_calls = []
    discard_calls = []

    class _RecordingArtifactsDirectory(workflow_execute_mod.WorkflowArtifactsDirectory):  # type: ignore[misc]
        def publish(self, producer_node_id: str, artifact_id: str, value: object) -> None:
            publish_calls.append((str(producer_node_id), str(artifact_id)))
            super(_RecordingArtifactsDirectory, self).publish(producer_node_id, artifact_id, value)

        def discard(self, producer_node_id: str, artifact_id: str) -> None:
            discard_calls.append((str(producer_node_id), str(artifact_id)))
            super(_RecordingArtifactsDirectory, self).discard(producer_node_id, artifact_id)

    monkeypatch.setattr(workflow_execute_mod, "WorkflowArtifactsDirectory", _RecordingArtifactsDirectory)

    a_out = tmp_path / "a_detail.csv"
    b_out = tmp_path / "b_detail.csv"
    c_out = tmp_path / "c_detail.csv"
    d_out = tmp_path / "d_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_raises",
        output_name="detail",
        output_path=b_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="c.yaml",
        name="c",
        loader_ref="tests.fixtures.workflow_loaders:load_table_raises",
        output_name="detail",
        output_path=c_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="d.yaml",
        name="d",
        loader_ref="tests.fixtures.workflow_loaders:load_table_b_fast",
        output_name="detail",
        output_path=d_out,
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml", "depends_on": ["a"], "main_rows_from": {"run": "a"}},
            {"id": "c", "demand": "c.yaml", "depends_on": ["a"], "main_rows_from": {"run": "a"}},
            {"id": "d", "demand": "d.yaml", "depends_on": ["b", "c"]},
        ],
        max_concurrency=2,
        failure_policy="primary_only",
    )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert not result.errors()
    assert _read_csv_rows(b_out) == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]
    assert _read_csv_rows(c_out) == [
        ["id", "value"],
        ["a1", "A1"],
        ["a2", "A2"],
    ]

    typed_rows_publishes = [c for c in publish_calls if c[1] == "in_memory_rows"]
    assert typed_rows_publishes == [("a", "in_memory_rows")]

    typed_rows_discards = [c for c in discard_calls if c[1] == "in_memory_rows"]
    assert typed_rows_discards == [("a", "in_memory_rows")]


def test_workflow_artifacts_directory_discard_all_in_memory_rows_removes_empty_producer_entry() -> None:
    from scalim.sinks.sink_rows import InMemoryRows
    from scalim.spec.ir.workflow import WorkflowArtifactsIr, WorkflowIr, WorkflowNodeIr, WorkflowNodeType, WorkflowOptionsIr

    workflow_ir = WorkflowIr(
        nodes=(WorkflowNodeIr(node_id="a", node_type=WorkflowNodeType.DEMAND, decl_order=0, deps=()),),
        edges=(),
        options=WorkflowOptionsIr(),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={}),
    )
    artifacts_dir = workflow_execute_mod.WorkflowArtifactsDirectory(workflow_ir)
    artifacts_dir.publish("a", "in_memory_rows", InMemoryRows(header=["id"], rows=[[1]]))
    artifacts_dir.discard_all_in_memory_rows()

    with pytest.raises(KeyError, match=r"Unknown artifact"):
        _ = artifacts_dir.get("a", "a", "in_memory_rows")


def test_workflow_main_rows_from_rejects_non_in_memory_rows_artifact(tmp_path: Path, monkeypatch) -> None:
    class _CorruptArtifactsDirectory(workflow_execute_mod.WorkflowArtifactsDirectory):  # type: ignore[misc]
        def publish(self, producer_node_id: str, artifact_id: str, value: object) -> None:
            if str(producer_node_id) == "a" and str(artifact_id) == "in_memory_rows":
                value = "not_in_memory_rows"
            super(_CorruptArtifactsDirectory, self).publish(producer_node_id, artifact_id, value)

    monkeypatch.setattr(workflow_execute_mod, "WorkflowArtifactsDirectory", _CorruptArtifactsDirectory)

    a_out = tmp_path / "a_detail.csv"
    b_out = tmp_path / "b_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=a_out,
        field_ids=["id", "value"],
    )
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="b.yaml",
        name="b",
        loader_ref="tests.fixtures.workflow_loaders:load_table_raises",
        output_name="detail",
        output_path=b_out,
        field_ids=["id", "value"],
    )

    wf = _write_workflow_yaml(
        tmp_path,
        runs=[
            {"id": "a", "demand": "a.yaml"},
            {"id": "b", "demand": "b.yaml", "depends_on": ["a"], "main_rows_from": {"run": "a"}},
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert result.errors()
    b_outcome = next(o for o in result.outcomes if o.run_id == "b")
    assert b_outcome.error is not None
    assert b_outcome.error.exc_type == "WorkflowWriteError"


def test_workflow_execute_release_main_rows_artifact_returns_when_missing_count_entry(tmp_path: Path) -> None:
    from typing import Any

    from scalim.execution.run_ir import ExecutionRequest, ExportLayout, OutputSpec
    from scalim.spec.ir.workflow import WorkflowArtifactsIr, WorkflowEdgeIr, WorkflowIr, WorkflowNodeIr, WorkflowNodeType, WorkflowOptionsIr
    from scalim.vendor.dataclassesx import dataclass

    @dataclass(frozen=True)
    class _Compilation:
        demand_ir: object
        request: ExecutionRequest

    class _Core:
        output_path = None
        total_rows = 0
        duration = 0.0
        outputs = None
        in_memory_csv_outputs = {}
        in_memory_rows = None

    def _compile_demand_node(demand_path: str, **kwargs: Any) -> object:
        _ = demand_path, kwargs
        request = ExecutionRequest(
            export_layout=ExportLayout(field_ids=(), header_names=None),
            output=OutputSpec(path=None),
            sink=None,
        )
        return _Compilation(demand_ir=object(), request=request)

    def _run_ir_fn(demand_ir: object, request: ExecutionRequest, **kwargs: Any) -> object:
        _ = demand_ir, request, kwargs
        return _Core()

    workflow_ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(
                node_id="a",
                node_type=WorkflowNodeType.DEMAND,
                decl_order=0,
                deps=(),
                demand_path="a.yaml",
            ),
            WorkflowNodeIr(
                node_id="b",
                node_type=WorkflowNodeType.DEMAND,
                decl_order=1,
                deps=("a",),
                demand_path="b.yaml",
                main_rows_from_run_id="a",
            ),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(max_concurrency=1, failure_policy="primary_only"),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={}),
    )

    prepared = workflow_execute_mod._prepare_workflow_run_ir(  # type: ignore[attr-defined]
        str(tmp_path / "workflow.yaml"),
        workflow_ir,
        components=None,
        bundle_viz_base_config=None,
        cache_pool_logical_keys_by_node_id=None,
        cache_pool_consumers_by_logical_key=None,
    )
    prepared.main_rows_consumers_remaining_by_run_id = {}
    try:
        outcomes, _failed, _exc = workflow_execute_mod._execute_workflow_run(  # type: ignore[attr-defined]
            prepared,
            compile_demand_fn=_compile_demand_node,
            build_demand_run_result_fn=None,
            run_ir_fn=_run_ir_fn,
        )
    finally:
        workflow_execute_mod._cleanup_workflow_finally(prepared, resources_finalized=False)  # type: ignore[attr-defined]

    b_outcome = next(o for o in outcomes if o.run_id == "b")
    assert b_outcome.error is not None
    assert b_outcome.error.exc_type == "KeyError"


def test_workflow_execute_release_main_rows_artifact_raises_on_negative_count(tmp_path: Path) -> None:
    from typing import Any

    from scalim.execution.run_ir import ExecutionRequest, ExportLayout, OutputSpec
    from scalim.spec.ir.workflow import WorkflowArtifactsIr, WorkflowEdgeIr, WorkflowIr, WorkflowNodeIr, WorkflowNodeType, WorkflowOptionsIr
    from scalim.vendor.dataclassesx import dataclass

    @dataclass(frozen=True)
    class _Compilation:
        demand_ir: object
        request: ExecutionRequest

    class _Core:
        output_path = None
        total_rows = 0
        duration = 0.0
        outputs = None
        in_memory_csv_outputs = {}
        in_memory_rows = None

    def _compile_demand_node(demand_path: str, **kwargs: Any) -> object:
        _ = demand_path, kwargs
        request = ExecutionRequest(
            export_layout=ExportLayout(field_ids=(), header_names=None),
            output=OutputSpec(path=None),
            sink=None,
        )
        return _Compilation(demand_ir=object(), request=request)

    def _run_ir_fn(demand_ir: object, request: ExecutionRequest, **kwargs: Any) -> object:
        _ = demand_ir, request, kwargs
        return _Core()

    workflow_ir = WorkflowIr(
        nodes=(
            WorkflowNodeIr(
                node_id="a",
                node_type=WorkflowNodeType.DEMAND,
                decl_order=0,
                deps=(),
                demand_path="a.yaml",
            ),
            WorkflowNodeIr(
                node_id="b",
                node_type=WorkflowNodeType.DEMAND,
                decl_order=1,
                deps=("a",),
                demand_path="b.yaml",
                main_rows_from_run_id="a",
            ),
        ),
        edges=(WorkflowEdgeIr(from_node_id="a", to_node_id="b"),),
        options=WorkflowOptionsIr(max_concurrency=1, failure_policy="primary_only"),
        resources=(),
        artifacts=WorkflowArtifactsIr(slots_by_node_id={}),
    )

    prepared = workflow_execute_mod._prepare_workflow_run_ir(  # type: ignore[attr-defined]
        str(tmp_path / "workflow.yaml"),
        workflow_ir,
        components=None,
        bundle_viz_base_config=None,
        cache_pool_logical_keys_by_node_id=None,
        cache_pool_consumers_by_logical_key=None,
    )
    prepared.main_rows_consumers_remaining_by_run_id = {"a": 0}
    try:
        with pytest.raises(RuntimeError, match=r"negative main_rows consumer count"):
            _ = workflow_execute_mod._execute_workflow_run(  # type: ignore[attr-defined]
                prepared,
                compile_demand_fn=_compile_demand_node,
                build_demand_run_result_fn=None,
                run_ir_fn=_run_ir_fn,
            )
    finally:
        workflow_execute_mod._cleanup_workflow_finally(prepared, resources_finalized=False)  # type: ignore[attr-defined]


def test_workflow_shared_write_node_validates_demand_outputs_mapping_and_output_id(tmp_path: Path) -> None:
    ok_out = tmp_path / "ok_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=ok_out,
        field_ids=["id", "value"],
    )
    no_outputs = _write_text(
        tmp_path / "no_outputs.yaml",
        (
            """
name: demo
main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {extract: id}
    value: {extract: value}
"""
        ).lstrip(),
    )

    workbook_path = tmp_path / "validate.xlsx"

    wf_missing_outputs = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path)}}},
        runs=[
            {"id": "a", "demand": str(no_outputs), "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}}]}
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(WorkflowConfigError, match="Unknown demand output id referenced by workflow writes"):
        _ = run_workflow(str(wf_missing_outputs), allowed_modules=_ALLOWED_MODULES)

    wf_unknown_output_id = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path)}}},
        runs=[{"id": "ok", "demand": "ok.yaml", "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "nope"}}]}],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(WorkflowConfigError, match="Unknown demand output id referenced by workflow writes"):
        _ = run_workflow(str(wf_unknown_output_id), allowed_modules=_ALLOWED_MODULES)


def test_workflow_shared_append_write_node_validates_outputs_mapping_and_output_id(tmp_path: Path) -> None:
    ok_out = tmp_path / "ok_detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=ok_out,
        field_ids=["id", "value"],
    )
    no_outputs = _write_text(
        tmp_path / "no_outputs.yaml",
        (
            """
name: demo
main_source:
  source_id: main
  loader: "tests.fixtures.workflow_loaders:load_table_a_fast"
  fields:
    id: {extract: id}
    value: {extract: value}
"""
        ).lstrip(),
    )

    workbook_path = tmp_path / "append_validate.xlsx"

    wf_missing_outputs = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path)}}},
        runs=[
            {
                "id": "a",
                "demand": str(no_outputs),
                "writes": [{"workbook_append": {"workbook": "report", "sheet": "S", "output": "detail"}}],
            }
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(WorkflowConfigError, match="Unknown demand output id referenced by workflow writes"):
        _ = run_workflow(str(wf_missing_outputs), allowed_modules=_ALLOWED_MODULES)

    wf_unknown_output_id = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path)}}},
        runs=[
            {
                "id": "ok",
                "demand": "ok.yaml",
                "writes": [{"workbook_append": {"workbook": "report", "sheet": "S", "output": "nope"}}],
            }
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(WorkflowConfigError, match="Unknown demand output id referenced by workflow writes"):
        _ = run_workflow(str(wf_unknown_output_id), allowed_modules=_ALLOWED_MODULES)


def test_workflow_shared_write_node_rejects_non_csv_output_paths(tmp_path: Path) -> None:
    out = tmp_path / "detail.txt"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="badext.yaml",
        name="badext",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=out,
        field_ids=["id", "value"],
    )

    workbook_path = tmp_path / "report.xlsx"
    wf_sheet = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path)}}},
        runs=[
            {
                "id": "a",
                "demand": "badext.yaml",
                "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}}],
            }
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    result_sheet = run_workflow(str(wf_sheet), allowed_modules=_ALLOWED_MODULES)
    assert result_sheet.errors()

    workbook_path2 = tmp_path / "append.xlsx"
    wf_append = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_path2)}}},
        runs=[
            {
                "id": "a",
                "demand": "badext.yaml",
                "writes": [{"workbook_append": {"workbook": "report", "sheet": "S", "output": "detail"}}],
            }
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    result_append = run_workflow(str(wf_append), allowed_modules=_ALLOWED_MODULES)
    assert result_append.errors()


@pytest.mark.parametrize(
    ("outputs_override", "expected_substrings"),
    [
        (
            None,
            (
                "write node requires demand outputs mapping",
                "append node requires demand outputs mapping",
            ),
        ),
        (
            {"other": "./other.csv"},
            ("Unknown demand output id: input_node_id='a', output_id='detail'",),
        ),
    ],
    ids=["missing-outputs-mapping", "unknown-output-id"],
)
def test_workflow_write_nodes_validate_outputs_mapping_runtime(
    tmp_path: Path,
    outputs_override: Optional[Dict[str, str]],
    expected_substrings: tuple,
) -> None:
    out = tmp_path / "detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="a.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=out,
        field_ids=["id", "value"],
    )

    workbook_path = tmp_path / "report.xlsx"
    merged_path = tmp_path / "merged.csv"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={
            "workbooks": {"report": {"path": str(workbook_path)}},
            "csvs": {"merged": {"path": str(merged_path)}},
        },
        runs=[
            {
                "id": "a",
                "demand": "a.yaml",
                "writes": [
                    {"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}},
                    {"csv_append": {"csv": "merged", "output": "detail"}},
                ],
            }
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )

    from scalim.execution.run_ir import ExecutionResult, run_ir as real_run_ir  # noqa: PLC0415

    def fake_run_ir(demand_ir, request, engine_factory=None, event_meta_defaults=None):  # type: ignore[no-untyped-def]
        real = real_run_ir(
            demand_ir,
            request,
            engine_factory=engine_factory,
            event_meta_defaults=event_meta_defaults,
        )
        return ExecutionResult(
            output_path=real.output_path,
            total_rows=real.total_rows,
            duration=real.duration,
            demand_ir=real.demand_ir,
            plan=real.plan,
            outputs=outputs_override,
            output_target_stats=real.output_target_stats,
        )

    result = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES, run_ir_fn=fake_run_ir)
    errors = result.errors()
    assert len(errors) == 2
    messages = sorted(str(e.message) for e in errors)
    for expected in expected_substrings:
        assert any(expected in msg for msg in messages)


def test_workflow_writes_rejects_non_csv_output_types(tmp_path: Path) -> None:
    workbook_out = tmp_path / "demand.xlsx"
    _ = _write_table_demand_yaml_with_workbook_output(
        tmp_path,
        file_name="a.xlsx.yaml",
        name="a",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=workbook_out,
        sheet="S",
        field_ids=["id", "value"],
    )

    report_path = tmp_path / "report.xlsx"
    wf = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(report_path)}}},
        runs=[
            {
                "id": "a",
                "demand": "a.xlsx.yaml",
                "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}}],
            }
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(WorkflowConfigError, match=r"only supports CSV outputs"):
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)


def test_workflow_shared_resource_commit_failure_raises_workflow_config_error(tmp_path: Path) -> None:
    out = tmp_path / "detail.csv"
    _ = _write_table_demand_yaml_with_csv_output(
        tmp_path,
        file_name="ok.yaml",
        name="ok",
        loader_ref="tests.fixtures.workflow_loaders:load_table_a_fast",
        output_name="detail",
        output_path=out,
        field_ids=["id", "value"],
    )

    workbook_dir = tmp_path / "outdir"
    workbook_dir.mkdir()
    lock_path = Path(str(workbook_dir) + ".scalim.lock")

    wf = _write_workflow_yaml(
        tmp_path,
        resources={"workbooks": {"report": {"path": str(workbook_dir)}}},
        runs=[
            {
                "id": "a",
                "demand": "ok.yaml",
                "writes": [{"workbook_sheet": {"workbook": "report", "sheet": "S", "output": "detail"}}],
            }
        ],
        max_concurrency=1,
        failure_policy="primary_only",
    )
    with pytest.raises(WorkflowConfigError, match="workflow.resources"):
        _ = run_workflow(str(wf), allowed_modules=_ALLOWED_MODULES)
    assert workbook_dir.exists()
    assert not lock_path.exists()
