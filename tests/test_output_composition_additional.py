from pathlib import Path

import pytest

from scalim.execution import output_composition as mod
from scalim.execution.output_contracts import ExportLayout, OutputSpec
from scalim.sinks.sink_base import BaseRowSink


class _CollectingRowSink(BaseRowSink):
    def __init__(self) -> None:
        self.rows = []
        self.closed = False

    def write_row(self, row) -> None:  # type: ignore[no-untyped-def]
        self.rows.append(dict(row))

    def close(self) -> None:
        self.closed = True


class _FailingWriteRowSink(BaseRowSink):
    def write_row(self, _row) -> None:  # type: ignore[no-untyped-def]
        raise ValueError("boom")

    def close(self) -> None:
        return None


class _FailingCloseSink(BaseRowSink):
    def write_row(self, _row) -> None:  # type: ignore[no-untyped-def]
        return None

    def close(self) -> None:
        raise ValueError("boom-close")


def test_required_demand_fields_includes_requires_fields() -> None:
    spec = mod.OutputCompositionSpec(
        targets=(
            mod.OutputTargetSpec(
                target_id="t1",
                layout=ExportLayout(field_ids=("a",), header_names=None),
                output=OutputSpec(format="csv", path="out.csv", streaming=True),
                requires=("req_a",),
            ),
        ),
        derived_targets=(
            mod.DerivedOutputTargetSpec(
                target_id="d1",
                derived=mod.DerivedGroupBySpec(
                    group_by=("g",),
                    metrics=(mod.AggMetricSpec(out_field_id="cnt", op="count", field_id=None),),
                ),
                output_layout=ExportLayout(field_ids=("g", "cnt"), header_names=None),
                output=OutputSpec(format="csv", path="out.csv", streaming=True),
                requires=("req_b",),
            ),
        ),
    )

    fields = mod.required_demand_fields(spec)
    assert "req_a" in fields
    assert "req_b" in fields


def test_create_row_sink_for_composed_output_validations(tmp_path: Path) -> None:
    layout = ExportLayout(field_ids=("a",), header_names=None)

    with pytest.raises(ValueError, match="OutputSpec.path is required"):
        _ = mod._create_row_sink_for_composed_output(  # noqa: SLF001
            target_id="t",
            output=OutputSpec(format="csv", path=None, streaming=True),
            layout=layout,
            workbook_by_path={},
        )

    with pytest.raises(ValueError, match="streaming=true"):
        _ = mod._create_row_sink_for_composed_output(  # noqa: SLF001
            target_id="t",
            output=OutputSpec(format="csv", path=str(tmp_path / "x.csv"), streaming=False),
            layout=layout,
            workbook_by_path={},
        )

    with pytest.raises(ValueError, match="streaming=true"):
        _ = mod._create_row_sink_for_composed_output(  # noqa: SLF001
            target_id="t",
            output=OutputSpec(format="excel", path=str(tmp_path / "x.xlsx"), streaming=False),
            layout=layout,
            workbook_by_path={},
        )

    with pytest.raises(ValueError, match="Unsupported output format"):
        _ = mod._create_row_sink_for_composed_output(  # noqa: SLF001
            target_id="t",
            output=OutputSpec(format="json", path=str(tmp_path / "x.json"), streaming=True),
            layout=layout,
            workbook_by_path={},
        )

    with pytest.raises(ValueError, match="In-memory composed output only supports format=csv"):
        _ = mod._create_row_sink_for_composed_output(  # noqa: SLF001
            target_id="t",
            output=OutputSpec(format="excel", path=None, streaming=True),
            layout=layout,
            workbook_by_path={},
            in_memory=True,
        )

    with pytest.raises(ValueError, match="streaming=true"):
        _ = mod._create_row_sink_for_composed_output(  # noqa: SLF001
            target_id="t",
            output=OutputSpec(format="csv", path=None, streaming=False),
            layout=layout,
            workbook_by_path={},
            in_memory=True,
        )


def test_create_row_sink_for_composed_output_csv_success_counts_and_writes(tmp_path: Path) -> None:
    out = tmp_path / "out.csv"
    sink, counter, _ = mod._create_row_sink_for_composed_output(  # noqa: SLF001
        target_id="t",
        output=OutputSpec(format="csv", path=str(out), streaming=True, include_header=True),
        layout=ExportLayout(field_ids=("a",), header_names=None),
        workbook_by_path={},
    )
    sink.write_row({"a": 1})
    sink.write_row({"a": 2})
    sink.close()

    assert counter.rows == 2
    assert out.exists()


def test_create_row_sink_for_composed_output_excel_single_sheet_defaults_sheet_name(tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    sink, counter, _ = mod._create_row_sink_for_composed_output(  # noqa: SLF001
        target_id="t",
        output=OutputSpec(format="excel", path=str(out), streaming=True, include_header=True, sheet_name=None),
        layout=ExportLayout(field_ids=("a",), header_names=None),
        workbook_by_path={},
    )
    sink.write_row({"a": 1})
    sink.close()

    assert counter.rows == 1
    assert out.exists()

    from openpyxl import load_workbook

    wb = load_workbook(str(out), read_only=True, data_only=True)
    try:
        assert wb.sheetnames == ["Sheet1"]
    finally:
        wb.close()


def test_validate_excel_workbook_sheet_names_continue_and_error(tmp_path: Path) -> None:
    single = mod.OutputCompositionSpec(
        targets=(
            mod.OutputTargetSpec(
                target_id="t1",
                layout=ExportLayout(field_ids=("a",), header_names=None),
                output=OutputSpec(format="excel", path=str(tmp_path / "one.xlsx"), streaming=True, sheet_name="S1"),
            ),
        ),
    )
    mod._validate_excel_workbook_sheet_names(single)  # noqa: SLF001

    shared = mod.OutputCompositionSpec(
        targets=(
            mod.OutputTargetSpec(
                target_id="t_ok",
                layout=ExportLayout(field_ids=("a",), header_names=None),
                output=OutputSpec(format="excel", path=str(tmp_path / "shared.xlsx"), streaming=True, sheet_name="S1"),
            ),
            mod.OutputTargetSpec(
                target_id="t_missing",
                layout=ExportLayout(field_ids=("a",), header_names=None),
                output=OutputSpec(format="excel", path=str(tmp_path / "shared.xlsx"), streaming=True, sheet_name=None),
            ),
        ),
    )
    with pytest.raises(ValueError, match="missing sheet_name"):
        mod._validate_excel_workbook_sheet_names(shared)  # noqa: SLF001


def test_normalize_failure_policy_rejects_unknown_value() -> None:
    with pytest.raises(ValueError, match="Unsupported failure_policy"):
        _ = mod._normalize_failure_policy("bad")  # noqa: SLF001


def test_ensure_primary_route_sets_first_when_missing() -> None:
    r1 = mod._RouteState(  # noqa: SLF001
        target_id="t1",
        sink=_CollectingRowSink(),
        predicate=None,
        is_primary=False,
        output_path=None,
        sheet_name=None,
        output_counter=mod._RowCounter(),  # noqa: SLF001
    )
    r2 = mod._RouteState(  # noqa: SLF001
        target_id="t2",
        sink=_CollectingRowSink(),
        predicate=None,
        is_primary=False,
        output_path=None,
        sheet_name=None,
        output_counter=mod._RowCounter(),  # noqa: SLF001
    )
    routes = [r1, r2]
    mod._ensure_primary_route(routes)  # noqa: SLF001
    assert routes[0].is_primary is True


def test_router_row_sink_skips_disabled_and_predicate() -> None:
    s1 = _CollectingRowSink()
    s2 = _CollectingRowSink()
    routes = [
        mod._RouteState(  # noqa: SLF001
            target_id="disabled",
            sink=s1,
            predicate=None,
            is_primary=False,
            output_path=None,
            sheet_name=None,
            disabled=True,
            output_counter=mod._RowCounter(),  # noqa: SLF001
        ),
        mod._RouteState(  # noqa: SLF001
            target_id="pred",
            sink=s2,
            predicate=lambda _row: False,
            is_primary=False,
            output_path=None,
            sheet_name=None,
            output_counter=mod._RowCounter(),  # noqa: SLF001
        ),
    ]
    router = mod.RouterRowSink(routes=routes, failure_policy="all_fail", workbook_resources=())
    router.write_row({"a": 1})
    assert s1.rows == []
    assert s2.rows == []


def test_router_row_sink_raises_output_target_write_error_on_write_exception() -> None:
    router = mod.RouterRowSink(
        routes=(
            mod._RouteState(  # noqa: SLF001
                target_id="t_fail",
                sink=_FailingWriteRowSink(),
                predicate=None,
                is_primary=True,
                output_path=None,
                sheet_name=None,
                output_counter=mod._RowCounter(),  # noqa: SLF001
            ),
        ),
        failure_policy="all_fail",
        workbook_resources=(),
    )
    with pytest.raises(mod.ScalimOutputTargetWriteError, match="t_fail"):
        router.write_row({"a": 1})


def test_router_row_sink_write_row_after_close_raises() -> None:
    route = mod._RouteState(  # noqa: SLF001
        target_id="t",
        sink=_CollectingRowSink(),
        predicate=None,
        is_primary=True,
        output_path=None,
        sheet_name=None,
        output_counter=mod._RowCounter(),  # noqa: SLF001
    )
    router = mod.RouterRowSink(routes=(route,), failure_policy="all_fail", workbook_resources=())
    router.close()
    with pytest.raises(RuntimeError, match="RouterRowSink is closed"):
        router.write_row({"a": 1})


def test_router_row_sink_close_handles_close_errors_primary_only_non_primary_disables() -> None:
    route = mod._RouteState(  # noqa: SLF001
        target_id="t",
        sink=_FailingCloseSink(),
        predicate=None,
        is_primary=False,
        output_path=None,
        sheet_name=None,
        output_counter=mod._RowCounter(),  # noqa: SLF001
    )
    router = mod.RouterRowSink(routes=(route,), failure_policy="primary_only", workbook_resources=())
    router.close()
    assert route.disabled is True


def test_router_row_sink_close_raises_on_close_error_all_fail() -> None:
    route = mod._RouteState(  # noqa: SLF001
        target_id="t",
        sink=_FailingCloseSink(),
        predicate=None,
        is_primary=False,
        output_path=None,
        sheet_name=None,
        output_counter=mod._RowCounter(),  # noqa: SLF001
    )
    router = mod.RouterRowSink(routes=(route,), failure_policy="all_fail", workbook_resources=())
    with pytest.raises(mod.ScalimOutputTargetWriteError, match="t"):
        router.close()


def test_output_target_write_error_str_includes_target_id() -> None:
    err = mod.ScalimOutputTargetWriteError("t", Exception("x"))
    assert "t" in str(err)
