import csv
import threading
import time
from contextlib import suppress
from pathlib import Path
from typing import Any, Dict, List, Optional

import pytest

from scalim.workflow import resources as resources_mod
from scalim.workflow import resources_base as resources_base_mod
from scalim.workflow import resources_csv as resources_csv_mod
from scalim.workflow import resources_sheetbook as resources_sheetbook_mod
from scalim.workflow import resources_workbook as resources_workbook_mod
from scalim.events.catalog import (
    EVENT_DIAGNOSTIC_WARNING,
    EVENT_WORKFLOW_RESOURCE_DISCARD,
    EVENT_WORKFLOW_RESOURCE_WRITE,
)

_TIMEOUT_S = 5.0


class _Instrumentation:
    def __init__(self) -> None:
        self.events: List[Dict[str, Any]] = []

    def emit(self, event_type: str, payload: Any, meta: Optional[Dict[str, Any]] = None) -> None:
        self.events.append({"event_type": str(event_type), "payload": payload, "meta": dict(meta or {})})


class _RaiseOnDiagnosticWarningInstrumentation(_Instrumentation):
    def emit(self, event_type: str, payload: Any, meta: Optional[Dict[str, Any]] = None) -> None:
        if str(event_type) == EVENT_DIAGNOSTIC_WARNING:
            raise RuntimeError("emit boom")
        return super(_RaiseOnDiagnosticWarningInstrumentation, self).emit(event_type, payload, meta=meta)


class _JoinableTestManager(resources_base_mod.WorkflowResourceManagerBase):
    def __init__(
        self,
        *,
        instrumentation: Any,
        wait_diagnostics: Optional[resources_base_mod.WorkflowResourceWaitDiagnostics] = None,
        max_wait_s: Optional[float] = None,
    ) -> None:
        super(_JoinableTestManager, self).__init__(
            workflow_exec_id="wf",
            instrumentation=instrumentation,
            workbook_defs={},
            csv_defs={},
            sheetbook_defs={},
            wait_diagnostics=wait_diagnostics,
            max_wait_s=max_wait_s,
        )
        self.commits: List[str] = []
        self.discards: List[str] = []

    def _commit_workbook(self, plan: object) -> None:  # type: ignore[override]
        _ = plan
        self.commits.append("workbook")

    def _commit_csv(self, plan: object) -> None:  # type: ignore[override]
        _ = plan
        self.commits.append("csv")

    def _commit_sheetbook(self, plan: object) -> None:  # type: ignore[override]
        _ = plan
        self.commits.append("sheetbook")

    def _discard_workbook(self, plan: object, *, workflow_node_id: str, reason: str) -> None:  # type: ignore[override]
        _ = plan, workflow_node_id, reason
        self.discards.append("workbook")

    def _discard_csv(self, plan: object, *, workflow_node_id: str, reason: str) -> None:  # type: ignore[override]
        _ = plan, workflow_node_id, reason
        self.discards.append("csv")

    def _discard_sheetbook(self, plan: object, *, workflow_node_id: str, reason: str) -> None:  # type: ignore[override]
        _ = plan, workflow_node_id, reason
        self.discards.append("sheetbook")


def _write_csv(path: Path, rows: List[List[str]]) -> Path:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)
    return path


def test_release_write_lock_is_best_effort(tmp_path: Path) -> None:
    missing = tmp_path / "missing.lock"
    resources_base_mod._release_write_lock(missing)

    bad_dir = tmp_path / "lockdir"
    bad_dir.mkdir()
    resources_base_mod._release_write_lock(bad_dir)


def test_clone_exception_for_reraise_handles_fallbacks() -> None:
    clone = resources_base_mod._clone_exception_for_reraise

    first = clone(resources_mod.WorkflowWriteError("boom"))
    assert isinstance(first, BaseException)

    class _CopyFails(Exception):
        def __reduce_ex__(self, _protocol: int) -> object:
            raise RuntimeError("no copy")

    exc2 = _CopyFails("x")
    second = clone(exc2)
    assert isinstance(second, _CopyFails)
    assert second is not exc2
    assert second.args == exc2.args

    class _CopyAndCtorFail(Exception):
        def __reduce_ex__(self, _protocol: int) -> object:
            raise RuntimeError("no copy")

        def __init__(self, code: int) -> None:
            if not isinstance(code, int):
                raise TypeError("code must be int")
            super(_CopyAndCtorFail, self).__init__(code)

    exc3 = _CopyAndCtorFail(1)
    exc3.args = ("bad",)  # type: ignore[assignment]
    third = clone(exc3)
    assert third is exc3


def test_best_effort_close_write_only_workbook_worksheets_handles_variants() -> None:
    class _NoWorksheets:
        pass

    class _BadWorksheets:
        worksheets = 123

    class _NoClosed:
        pass

    class _ClosedTrue:
        closed = True

    class _ClosedFalse:
        def __init__(self) -> None:
            self.closed = False

        def close(self) -> None:
            self.closed = True

    ws = _ClosedFalse()

    class _Workbook:
        worksheets = [_NoClosed(), _ClosedTrue(), ws]

    resources_workbook_mod._best_effort_close_write_only_workbook_worksheets(_NoWorksheets())
    resources_workbook_mod._best_effort_close_write_only_workbook_worksheets(_BadWorksheets())
    resources_workbook_mod._best_effort_close_write_only_workbook_worksheets(_Workbook())
    assert ws.closed is True


def test_read_csv_header_errors(tmp_path: Path) -> None:
    with pytest.raises(resources_mod.WorkflowWriteError, match="Missing input CSV"):
        _ = resources_csv_mod._read_csv_header(str(tmp_path / "nope.csv"))

    empty = _write_csv(tmp_path / "empty.csv", [])
    with pytest.raises(resources_mod.WorkflowWriteError, match="empty"):
        _ = resources_csv_mod._read_csv_header(str(empty))

    invalid = _write_csv(tmp_path / "invalid.csv", [["", "ok"], ["x", "y"]])
    with pytest.raises(resources_mod.WorkflowWriteError, match="invalid header"):
        _ = resources_csv_mod._read_csv_header(str(invalid))


def test_read_csv_header_in_memory_invalid_header_raises() -> None:
    from scalim.sinks.sink_csv import InMemoryCsv

    bad = InMemoryCsv(header=["", "ok"], rows=[])
    with pytest.raises(resources_mod.WorkflowWriteError, match="<in_memory>"):
        _ = resources_csv_mod._read_csv_header(bad)


def test_resource_manager_unknown_resource_ids_raise(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a"]])
    with pytest.raises(resources_mod.WorkflowWriteError, match="Unknown workbook resource id"):
        manager.apply_workbook_sheet(
            workflow_node_id="n",
            workbook_id="nope",
            sheet="S",
            input_node_id="a",
            input_output_id="out",
            input_csv=str(csv_path),
            on_conflict="error",
        )

    with pytest.raises(resources_mod.WorkflowWriteError, match="Unknown csv resource id"):
        manager.apply_csv_append(
            workflow_node_id="n",
            csv_id="nope",
            input_node_id="a",
            input_output_id="out",
            input_csv=str(csv_path),
            header_policy="once",
            on_mismatch="error",
        )

    with pytest.raises(resources_mod.WorkflowWriteError, match="Unknown sheetbook resource id"):
        manager.apply_sheetbook_sheet(
            workflow_node_id="n",
            sheetbook_id="nope",
            sheet="S",
            input_node_id="a",
            input_output_id="out",
            input_csv=str(csv_path),
            on_conflict="error",
        )


def test_resource_manager_workbook_append_mismatch_error_warn_skip(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    workbook_path = tmp_path / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    mismatch = _write_csv(tmp_path / "b.csv", [["id", "other"], ["a1", "A1"]])

    manager.apply_workbook_append(
        workflow_node_id="n0",
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.WorkflowWriteError, match="Field alignment mismatch"):
        manager.apply_workbook_append(
            workflow_node_id="n1",
            workbook_id="report",
            sheet="S",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(mismatch),
            align_by="field_id",
            header_policy="once",
            on_mismatch="error",
        )
    manager.discard_all(workflow_node_id="n_discard", reason="test")

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(tmp_path / "warn.xlsx")},
        csv_defs={},
        sheetbook_defs={},
    )
    manager.apply_workbook_append(
        workflow_node_id="n0",
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_workbook_append(
        workflow_node_id="n1",
        workbook_id="report",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(mismatch),
        align_by="field_id",
        header_policy="once",
        on_mismatch="warn",
    )
    assert [e for e in instrumentation.events if e["event_type"] == EVENT_DIAGNOSTIC_WARNING]
    manager.discard_all(workflow_node_id="n_discard", reason="test")

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(tmp_path / "skip.xlsx")},
        csv_defs={},
        sheetbook_defs={},
    )
    manager.apply_workbook_append(
        workflow_node_id="n0",
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_workbook_append(
        workflow_node_id="n1",
        workbook_id="report",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(mismatch),
        align_by="field_id",
        header_policy="once",
        on_mismatch="skip",
    )
    skip_events = [e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_WRITE and e["payload"].action == "skip"]
    assert skip_events
    manager.discard_all(workflow_node_id="n_discard", reason="test")


def test_resource_manager_csv_append_mismatch_error_and_discard(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    csv_path = tmp_path / "merged.csv"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={"merged": str(csv_path)},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    mismatch = _write_csv(tmp_path / "b.csv", [["id", "other"], ["a1", "A1"]])

    manager.apply_csv_append(
        workflow_node_id="n0",
        csv_id="merged",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.WorkflowWriteError, match="Field alignment mismatch"):
        manager.apply_csv_append(
            workflow_node_id="n1",
            csv_id="merged",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(mismatch),
            header_policy="once",
            on_mismatch="error",
        )

    manager.discard_all(workflow_node_id="n_discard", reason="test")


def test_resource_manager_sheetbook_sheet_conflict_skip_error_overwrite(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
                export_write_lock=False,
            )
        },
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])

    manager.apply_sheetbook_sheet(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        on_conflict="error",
    )

    manager.apply_sheetbook_sheet(
        workflow_node_id="n1",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(second),
        on_conflict="skip",
    )
    assert [e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_WRITE and e["payload"].action == "skip"]

    with pytest.raises(resources_mod.WorkflowWriteError, match="Sheet conflict"):
        manager.apply_sheetbook_sheet(
            workflow_node_id="n2",
            sheetbook_id="sb",
            sheet="S",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(second),
            on_conflict="error",
        )

    manager.apply_sheetbook_sheet(
        workflow_node_id="n3",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="c",
        input_output_id="detail",
        input_csv=str(second),
        on_conflict="overwrite",
    )
    overwrite_events = [
        e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_WRITE and e["payload"].action == "overwrite"
    ]
    assert overwrite_events

    rows = list(
        manager.iter_sheetbook_sheet_rows(
            consumer_node_id="consumer",
            visible_producer_node_ids=frozenset(),
            producer_node_id="c",
            sheetbook_id="sb",
            sheet="S",
        )
    )
    assert rows == [{"id": "b1", "value": "B1"}]


def test_resource_manager_sheetbook_append_mismatch_error_warn_skip_and_budget(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
                export_write_lock=False,
            )
        },
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    mismatch = _write_csv(tmp_path / "b.csv", [["id", "other"], ["b1", "B1"]])

    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.WorkflowWriteError, match="Field alignment mismatch"):
        manager.apply_sheetbook_append(
            workflow_node_id="n1",
            sheetbook_id="sb",
            sheet="S",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(mismatch),
            align_by="field_id",
            header_policy="once",
            on_mismatch="error",
        )
    manager.discard_all(workflow_node_id="n_discard", reason="test")

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
                export_write_lock=False,
            )
        },
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n1",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(mismatch),
        align_by="header",
        header_policy="once",
        on_mismatch="warn",
    )
    assert [e for e in instrumentation.events if e["event_type"] == EVENT_DIAGNOSTIC_WARNING]
    manager.discard_all(workflow_node_id="n_discard", reason="test")

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
                export_write_lock=False,
            )
        },
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n1",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(mismatch),
        align_by="header",
        header_policy="once",
        on_mismatch="skip",
    )
    skip_events = [e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_WRITE and e["payload"].action == "skip"]
    assert skip_events
    manager.discard_all(workflow_node_id="n_discard", reason="test")

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=1,
                budget_max_total_cells=1000,
                export_path=None,
                export_write_lock=False,
            )
        },
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S1",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.WorkflowWriteError, match="max_sheets"):
        manager.apply_sheetbook_append(
            workflow_node_id="n1",
            sheetbook_id="sb",
            sheet="S2",
            input_node_id="b",
            input_output_id="detail",
            input_csv=str(first),
            align_by="field_id",
            header_policy="once",
            on_mismatch="error",
        )
    manager.discard_all(workflow_node_id="n_discard", reason="test")


def test_resource_manager_sheetbook_append_duplicate_producer_is_rejected(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
                export_write_lock=False,
            )
        },
    )
    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])

    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.WorkflowWriteError, match="Duplicate sheetbook write"):
        manager.apply_sheetbook_append(
            workflow_node_id="n1",
            sheetbook_id="sb",
            sheet="S",
            input_node_id="a",
            input_output_id="detail",
            input_csv=str(first),
            align_by="field_id",
            header_policy="once",
            on_mismatch="error",
        )


def test_workflow_resource_manager_emit_does_not_deadlock_on_reentry(tmp_path: Path) -> None:
    class _ReentrantInstrumentation:
        def __init__(self) -> None:
            self._reentered = False
            self.manager = None
            self.csv_input_path = None

        def emit(self, event_type: str, payload: Any, meta: Optional[Dict[str, Any]] = None) -> None:
            _ = event_type, meta
            if self._reentered:
                return
            if str(event_type) != EVENT_WORKFLOW_RESOURCE_WRITE:
                return
            if getattr(payload, "action", None) != "skip":
                return
            self._reentered = True
            self.manager.apply_csv_append(  # type: ignore[union-attr]
                workflow_node_id="reentry",
                csv_id="merged",
                input_node_id="r",
                input_output_id="out",
                input_csv=str(self.csv_input_path),  # type: ignore[arg-type]
                header_policy="once",
                on_mismatch="error",
            )

    instrumentation = _ReentrantInstrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(tmp_path / "report.xlsx")},
        csv_defs={"merged": str(tmp_path / "merged.csv")},
        sheetbook_defs={},
    )
    instrumentation.manager = manager

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    instrumentation.csv_input_path = str(first)

    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        on_conflict="error",
    )

    runner_done = threading.Event()
    runner_errors: List[BaseException] = []

    def _run() -> None:
        try:
            manager.apply_workbook_sheet(
                workflow_node_id="n1",
                workbook_id="report",
                sheet="S",
                input_node_id="b",
                input_output_id="detail",
                input_csv=str(first),
                on_conflict="skip",
            )
        except BaseException as exc:
            runner_errors.append(exc)
        finally:
            runner_done.set()

    runner = threading.Thread(target=_run, daemon=True)
    runner.start()
    if not runner_done.wait(timeout=_TIMEOUT_S):
        pytest.fail("WorkflowResourceManager.emit appears to be called under internal locks (reentry deadlock)")
    runner.join(timeout=_TIMEOUT_S)
    assert not runner_errors
    assert instrumentation._reentered is True


def test_workflow_resource_wait_diagnostics_rejects_invalid_values() -> None:
    with pytest.raises(ValueError, match=r"warn_after_s"):
        _ = resources_base_mod.WorkflowResourceWaitDiagnostics(enabled=True, warn_after_s=-1.0)
    with pytest.raises(ValueError, match=r"repeat_every_s"):
        _ = resources_base_mod.WorkflowResourceWaitDiagnostics(enabled=True, repeat_every_s=0.0)

    diagnostics = resources_base_mod.WorkflowResourceWaitDiagnostics(
        enabled=True,
        warn_after_s=0.0,
        repeat_every_s=0.01,
        capture_owner_callsite=True,
    )
    assert diagnostics.enabled is True
    assert diagnostics.warn_after_s == 0.0
    assert diagnostics.repeat_every_s == 0.01
    assert diagnostics.capture_owner_callsite is True


def test_workflow_resource_manager_rejects_invalid_max_wait_s() -> None:
    with pytest.raises(ValueError, match=r"max_wait_s"):
        _ = _JoinableTestManager(instrumentation=_Instrumentation(), max_wait_s=-1.0)


def test_capture_owner_callsite_returns_unknown_when_stack_has_no_frames(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(resources_base_mod.traceback, "extract_stack", lambda limit=12: [object()])
    assert resources_base_mod._capture_owner_callsite() == "(unknown)"  # noqa: SLF001


def test_wait_for_inflight_done_does_not_emit_warning_before_warn_after() -> None:
    instrumentation = _Instrumentation()
    diagnostics = resources_base_mod.WorkflowResourceWaitDiagnostics(
        enabled=True,
        warn_after_s=0.3,
        repeat_every_s=0.01,
        capture_owner_callsite=False,
    )
    manager = _JoinableTestManager(instrumentation=instrumentation, wait_diagnostics=diagnostics)

    inflight_state = resources_base_mod._InFlightCreate(owner_thread_ident=123, owner_callsite=None)  # noqa: SLF001

    def _set_done() -> None:
        time.sleep(0.05)
        inflight_state.done.set()

    t1 = threading.Thread(target=_set_done, daemon=True)
    t1.start()
    manager._wait_for_inflight_done(  # noqa: SLF001
        inflight_state=inflight_state,
        resource_type="workbook",
        resource_id="report",
        wait_kind="join",
    )
    t1.join(timeout=_TIMEOUT_S)
    assert not t1.is_alive()
    assert not [e for e in instrumentation.events if e["event_type"] == EVENT_DIAGNOSTIC_WARNING]


def test_joinable_wait_diagnostics_emits_warning_and_includes_required_fields() -> None:
    instrumentation = _Instrumentation()
    diagnostics = resources_base_mod.WorkflowResourceWaitDiagnostics(
        enabled=True,
        warn_after_s=0.01,
        repeat_every_s=None,
        capture_owner_callsite=True,
    )
    manager = _JoinableTestManager(instrumentation=instrumentation, wait_diagnostics=diagnostics)

    plans: Dict[str, object] = {}
    inflight: Dict[str, object] = {}
    owner_in_create = threading.Event()
    continue_create = threading.Event()
    errors: List[BaseException] = []

    def _create() -> object:
        owner_in_create.set()
        if not continue_create.wait(timeout=_TIMEOUT_S):
            raise RuntimeError("test timeout waiting to continue create")
        return object()

    def _owner() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="workbook",
                resource_id="report",
                plans=plans,
                inflight=inflight,  # type: ignore[arg-type]
                create_fn=_create,
                on_create=lambda _p: None,
            )
        except BaseException as exc:
            errors.append(exc)

    def _waiter() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="workbook",
                resource_id="report",
                plans=plans,
                inflight=inflight,  # type: ignore[arg-type]
                create_fn=_create,
                on_create=lambda _p: None,
            )
        except BaseException as exc:
            errors.append(exc)

    t1 = threading.Thread(target=_owner)
    t1.start()
    assert owner_in_create.wait(timeout=_TIMEOUT_S)

    t2 = threading.Thread(target=_waiter)
    t2.start()

    deadline = time.monotonic() + 2.0
    warnings: List[Dict[str, Any]] = []
    while time.monotonic() < deadline:
        warnings = [e for e in instrumentation.events if e["event_type"] == EVENT_DIAGNOSTIC_WARNING]
        if warnings:
            break
        time.sleep(0.01)
    assert warnings

    continue_create.set()
    t1.join(timeout=_TIMEOUT_S)
    t2.join(timeout=_TIMEOUT_S)
    assert not t1.is_alive()
    assert not t2.is_alive()
    assert errors == []

    payload = warnings[0]["payload"]
    lookup = getattr(payload, "lookup_key", None)
    assert isinstance(lookup, dict)
    assert lookup.get("resource_id") == "report"
    assert lookup.get("resource_type") == "workbook"
    assert isinstance(lookup.get("owner_thread_ident"), int)
    assert isinstance(lookup.get("waiter_thread_ident"), int)
    assert isinstance(lookup.get("wait_s"), float)
    assert lookup.get("owner_callsite")


def test_joinable_on_create_failure_propagates_error_to_waiters() -> None:
    manager = _JoinableTestManager(instrumentation=_Instrumentation())

    plans: Dict[str, object] = {}
    inflight: Dict[str, object] = {}
    owner_in_create = threading.Event()
    continue_create = threading.Event()
    owner_errors: List[BaseException] = []
    waiter_errors: List[BaseException] = []

    def _create() -> object:
        owner_in_create.set()
        if not continue_create.wait(timeout=_TIMEOUT_S):
            raise RuntimeError("test timeout waiting to continue create")
        return object()

    def _on_create(_plan: object) -> None:
        raise RuntimeError("on_create boom")

    def _owner() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="workbook",
                resource_id="report",
                plans=plans,
                inflight=inflight,  # type: ignore[arg-type]
                create_fn=_create,
                on_create=_on_create,
            )
        except BaseException as exc:
            owner_errors.append(exc)

    def _waiter() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="workbook",
                resource_id="report",
                plans=plans,
                inflight=inflight,  # type: ignore[arg-type]
                create_fn=_create,
                on_create=_on_create,
            )
        except BaseException as exc:
            waiter_errors.append(exc)

    t1 = threading.Thread(target=_owner)
    t1.start()
    assert owner_in_create.wait(timeout=_TIMEOUT_S)

    t2 = threading.Thread(target=_waiter)
    t2.start()

    continue_create.set()
    t1.join(timeout=_TIMEOUT_S)
    t2.join(timeout=_TIMEOUT_S)
    assert not t1.is_alive()
    assert not t2.is_alive()
    assert len(owner_errors) == 1
    assert len(waiter_errors) == 1
    assert isinstance(owner_errors[0], RuntimeError)
    assert isinstance(waiter_errors[0], RuntimeError)
    assert "on_create boom" in str(owner_errors[0])
    assert "on_create boom" in str(waiter_errors[0])
    assert plans == {}


def test_joinable_wait_diagnostics_repeats_warning() -> None:
    instrumentation = _Instrumentation()
    diagnostics = resources_base_mod.WorkflowResourceWaitDiagnostics(
        enabled=True,
        warn_after_s=0.01,
        repeat_every_s=0.01,
    )
    manager = _JoinableTestManager(instrumentation=instrumentation, wait_diagnostics=diagnostics)

    plans: Dict[str, object] = {}
    inflight: Dict[str, object] = {}
    owner_in_create = threading.Event()
    continue_create = threading.Event()
    errors: List[BaseException] = []

    def _create() -> object:
        owner_in_create.set()
        if not continue_create.wait(timeout=_TIMEOUT_S):
            raise RuntimeError("test timeout waiting to continue create")
        return object()

    def _owner() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="csv",
                resource_id="merged",
                plans=plans,
                inflight=inflight,  # type: ignore[arg-type]
                create_fn=_create,
                on_create=lambda _p: None,
            )
        except BaseException as exc:
            errors.append(exc)

    def _waiter() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="csv",
                resource_id="merged",
                plans=plans,
                inflight=inflight,  # type: ignore[arg-type]
                create_fn=_create,
                on_create=lambda _p: None,
            )
        except BaseException as exc:
            errors.append(exc)

    t1 = threading.Thread(target=_owner)
    t1.start()
    assert owner_in_create.wait(timeout=_TIMEOUT_S)

    t2 = threading.Thread(target=_waiter)
    t2.start()

    deadline = time.monotonic() + 2.0
    while time.monotonic() < deadline:
        warnings = [e for e in instrumentation.events if e["event_type"] == EVENT_DIAGNOSTIC_WARNING]
        if len(warnings) >= 2:
            break
        time.sleep(0.01)

    warnings = [e for e in instrumentation.events if e["event_type"] == EVENT_DIAGNOSTIC_WARNING]
    assert len(warnings) >= 2

    continue_create.set()
    t1.join(timeout=_TIMEOUT_S)
    t2.join(timeout=_TIMEOUT_S)
    assert not t1.is_alive()
    assert not t2.is_alive()
    assert errors == []


def test_wait_diagnostics_warning_falls_back_to_logger_on_emit_failure(caplog: pytest.LogCaptureFixture) -> None:
    instrumentation = _RaiseOnDiagnosticWarningInstrumentation()
    diagnostics = resources_base_mod.WorkflowResourceWaitDiagnostics(enabled=True, warn_after_s=0.01)
    manager = _JoinableTestManager(instrumentation=instrumentation, wait_diagnostics=diagnostics)

    caplog.set_level("WARNING", logger="scalim.workflow-resources")

    plans: Dict[str, object] = {}
    inflight: Dict[str, object] = {}
    owner_in_create = threading.Event()
    continue_create = threading.Event()
    errors: List[BaseException] = []

    def _create() -> object:
        owner_in_create.set()
        if not continue_create.wait(timeout=_TIMEOUT_S):
            raise RuntimeError("test timeout waiting to continue create")
        return object()

    def _owner() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="workbook",
                resource_id="report",
                plans=plans,
                inflight=inflight,  # type: ignore[arg-type]
                create_fn=_create,
                on_create=lambda _p: None,
            )
        except BaseException as exc:
            errors.append(exc)

    def _waiter() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="workbook",
                resource_id="report",
                plans=plans,
                inflight=inflight,  # type: ignore[arg-type]
                create_fn=_create,
                on_create=lambda _p: None,
            )
        except BaseException as exc:
            errors.append(exc)

    t1 = threading.Thread(target=_owner)
    t1.start()
    assert owner_in_create.wait(timeout=_TIMEOUT_S)

    t2 = threading.Thread(target=_waiter)
    t2.start()

    deadline = time.monotonic() + 2.0
    while time.monotonic() < deadline:
        if any("inflight wait slow" in r.message for r in caplog.records):
            break
        time.sleep(0.01)

    continue_create.set()
    t1.join(timeout=_TIMEOUT_S)
    t2.join(timeout=_TIMEOUT_S)
    assert not t1.is_alive()
    assert not t2.is_alive()
    assert errors == []
    assert any("inflight wait slow" in r.message for r in caplog.records)


def test_joinable_wait_timeout_raises_workflow_write_error_for_waiter_only() -> None:
    instrumentation = _Instrumentation()
    manager = _JoinableTestManager(
        instrumentation=instrumentation,
        max_wait_s=0.05,
    )

    plans: Dict[str, object] = {}
    inflight: Dict[str, object] = {}
    owner_in_create = threading.Event()
    continue_create = threading.Event()
    owner_errors: List[BaseException] = []
    waiter_errors: List[BaseException] = []

    def _create() -> object:
        owner_in_create.set()
        if not continue_create.wait(timeout=_TIMEOUT_S):
            raise RuntimeError("test timeout waiting to continue create")
        return object()

    def _owner() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="workbook",
                resource_id="report",
                plans=plans,
                inflight=inflight,  # type: ignore[arg-type]
                create_fn=_create,
                on_create=lambda _p: None,
            )
        except BaseException as exc:
            owner_errors.append(exc)

    def _waiter() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="workbook",
                resource_id="report",
                plans=plans,
                inflight=inflight,  # type: ignore[arg-type]
                create_fn=_create,
                on_create=lambda _p: None,
            )
        except BaseException as exc:
            waiter_errors.append(exc)

    t1 = threading.Thread(target=_owner)
    t1.start()
    assert owner_in_create.wait(timeout=_TIMEOUT_S)

    t2 = threading.Thread(target=_waiter)
    t2.start()
    t2.join(timeout=_TIMEOUT_S)
    assert not t2.is_alive()

    continue_create.set()
    t1.join(timeout=_TIMEOUT_S)
    assert not t1.is_alive()
    assert owner_errors == []
    assert len(waiter_errors) == 1
    assert isinstance(waiter_errors[0], resources_mod.WorkflowWriteError)
    assert "resource_id=report" in str(waiter_errors[0])
    assert "owner_thread_ident" in str(waiter_errors[0])


def test_commit_all_and_discard_all_drain_inflight_workbook_and_csv() -> None:
    manager = _JoinableTestManager(instrumentation=_Instrumentation())

    wb_started = threading.Event()
    wb_continue = threading.Event()
    csv_started = threading.Event()
    csv_continue = threading.Event()
    errors: List[BaseException] = []

    def _create_wb() -> object:
        wb_started.set()
        if not wb_continue.wait(timeout=_TIMEOUT_S):
            raise RuntimeError("test timeout waiting to continue workbook create")
        return object()

    def _create_csv() -> object:
        csv_started.set()
        if not csv_continue.wait(timeout=_TIMEOUT_S):
            raise RuntimeError("test timeout waiting to continue csv create")
        return object()

    def _owner_wb() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="workbook",
                resource_id="report",
                plans=manager._workbooks,  # noqa: SLF001
                inflight=manager._inflight_workbooks,  # noqa: SLF001
                create_fn=_create_wb,
                on_create=lambda _p: None,
            )
        except BaseException as exc:
            errors.append(exc)

    def _owner_csv() -> None:
        try:
            _ = manager._get_or_create_joinable_plan(  # noqa: SLF001
                resource_type="csv",
                resource_id="merged",
                plans=manager._csvs,  # noqa: SLF001
                inflight=manager._inflight_csvs,  # noqa: SLF001
                create_fn=_create_csv,
                on_create=lambda _p: None,
            )
        except BaseException as exc:
            errors.append(exc)

    owner1 = threading.Thread(target=_owner_wb)
    owner2 = threading.Thread(target=_owner_csv)
    owner1.start()
    owner2.start()
    assert wb_started.wait(timeout=_TIMEOUT_S)
    assert csv_started.wait(timeout=_TIMEOUT_S)

    committer = threading.Thread(target=manager.commit_all)
    committer.start()
    time.sleep(0.05)
    assert committer.is_alive()

    wb_continue.set()
    csv_continue.set()

    owner1.join(timeout=_TIMEOUT_S)
    owner2.join(timeout=_TIMEOUT_S)
    committer.join(timeout=_TIMEOUT_S)
    assert not owner1.is_alive()
    assert not owner2.is_alive()
    assert not committer.is_alive()
    assert errors == []
    assert manager.commits[:2] == ["workbook", "csv"]

    # discard_all should also drain (no inflight now, but cover path).
    manager.discard_all(workflow_node_id="n_discard", reason="test")
    assert "workbook" in manager.discards
    assert "csv" in manager.discards


def test_resource_manager_concurrent_first_workbook_write_joins_single_plan_and_acquires_lock_once(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from openpyxl import load_workbook

    lock_started = threading.Event()
    lock_continue = threading.Event()
    lock_calls: List[str] = []
    orig_acquire_write_lock = resources_workbook_mod.acquire_write_lock

    def _delayed_acquire_write_lock(output_path: str) -> Path:
        lock_calls.append(str(output_path))
        lock_started.set()
        if not lock_continue.wait(timeout=5.0):
            raise RuntimeError("test timeout waiting to continue acquire_write_lock")
        return orig_acquire_write_lock(output_path)

    monkeypatch.setattr(resources_workbook_mod, "acquire_write_lock", _delayed_acquire_write_lock)

    instrumentation = _Instrumentation()
    workbook_path = tmp_path / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])

    errors: List[BaseException] = []

    def _worker(*, workflow_node_id: str, sheet: str, csv_path: Path) -> None:
        try:
            manager.apply_workbook_sheet(
                workflow_node_id=str(workflow_node_id),
                workbook_id="report",
                sheet=str(sheet),
                input_node_id=str(workflow_node_id),
                input_output_id="detail",
                input_csv=str(csv_path),
                on_conflict="error",
            )
        except BaseException as exc:
            errors.append(exc)

    t1 = threading.Thread(target=lambda: _worker(workflow_node_id="n1", sheet="S1", csv_path=first))
    t1.start()
    assert lock_started.wait(timeout=5.0)

    t2 = threading.Thread(target=lambda: _worker(workflow_node_id="n2", sheet="S2", csv_path=second))
    t2.start()
    lock_continue.set()

    t1.join(timeout=5.0)
    t2.join(timeout=5.0)
    assert not t1.is_alive()
    assert not t2.is_alive()
    assert errors == []
    assert len(lock_calls) == 1

    manager.commit_all()

    wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
    try:
        assert "S1" in wb.sheetnames
        assert "S2" in wb.sheetnames
        assert list(wb["S1"].iter_rows(values_only=True)) == [("id", "value"), ("a1", "A1")]
        assert list(wb["S2"].iter_rows(values_only=True)) == [("id", "value"), ("b1", "B1")]
    finally:
        with suppress(Exception):
            wb.close()


def test_resource_manager_concurrent_first_csv_write_joins_single_plan_and_acquires_lock_once(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    lock_started = threading.Event()
    lock_continue = threading.Event()
    lock_calls: List[str] = []
    orig_acquire_write_lock = resources_csv_mod.acquire_write_lock

    def _delayed_acquire_write_lock(output_path: str) -> Path:
        lock_calls.append(str(output_path))
        lock_started.set()
        if not lock_continue.wait(timeout=5.0):
            raise RuntimeError("test timeout waiting to continue acquire_write_lock")
        return orig_acquire_write_lock(output_path)

    monkeypatch.setattr(resources_csv_mod, "acquire_write_lock", _delayed_acquire_write_lock)

    instrumentation = _Instrumentation()
    output_path = tmp_path / "merged.csv"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={"merged": str(output_path)},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])

    errors: List[BaseException] = []

    def _worker(*, workflow_node_id: str, csv_path: Path) -> None:
        try:
            manager.apply_csv_append(
                workflow_node_id=str(workflow_node_id),
                csv_id="merged",
                input_node_id=str(workflow_node_id),
                input_output_id="detail",
                input_csv=str(csv_path),
                header_policy="once",
                on_mismatch="error",
            )
        except BaseException as exc:
            errors.append(exc)

    t1 = threading.Thread(target=lambda: _worker(workflow_node_id="n1", csv_path=first))
    t1.start()
    assert lock_started.wait(timeout=5.0)

    t2 = threading.Thread(target=lambda: _worker(workflow_node_id="n2", csv_path=second))
    t2.start()
    lock_continue.set()

    t1.join(timeout=5.0)
    t2.join(timeout=5.0)
    assert not t1.is_alive()
    assert not t2.is_alive()
    assert errors == []
    assert len(lock_calls) == 1

    manager.commit_all()

    with output_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    assert rows[0] == ["id", "value"]
    assert {tuple(row) for row in rows[1:]} == {("a1", "A1"), ("b1", "B1")}


def test_resource_manager_concurrent_first_csv_write_lock_failure_wakes_waiters_and_acquires_lock_once(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    lock_started = threading.Event()
    lock_continue = threading.Event()
    lock_calls: List[str] = []

    def _fail_acquire_write_lock(output_path: str) -> Path:
        lock_calls.append(str(output_path))
        lock_started.set()
        if not lock_continue.wait(timeout=5.0):
            raise RuntimeError("test timeout waiting to continue acquire_write_lock")
        raise resources_mod.WorkflowWriteError("boom")

    monkeypatch.setattr(resources_csv_mod, "acquire_write_lock", _fail_acquire_write_lock)

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={"merged": str(tmp_path / "merged.csv")},
        sheetbook_defs={},
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])

    errors: List[BaseException] = []

    def _worker(*, workflow_node_id: str, csv_path: Path) -> None:
        try:
            manager.apply_csv_append(
                workflow_node_id=str(workflow_node_id),
                csv_id="merged",
                input_node_id=str(workflow_node_id),
                input_output_id="detail",
                input_csv=str(csv_path),
                header_policy="once",
                on_mismatch="error",
            )
        except BaseException as exc:
            errors.append(exc)

    t1 = threading.Thread(target=lambda: _worker(workflow_node_id="n1", csv_path=first))
    t1.start()
    assert lock_started.wait(timeout=5.0)

    t2 = threading.Thread(target=lambda: _worker(workflow_node_id="n2", csv_path=second))
    t2.start()
    lock_continue.set()

    t1.join(timeout=5.0)
    t2.join(timeout=5.0)
    assert not t1.is_alive()
    assert not t2.is_alive()
    assert len(lock_calls) == 1
    assert len(errors) == 2
    assert all(isinstance(err, resources_mod.WorkflowWriteError) for err in errors)
    assert all("boom" in str(err) for err in errors)


def test_resource_manager_concurrent_first_sheetbook_write_creates_single_plan(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    init_started = threading.Event()
    init_continue = threading.Event()
    plan_calls: List[int] = []
    orig_plan_cls = resources_sheetbook_mod.SheetBookPlan

    def _delayed_plan(*args: object, **kwargs: object) -> Any:  # noqa: ANN401
        plan_calls.append(1)
        init_started.set()
        if not init_continue.wait(timeout=5.0):
            raise RuntimeError("test timeout waiting to continue SheetBookPlan init")
        return orig_plan_cls(*args, **kwargs)

    monkeypatch.setattr(resources_sheetbook_mod, "SheetBookPlan", _delayed_plan)

    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
                export_write_lock=False,
            )
        },
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])

    errors: List[BaseException] = []

    def _worker(*, workflow_node_id: str, sheet: str, csv_path: Path) -> None:
        try:
            manager.apply_sheetbook_sheet(
                workflow_node_id=str(workflow_node_id),
                sheetbook_id="sb",
                sheet=str(sheet),
                input_node_id=str(workflow_node_id),
                input_output_id="detail",
                input_csv=str(csv_path),
                on_conflict="error",
            )
        except BaseException as exc:
            errors.append(exc)

    t1 = threading.Thread(target=lambda: _worker(workflow_node_id="n1", sheet="S1", csv_path=first))
    t1.start()
    assert init_started.wait(timeout=5.0)

    t2 = threading.Thread(target=lambda: _worker(workflow_node_id="n2", sheet="S2", csv_path=second))
    t2.start()
    init_continue.set()

    t1.join(timeout=5.0)
    t2.join(timeout=5.0)
    assert not t1.is_alive()
    assert not t2.is_alive()
    assert errors == []
    assert len(plan_calls) == 1

    plan = manager._sheetbooks["sb"]
    assert set(getattr(plan, "sheets", {}).keys()) == {"S1", "S2"}


def test_resource_manager_sheetbook_iter_rows_visibility_and_errors(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=None,
                export_write_lock=False,
            )
        },
    )

    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    second = _write_csv(tmp_path / "b.csv", [["id", "value"], ["b1", "B1"]])

    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n1",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="b",
        input_output_id="detail",
        input_csv=str(second),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )

    rows = list(
        manager.iter_sheetbook_sheet_rows(
            consumer_node_id="consumer",
            visible_producer_node_ids=frozenset(),
            producer_node_id="b",
            sheetbook_id="sb",
            sheet="S",
        )
    )
    assert rows == [{"id": "b1", "value": "B1"}]

    with pytest.raises(ValueError, match="Unknown sheetbook resource id"):
        _ = list(
            manager.iter_sheetbook_sheet_rows(
                consumer_node_id="consumer",
                visible_producer_node_ids=frozenset(),
                producer_node_id="a",
                sheetbook_id="nope",
                sheet="S",
            )
        )

    with pytest.raises(ValueError, match="Unknown sheetbook sheet"):
        _ = list(
            manager.iter_sheetbook_sheet_rows(
                consumer_node_id="consumer",
                visible_producer_node_ids=frozenset(),
                producer_node_id="a",
                sheetbook_id="sb",
                sheet="Other",
            )
        )

    with pytest.raises(ValueError, match="Unknown sheetbook ref node"):
        _ = list(
            manager.iter_sheetbook_sheet_rows(
                consumer_node_id="consumer",
                visible_producer_node_ids=frozenset(),
                producer_node_id="nope",
                sheetbook_id="sb",
                sheet="S",
            )
        )


def test_resource_manager_sheetbook_commit_import_error_and_save_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    export_path = tmp_path / "report.xlsx"
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(export_path),
                export_write_lock=True,
            )
        },
    )
    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )

    def _raise_import(*_args: object, **_kwargs: object) -> Any:  # noqa: ANN401
        raise ImportError("no openpyxl")

    monkeypatch.setattr(resources_workbook_mod, "require_optional_dependency", _raise_import)
    with pytest.raises(resources_mod.WorkflowWriteError, match="openpyxl"):
        manager.commit_all()
    assert not Path(str(export_path) + resources_base_mod._WRITE_LOCK_SUFFIX).exists()

    class _FakeWS:
        def append(self, row: List[object]) -> None:
            _ = row

    class _FakeWB:
        def __init__(self, write_only: bool = True) -> None:
            _ = write_only

        def create_sheet(self, name: str) -> _FakeWS:
            _ = name
            return _FakeWS()

        def save(self, path: Path) -> None:
            path.write_text("tmp", encoding="utf-8")
            raise RuntimeError("boom")

        def close(self) -> None:
            return

    export_path2 = tmp_path / "report2.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(export_path2),
                export_write_lock=False,
            )
        },
    )
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )

    temp_path = tmp_path / "temp.xlsx.tmp"
    monkeypatch.setattr(
        resources_workbook_mod,
        "require_optional_dependency",
        lambda *_args, **_kwargs: type("_M", (), {"Workbook": _FakeWB})(),
    )
    monkeypatch.setattr(resources_sheetbook_mod, "create_temp_path", lambda _path, _suffix: str(temp_path))
    with pytest.raises(resources_mod.WorkflowWriteError, match="Sheetbook export failed"):
        manager.commit_all()
    assert not temp_path.exists()


def test_resource_manager_sheetbook_discard_releases_lock_if_present(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(tmp_path / "report.xlsx"),
                export_write_lock=True,
            )
        },
    )
    first = _write_csv(tmp_path / "a.csv", [["id", "value"], ["a1", "A1"]])
    manager.apply_sheetbook_append(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(first),
        align_by="field_id",
        header_policy="once",
        on_mismatch="error",
    )

    plan = manager._sheetbooks["sb"]
    lock_path = tmp_path / "forced.lock"
    lock_path.write_text("lock", encoding="utf-8")
    plan.lock_path = lock_path

    manager.discard_all(workflow_node_id="n_discard", reason="test")
    assert not lock_path.exists()
    assert [e for e in instrumentation.events if e["event_type"] == EVENT_WORKFLOW_RESOURCE_DISCARD]


def test_resource_manager_commit_all_releases_locks_when_no_segments(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()

    workbook_path = tmp_path / "empty.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )
    _ = manager._get_or_create_workbook("report", workflow_node_id="n")  # noqa: SLF001
    assert Path(str(workbook_path) + resources_base_mod._WRITE_LOCK_SUFFIX).exists()
    manager.commit_all()
    assert not Path(str(workbook_path) + resources_base_mod._WRITE_LOCK_SUFFIX).exists()
    assert not workbook_path.exists()

    csv_output_path = tmp_path / "empty.csv"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=_Instrumentation(),
        workbook_defs={},
        csv_defs={"merged": str(csv_output_path)},
        sheetbook_defs={},
    )
    _ = manager._get_or_create_csv("merged", workflow_node_id="n")  # noqa: SLF001
    assert Path(str(csv_output_path) + resources_base_mod._WRITE_LOCK_SUFFIX).exists()
    manager.commit_all()
    assert not Path(str(csv_output_path) + resources_base_mod._WRITE_LOCK_SUFFIX).exists()
    assert not csv_output_path.exists()


def test_resource_manager_commit_workbook_escapes_excel_formulas_by_default(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    instrumentation = _Instrumentation()
    workbook_path = tmp_path / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(
        tmp_path / "input.csv",
        [
            ["=h", "eq", "plus", "minus", "at"],
            ["ok", "=1+1", "  +SUM(A1:A2)", "-1+2", "@X"],
        ],
    )
    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )
    manager.commit_all()

    wb = load_workbook(str(workbook_path), data_only=False)
    try:
        ws = wb["S"]
        assert ws["A1"].value == "'=h"
        assert ws["A2"].value == "ok"

        eq_value = ws["B2"].value
        eq_type = ws["B2"].data_type
        assert eq_type != "f"
        assert eq_value == "'=1+1"

        assert ws["C2"].value == "'  +SUM(A1:A2)"
        assert ws["D2"].value == "'-1+2"
        assert ws["E2"].value == "'@X"
    finally:
        wb.close()


def test_resource_manager_commit_workbook_allow_formulas_preserves_raw_strings(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    instrumentation = _Instrumentation()
    workbook_path = tmp_path / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        workbook_allow_formulas={"report": True},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(
        tmp_path / "input.csv",
        [
            ["=h", "eq", "plus", "minus", "at"],
            ["ok", "=1+1", "  +SUM(A1:A2)", "-1+2", "@X"],
        ],
    )
    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )
    manager.commit_all()

    wb = load_workbook(str(workbook_path), data_only=False)
    try:
        ws = wb["S"]
        assert ws["A1"].value == "=h"
        assert ws["A2"].value == "ok"

        eq_value = ws["B2"].value
        eq_type = ws["B2"].data_type
        assert eq_type == "f"
        assert eq_value == "=1+1"

        assert ws["C2"].value == "  +SUM(A1:A2)"
        assert ws["D2"].value == "-1+2"
        assert ws["E2"].value == "@X"
    finally:
        wb.close()


def test_resource_manager_commit_sheetbook_escapes_excel_formulas_by_default(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    instrumentation = _Instrumentation()
    export_path = tmp_path / "sheetbook.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(export_path),
                export_write_lock=False,
            )
        },
    )

    csv_path = _write_csv(
        tmp_path / "input.csv",
        [
            ["=h", "eq", "plus", "minus", "at"],
            ["ok", "=1+1", "  +SUM(A1:A2)", "-1+2", "@X"],
        ],
    )
    manager.apply_sheetbook_sheet(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )
    manager.commit_all()

    wb = load_workbook(str(export_path), data_only=False)
    try:
        ws = wb["S"]
        assert ws["A1"].value == "'=h"
        assert ws["A2"].value == "ok"

        eq_value = ws["B2"].value
        eq_type = ws["B2"].data_type
        assert eq_type != "f"
        assert eq_value == "'=1+1"

        assert ws["C2"].value == "'  +SUM(A1:A2)"
        assert ws["D2"].value == "'-1+2"
        assert ws["E2"].value == "'@X"
    finally:
        wb.close()


def test_resource_manager_commit_sheetbook_allow_formulas_preserves_raw_strings(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    instrumentation = _Instrumentation()
    export_path = tmp_path / "sheetbook.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={},
        sheetbook_defs={
            "sb": resources_mod.SheetBookDef(
                resource_id="sb",
                budget_max_sheets=4,
                budget_max_total_cells=1000,
                export_path=str(export_path),
                export_write_lock=False,
                export_allow_formulas=True,
            )
        },
    )

    csv_path = _write_csv(
        tmp_path / "input.csv",
        [
            ["=h", "eq", "plus", "minus", "at"],
            ["ok", "=1+1", "  +SUM(A1:A2)", "-1+2", "@X"],
        ],
    )
    manager.apply_sheetbook_sheet(
        workflow_node_id="n0",
        sheetbook_id="sb",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )
    manager.commit_all()

    wb = load_workbook(str(export_path), data_only=False)
    try:
        ws = wb["S"]
        assert ws["A1"].value == "=h"
        assert ws["A2"].value == "ok"

        eq_value = ws["B2"].value
        eq_type = ws["B2"].data_type
        assert eq_type == "f"
        assert eq_value == "=1+1"

        assert ws["C2"].value == "  +SUM(A1:A2)"
        assert ws["D2"].value == "-1+2"
        assert ws["E2"].value == "@X"
    finally:
        wb.close()


def test_resource_manager_commit_workbook_handles_missing_sheet_plan(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    workbook_path = tmp_path / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a1"]])
    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )

    plan = manager._workbooks["report"]  # noqa: SLF001
    plan.sheet_order.insert(0, "MISSING")

    manager.commit_all()
    assert workbook_path.exists()
    assert not Path(str(workbook_path) + resources_base_mod._WRITE_LOCK_SUFFIX).exists()


def test_resource_manager_commit_workbook_commit_failure_raises(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    output_dir = tmp_path / "out"
    output_dir.mkdir()

    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(output_dir)},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a1"]])
    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )

    with pytest.raises(resources_mod.WorkflowWriteError, match="Workbook commit failed"):
        manager.commit_all()
    assert not Path(str(output_dir) + resources_base_mod._WRITE_LOCK_SUFFIX).exists()


def test_resource_manager_commit_csv_commit_failure_raises(tmp_path: Path) -> None:
    instrumentation = _Instrumentation()
    output_dir = tmp_path / "csv_out"
    output_dir.mkdir()

    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={},
        csv_defs={"merged": str(output_dir)},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a1"]])
    manager.apply_csv_append(
        workflow_node_id="n0",
        csv_id="merged",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        header_policy="once",
        on_mismatch="error",
    )
    with pytest.raises(resources_mod.WorkflowWriteError, match="CSV commit failed"):
        manager.commit_all()
    assert not Path(str(output_dir) + resources_base_mod._WRITE_LOCK_SUFFIX).exists()


def test_resource_manager_commit_workbook_missing_openpyxl_releases_lock(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    instrumentation = _Instrumentation()
    workbook_path = tmp_path / "report.xlsx"
    manager = resources_mod.WorkflowResourceManager(
        workflow_exec_id="wf",
        instrumentation=instrumentation,
        workbook_defs={"report": str(workbook_path)},
        csv_defs={},
        sheetbook_defs={},
    )

    csv_path = _write_csv(tmp_path / "input.csv", [["id"], ["a1"]])
    manager.apply_workbook_sheet(
        workflow_node_id="n0",
        workbook_id="report",
        sheet="S",
        input_node_id="a",
        input_output_id="detail",
        input_csv=str(csv_path),
        on_conflict="error",
    )

    def _raise_import_error(*args: object, **kwargs: object) -> object:
        raise ImportError("missing openpyxl")

    monkeypatch.setattr(resources_workbook_mod, "require_optional_dependency", _raise_import_error)
    with pytest.raises(resources_mod.WorkflowWriteError, match="missing openpyxl"):
        manager.commit_all()
    assert not Path(str(workbook_path) + resources_base_mod._WRITE_LOCK_SUFFIX).exists()
    assert not workbook_path.exists()
