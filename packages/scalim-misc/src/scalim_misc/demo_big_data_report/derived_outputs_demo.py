"""多输出组合/派生汇总示例的共享执行与对拍逻辑."""

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Tuple

from scalim.execution.output_composition import (
    AggMetricSpec,
    AuditSheetSpec,
    DerivedGroupBySpec,
    DerivedOutputTargetSpec,
    MetaSheetSpec,
    OutputCompositionSpec,
    OutputTargetSpec,
)
from scalim.execution.run_ir import ExecutionRequest, ExportLayout, OutputSpec, export_layout_from_demand_ir, run_ir

from .cases import build_test_config_small
from .loaders import get_config, set_config
from .shared import build_ecommerce_model
from .verification import VerificationResult, verify_scalim_output

DETAIL_FIELDS: Tuple[str, ...] = (
    "order_id",
    "customer_name",
    "payment_method_name",
    "order_amount",
    "profit",
    "final_price",
)

SUMMARY_FIELDS: Tuple[str, ...] = (
    "payment_method_name",
    "order_cnt",
    "sum_amount",
    "sum_profit",
    "rank",
)

_SHEET_DETAIL = "Detail"
_SHEET_SUMMARY = "Summary"
_SHEET_META = "Meta"
_SHEET_AUDIT = "Audit"
_DECIMAL_ZERO = Decimal(0)


@dataclass
class DerivedOutputsDemoResult:
    workbook_path: str
    outputs: Dict[str, str]
    sheet_names: List[str]
    total_rows: int
    detail_rows: List[Dict[str, Any]]
    summary_rows: List[Dict[str, Any]]
    expected_summary_rows: List[Dict[str, Any]]
    detail_verification: VerificationResult
    summary_ok: bool
    summary_message: str

    @property
    def passed(self) -> bool:
        return self.detail_verification.passed and self.summary_ok

    def raise_if_failed(self) -> None:
        if self.detail_verification.passed and self.summary_ok:
            return
        msg = "多输出组合示例对拍失败:\n{}\n{}".format(self.detail_verification.summary, self.summary_message)
        raise AssertionError(msg)


def run_derived_outputs_demo(output_path: str) -> DerivedOutputsDemoResult:
    workbook_path = str(output_path)
    prev_config = get_config()
    set_config(build_test_config_small())
    try:
        demand_ir = build_ecommerce_model()

        detail_layout = export_layout_from_demand_ir(demand_ir, DETAIL_FIELDS)
        summary_layout = ExportLayout(field_ids=SUMMARY_FIELDS, header_names=None)

        composition = OutputCompositionSpec(
            targets=(
                OutputTargetSpec(
                    target_id="detail",
                    layout=detail_layout,
                    output=OutputSpec(
                        format="excel",
                        path=workbook_path,
                        streaming=True,
                        include_header=True,
                        sheet_name=_SHEET_DETAIL,
                    ),
                    is_primary=True,
                ),
            ),
            derived_targets=(
                DerivedOutputTargetSpec(
                    target_id="summary_by_payment",
                    derived=DerivedGroupBySpec(
                        group_by=("payment_method_name",),
                        metrics=(
                            AggMetricSpec(out_field_id="order_cnt", op="count", field_id="order_id"),
                            AggMetricSpec(out_field_id="sum_amount", op="sum", field_id="order_amount"),
                            AggMetricSpec(out_field_id="sum_profit", op="sum", field_id="profit"),
                        ),
                        rank_by="sum_profit",
                        rank_order="desc",
                        max_groups=100,
                    ),
                    output_layout=summary_layout,
                    output=OutputSpec(
                        format="excel",
                        path=workbook_path,
                        streaming=True,
                        include_header=True,
                        sheet_name=_SHEET_SUMMARY,
                    ),
                ),
            ),
            meta_sheet=MetaSheetSpec(
                target_id="meta",
                output=OutputSpec(format="excel", path=workbook_path, streaming=True, include_header=True),
                sheet_name=_SHEET_META,
            ),
            audit_sheet=AuditSheetSpec(
                target_id="audit",
                output=OutputSpec(format="excel", path=workbook_path, streaming=True, include_header=True),
                sheet_name=_SHEET_AUDIT,
            ),
            failure_policy="all_fail",
        )

        result = run_ir(
            demand_ir,
            ExecutionRequest(
                export_layout=detail_layout,
                output=OutputSpec(path=None),
                sink=None,
                output_composition=composition,
                parallel_mode="seq",
                batch_size=10,
            ),
        )

        detail_rows = _read_sheet_rows_as_dicts(workbook_path, _SHEET_DETAIL)
        summary_rows = _read_sheet_rows_as_dicts(workbook_path, _SHEET_SUMMARY)
        expected_summary_rows = _build_expected_summary_rows(detail_rows)
        detail_verification = verify_scalim_output(detail_rows, fields_to_check=list(DETAIL_FIELDS))
        summary_ok, summary_message = _compare_summary_rows(summary_rows, expected_summary_rows)

        demo_result = DerivedOutputsDemoResult(
            workbook_path=workbook_path,
            outputs=dict(result.outputs or {}),
            sheet_names=_read_workbook_sheet_names(workbook_path),
            total_rows=int(result.total_rows),
            detail_rows=detail_rows,
            summary_rows=summary_rows,
            expected_summary_rows=expected_summary_rows,
            detail_verification=detail_verification,
            summary_ok=summary_ok,
            summary_message=summary_message,
        )
        demo_result.raise_if_failed()
        return demo_result
    finally:
        set_config(prev_config)


def _read_workbook_sheet_names(path: str) -> List[str]:
    from openpyxl import load_workbook  # noqa: PLC0415

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _read_sheet_rows_as_dicts(path: str, sheet_name: str) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook  # noqa: PLC0415

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not rows:
        return []

    header = [str(item) if item is not None else "" for item in rows[0]]
    payload: List[Dict[str, Any]] = []
    for row in rows[1:]:
        payload.append({header[idx]: row[idx] if idx < len(row) else None for idx in range(len(header))})
    return payload


def _build_expected_summary_rows(detail_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in detail_rows:
        group_key = str(row.get("payment_method_name") or "")
        bucket = grouped.setdefault(
            group_key,
            {
                "payment_method_name": group_key,
                "order_cnt": 0,
                "sum_amount": _DECIMAL_ZERO,
                "sum_profit": _DECIMAL_ZERO,
            },
        )
        bucket["order_cnt"] += 1
        bucket["sum_amount"] += _to_decimal(row.get("order_amount"))
        bucket["sum_profit"] += _to_decimal(row.get("profit"))

    ordered = list(grouped.values())
    ordered.sort(key=lambda item: (-_to_decimal(item["sum_profit"]), str(item["payment_method_name"])))
    for idx, row in enumerate(ordered, start=1):
        row["rank"] = idx
    return ordered


def _compare_summary_rows(actual_rows: List[Dict[str, Any]], expected_rows: List[Dict[str, Any]]) -> Tuple[bool, str]:
    if len(actual_rows) != len(expected_rows):
        msg = "汇总行数不一致: actual={} expected={}".format(len(actual_rows), len(expected_rows))
        return False, msg

    for idx, (actual, expected) in enumerate(zip(actual_rows, expected_rows), start=1):
        for field_name in SUMMARY_FIELDS:
            actual_value = actual.get(field_name)
            expected_value = expected.get(field_name)
            if field_name in ("sum_amount", "sum_profit"):
                if _to_decimal(actual_value) != _to_decimal(expected_value):
                    msg = "汇总第 {} 行字段 '{}' 不一致: actual={} expected={}".format(idx, field_name, actual_value, expected_value)
                    return False, msg
                continue
            if field_name in ("order_cnt", "rank"):
                if int(actual_value or 0) != int(expected_value or 0):
                    msg = "汇总第 {} 行字段 '{}' 不一致: actual={} expected={}".format(idx, field_name, actual_value, expected_value)
                    return False, msg
                continue
            if actual_value != expected_value:
                msg = "汇总第 {} 行字段 '{}' 不一致: actual={} expected={}".format(idx, field_name, actual_value, expected_value)
                return False, msg

    return True, "汇总 sheet 与明细手工聚合结果一致"


def _to_decimal(value: Any) -> Decimal:
    if value is None:
        return _DECIMAL_ZERO
    if isinstance(value, Decimal):
        return value
    try:
        dec = Decimal(str(value).strip())
    except (InvalidOperation, AttributeError, TypeError, ValueError):
        return _DECIMAL_ZERO
    if not dec.is_finite():
        return _DECIMAL_ZERO
    return dec


__all__ = [
    "DETAIL_FIELDS",
    "SUMMARY_FIELDS",
    "DerivedOutputsDemoResult",
    "run_derived_outputs_demo",
]
