"""派生聚合 set 口径示例的共享执行与对拍逻辑.

BREAKING (c10-remove-dedup-and-two-stage-derived):
- `DedupBySpec` / `DerivedDedupByGroupBySpec` / `TwoStageGroupBySpec` 已移除.
- 去重: 在 loader / 上游先去重, 或接受重复行后只用 `DerivedGroupBySpec`.
- 两阶段聚合: 用 workflow 两个 demand/run(中间表 → 再聚合), 不再单进程 `TwoStageGroupBySpec`.

本 demo 仅保留仍支持的 `count_distinct` + `DerivedGroupBySpec`.
"""

from dataclasses import dataclass
from typing import Any, Dict, List, Tuple

from scalim.execution import ExecutionRequest, ExportLayout, OutputSpec, export_layout_from_demand_ir, run_ir
from scalim.execution.output_composition import (
    AggMetricSpec,
    AuditSheetSpec,
    DerivedGroupBySpec,
    DerivedOutputTargetSpec,
    MetaSheetSpec,
    OutputCompositionSpec,
    OutputTargetSpec,
)

from .cases import build_test_config_small
from .loaders import get_config, set_config
from .shared import build_ecommerce_model, build_ecommerce_runtime_bindings

_SHEET_DETAIL = "Detail"
_SHEET_DISTINCT = "Distinct"
_SHEET_META = "Meta"
_SHEET_AUDIT = "Audit"


@dataclass
class DerivedSetAggregationsDemoResult:
    workbook_path: str
    sheet_names: List[str]
    detail_rows: List[Dict[str, Any]]
    distinct_rows: List[Dict[str, Any]]
    expected_distinct_rows: List[Dict[str, Any]]
    passed: bool
    message: str

    def raise_if_failed(self) -> None:
        if self.passed:
            return
        raise AssertionError(self.message)


def verify_derived_set_aggregations_workbook(workbook_path: str) -> "DerivedSetAggregationsDemoResult":
    """对拍验证 `run_ir(..., output_composition=...)` 写出的 workbook 内容(不执行引擎)."""
    workbook_path = str(workbook_path)
    detail_rows = _read_sheet_rows_as_dicts(workbook_path, _SHEET_DETAIL)
    distinct_rows = _read_sheet_rows_as_dicts(workbook_path, _SHEET_DISTINCT)

    expected_distinct_rows = _expected_distinct_by_payment(detail_rows)

    ok1, msg1 = _compare_rows(distinct_rows, expected_distinct_rows, ("payment_method_name", "customer_cnt"))

    passed = bool(ok1)
    message = str(msg1)

    return DerivedSetAggregationsDemoResult(
        workbook_path=workbook_path,
        sheet_names=_read_workbook_sheet_names(workbook_path),
        detail_rows=detail_rows,
        distinct_rows=distinct_rows,
        expected_distinct_rows=expected_distinct_rows,
        passed=passed,
        message=message,
    )


def run_derived_set_aggregations_demo(output_path: str) -> DerivedSetAggregationsDemoResult:
    workbook_path = str(output_path)
    prev_config = get_config()
    set_config(build_test_config_small())
    try:
        demand_ir = build_ecommerce_model()
        runtime_bindings = build_ecommerce_runtime_bindings()

        detail_fields: Tuple[str, ...] = (
            "order_id",
            "customer_name",
            "product_name",
            "payment_method_name",
        )
        detail_layout = export_layout_from_demand_ir(demand_ir, detail_fields)

        distinct_layout = ExportLayout(field_ids=("payment_method_name", "customer_cnt"), header_names=None)

        composition = OutputCompositionSpec(
            targets=(
                OutputTargetSpec(
                    target_id="detail",
                    layout=detail_layout,
                    output=OutputSpec(
                        format="excel",
                        path=workbook_path,
                        streaming=True,
                        include_header=True,
                        sheet_name=_SHEET_DETAIL,
                    ),
                    is_primary=True,
                ),
            ),
            derived_targets=(
                DerivedOutputTargetSpec(
                    target_id="distinct_by_payment",
                    derived=DerivedGroupBySpec(
                        group_by=("payment_method_name",),
                        metrics=(AggMetricSpec(out_field_id="customer_cnt", op="count_distinct", field_id="customer_name"),),
                    ),
                    output_layout=distinct_layout,
                    output=OutputSpec(
                        format="excel",
                        path=workbook_path,
                        streaming=True,
                        include_header=True,
                        sheet_name=_SHEET_DISTINCT,
                    ),
                ),
            ),
            meta_sheet=MetaSheetSpec(
                target_id="meta",
                output=OutputSpec(format="excel", path=workbook_path, streaming=True, include_header=True),
                sheet_name=_SHEET_META,
            ),
            audit_sheet=AuditSheetSpec(
                target_id="audit",
                output=OutputSpec(format="excel", path=workbook_path, streaming=True, include_header=True),
                sheet_name=_SHEET_AUDIT,
            ),
            failure_policy="all_fail",
        )

        _ = run_ir(
            demand_ir,
            ExecutionRequest(
                export_layout=detail_layout,
                output=OutputSpec(path=None),
                sink=None,
                output_composition=composition,
                runtime_bindings=runtime_bindings,
                parallel_mode="seq",
                batch_size=10,
            ),
        )

        result = verify_derived_set_aggregations_workbook(workbook_path)
        result.raise_if_failed()
        return result
    finally:
        set_config(prev_config)


def _read_workbook_sheet_names(path: str) -> List[str]:
    from openpyxl import load_workbook  # noqa: PLC0415

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _read_sheet_rows_as_dicts(path: str, sheet_name: str) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook  # noqa: PLC0415

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not rows:
        return []

    header = [str(item) if item is not None else "" for item in rows[0]]
    payload: List[Dict[str, Any]] = []
    for row in rows[1:]:
        payload.append({header[idx]: row[idx] if idx < len(row) else None for idx in range(len(header))})
    return payload


def _expected_distinct_by_payment(detail_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups: Dict[str, set] = {}
    for row in detail_rows:
        payment = str(row.get("payment_method_name") or "")
        customer_name = row.get("customer_name")
        groups.setdefault(payment, set()).add(customer_name)
    out = [{"payment_method_name": k, "customer_cnt": len(v)} for k, v in groups.items()]
    out.sort(key=lambda r: str(r.get("payment_method_name") or ""))
    return out


def _compare_rows(actual: List[Dict[str, Any]], expected: List[Dict[str, Any]], fields: Tuple[str, ...]) -> Tuple[bool, str]:
    if len(actual) != len(expected):
        return False, "row count mismatch: actual={} expected={}".format(len(actual), len(expected))
    for idx, (a, e) in enumerate(zip(actual, expected), start=1):
        for f in fields:
            if a.get(f) != e.get(f):
                return False, "row {} field '{}' mismatch: actual={} expected={}".format(idx, f, a.get(f), e.get(f))
    return True, "ok: {}".format(",".join(fields))
