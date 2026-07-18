"""FieldValue temporal types: loader + workbook oracle (notebook / just examples SSOT).

对应 change: `c0-add-field-value-datetime`.
"""

from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import openpyxl

from scalim_misc.examples._types import EXAMPLE_KIND_ORACLE, ExampleResult

TEMPORAL_FIELD_IDS: Tuple[str, ...] = (
    "order_id",
    "created",
    "day",
    "clock",
    "span",
    "label",
)
_ORDER_ID = 1001


def load_table_temporal_values() -> Iterable[Mapping[str, object]]:
    """One-row fixture covering openpyxl `TIME_TYPES` plus numeric/str regression."""
    return [
        {
            "order_id": _ORDER_ID,
            # force-en: intentional naive datetime for Excel date cell demo
            "created": datetime(2024, 1, 2, 3, 4, 5),  # noqa: DTZ001
            "day": date(2024, 1, 2),
            "clock": time(3, 4, 5),
            "span": timedelta(days=1, seconds=30),
            "label": "ok",
        }
    ]


def _check_temporal_cells(by_name: Mapping[str, Any]) -> List[str]:
    checks: List[str] = []
    order_id = by_name["order_id"]
    if order_id.value != _ORDER_ID or order_id.data_type != "n":
        checks.append("order_id not typed numeric")
    if by_name["label"].value != "ok":
        checks.append("label mismatch")

    created = by_name["created"]
    if created.data_type != "d" or not isinstance(created.value, datetime) or isinstance(created.value, str):
        checks.append("created not datetime/d")

    day = by_name["day"]
    if day.data_type != "d" or not isinstance(day.value, (datetime, date)) or isinstance(day.value, str):
        checks.append("day not date-like/d")

    clock = by_name["clock"]
    if clock.data_type != "d" or not isinstance(clock.value, time):
        checks.append("clock not time/d")

    span = by_name["span"]
    if span.data_type != "d" or not isinstance(span.value, timedelta):
        checks.append("span not timedelta/d")
    return checks


def inspect_temporal_workbook_sheet(
    book_path: Path,
    *,
    sheet: str = "Types",
    field_ids: Sequence[str] = TEMPORAL_FIELD_IDS,
) -> Tuple[bool, str, Dict[str, Any]]:
    """Read xlsx and assert temporal columns stay Excel date cells (not `str`)."""
    details: Dict[str, Any] = {"book_path": str(book_path), "sheet": sheet}
    if not book_path.exists():
        return False, "book missing: {}".format(book_path), details

    wb = openpyxl.load_workbook(str(book_path), data_only=False)
    try:
        if sheet not in wb.sheetnames:
            return False, "sheet missing: {}".format(sheet), details
        ws = wb[sheet]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        details["headers"] = list(headers)
        if list(headers) != list(field_ids):
            return False, "header mismatch: {} != {}".format(headers, list(field_ids)), details

        cells = list(next(ws.iter_rows(min_row=2, max_row=2)))
        by_name = {str(headers[i]): cells[i] for i in range(len(headers))}
        checks = _check_temporal_cells(by_name)
        details["cell_types"] = {k: {"value_type": type(v.value).__name__, "data_type": v.data_type} for k, v in by_name.items()}
        if checks:
            return False, "; ".join(checks), details
        return True, "temporal FieldValues preserved as Excel date cells", details
    finally:
        wb.close()


def verify_temporal_field_values_example(
    *,
    example_id: str,
    book_path: Optional[Path],
    errors: Sequence[object] = (),
) -> ExampleResult:
    if errors:
        return ExampleResult(
            example_id=example_id,
            passed=False,
            kind=EXAMPLE_KIND_ORACLE,
            summary="workflow errors: {}".format(len(errors)),
            details={"errors": list(errors)},
        )
    if book_path is None:
        return ExampleResult(
            example_id=example_id,
            passed=False,
            kind=EXAMPLE_KIND_ORACLE,
            summary="book_path is None",
            details={},
        )
    ok, summary, details = inspect_temporal_workbook_sheet(book_path)
    return ExampleResult(
        example_id=example_id,
        passed=ok,
        kind=EXAMPLE_KIND_ORACLE,
        summary=summary,
        details=details,
    )
