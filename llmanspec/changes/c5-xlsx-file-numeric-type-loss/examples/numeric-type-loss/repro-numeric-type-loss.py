#!/usr/bin/env python3
"""
MVP: 复现 xlsx_file 路径的数字类型丢失

问题
  xlsx_file book 在 workflow 中经过 CSV 中间层,所有 Python 数字(int/float/Decimal)
  被 stringify,导致最终 Excel 中数字列表现为字符串。

对比
  - xlsx_file: 通过 InMemoryCsv → resources_workbook._iter_workbook_sheet_rows → openpyxl
              数值被 str() 转换,输出为文本单元格
  - xlsx_memory: 通过 InMemoryRows → resources_sheetbook._write_sheetbook_plan → openpyxl
                 数值保留原始类型,输出为数字单元格

用法
  uv run python3 llmanspec/changes/c5-xlsx-file-numeric-type-loss/examples/numeric-type-loss/repro-numeric-type-loss.py

输出
  - 控制台: 各阶段值的类型信息
  - .tmp/repro/numeric_type_loss/output/: xlsx_file.xlsx + xlsx_memory.xlsx
"""

import os
import sys
import tempfile
from pathlib import Path

# ---------------------------------------------------------------------------
# 确保 scalim 可导入
# ---------------------------------------------------------------------------
_REPO_ROOT = Path(__file__).resolve()
for _ in range(10):
    if (_REPO_ROOT / "pyproject.toml").exists():
        break
    _REPO_ROOT = _REPO_ROOT.parent
else:
    raise RuntimeError("Cannot find repo root (pyproject.toml)")
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

# ---------------------------------------------------------------------------
# 导入 scalim 组件
# ---------------------------------------------------------------------------
from scalim.sinks.memory import InMemoryCsv, InMemoryCsvSink
from scalim.sinks.rows import InMemoryRows, InMemoryRowsSink, in_memory_rows_to_in_memory_csv
from scalim.workflow.resources_csv import iter_csv_rows
from scalim.workflow.resources_workbook import (
    SheetPlan,
    _iter_workbook_sheet_rows,
)
from scalim.workflow.resources_sheetbook import (
    SheetBookSheetPlan,
    SheetBookSegment,
    _write_sheetbook_plan_to_openpyxl_workbook,
    SheetBookPlan,
)
from scalim.workflow.resources_csv import AppendSegment
from scalim._internal.utils.excel import escape_excel_formula

# ---------------------------------------------------------------------------
# Helper: openpyxl lazy import
# ---------------------------------------------------------------------------
def _openpyxl_Workbook():
    import openpyxl as _m
    return _m.Workbook

def _load_workbook(path):
    import openpyxl as _m
    return _m.load_workbook(path)

# ===================================================================
# 阶段 1: 模拟 demand 执行 → 产生中间产物
# ===================================================================
print("=" * 60)
print("阶段 1: demand 执行 → 中间产物")
print("=" * 60)

# 原始数据（demand 计算完成后,行数据中包含 Python 数字）
raw_rows = [
    {"name": "商品A", "amount": 123.45, "rate": 0.95, "quantity": 10},
    {"name": "商品B", "amount": 678.90, "rate": 0.85, "quantity": 3},
    {"name": "商品C", "amount": 999.99, "rate": 0.50, "quantity": 0},
]
print("\n原始数据类型:")
for r in raw_rows:
    for k, v in r.items():
        print(f"  {k}: {v!r} ({type(v).__name__})")
    print()

# ---- 1a: InMemoryCsvSink（xlsx_file 路径） ----
csv_sink = InMemoryCsvSink(
    field_names=["name", "amount", "rate", "quantity"],
    header_names=["名称", "金额", "比率", "数量"],
)
for row in raw_rows:
    csv_sink.write_row(row)
csv_artifact = csv_sink.to_artifact()  # InMemoryCsv

print("--- xlsx_file 中间产物 (InMemoryCsv) ---")
print(f"header: {csv_artifact.header}")
for row in csv_artifact.rows:
    for i, v in enumerate(row):
        print(f"  {csv_artifact.header[i]}: {v!r} ({type(v).__name__})")
    print()

# ---- 1b: InMemoryRowsSink（xlsx_memory 路径） ----
rows_sink = InMemoryRowsSink(field_ids=["name", "amount", "rate", "quantity"])
for row in raw_rows:
    rows_sink.write_row(row)
rows_artifact = rows_sink.to_artifact()  # InMemoryRows

print("--- xlsx_memory 中间产物 (InMemoryRows) ---")
print(f"header: {rows_artifact.header}")
for row in rows_artifact.rows:
    for i, v in enumerate(row):
        print(f"  {rows_artifact.header[i]}: {v!r} ({type(v).__name__})")
    print()

# ===================================================================
# 阶段 2: 模拟 workflow 资源管理器合成 xlsx
# ===================================================================
print("=" * 60)
print("阶段 2: 资源管理器合成 xlsx")
print("=" * 60)

OUTPUT_DIR = _REPO_ROOT / ".tmp" / "repro" / "numeric_type_loss" / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ---- 2a: xlsx_file 路径（通过 _iter_workbook_sheet_rows） ----
seg = AppendSegment(
    decl_order=0,
    input_csv=csv_artifact,  # InMemoryCsv（已 stringify）
    header_policy="once",
    mapping=[0, 1, 2, 3],
    on_mismatch="error",
    align_by="header",
    input_header=list(csv_artifact.header),
)
sheet_plan = SheetPlan(
    sheet="数据",
    baseline_header=list(csv_artifact.header),
    segments=[seg],
)

xlsx_file_path = str(OUTPUT_DIR / "xlsx_file_output.xlsx")
Workbook = _openpyxl_Workbook()
wb = Workbook(write_only=True)
ws = wb.create_sheet("数据")

# 此行读取逻辑 = resources_workbook._write_workbook_plan_to_openpyxl_workbook 的核心
for row in _iter_workbook_sheet_rows(sheet_plan, allow_formulas=True):
    ws.append(row)

wb.save(xlsx_file_path)
print(f"\nxlsx_file 已写入: {xlsx_file_path}")

# ---- 2b: xlsx_memory 路径（通过 _write_sheetbook_plan_to_openpyxl_workbook） ----
seg_mem = SheetBookSegment(
    producer_node_id="demand_A",
    decl_order=0,
    rows=rows_artifact.rows,  # List[List[FieldValue]]（保留类型）
    header_policy="once",
)
sheet_mem_plan = SheetBookSheetPlan(
    sheet="数据",
    baseline_header=list(rows_artifact.header),
    export_header=["名称", "金额", "比率", "数量"],
    segments=[seg_mem],
)
book_plan = SheetBookPlan(
    resource_id="test_book",
    budget_max_sheets=10,
    budget_max_total_cells=100000,
    export_path=str(OUTPUT_DIR / "xlsx_memory_output.xlsx"),
    export_allow_formulas=True,
    sheet_decl_order={"数据": 0},
    sheet_order=["数据"],
    sheets={"数据": sheet_mem_plan},
)

xlsx_memory_path = str(OUTPUT_DIR / "xlsx_memory_output.xlsx")
wb2 = Workbook(write_only=True)
_write_sheetbook_plan_to_openpyxl_workbook(wb2, book_plan)
wb2.save(xlsx_memory_path)
print(f"xlsx_memory 已写入: {xlsx_memory_path}")

# ===================================================================
# 阶段 3: 验证结果 — 读取 xlsx 并检查单元格类型
# ===================================================================
print("=" * 60)
print("阶段 3: 验证结果 — 单元格类型对比")
print("=" * 60)

def inspect_xlsx(path: str, label: str):
    print(f"\n--- {label} ---")
    wb = _load_workbook(path)
    ws = wb.active
    print(f"表头: {[c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]}")
    for row in ws.iter_rows(min_row=2, values_only=False):
        for cell in row:
            type_name = type(cell.value).__name__
            marker = " <<< 数字类型丢失!" if (isinstance(cell.value, str) and cell.column > 1) else ""
            print(f"  {cell.coordinate}: {cell.value!r} (type={type_name}, number_format={cell.number_format!r}){marker}")
    print()

inspect_xlsx(xlsx_file_path, "xlsx_file 输出")
inspect_xlsx(xlsx_memory_path, "xlsx_memory 输出")

# ===================================================================
# 阶段 4: 汇总
# ===================================================================
print("=" * 60)
print("汇总")
print("=" * 60)

# 读取 xlsx_file 的 amount 列第二个数据单元格
wb_f = _load_workbook(xlsx_file_path)
ws_f = wb_f.active
file_amount_cell = list(ws_f.iter_rows(min_row=2, max_row=2))[0][1]  # 第2行, B列

wb_m = _load_workbook(xlsx_memory_path)
ws_m = wb_m.active
mem_amount_cell = list(ws_m.iter_rows(min_row=2, max_row=2))[0][1]  # 第2行, B列

print(f"\nxlsx_file.amount  cell type: {type(file_amount_cell.value).__name__},  value={file_amount_cell.value!r}")
print(f"xlsx_memory.amount cell type: {type(mem_amount_cell.value).__name__}, value={mem_amount_cell.value!r}")
print()
if isinstance(file_amount_cell.value, str):
    print("✅ 确认: xlsx_file 路径数字类型丢失 (amount 列是字符串)")
else:
    print("❌ 意外: xlsx_file 路径保留了数字类型")
if isinstance(mem_amount_cell.value, (int, float)):
    print("✅ 确认: xlsx_memory 路径保留了数字类型")
else:
    print("❌ 意外: xlsx_memory 路径数字类型丢失")

print(f"\n详细输出路径: {OUTPUT_DIR}")
print("可直接打开 .xlsx 文件在 Excel 中查看单元格格式确认")
