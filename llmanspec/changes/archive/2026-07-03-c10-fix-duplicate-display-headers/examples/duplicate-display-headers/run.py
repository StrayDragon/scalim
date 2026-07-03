"""MVP: 重复展示列名导出回归 demo(独立可运行,脱敏)。

场景:
  底层 field_id 唯一,展示名(name)可重复(多个「人数」/「金额」指标块)。
  workflow 模式导出 xlsx_file / xlsx_memory / csv 三路径,验证数据按 field_id
  正确对齐,后续同名列不被首列值填充。

运行:
  cd <repo-root>
  uv run python llmanspec/changes/c10-fix-duplicate-display-headers/examples/duplicate-display-headers/run.py

期望(修复后):
  三路径表头均为 [人数, 金额, 人数, 金额](展示名,可重复)
  三路径数据均为  [7, 13111.26, 1, 10510.00](按 field_id 位置正确,不错位)

bug 复现(修复前):
  xlsx_file / csv 数据为 [7, 13111.26, 7, 13111.26](第 3/4 列被首列值填充)
  xlsx_memory 数据正确(本就不受影响)
"""

import csv
import sys
import tempfile
from pathlib import Path
from typing import List

from openpyxl import load_workbook

from scalim.dsl.yaml_dsl import WorkflowRunOptions, run_workflow
from scalim.dsl.yaml_dsl.runtime.contracts import (
    DemandDiagnosticsOverride,
    DemandRunOptions,
    DemandRunRuntimeOptions,
    DemandRunSecurityOptions,
    UNSET,
)

DEMO_DIR = Path(__file__).resolve().parent

# 展示名(可重复);底层 field_id 唯一,数据按 field_id 位置正确对齐。
# 第 3/4 列应为 1 / 10510(若被首列值填充则为 7 / 13111.26 → 错位 bug)。
_EXPECTED_HEADER = ["人数", "金额", "人数", "金额"]
_EXPECTED_DATA = ["7", "13111.26", "1", "10510.0"]


def _norm_cell(value: object) -> object:
    """归一化为可比较标量: 数字按 float 比较(str/int/float 等价),其余按原 str。

    workbook(`xlsx_file`) 路径经 `InMemoryCsv` 文本传输(值为 str),
    sheetbook(`xlsx_memory`) 路径保留原始数值类型;两者数值等价即可。
    """
    s = value if isinstance(value, str) else str(value)
    try:
        return float(s)
    except (TypeError, ValueError):
        return s


def _check_xlsx(path: Path, sheet: str, label: str) -> None:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        rows: List[List[object]] = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    assert len(rows) >= 2, "{}: missing rows: {}".format(label, rows)
    print("{} header: {}".format(label, rows[0]))
    print("{} data  : {}".format(label, rows[1]))
    assert list(rows[0]) == _EXPECTED_HEADER, "{}: header mismatch: {}".format(label, rows[0])
    got = [_norm_cell(x) for x in rows[1]]
    want = [_norm_cell(x) for x in _EXPECTED_DATA]
    assert got == want, "{}: data corrupted (duplicate display headers): {}".format(label, rows[1])


def _check_csv(path: Path, label: str) -> None:
    with path.open(encoding="utf-8") as handle:
        rows = list(csv.reader(handle))
    assert len(rows) >= 2, "{}: missing rows: {}".format(label, rows)
    print("{} header: {}".format(label, rows[0]))
    print("{} data  : {}".format(label, rows[1]))
    assert rows[0] == _EXPECTED_HEADER, "{}: header mismatch: {}".format(label, rows[0])
    assert rows[1] == _EXPECTED_DATA, "{}: data corrupted (duplicate display headers): {}".format(label, rows[1])


def _find_versioned(out_root: Path, pattern: str, label: str) -> Path:
    # 每个 book/file 各自有 output root(`path`),版本化在各自 root 下:
    #   <root>/versions/<version_id>/books|files/<filename>
    matches = sorted(out_root.glob(pattern))
    assert len(matches) == 1, "{}: expected 1 versioned output for {}, got {}".format(label, pattern, matches)
    return matches[0]


def main() -> int:
    sys.path.insert(0, str(DEMO_DIR))

    with tempfile.TemporaryDirectory(prefix="scalim-mvp-dup-headers-") as temp_dir:
        out_dir = Path(temp_dir).resolve()
        wf_copy = out_dir / "workflow.yaml"
        wf_copy.write_text((DEMO_DIR / "workflow.yaml").read_text(encoding="utf-8"), encoding="utf-8")
        (out_dir / "demand.yaml").write_text((DEMO_DIR / "demand.yaml").read_text(encoding="utf-8"), encoding="utf-8")

        result = run_workflow(
            str(wf_copy),
            options=WorkflowRunOptions(
                demand=DemandRunOptions(
                    security=DemandRunSecurityOptions(
                        allowed_modules=frozenset(["data_loader"]),
                    ),
                    runtime=DemandRunRuntimeOptions(
                        batch_size=10,
                        demand_diagnostics=DemandDiagnosticsOverride(validate_unique_field_names=False),
                    ),
                ),
            ),
        )
        errors = result.errors()
        if errors:
            print("workflow errors:", errors)
            return 1

        # 输出 `path` 相对声明该路径的 YAML 文件目录解析(此处 = 临时目录);
        # 每个 book/file 各自有 output root,版本化结构为
        #   <root>/versions/<version_id>/books|files/<filename>
        out_root = out_dir / "out"
        _check_xlsx(_find_versioned(out_root, "report_workbook/versions/*/books/report_workbook.xlsx", "xlsx_file"), "Detail", "xlsx_file")
        _check_xlsx(
            _find_versioned(out_root, "report_sheetbook/versions/*/books/report_sheetbook.xlsx", "xlsx_memory"), "Detail", "xlsx_memory"
        )
        _check_csv(_find_versioned(out_root, "metrics_csv/versions/*/files/metrics_csv.csv", "csv"), "csv")

    print("\n全部通过: 重复展示列名下三路径数据均按 field_id 正确对齐,未被首列值填充。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
