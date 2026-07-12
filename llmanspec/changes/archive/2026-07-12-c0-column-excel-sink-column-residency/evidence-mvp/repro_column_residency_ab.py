#!/usr/bin/env python3
"""Pinned A/B: ColumnExcelSink 列驻留 hold vs close 阶段分块释放.

- hold: 生产路径, `_columns` 在整个 close(append+save) 期间驻留
- chunk_release: close 时按行块 append 到 write_only sheet,每块后从各列 dict 删除已写出 row_id

小 shape 做单元格正确性对拍;大 shape 在独立子进程测 RSS.

输出: `.tmp/evidence/column-excel-column-residency-ab/<ts>/result.json`
"""

from __future__ import annotations

import argparse
import gc
import json
import os
import subprocess
import sys
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional


def _rss_kb() -> int:
    try:
        with open("/proc/self/statm", "r", encoding="utf-8") as handle:
            parts = handle.read().strip().split()
        if not parts:
            return 0
        return int(int(parts[1]) * os.sysconf("SC_PAGE_SIZE") / 1024)
    except Exception:
        return 0


def _rss_gb() -> float:
    return _rss_kb() / (1024.0 * 1024.0)


def _build(n_rows: int, n_cols: int):
    field_names = ["f{}".format(i) for i in range(n_cols)]
    row_ids = list(range(n_rows))
    columns = {name: {r: r * 10 + i for r in row_ids} for i, name in enumerate(field_names)}
    return field_names, row_ids, columns


def _read_xlsx_matrix(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _close_hold(sink: Any) -> None:
    sink.close()


def _close_chunk_release(sink: Any, *, chunk_rows: int) -> None:
    """模拟 B/C: write_only append 过程中释放已写出行的列切片."""

    from contextlib import suppress
    from pathlib import Path as PathCls

    from scalim._internal.utils.excel import escape_excel_formula
    from scalim.sinks._internal.base import (
        atomic_replace_temp_path,
        best_effort_cleanup_temp_path_dir,
        create_temp_path,
    )
    from scalim.sinks._internal.excel import (
        COLUMN_EXCEL_SINK_REMOVE_TEMP_FILE_FAILED_LOG,
        COLUMN_EXCEL_SINK_SAVE_FAILED_LOG,
        Workbook,
        _best_effort_close_write_only_workbook_worksheets,
        _excel_atomic_save_errors,
        _LOGGER,
    )

    if sink._closed:
        return

    output_dir = PathCls(sink.output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)
    temp_path = create_temp_path(sink.output_path, ".xlsx.tmp")
    temp_path_obj = PathCls(temp_path)
    wb = None
    try:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(sink.sheet_name)
        if sink.include_header:
            _ = ws.append([escape_excel_formula(x, allow_formulas=sink._allow_formulas) for x in sink.header_names])

        row_ids = list(sink._row_ids)
        n = len(row_ids)
        chunk = max(1, int(chunk_rows))
        for start in range(0, n, chunk):
            end = min(n, start + chunk)
            for pk in row_ids[start:end]:
                row_values = []
                for field_name in sink.field_names:
                    column_data = sink._columns.get(field_name, {})
                    row_values.append(column_data.get(pk))
                _ = ws.append([escape_excel_formula(x, allow_formulas=sink._allow_formulas) for x in row_values])
            # 释放本块 row_id 在各列中的驻留
            for pk in row_ids[start:end]:
                for field_name in sink.field_names:
                    col = sink._columns.get(field_name)
                    if col is not None:
                        _ = col.pop(pk, None)
            gc.collect()

        wb.save(temp_path_obj)
        atomic_replace_temp_path(temp_path, sink.output_path)
    except _excel_atomic_save_errors():
        _LOGGER.exception(COLUMN_EXCEL_SINK_SAVE_FAILED_LOG, sink.output_path)
        try:
            temp_path_obj.unlink()
        except FileNotFoundError:
            pass
        except OSError:
            _LOGGER.warning(COLUMN_EXCEL_SINK_REMOVE_TEMP_FILE_FAILED_LOG, temp_path_obj, exc_info=True)
        best_effort_cleanup_temp_path_dir(temp_path)
        if wb is not None:
            _best_effort_close_write_only_workbook_worksheets(wb)
        raise
    finally:
        if wb is not None:
            with suppress(Exception):
                wb.close()
    sink._closed = True


def run_arm(
    *,
    label: str,
    out_path: Path,
    n_rows: int,
    n_cols: int,
    chunk_rows: int,
) -> Dict[str, Any]:
    from scalim.sinks import ColumnExcelSink

    gc.collect()
    field_names, row_ids, columns = _build(n_rows, n_cols)
    rss0 = _rss_gb()
    sink = ColumnExcelSink(str(out_path), field_names=field_names, include_header=True)
    sink.set_row_ids(row_ids)
    for name in field_names:
        sink.write_column(name, columns[name])
    del columns
    gc.collect()
    rss_pre = _rss_gb()
    peak = {"v": max(rss0, rss_pre)}
    stop = {"v": False}

    def _sampler() -> None:
        while not stop["v"]:
            peak["v"] = max(peak["v"], _rss_gb())
            time.sleep(0.05)

    thr = threading.Thread(target=_sampler)
    thr.daemon = True
    thr.start()
    t0 = time.perf_counter()
    try:
        if label == "chunk_release":
            _close_chunk_release(sink, chunk_rows=chunk_rows)
        else:
            _close_hold(sink)
    finally:
        stop["v"] = True
        thr.join(timeout=2.0)
    close_s = time.perf_counter() - t0
    rss_post = _rss_gb()
    peak["v"] = max(peak["v"], rss_post)
    return {
        "label": label,
        "n_rows": n_rows,
        "n_cols": n_cols,
        "cells": n_rows * n_cols,
        "chunk_rows": chunk_rows if label == "chunk_release" else None,
        "rss_gb_pre_close": rss_pre,
        "rss_gb_post_close": rss_post,
        "peak_rss_gb_observed": peak["v"],
        "duration_close_s": close_s,
        "bytes_on_disk": out_path.stat().st_size if out_path.exists() else 0,
    }


def _child(mode: str, out_path: Path, n_rows: int, n_cols: int, chunk_rows: int) -> Dict[str, Any]:
    script = (
        "import json,sys;\n"
        "from importlib.machinery import SourceFileLoader;\n"
        "from pathlib import Path;\n"
        "m=SourceFileLoader('ab', sys.argv[1]).load_module();\n"
        "r=m.run_arm(label=sys.argv[2], out_path=Path(sys.argv[3]), n_rows=int(sys.argv[4]), "
        "n_cols=int(sys.argv[5]), chunk_rows=int(sys.argv[6]));\n"
        "print(json.dumps(r));\n"
    )
    here = str(Path(__file__).resolve())
    proc = subprocess.run(
        [
            sys.executable,
            "-c",
            script,
            here,
            mode,
            str(out_path),
            str(n_rows),
            str(n_cols),
            str(chunk_rows),
        ],
        cwd=str(Path.cwd()),
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        return {"error": "child_failed", "stderr": (proc.stderr or "")[-1000:], "code": proc.returncode}
    try:
        return json.loads(proc.stdout.strip().splitlines()[-1])
    except Exception as exc:
        return {"error": str(exc), "stdout": (proc.stdout or "")[-500:]}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", type=int, default=30000)
    parser.add_argument("--cols", type=int, default=300)
    parser.add_argument("--chunk-rows", type=int, default=2000)
    parser.add_argument("--correctness-rows", type=int, default=200)
    parser.add_argument("--out-dir", type=str, default="")
    args = parser.parse_args()

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_dir = Path(args.out_dir or os.path.join(".tmp", "evidence", "column-excel-column-residency-ab", ts))
    out_dir.mkdir(parents=True, exist_ok=True)
    work = out_dir / "work"
    work.mkdir(parents=True, exist_ok=True)

    # Correctness
    hold_small = work / "small_hold.xlsx"
    release_small = work / "small_release.xlsx"
    r_hold = run_arm(
        label="hold",
        out_path=hold_small,
        n_rows=args.correctness_rows,
        n_cols=min(20, args.cols),
        chunk_rows=args.chunk_rows,
    )
    r_rel = run_arm(
        label="chunk_release",
        out_path=release_small,
        n_rows=args.correctness_rows,
        n_cols=min(20, args.cols),
        chunk_rows=max(1, args.correctness_rows // 4),
    )
    m_hold = _read_xlsx_matrix(hold_small)
    m_rel = _read_xlsx_matrix(release_small)
    correctness = {
        "ok": m_hold == m_rel,
        "hold_rows": len(m_hold),
        "release_rows": len(m_rel),
        "hold_arm": r_hold,
        "release_arm": r_rel,
    }

    hold = _child("hold", work / "hold.xlsx", args.rows, args.cols, args.chunk_rows)
    release = _child("chunk_release", work / "release.xlsx", args.rows, args.cols, args.chunk_rows)

    result = {
        "ts": ts,
        "shape": {"rows": args.rows, "cols": args.cols, "chunk_rows": args.chunk_rows},
        "correctness": correctness,
        "hold": hold,
        "chunk_release": release,
    }
    if isinstance(hold, dict) and isinstance(release, dict) and "peak_rss_gb_observed" in hold and "peak_rss_gb_observed" in release:
        result["delta_peak_rss_gb"] = float(hold["peak_rss_gb_observed"]) - float(release["peak_rss_gb_observed"])
        result["delta_close_s"] = float(hold["duration_close_s"]) - float(release["duration_close_s"])

    out_path = out_dir / "result.json"
    out_path.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(result, indent=2, ensure_ascii=False))
    print("wrote", out_path)


if __name__ == "__main__":
    main()
