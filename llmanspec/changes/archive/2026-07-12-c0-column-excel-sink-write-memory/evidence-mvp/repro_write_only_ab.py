#!/usr/bin/env python3
"""Pinned A/B: ColumnExcelSink close write_only=False vs production write_only=True.

Baseline arm monkeypatches openpyxl Workbook to force write_only=False.
Current arm uses production ColumnExcelSink.close().

Also checks small-shape cell equality between arms.

Output: .tmp/evidence/column-excel-write-only-ab/<ts>/result.json
"""

from __future__ import annotations

import argparse
import gc
import json
import os
import subprocess
import sys
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Optional


def _rss_kb() -> int:
    try:
        with open("/proc/self/statm", "r", encoding="utf-8") as handle:
            parts = handle.read().strip().split()
        if not parts:
            return 0
        return int(int(parts[1]) * os.sysconf("SC_PAGE_SIZE") / 1024)
    except Exception:
        return 0


def _rss_gb() -> float:
    return _rss_kb() / (1024.0 * 1024.0)


def _build(n_rows: int, n_cols: int):
    field_names = ["f{}".format(i) for i in range(n_cols)]
    row_ids = list(range(n_rows))
    columns = {name: {r: r * 10 + i for r in row_ids} for i, name in enumerate(field_names)}
    return field_names, row_ids, columns


def _read_xlsx_matrix(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def run_arm(
    *,
    label: str,
    out_path: Path,
    n_rows: int,
    n_cols: int,
    force_non_write_only: bool,
) -> Dict[str, Any]:
    import scalim.sinks._internal.excel as excel_mod
    from openpyxl import Workbook as RealWorkbook
    from scalim.sinks import ColumnExcelSink

    if force_non_write_only:

        def _wb(*args, **kwargs):  # type: ignore[no-untyped-def]
            kwargs = dict(kwargs)
            kwargs["write_only"] = False
            return RealWorkbook(*args, **kwargs)

        excel_mod.Workbook = _wb  # type: ignore[assignment]
    else:
        excel_mod.Workbook = RealWorkbook  # type: ignore[assignment]

    gc.collect()
    field_names, row_ids, columns = _build(n_rows, n_cols)
    rss0 = _rss_gb()
    sink = ColumnExcelSink(str(out_path), field_names=field_names, include_header=True)
    sink.set_row_ids(row_ids)
    for name in field_names:
        sink.write_column(name, columns[name])
    del columns
    gc.collect()
    rss_pre = _rss_gb()
    peak = {"v": max(rss0, rss_pre)}
    stop = {"v": False}

    def _sampler() -> None:
        while not stop["v"]:
            peak["v"] = max(peak["v"], _rss_gb())
            time.sleep(0.05)

    thr = threading.Thread(target=_sampler)
    thr.daemon = True
    thr.start()
    t0 = time.perf_counter()
    try:
        sink.close()
    finally:
        stop["v"] = True
        thr.join(timeout=2.0)
    close_s = time.perf_counter() - t0
    rss_post = _rss_gb()
    peak["v"] = max(peak["v"], rss_post)
    return {
        "label": label,
        "force_non_write_only": force_non_write_only,
        "n_rows": n_rows,
        "n_cols": n_cols,
        "cells": n_rows * n_cols,
        "rss_gb_pre_close": rss_pre,
        "rss_gb_post_close": rss_post,
        "peak_rss_gb_observed": peak["v"],
        "duration_close_s": close_s,
        "bytes_on_disk": out_path.stat().st_size if out_path.exists() else 0,
    }


def _child(mode: str, out_path: Path, n_rows: int, n_cols: int) -> Dict[str, Any]:
    script = (
        "import json,sys;\n"
        "from importlib.machinery import SourceFileLoader;\n"
        "from pathlib import Path;\n"
        "m=SourceFileLoader('ab', sys.argv[1]).load_module();\n"
        "r=m.run_arm(label=sys.argv[2], out_path=Path(sys.argv[3]), n_rows=int(sys.argv[4]), "
        "n_cols=int(sys.argv[5]), force_non_write_only=(sys.argv[6]=='1'));\n"
        "print(json.dumps(r));\n"
    )
    here = str(Path(__file__).resolve())
    proc = subprocess.run(
        [
            sys.executable,
            "-c",
            script,
            here,
            mode,
            str(out_path),
            str(n_rows),
            str(n_cols),
            "1" if mode == "baseline" else "0",
        ],
        cwd=str(Path.cwd()),
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        return {"error": "child_failed", "stderr": (proc.stderr or "")[-1000:], "code": proc.returncode}
    try:
        return json.loads(proc.stdout.strip().splitlines()[-1])
    except Exception as exc:
        return {"error": str(exc), "stdout": (proc.stdout or "")[-500:]}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", type=int, default=50000)
    parser.add_argument("--cols", type=int, default=300)
    parser.add_argument("--correctness-rows", type=int, default=200)
    parser.add_argument("--out-dir", type=str, default="")
    args = parser.parse_args()

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_dir = Path(args.out_dir or os.path.join(".tmp", "evidence", "column-excel-write-only-ab", ts))
    out_dir.mkdir(parents=True, exist_ok=True)
    work = out_dir / "work"
    work.mkdir(parents=True, exist_ok=True)

    # Correctness on small shape (same process ok)
    base_small = work / "small_baseline.xlsx"
    cur_small = work / "small_current.xlsx"
    small_b = run_arm(
        label="small_baseline",
        out_path=base_small,
        n_rows=int(args.correctness_rows),
        n_cols=min(20, int(args.cols)),
        force_non_write_only=True,
    )
    small_c = run_arm(
        label="small_current",
        out_path=cur_small,
        n_rows=int(args.correctness_rows),
        n_cols=min(20, int(args.cols)),
        force_non_write_only=False,
    )
    cells_equal = _read_xlsx_matrix(base_small) == _read_xlsx_matrix(cur_small)

    baseline = _child("baseline", work / "baseline.xlsx", int(args.rows), int(args.cols))
    current = _child("current", work / "current.xlsx", int(args.rows), int(args.cols))

    def _f(d: Optional[Dict[str, Any]], key: str):
        if not isinstance(d, dict) or d.get("error"):
            return None
        return d.get(key)

    report = {
        "topic": "column-excel-write-only-ab",
        "pinned_script": str(Path(__file__).as_posix()),
        "shape": {"n_rows": int(args.rows), "n_cols": int(args.cols)},
        "correctness": {
            "rows": int(args.correctness_rows),
            "cols": min(20, int(args.cols)),
            "cells_equal": cells_equal,
            "baseline": small_b,
            "current": small_c,
        },
        "fresh_process": {"baseline_force_non_write_only": baseline, "current_write_only": current},
        "summary": {
            "cells_equal": cells_equal,
            "peak_rss_baseline_gb": _f(baseline, "peak_rss_gb_observed"),
            "peak_rss_current_gb": _f(current, "peak_rss_gb_observed"),
            "peak_rss_saved_gb": (
                None
                if _f(baseline, "peak_rss_gb_observed") is None or _f(current, "peak_rss_gb_observed") is None
                else float(baseline["peak_rss_gb_observed"]) - float(current["peak_rss_gb_observed"])
            ),
            "close_s_baseline": _f(baseline, "duration_close_s"),
            "close_s_current": _f(current, "duration_close_s"),
        },
    }
    path = out_dir / "result.json"
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2, sort_keys=True))
    print("report -> {}".format(path))
    if not cells_equal:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
