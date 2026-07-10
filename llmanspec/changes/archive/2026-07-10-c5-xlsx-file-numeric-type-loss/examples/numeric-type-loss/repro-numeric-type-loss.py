#!/usr/bin/env python3
"""
MVP: 验证 xlsx_file / xlsx_memory 在 typed-ROWS 路径下保留数字类型。

修复后契约
  - 两条路径都以 InMemoryRows（FieldValue）为中间产物 SSOT
  - workbook / sheetbook 都持有 typed segments，commit 时仅对 str 做 formula escape
  - 最终 Excel 数字列应为 int/float，而非 str

用法
  uv run python3 llmanspec/changes/c5-xlsx-file-numeric-type-loss/examples/numeric-type-loss/repro-numeric-type-loss.py

输出
  - 控制台: 各阶段值的类型信息
  - .tmp/repro/numeric_type_loss/output/: xlsx_file.xlsx + xlsx_memory.xlsx
"""

import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# 确保 scalim 可导入
# ---------------------------------------------------------------------------
_REPO_ROOT = Path(__file__).resolve()
for _ in range(10):
    if (_REPO_ROOT / "pyproject.toml").exists():
        break
    _REPO_ROOT = _REPO_ROOT.parent
else:
    raise RuntimeError("Cannot find repo root (pyproject.toml)")
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from scalim.sinks.rows import InMemoryRowsSink
from scalim.workflow.resources_workbook import (
    SheetPlan,
    _WorkbookSegment,
    _iter_workbook_sheet_rows,
)
from scalim.workflow.resources_sheetbook import (
    SheetBookPlan,
    SheetBookSegment,
    SheetBookSheetPlan,
    _write_sheetbook_plan_to_openpyxl_workbook,
)


def _openpyxl_Workbook():
    import openpyxl as _m

    return _m.Workbook


def _load_workbook(path):
    import openpyxl as _m

    return _m.load_workbook(path)


# ===================================================================
# 阶段 1: 模拟 demand 执行 → 产生中间产物（两条路径均为 InMemoryRows）
# ===================================================================
print("=" * 60)
print("阶段 1: demand 执行 → 中间产物")
print("=" * 60)

raw_rows = [
    {"name": "商品A", "amount": 123.45, "rate": 0.95, "quantity": 10},
    {"name": "商品B", "amount": 678.90, "rate": 0.85, "quantity": 3},
    {"name": "商品C", "amount": 999.99, "rate": 0.50, "quantity": 0},
]
print("\n原始数据类型:")
for r in raw_rows:
    for k, v in r.items():
        print("  {}: {!r} ({})".format(k, v, type(v).__name__))
    print()

# ---- 1a: InMemoryRowsSink（xlsx_file managed book 路径） ----
file_sink = InMemoryRowsSink(field_ids=["name", "amount", "rate", "quantity"])
for row in raw_rows:
    file_sink.write_row(row)
file_artifact = file_sink.to_artifact()

print("--- xlsx_file 中间产物 (InMemoryRows) ---")
print("header: {}".format(file_artifact.header))
for row in file_artifact.rows:
    for i, v in enumerate(row):
        print("  {}: {!r} ({})".format(file_artifact.header[i], v, type(v).__name__))
    print()

# ---- 1b: InMemoryRowsSink（xlsx_memory 路径） ----
mem_sink = InMemoryRowsSink(field_ids=["name", "amount", "rate", "quantity"])
for row in raw_rows:
    mem_sink.write_row(row)
mem_artifact = mem_sink.to_artifact()

print("--- xlsx_memory 中间产物 (InMemoryRows) ---")
print("header: {}".format(mem_artifact.header))
for row in mem_artifact.rows:
    for i, v in enumerate(row):
        print("  {}: {!r} ({})".format(mem_artifact.header[i], v, type(v).__name__))
    print()

# ===================================================================
# 阶段 2: 模拟 workflow 资源管理器合成 xlsx
# ===================================================================
print("=" * 60)
print("阶段 2: 资源管理器合成 xlsx")
print("=" * 60)

OUTPUT_DIR = _REPO_ROOT / ".tmp" / "repro" / "numeric_type_loss" / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

export_header = ["名称", "金额", "比率", "数量"]

# ---- 2a: xlsx_file 路径（typed workbook segments） ----
seg_file = _WorkbookSegment(
    producer_node_id="demand_A",
    decl_order=0,
    rows=list(file_artifact.rows),
    header_policy="once",
)
sheet_plan = SheetPlan(
    sheet="数据",
    baseline_header=list(file_artifact.header),
    export_header=list(export_header),
    segments=[seg_file],
)

xlsx_file_path = str(OUTPUT_DIR / "xlsx_file_output.xlsx")
Workbook = _openpyxl_Workbook()
wb = Workbook(write_only=True)
ws = wb.create_sheet("数据")
for row in _iter_workbook_sheet_rows(sheet_plan, allow_formulas=True):
    ws.append(row)
wb.save(xlsx_file_path)
print("\nxlsx_file 已写入: {}".format(xlsx_file_path))

# ---- 2b: xlsx_memory 路径（typed sheetbook segments） ----
seg_mem = SheetBookSegment(
    producer_node_id="demand_A",
    decl_order=0,
    rows=list(mem_artifact.rows),
    header_policy="once",
)
sheet_mem_plan = SheetBookSheetPlan(
    sheet="数据",
    baseline_header=list(mem_artifact.header),
    export_header=list(export_header),
    segments=[seg_mem],
)
book_plan = SheetBookPlan(
    resource_id="test_book",
    budget_max_sheets=10,
    budget_max_total_cells=100000,
    export_path=str(OUTPUT_DIR / "xlsx_memory_output.xlsx"),
    export_allow_formulas=True,
    sheet_decl_order={"数据": 0},
    sheet_order=["数据"],
    sheets={"数据": sheet_mem_plan},
)

xlsx_memory_path = str(OUTPUT_DIR / "xlsx_memory_output.xlsx")
wb2 = Workbook(write_only=True)
_write_sheetbook_plan_to_openpyxl_workbook(wb2, book_plan)
wb2.save(xlsx_memory_path)
print("xlsx_memory 已写入: {}".format(xlsx_memory_path))

# ===================================================================
# 阶段 3: 验证结果 — 读取 xlsx 并检查单元格类型
# ===================================================================
print("=" * 60)
print("阶段 3: 验证结果 — 单元格类型对比")
print("=" * 60)


def inspect_xlsx(path, label):
    print("\n--- {} ---".format(label))
    wb_local = _load_workbook(path)
    ws_local = wb_local.active
    print("表头: {}".format([c.value for c in next(ws_local.iter_rows(min_row=1, max_row=1))]))
    for row in ws_local.iter_rows(min_row=2, values_only=False):
        for cell in row:
            type_name = type(cell.value).__name__
            marker = ""
            if isinstance(cell.value, str) and cell.column > 1:
                marker = " <<< 数字类型丢失!"
            print(
                "  {}: {!r} (type={}, number_format={!r}){}".format(
                    cell.coordinate,
                    cell.value,
                    type_name,
                    cell.number_format,
                    marker,
                )
            )
    print()


inspect_xlsx(xlsx_file_path, "xlsx_file 输出")
inspect_xlsx(xlsx_memory_path, "xlsx_memory 输出")

# ===================================================================
# 阶段 4: 汇总
# ===================================================================
print("=" * 60)
print("汇总")
print("=" * 60)

wb_f = _load_workbook(xlsx_file_path)
ws_f = wb_f.active
file_amount_cell = list(ws_f.iter_rows(min_row=2, max_row=2))[0][1]

wb_m = _load_workbook(xlsx_memory_path)
ws_m = wb_m.active
mem_amount_cell = list(ws_m.iter_rows(min_row=2, max_row=2))[0][1]

print(
    "\nxlsx_file.amount  cell type: {},  value={!r}".format(
        type(file_amount_cell.value).__name__,
        file_amount_cell.value,
    )
)
print(
    "xlsx_memory.amount cell type: {}, value={!r}".format(
        type(mem_amount_cell.value).__name__,
        mem_amount_cell.value,
    )
)
print()

ok = True
if isinstance(file_amount_cell.value, (int, float)):
    print("✅ 确认: xlsx_file 路径保留了数字类型")
else:
    print("❌ 回归: xlsx_file 路径数字类型丢失 (amount 列是 {!r})".format(type(file_amount_cell.value).__name__))
    ok = False
if isinstance(mem_amount_cell.value, (int, float)):
    print("✅ 确认: xlsx_memory 路径保留了数字类型")
else:
    print("❌ 回归: xlsx_memory 路径数字类型丢失")
    ok = False

print("\n详细输出路径: {}".format(OUTPUT_DIR))
print("可直接打开 .xlsx 文件在 Excel 中查看单元格格式确认")
sys.exit(0 if ok else 1)
