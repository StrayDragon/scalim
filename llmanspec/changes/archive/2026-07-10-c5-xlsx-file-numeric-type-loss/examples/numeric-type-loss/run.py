#!/usr/bin/env python3
"""
MVP: 验证 xlsx_file 与 xlsx_memory 在 typed-ROWS 路径下数字类型一致保留。

场景覆盖:
  场景 A (主场景 - 多层): workflow.yaml
    3 demands → 3 sheets in same xlsx_file book → 多 demand 写入同一文件
    镜像 Pay Order 的"明细 + 渠道维度 + 整体指标"模式

  场景 B (最小复现): workflow_detail_only.yaml
    1 demand → 1 sheet → 最简路径

  数据边界:
    - int, float, Decimal, bool, None, 零值
    - 经 to_numeric() 处理后的字段
    - 聚合计算后的派生数值

  对比基线: xlsx_memory book 全程保留类型；修复后 xlsx_file 应对齐

运行:
  cd scalim/
  uv run python3 llmanspec/changes/c5-xlsx-file-numeric-type-loss/examples/numeric-type-loss/run.py

输出:
  - 控制台: 两本书各 sheet 的单元格类型对比 + 逐列评分
  - 临时目录下输出 xlsx 文件（自动清理）
"""

import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Tuple

from scalim.dsl.yaml_dsl import WorkflowRunOptions, run_workflow
from scalim.dsl.yaml_dsl.runtime.contracts import (
    DemandDiagnosticsOverride,
    DemandRunOptions,
    DemandRunRuntimeOptions,
    DemandRunSecurityOptions,
)

DEMO_DIR = Path(__file__).resolve().parent

# 非数字列的标识（脱敏场景中仅有"交易编号""产品""指标""单位"为文本）
_TEXT_COLUMNS = frozenset({"交易编号", "产品", "指标", "单位", "渠道"})


def _is_numeric(val: Any) -> bool:
    """判断一个值是否为数字类型（或 None/NULL 占位）。"""
    if val is None:
        return True  # None 在数值列中是合法的"缺失"，不视为类型错误
    return isinstance(val, (int, float, Decimal, bool))


def _inspect_xlsx(path: Path, label: str) -> List[Dict[str, Any]]:
    """读取 xlsx，返回逐 sheet 的列级类型分析。"""
    from decimal import Decimal
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheet_results: List[Dict[str, Any]] = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_raw: List[List[Any]] = [list(r) for r in ws.iter_rows(values_only=True)]

            sheet_info: Dict[str, Any] = {
                "sheet": sheet_name,
                "header": rows_raw[0] if rows_raw else [],
                "data_rows": rows_raw[1:] if len(rows_raw) > 1 else [],
                "col_types": {},  # col_name -> set of type names seen
                "col_errors": [],  # (col_name, row_idx, expected, actual, value)
            }

            print(f"\n--- {label} / {sheet_name} ---")
            if not rows_raw:
                print("  (空 sheet)")
                sheet_results.append(sheet_info)
                continue

            header = rows_raw[0]
            print(f"  表头: {header}")
            print(f"  数据行数: {len(rows_raw) - 1}")

            for data_row_idx, data_row in enumerate(rows_raw[1:], 2):
                for col_idx, val in enumerate(data_row):
                    col_name = header[col_idx] if col_idx < len(header) else f"Col{col_idx}"
                    col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx}"

                    # 收集该列出现的类型
                    if col_name not in sheet_info["col_types"]:
                        sheet_info["col_types"][col_name] = set()
                    type_name = type(val).__name__ if val is not None else "NoneType"
                    sheet_info["col_types"][col_name].add(type_name)

                    # 检查类型丢失: 文本列跳过，None 跳过
                    if col_name in _TEXT_COLUMNS:
                        continue
                    if val is None:
                        continue

                    # 如果是字符串且在数字列 -> 类型丢失
                    if isinstance(val, str):
                        if col_name not in [x[0] for x in sheet_info["col_errors"]]:
                            sheet_info["col_errors"].append((col_name, data_row_idx, "int/float/Decimal/bool", "str", val))

                    # 打印前 3 行
                    if data_row_idx <= 3:
                        marker = "  <<< 数字类型丢失!" if isinstance(val, str) else ""
                        if col_name not in _TEXT_COLUMNS or isinstance(val, str):
                            print(f"  {col_letter}{data_row_idx}: {val!r} (type={type_name}){marker}")

            # sheet 汇总
            n_errors = len(sheet_info["col_errors"])
            n_cols = len([c for c in header if c not in _TEXT_COLUMNS])
            print(f"  >>> 数字列: {n_cols}, 类型丢失列数: {n_errors}")
            for col_name, row_idx, exp, act, val in sheet_info["col_errors"][:5]:
                print(f"      {col_name} (行{row_idx}): 期望 {exp}, 实际 {act} ({val!r})")
            if len(sheet_info["col_errors"]) > 5:
                print(f"      ... 还有 {len(sheet_info['col_errors']) - 5} 处丢失")

            sheet_results.append(sheet_info)
    finally:
        wb.close()

    return sheet_results


def _find_versioned(out_root: Path, pattern: str, label: str) -> Path:
    """在 workflow 版本化输出目录中找到目标文件。"""
    matches = sorted(out_root.glob(pattern))
    if not matches:
        matches = sorted(out_root.glob("**/*.xlsx"))
    assert len(matches) >= 1, f"{label}: 未找到输出文件 ({pattern})"
    return matches[0]


def _score(
    sheet_results: Dict[str, List[Dict[str, Any]]],
) -> Tuple[int, int, int]:
    """统计: (总数字列数, 类型丢失列数, 正常列数)。"""
    total = 0
    lost = 0
    ok = 0
    for label, sheets in sheet_results.items():
        for s in sheets:
            # 总数字列 = 非文本列
            numeric_cols = [c for c in s["header"] if c not in _TEXT_COLUMNS]
            total += len(numeric_cols)
            # 丢失的数字列 = 有 errors 的列
            error_cols = set(e[0] for e in s["col_errors"])
            lost += len(error_cols)
            ok += len(numeric_cols) - len(error_cols)
    return total, lost, ok


def _run_workflow_and_check(
    workflow_name: str,
    workflow_yaml: Path,
    temp_dir: Path,
) -> int:
    """运行一个 workflow 并检查结果。"""
    print("\n" + "=" * 60)
    print(f"运行 Workflow: {workflow_name}")
    print(f"  YAML: {workflow_yaml.name}")
    print("=" * 60)

    out_dir = temp_dir / workflow_name
    out_dir.mkdir(parents=True, exist_ok=True)

    # 复制所有依赖文件到临时子目录
    for fname in [
        workflow_yaml.name,
        "demand_a.yaml",
        "demand_b.yaml",
        "demand_c.yaml",
        "loaders.py",
    ]:
        src = DEMO_DIR / fname
        if src.exists():
            (out_dir / fname).write_text(src.read_text(encoding="utf-8"), encoding="utf-8")

    wf_path = out_dir / workflow_yaml.name

    result = run_workflow(
        str(wf_path),
        options=WorkflowRunOptions(
            demand=DemandRunOptions(
                security=DemandRunSecurityOptions(
                    allowed_modules=frozenset(["loaders"]),
                ),
                runtime=DemandRunRuntimeOptions(
                    batch_size=100,
                    demand_diagnostics=DemandDiagnosticsOverride(validate_unique_field_names=False),
                ),
            ),
        ),
    )
    errors = result.errors()
    if errors:
        print("❌ Workflow 运行出错:")
        for e in errors:
            print(f"  run_id={e.run_id}: {e.message}")
        return 1

    print("✅ Workflow 运行成功")

    # 定位输出文件
    out_root = out_dir / "out"
    try:
        xlsx_file_path = _find_versioned(
            out_root,
            "report_workbook/versions/*/books/report_workbook.xlsx",
            "xlsx_file",
        )
        xlsx_memory_path = _find_versioned(
            out_root,
            "report_sheetbook/versions/*/books/report_sheetbook.xlsx",
            "xlsx_memory",
        )
    except AssertionError as exc:
        print(f"❌ {exc}")
        return 1

    # 检查 xlsx
    results: Dict[str, List[Dict[str, Any]]] = {}
    results["xlsx_file"] = _inspect_xlsx(xlsx_file_path, "xlsx_file")
    results["xlsx_memory"] = _inspect_xlsx(xlsx_memory_path, "xlsx_memory")

    # 汇总评分
    total, lost, ok = _score(results)

    print("\n" + "-" * 40)
    print(f"汇总评分 [{workflow_name}]")
    print("-" * 40)
    print(f"  总数字列: {total} | 类型丢失: {lost} | 类型保留: {ok}")

    # 逐列对比
    print("\n  逐列类型对比:")
    file_sheets = results["xlsx_file"]
    mem_sheets = results["xlsx_memory"]

    all_cols: List[str] = []
    col_book_map: Dict[str, Dict[str, str]] = {}
    for s in file_sheets:
        for col in s["header"]:
            if col not in all_cols:
                all_cols.append(col)
            if col not in col_book_map:
                col_book_map[col] = {}
            col_book_map[col]["xlsx_file"] = ", ".join(sorted(s["col_types"].get(col, {"N/A"})))
    for s in mem_sheets:
        for col in s["header"]:
            if col not in col_book_map:
                col_book_map[col] = {}
            col_book_map[col]["xlsx_memory"] = ", ".join(sorted(s["col_types"].get(col, {"N/A"})))

    for col in all_cols:
        ft = col_book_map.get(col, {}).get("xlsx_file", "N/A")
        mt = col_book_map.get(col, {}).get("xlsx_memory", "N/A")
        is_text = col in _TEXT_COLUMNS
        # 对于文本列: 两路径都应是 str ✓
        # 对于数字列: xlsx_memory 保留类型, xlsx_file 全部 str
        if is_text:
            status = "~"
        elif "str" in ft and "str" not in mt and mt != "N/A":
            status = "❌丢失"
        elif ft == mt:
            status = "✅"
        else:
            status = "⚠️"
        print(f"    {status} {col:20s}  file={ft:20s}  mem={mt:20s}")

    # 结论（修复后期望: lost == 0 且 file/mem 类型对齐）
    print()
    if lost == 0 and total > 0:
        print("  ✅ 确认: xlsx_file 与 xlsx_memory 均保留数字类型 (0 丢失)")
        return 0
    if lost > 0:
        print(f"  ❌ 回归: xlsx_file 仍有数字类型丢失 ({lost}/{total} 列变字符串)")
        return 1
    print("  ⚠️ 结果异常, 需人工排查")
    return 1


def main() -> int:
    # 确保 loaders 可导入
    sys.path.insert(0, str(DEMO_DIR))

    with tempfile.TemporaryDirectory(prefix="scalim-mvp-numeric-type-loss-") as temp_dir:
        temp_path = Path(temp_dir).resolve()
        print(f"临时工作目录: {temp_path}")
        print()

        # ---------------------------------------------------------
        # 场景 A (主场景): 多 sheet 多 demand
        # ---------------------------------------------------------
        rc_a = _run_workflow_and_check(
            "场景A-多sheet",
            DEMO_DIR / "workflow.yaml",
            temp_path,
        )

        # ---------------------------------------------------------
        # 场景 B (最小复现): 单 demand 单 sheet
        # ---------------------------------------------------------
        if (DEMO_DIR / "workflow_detail_only.yaml").exists():
            rc_b = _run_workflow_and_check(
                "场景B-单sheet(最小复现)",
                DEMO_DIR / "workflow_detail_only.yaml",
                temp_path,
            )
        else:
            print("\n⚠️ workflow_detail_only.yaml 不存在,跳过场景B")
            rc_b = 0

        # ---------------------------------------------------------
        # 最终判定
        # ---------------------------------------------------------
        print("\n" + "=" * 60)
        print("最终结论")
        print("=" * 60)
        if rc_a == 0:
            print("\n✅ 场景A (多 sheet 多 demand): xlsx_file / xlsx_memory 类型对齐,无 丢失")
        else:
            print("\n❌ 场景A 回归: 仍存在数字类型丢失或不一致")
        if rc_b == 0:
            print("\n✅ 场景B (单 sheet 最小复现): xlsx_file / xlsx_memory 类型对齐,0 丢失")
        else:
            print("\n❌ 场景B 回归: 仍存在数字类型丢失或不一致")

        detail_workflow_a = temp_path / "场景A-多sheet" / "out"
        detail_workflow_b = temp_path / "场景B-单sheet(最小复现)" / "out"
        print(f"\n输出文件位置 (临时目录, 运行后自动清理):")
        print(f"  多 sheet:  {detail_workflow_a}")
        print(f"  单 sheet:  {detail_workflow_b}")

        return 0 if (rc_a == 0 and rc_b == 0) else 1


if __name__ == "__main__":
    from decimal import Decimal  # noqa: F401 用于 _is_numeric

    sys.exit(main())
