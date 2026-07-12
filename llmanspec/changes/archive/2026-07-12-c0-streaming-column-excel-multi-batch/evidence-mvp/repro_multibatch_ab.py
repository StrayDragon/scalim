#!/usr/bin/env python3
"""Pinned A/B: hold vs StreamingColumnExcelSink multi-batch set_row_ids.

- hold: ColumnExcelSink 全量列再 close
- streaming_multibatch: 每窗 set_row_ids(本窗) → 写满全部列(对齐 pipeline 列模式)
- streaming_window(对照): 一次 set_row_ids(全量) + 按窗 write_column

输出: `.tmp/evidence/streaming-column-excel-multibatch/<ts>/result.json`
"""

from __future__ import annotations

import argparse
import gc
import json
import os
import subprocess
import sys
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Sequence


def _rss_kb() -> int:
    try:
        with open("/proc/self/statm", "r", encoding="utf-8") as handle:
            parts = handle.read().strip().split()
        if not parts:
            return 0
        return int(int(parts[1]) * os.sysconf("SC_PAGE_SIZE") / 1024)
    except Exception:
        return 0


def _rss_gb() -> float:
    return _rss_kb() / (1024.0 * 1024.0)


def _build_window_columns(
    *,
    field_names: Sequence[str],
    row_ids: Sequence[int],
) -> Dict[str, Dict[int, int]]:
    columns: Dict[str, Dict[int, int]] = {}
    for i, name in enumerate(field_names):
        columns[name] = {r: r * 10 + i for r in row_ids}
    return columns


def _read_xlsx_matrix(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _sample_peak(fn):
    gc.collect()
    rss0 = _rss_gb()
    peak = {"v": rss0}
    stop = {"v": False}

    def _sampler() -> None:
        while not stop["v"]:
            peak["v"] = max(peak["v"], _rss_gb())
            time.sleep(0.05)

    thr = threading.Thread(target=_sampler)
    thr.daemon = True
    thr.start()
    t0 = time.perf_counter()
    try:
        result = fn(peak)
    finally:
        stop["v"] = True
        thr.join(timeout=2.0)
    elapsed = time.perf_counter() - t0
    rss_post = _rss_gb()
    peak["v"] = max(peak["v"], rss_post)
    result["duration_s"] = elapsed
    result["rss_gb_post_close"] = rss_post
    result["peak_rss_gb_observed"] = peak["v"]
    return result


def run_hold(*, out_path: Path, n_rows: int, n_cols: int) -> Dict[str, Any]:
    from scalim.sinks import ColumnExcelSink

    field_names = ["f{}".format(i) for i in range(n_cols)]
    row_ids = list(range(n_rows))

    def _body(peak):
        columns = _build_window_columns(field_names=field_names, row_ids=row_ids)
        sink = ColumnExcelSink(str(out_path), field_names=field_names, include_header=True)
        sink.set_row_ids(row_ids)
        for name in field_names:
            sink.write_column(name, columns[name])
        del columns
        gc.collect()
        rss_pre = _rss_gb()
        peak["v"] = max(peak["v"], rss_pre)
        sink.close()
        return {
            "label": "hold",
            "n_rows": n_rows,
            "n_cols": n_cols,
            "cells": n_rows * n_cols,
            "window_rows": None,
            "rss_gb_pre_close": rss_pre,
            "bytes_on_disk": out_path.stat().st_size if out_path.exists() else 0,
            "flushed_rows": n_rows,
        }

    return _sample_peak(_body)


def run_streaming_window(*, out_path: Path, n_rows: int, n_cols: int, window_rows: int) -> Dict[str, Any]:
    from scalim.sinks import StreamingColumnExcelSink

    field_names = ["f{}".format(i) for i in range(n_cols)]
    row_ids = list(range(n_rows))
    window = max(1, int(window_rows))

    def _body(peak):
        sink = StreamingColumnExcelSink(str(out_path), field_names=field_names, include_header=True)
        sink.set_row_ids(row_ids)
        for start in range(0, n_rows, window):
            end = min(n_rows, start + window)
            win_ids = row_ids[start:end]
            columns = _build_window_columns(field_names=field_names, row_ids=win_ids)
            for name in field_names:
                sink.write_column(name, columns[name])
            del columns
            gc.collect()
            peak["v"] = max(peak["v"], _rss_gb())
        rss_pre = _rss_gb()
        peak["v"] = max(peak["v"], rss_pre)
        sink.close()
        return {
            "label": "streaming_window",
            "n_rows": n_rows,
            "n_cols": n_cols,
            "cells": n_rows * n_cols,
            "window_rows": window,
            "rss_gb_pre_close": rss_pre,
            "bytes_on_disk": out_path.stat().st_size if out_path.exists() else 0,
            "flushed_rows": sink._flushed_rows,
            "impl": "StreamingColumnExcelSink+single_set_row_ids",
        }

    return _sample_peak(_body)


def run_streaming_multibatch(*, out_path: Path, n_rows: int, n_cols: int, window_rows: int) -> Dict[str, Any]:
    from scalim.sinks import StreamingColumnExcelSink

    field_names = ["f{}".format(i) for i in range(n_cols)]
    row_ids = list(range(n_rows))
    window = max(1, int(window_rows))

    def _body(peak):
        sink = StreamingColumnExcelSink(str(out_path), field_names=field_names, include_header=True)
        for start in range(0, n_rows, window):
            end = min(n_rows, start + window)
            win_ids = row_ids[start:end]
            sink.set_row_ids(win_ids)
            columns = _build_window_columns(field_names=field_names, row_ids=win_ids)
            for name in field_names:
                sink.write_column(name, columns[name])
            del columns
            gc.collect()
            peak["v"] = max(peak["v"], _rss_gb())
        rss_pre = _rss_gb()
        peak["v"] = max(peak["v"], rss_pre)
        sink.close()
        return {
            "label": "streaming_multibatch",
            "n_rows": n_rows,
            "n_cols": n_cols,
            "cells": n_rows * n_cols,
            "window_rows": window,
            "rss_gb_pre_close": rss_pre,
            "bytes_on_disk": out_path.stat().st_size if out_path.exists() else 0,
            "flushed_rows": sink._flushed_rows,
            "impl": "StreamingColumnExcelSink+append_set_row_ids",
        }

    return _sample_peak(_body)


def _child(mode: str, out_path: Path, n_rows: int, n_cols: int, window_rows: int) -> Dict[str, Any]:
    script = (
        "import json,sys;\n"
        "from importlib.machinery import SourceFileLoader;\n"
        "from pathlib import Path;\n"
        "m=SourceFileLoader('ab', sys.argv[1]).load_module();\n"
        "mode=sys.argv[2]; out=Path(sys.argv[3]); rows=int(sys.argv[4]); cols=int(sys.argv[5]); win=int(sys.argv[6]);\n"
        "if mode=='hold':\n"
        " r=m.run_hold(out_path=out, n_rows=rows, n_cols=cols)\n"
        "elif mode=='streaming_window':\n"
        " r=m.run_streaming_window(out_path=out, n_rows=rows, n_cols=cols, window_rows=win)\n"
        "else:\n"
        " r=m.run_streaming_multibatch(out_path=out, n_rows=rows, n_cols=cols, window_rows=win)\n"
        "print(json.dumps(r));\n"
    )
    here = str(Path(__file__).resolve())
    proc = subprocess.run(
        [sys.executable, "-c", script, here, mode, str(out_path), str(n_rows), str(n_cols), str(window_rows)],
        cwd=str(Path.cwd()),
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        return {"error": "child_failed", "stderr": (proc.stderr or "")[-1200:], "code": proc.returncode}
    try:
        return json.loads(proc.stdout.strip().splitlines()[-1])
    except Exception as exc:
        return {"error": str(exc), "stdout": (proc.stdout or "")[-500:]}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", type=int, default=20000)
    parser.add_argument("--cols", type=int, default=200)
    parser.add_argument("--window-rows", type=int, default=2000)
    parser.add_argument("--correctness-rows", type=int, default=200)
    parser.add_argument("--correctness-cols", type=int, default=20)
    parser.add_argument("--out-dir", type=str, default="")
    args = parser.parse_args()

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_dir = Path(args.out_dir or os.path.join(".tmp", "evidence", "streaming-column-excel-multibatch", ts))
    out_dir.mkdir(parents=True, exist_ok=True)
    work = out_dir / "work"
    work.mkdir(parents=True, exist_ok=True)

    hold_small = work / "small_hold.xlsx"
    mb_small = work / "small_multibatch.xlsx"
    win_small = work / "small_window.xlsx"
    r_hold_s = run_hold(out_path=hold_small, n_rows=args.correctness_rows, n_cols=args.correctness_cols)
    r_mb_s = run_streaming_multibatch(
        out_path=mb_small,
        n_rows=args.correctness_rows,
        n_cols=args.correctness_cols,
        window_rows=max(1, args.correctness_rows // 4),
    )
    r_win_s = run_streaming_window(
        out_path=win_small,
        n_rows=args.correctness_rows,
        n_cols=args.correctness_cols,
        window_rows=max(1, args.correctness_rows // 4),
    )
    m_hold = _read_xlsx_matrix(hold_small)
    m_mb = _read_xlsx_matrix(mb_small)
    m_win = _read_xlsx_matrix(win_small)
    correctness = {
        "ok": m_hold == m_mb == m_win,
        "hold_rows": len(m_hold),
        "multibatch_rows": len(m_mb),
        "window_rows": len(m_win),
        "hold_arm": r_hold_s,
        "multibatch_arm": r_mb_s,
        "window_arm": r_win_s,
    }

    hold = _child("hold", work / "hold.xlsx", args.rows, args.cols, args.window_rows)
    multibatch = _child("streaming_multibatch", work / "multibatch.xlsx", args.rows, args.cols, args.window_rows)
    window = _child("streaming_window", work / "window.xlsx", args.rows, args.cols, args.window_rows)

    result: Dict[str, Any] = {
        "ts": ts,
        "mvp": "StreamingColumnExcelSink+append_set_row_ids",
        "shape": {"rows": args.rows, "cols": args.cols, "window_rows": args.window_rows},
        "correctness": correctness,
        "hold": hold,
        "streaming_multibatch": multibatch,
        "streaming_window": window,
    }

    def _ratio(base, arm):
        if not isinstance(base, dict) or not isinstance(arm, dict):
            return
        if "peak_rss_gb_observed" not in base or "peak_rss_gb_observed" not in arm:
            return
        bp = float(base["peak_rss_gb_observed"])
        ap = float(arm["peak_rss_gb_observed"])
        return {
            "delta_peak_rss_gb": bp - ap,
            "peak_reduction_ratio": (bp - ap) / bp if bp > 0 else None,
            "gate_peak_reduced": bool(ap < bp * 0.8 or (bp - ap) >= 0.2),
        }

    result["vs_hold_multibatch"] = _ratio(hold, multibatch)
    result["vs_hold_window"] = _ratio(hold, window)
    result["multibatch_vs_window"] = _ratio(window, multibatch)

    out_path = out_dir / "result.json"
    out_path.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(result, indent=2, ensure_ascii=False))
    print("wrote", out_path)


if __name__ == "__main__":
    main()
