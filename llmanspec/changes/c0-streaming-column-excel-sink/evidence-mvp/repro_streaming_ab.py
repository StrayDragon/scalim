#!/usr/bin/env python3
"""Pinned A/B: ColumnExcelSink hold vs evidence-MVP streaming (row-window).

生产代码不改动. 流式臂实现见本文件 `MvpStreamingColumnExcel`.

- hold: 全量列写入 ColumnExcelSink 后 close
- streaming_window: 按行窗写入全部列; 行字段齐备即 append 到 write_only 并释放行缓冲

小 shape 单元格对拍; 大 shape 独立子进程测 RSS.

输出: `.tmp/evidence/streaming-column-excel-ab/<ts>/result.json`
"""

from __future__ import annotations

import argparse
import gc
import json
import os
import subprocess
import sys
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Hashable, List, Mapping, Optional, Sequence, Set


def _rss_kb() -> int:
    try:
        with open("/proc/self/statm", "r", encoding="utf-8") as handle:
            parts = handle.read().strip().split()
        if not parts:
            return 0
        return int(int(parts[1]) * os.sysconf("SC_PAGE_SIZE") / 1024)
    except Exception:
        return 0


def _rss_gb() -> float:
    return _rss_kb() / (1024.0 * 1024.0)


def _build_window_columns(
    *,
    field_names: Sequence[str],
    row_ids: Sequence[int],
) -> Dict[str, Dict[int, int]]:
    columns: Dict[str, Dict[int, int]] = {}
    for i, name in enumerate(field_names):
        columns[name] = {r: r * 10 + i for r in row_ids}
    return columns


def _read_xlsx_matrix(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


class MvpStreamingColumnExcel:
    """证据用 MVP: 行字段齐备即 write_only append 并释放行缓冲."""

    def __init__(
        self,
        output_path: str,
        field_names: List[str],
        *,
        header_names: Optional[List[str]] = None,
        sheet_name: str = "Sheet1",
        include_header: bool = True,
        allow_formulas: bool = True,
    ) -> None:
        self.output_path = str(output_path)
        self.field_names = list(field_names)
        self.header_names = list(header_names) if header_names is not None else list(field_names)
        self.sheet_name = str(sheet_name)
        self.include_header = bool(include_header)
        self._allow_formulas = bool(allow_formulas)
        self._row_ids: List[Hashable] = []
        self._row_index: Dict[Hashable, int] = {}
        self._pending: List[Optional[Set[str]]] = []
        self._values: List[Optional[List[Any]]] = []
        self._field_index = {name: idx for idx, name in enumerate(self.field_names)}
        self._wb = None
        self._ws = None
        self._closed = False
        self._flushed_rows = 0

    def set_row_ids(self, row_ids: Sequence[Hashable]) -> None:
        if self._row_ids:
            raise RuntimeError("row_ids already set")
        from scalim._internal.utils.excel import escape_excel_formula
        from scalim.sinks._internal.excel import Workbook

        self._row_ids = list(row_ids)
        self._row_index = {pk: i for i, pk in enumerate(self._row_ids)}
        n = len(self._row_ids)
        self._pending = [set(self.field_names) for _ in range(n)]
        self._values = [[None] * len(self.field_names) for _ in range(n)]

        self._wb = Workbook(write_only=True)
        self._ws = self._wb.create_sheet(self.sheet_name)
        if self.include_header:
            _ = self._ws.append([escape_excel_formula(x, allow_formulas=self._allow_formulas) for x in self.header_names])

    def write_column(self, field_key: str, values: Mapping[Hashable, Any]) -> None:
        if self._closed:
            raise RuntimeError("sink closed")
        if not self._row_ids:
            raise RuntimeError("set_row_ids required")
        field = str(field_key)
        if field not in self._field_index:
            raise KeyError(field)
        from scalim._internal.utils.excel import escape_excel_formula

        fidx = self._field_index[field]
        completed: List[int] = []
        for pk, value in values.items():
            ridx = self._row_index.get(pk)
            if ridx is None:
                continue
            pending = self._pending[ridx]
            row_vals = self._values[ridx]
            if pending is None or row_vals is None:
                continue
            row_vals[fidx] = value
            pending.discard(field)
            if not pending:
                completed.append(ridx)

        for ridx in completed:
            row_vals = self._values[ridx]
            assert row_vals is not None
            assert self._ws is not None
            _ = self._ws.append([escape_excel_formula(x, allow_formulas=self._allow_formulas) for x in row_vals])
            self._values[ridx] = None
            self._pending[ridx] = None
            self._flushed_rows += 1

    def close(self) -> None:
        if self._closed:
            return
        from contextlib import suppress
        from pathlib import Path as PathCls

        from scalim._internal.utils.openpyxl_helpers import (
            best_effort_close_write_only_workbook_worksheets,
            save_openpyxl_workbook_atomic,
        )

        leftover = sum(1 for p in self._pending if p is not None)
        if leftover:
            raise RuntimeError("incomplete rows at close: {}".format(leftover))

        output_dir = PathCls(self.output_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)
        assert self._wb is not None
        try:
            save_openpyxl_workbook_atomic(self._wb, output_path=self.output_path)
        except Exception:
            best_effort_close_write_only_workbook_worksheets(self._wb)
            raise
        finally:
            with suppress(Exception):
                self._wb.close()
        self._closed = True


def run_hold(*, out_path: Path, n_rows: int, n_cols: int) -> Dict[str, Any]:
    from scalim.sinks import ColumnExcelSink

    field_names = ["f{}".format(i) for i in range(n_cols)]
    row_ids = list(range(n_rows))
    gc.collect()
    rss0 = _rss_gb()
    peak = {"v": rss0}
    stop = {"v": False}

    def _sampler() -> None:
        while not stop["v"]:
            peak["v"] = max(peak["v"], _rss_gb())
            time.sleep(0.05)

    thr = threading.Thread(target=_sampler)
    thr.daemon = True
    thr.start()
    t0 = time.perf_counter()
    try:
        columns = _build_window_columns(field_names=field_names, row_ids=row_ids)
        sink = ColumnExcelSink(str(out_path), field_names=field_names, include_header=True)
        sink.set_row_ids(row_ids)
        for name in field_names:
            sink.write_column(name, columns[name])
        del columns
        gc.collect()
        rss_pre = _rss_gb()
        peak["v"] = max(peak["v"], rss_pre)
        sink.close()
    finally:
        stop["v"] = True
        thr.join(timeout=2.0)
    elapsed = time.perf_counter() - t0
    rss_post = _rss_gb()
    peak["v"] = max(peak["v"], rss_post)
    return {
        "label": "hold",
        "n_rows": n_rows,
        "n_cols": n_cols,
        "cells": n_rows * n_cols,
        "window_rows": None,
        "rss_gb_pre_close": rss_pre,
        "rss_gb_post_close": rss_post,
        "peak_rss_gb_observed": peak["v"],
        "duration_s": elapsed,
        "bytes_on_disk": out_path.stat().st_size if out_path.exists() else 0,
        "flushed_rows": n_rows,
    }


def run_streaming_window(
    *,
    out_path: Path,
    n_rows: int,
    n_cols: int,
    window_rows: int,
) -> Dict[str, Any]:
    field_names = ["f{}".format(i) for i in range(n_cols)]
    row_ids = list(range(n_rows))
    window = max(1, int(window_rows))
    gc.collect()
    rss0 = _rss_gb()
    peak = {"v": rss0}
    stop = {"v": False}

    def _sampler() -> None:
        while not stop["v"]:
            peak["v"] = max(peak["v"], _rss_gb())
            time.sleep(0.05)

    thr = threading.Thread(target=_sampler)
    thr.daemon = True
    thr.start()
    t0 = time.perf_counter()
    rss_pre = rss0
    try:
        sink = MvpStreamingColumnExcel(str(out_path), field_names=field_names, include_header=True)
        sink.set_row_ids(row_ids)
        for start in range(0, n_rows, window):
            end = min(n_rows, start + window)
            win_ids = row_ids[start:end]
            columns = _build_window_columns(field_names=field_names, row_ids=win_ids)
            for name in field_names:
                sink.write_column(name, columns[name])
            del columns
            gc.collect()
            peak["v"] = max(peak["v"], _rss_gb())
        rss_pre = _rss_gb()
        peak["v"] = max(peak["v"], rss_pre)
        sink.close()
        flushed = sink._flushed_rows
    finally:
        stop["v"] = True
        thr.join(timeout=2.0)
    elapsed = time.perf_counter() - t0
    rss_post = _rss_gb()
    peak["v"] = max(peak["v"], rss_post)
    return {
        "label": "streaming_window",
        "n_rows": n_rows,
        "n_cols": n_cols,
        "cells": n_rows * n_cols,
        "window_rows": window,
        "rss_gb_pre_close": rss_pre,
        "rss_gb_post_close": rss_post,
        "peak_rss_gb_observed": peak["v"],
        "duration_s": elapsed,
        "bytes_on_disk": out_path.stat().st_size if out_path.exists() else 0,
        "flushed_rows": flushed,
    }


def _child(mode: str, out_path: Path, n_rows: int, n_cols: int, window_rows: int) -> Dict[str, Any]:
    script = (
        "import json,sys;\n"
        "from importlib.machinery import SourceFileLoader;\n"
        "from pathlib import Path;\n"
        "m=SourceFileLoader('ab', sys.argv[1]).load_module();\n"
        "mode=sys.argv[2]; out=Path(sys.argv[3]); rows=int(sys.argv[4]); cols=int(sys.argv[5]); win=int(sys.argv[6]);\n"
        "r=(m.run_hold(out_path=out, n_rows=rows, n_cols=cols) if mode=='hold' "
        "else m.run_streaming_window(out_path=out, n_rows=rows, n_cols=cols, window_rows=win));\n"
        "print(json.dumps(r));\n"
    )
    here = str(Path(__file__).resolve())
    proc = subprocess.run(
        [sys.executable, "-c", script, here, mode, str(out_path), str(n_rows), str(n_cols), str(window_rows)],
        cwd=str(Path.cwd()),
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        return {"error": "child_failed", "stderr": (proc.stderr or "")[-1200:], "code": proc.returncode}
    try:
        return json.loads(proc.stdout.strip().splitlines()[-1])
    except Exception as exc:
        return {"error": str(exc), "stdout": (proc.stdout or "")[-500:]}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", type=int, default=20000)
    parser.add_argument("--cols", type=int, default=200)
    parser.add_argument("--window-rows", type=int, default=2000)
    parser.add_argument("--correctness-rows", type=int, default=200)
    parser.add_argument("--correctness-cols", type=int, default=20)
    parser.add_argument("--out-dir", type=str, default="")
    args = parser.parse_args()

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_dir = Path(args.out_dir or os.path.join(".tmp", "evidence", "streaming-column-excel-ab", ts))
    out_dir.mkdir(parents=True, exist_ok=True)
    work = out_dir / "work"
    work.mkdir(parents=True, exist_ok=True)

    hold_small = work / "small_hold.xlsx"
    stream_small = work / "small_stream.xlsx"
    r_hold_s = run_hold(out_path=hold_small, n_rows=args.correctness_rows, n_cols=args.correctness_cols)
    r_stream_s = run_streaming_window(
        out_path=stream_small,
        n_rows=args.correctness_rows,
        n_cols=args.correctness_cols,
        window_rows=max(1, args.correctness_rows // 4),
    )
    m_hold = _read_xlsx_matrix(hold_small)
    m_stream = _read_xlsx_matrix(stream_small)
    correctness = {
        "ok": m_hold == m_stream,
        "hold_rows": len(m_hold),
        "stream_rows": len(m_stream),
        "hold_arm": r_hold_s,
        "stream_arm": r_stream_s,
    }

    hold = _child("hold", work / "hold.xlsx", args.rows, args.cols, args.window_rows)
    stream = _child("streaming_window", work / "stream.xlsx", args.rows, args.cols, args.window_rows)

    result: Dict[str, Any] = {
        "ts": ts,
        "mvp": "MvpStreamingColumnExcel+row_window",
        "shape": {"rows": args.rows, "cols": args.cols, "window_rows": args.window_rows},
        "correctness": correctness,
        "hold": hold,
        "streaming_window": stream,
    }
    if isinstance(hold, dict) and isinstance(stream, dict) and "peak_rss_gb_observed" in hold and "peak_rss_gb_observed" in stream:
        hold_peak = float(hold["peak_rss_gb_observed"])
        stream_peak = float(stream["peak_rss_gb_observed"])
        result["delta_peak_rss_gb"] = hold_peak - stream_peak
        result["peak_reduction_ratio"] = (hold_peak - stream_peak) / hold_peak if hold_peak > 0 else None
        result["gate_peak_reduced"] = bool(stream_peak < hold_peak * 0.8 or (hold_peak - stream_peak) >= 0.2)

    out_path = out_dir / "result.json"
    out_path.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(result, indent=2, ensure_ascii=False))
    print("wrote", out_path)


if __name__ == "__main__":
    main()
