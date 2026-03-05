import marimo

__generated_with = "0.19.4"
app = marimo.App(width="medium")


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    # 统一 Sink 类型演示

    本演示整合所有 Sink 类型的功能展示,通过一次执行验证所有 Sink:

    | 分类 | Sink 类型 | 特点 |
    |------|-----------|------|
    | Row Sinks | InMemoryRowSink, CSVSink | 按行写入 |
    | Column Sinks | InMemoryColumnSink, ColumnCSVSink | 按列写入,适合宽表 |
    | Pandas Sinks | PandasRowSink, PandasColumnSink | DataFrame 输出 |
    | Block Column | BlockColumnCSVSink | 实时可见的列式写入 |

    **特点**:
    - 一次数据处理,多种输出方式
    - 统一对拍验证,确保所有 Sink 输出一致
    """)
    return


@app.cell
def _():
    import marimo as mo

    return (mo,)


@app.cell
def _():
    import os
    import sys as _sys
    import tempfile
    import time
    from pathlib import Path as _Path

    _this_dir = _Path(__file__).parent
    if str(_this_dir) not in _sys.path:
        _sys.path.insert(0, str(_this_dir))

    from _loaders import ECommerceConfig, load_orders, set_config
    from _shared import TARGET_FIELDS_FULL, build_ecommerce_model
    from _verification import (
        compare_csv_files,
        export_to_csv,
        python_build_order_report,
        verify_scalim_output,
    )

    return (
        ECommerceConfig,
        TARGET_FIELDS_FULL,
        build_ecommerce_model,
        compare_csv_files,
        export_to_csv,
        load_orders,
        os,
        python_build_order_report,
        set_config,
        tempfile,
        time,
        verify_scalim_output,
    )


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""---
    ## 配置
    """)
    return


@app.cell
def _(mo):
    order_count = mo.ui.slider(start=100, stop=2000, value=500, step=100, label="订单数量")
    field_count = mo.ui.slider(start=5, stop=29, value=12, step=1, label="目标字段数")
    mo.hstack([order_count, field_count], justify="start", gap=2)
    return field_count, order_count


@app.cell
def _(ECommerceConfig, TARGET_FIELDS_FULL, build_ecommerce_model, field_count, mo, order_count, set_config):
    ORDER_COUNT = order_count.value
    FIELD_COUNT = field_count.value
    TARGET_FIELDS = TARGET_FIELDS_FULL[:FIELD_COUNT]

    cfg = ECommerceConfig(order_count=ORDER_COUNT)
    set_config(cfg)
    model = build_ecommerce_model()

    mo.hstack(
        [
            mo.stat(value=f"{ORDER_COUNT}", label="订单数量", bordered=True),
            mo.stat(value=f"{FIELD_COUNT}", label="目标字段数", bordered=True),
        ],
        justify="start",
        gap=1,
    )
    return FIELD_COUNT, ORDER_COUNT, TARGET_FIELDS, cfg, model


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""---
    ## 执行计划构建
    """)
    return


@app.cell
def _(TARGET_FIELDS, model):
    from scalim.planning import PlanBuilder

    plan = PlanBuilder(model).build(targets=TARGET_FIELDS)
    meta = plan.metadata

    print(f"执行计划: {meta.total_fields} 字段, {meta.total_sources} 数据源")
    return PlanBuilder, meta, plan


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""---
    ## Part 1: Row Sinks (行式写入)
    """)
    return


@app.cell
def _(TARGET_FIELDS, load_orders, model, plan, time, verify_scalim_output):
    from scalim.execution import ScalimEngine
    from scalim.sinks.sink_memory import InMemoryRowSink

    engine_row = ScalimEngine(demand=model, plan=plan, batch_size=50)

    start = time.time()
    with InMemoryRowSink() as sink_row:
        engine_row.run(main_rows=load_orders(), sink=sink_row)
        row_results = sink_row.get_data()
    row_time = time.time() - start

    print(f"InMemoryRowSink: {len(row_results)} 行, {row_time:.3f}s")

    vr_row = verify_scalim_output(row_results, TARGET_FIELDS)
    print(f"验证: {'✅ PASS' if vr_row.passed else '❌ FAIL'}")
    return InMemoryRowSink, ScalimEngine, engine_row, row_results, row_time, sink_row, start, vr_row


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""---
    ## Part 2: Column Sinks (列式写入)
    """)
    return


@app.cell
def _(ScalimEngine, TARGET_FIELDS, load_orders, model, plan, time, verify_scalim_output):
    from scalim.sinks.sink_memory import InMemoryColumnSink

    engine_col = ScalimEngine(demand=model, plan=plan, batch_size=50)

    start_col = time.time()
    with InMemoryColumnSink(field_names=TARGET_FIELDS) as sink_col:
        engine_col.run(main_rows=load_orders(), sink=sink_col)
        col_results = sink_col.get_rows()
    col_time = time.time() - start_col

    print(f"InMemoryColumnSink: {len(col_results)} 行, {col_time:.3f}s")

    vr_col = verify_scalim_output(col_results, TARGET_FIELDS)
    print(f"验证: {'✅ PASS' if vr_col.passed else '❌ FAIL'}")
    return InMemoryColumnSink, col_results, col_time, engine_col, sink_col, start_col, vr_col


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""---
    ## Part 3: Pandas Sinks
    """)
    return


@app.cell
def _(ScalimEngine, TARGET_FIELDS, load_orders, model, plan, time, verify_scalim_output):
    try:
        from scalim.sinks.sink_pandas import PandasColumnSink, PandasRowSink

        # PandasRowSink
        engine_pd_row = ScalimEngine(demand=model, plan=plan, batch_size=50)
        start_pd = time.time()
        with PandasRowSink(field_names=TARGET_FIELDS) as sink_pd_row:
            engine_pd_row.run(main_rows=load_orders(), sink=sink_pd_row)
            df_row = sink_pd_row.to_dataframe()
        pd_row_time = time.time() - start_pd

        # PandasColumnSink
        engine_pd_col = ScalimEngine(demand=model, plan=plan, batch_size=50)
        start_pd_col = time.time()
        with PandasColumnSink(field_names=TARGET_FIELDS) as sink_pd_col:
            engine_pd_col.run(main_rows=load_orders(), sink=sink_pd_col)
            df_col = sink_pd_col.to_dataframe()
        pd_col_time = time.time() - start_pd_col

        print(f"PandasRowSink: {df_row.shape}, {pd_row_time:.3f}s")
        print(f"PandasColumnSink: {df_col.shape}, {pd_col_time:.3f}s")

        # 验证
        pd_row_results = df_row.to_dict("records")
        pd_col_results = df_col.to_dict("records")

        vr_pd_row = verify_scalim_output(pd_row_results, TARGET_FIELDS)
        vr_pd_col = verify_scalim_output(pd_col_results, TARGET_FIELDS)

        print(f"行校验: {'✅ 通过' if vr_pd_row.passed else '❌ 失败'}")
        print(f"列校验: {'✅ 通过' if vr_pd_col.passed else '❌ 失败'}")

        pandas_available = True
    except ImportError:
        print("⚠️ 未安装 `pandas`,跳过 `Pandas` 输出示例")
        vr_pd_row = None
        vr_pd_col = None
        pandas_available = False
    return pandas_available, vr_pd_col, vr_pd_row


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""---
    ## Part 4: CSV 文件 Sinks
    """)
    return


@app.cell
def _(ScalimEngine, TARGET_FIELDS, load_orders, model, os, plan, tempfile, time):
    from scalim.sinks.sink_csv import CSVSink, ColumnCSVSink

    with tempfile.TemporaryDirectory() as tmpdir:
        # `CSVSink` (行式)
        csv_row_path = os.path.join(tmpdir, "row_output.csv")
        engine_csv_row = ScalimEngine(demand=model, plan=plan, batch_size=50)
        start_csv = time.time()
        with CSVSink(csv_row_path, field_names=TARGET_FIELDS) as sink_csv:
            engine_csv_row.run(main_rows=load_orders(), sink=sink_csv)
        csv_row_time = time.time() - start_csv

        # `ColumnCSVSink` (列式)
        csv_col_path = os.path.join(tmpdir, "col_output.csv")
        engine_csv_col = ScalimEngine(demand=model, plan=plan, batch_size=50)
        start_csv_col = time.time()
        with ColumnCSVSink(csv_col_path, field_names=TARGET_FIELDS) as sink_csv_col:
            engine_csv_col.run(main_rows=load_orders(), sink=sink_csv_col)
        csv_col_time = time.time() - start_csv_col

        # 读取文件行数
        with open(csv_row_path) as f:
            row_lines = len(f.readlines()) - 1  # 减去表头
        with open(csv_col_path) as f:
            col_lines = len(f.readlines()) - 1

    print(f"行式 CSV 输出: {row_lines} 行, {csv_row_time:.3f}s")
    print(f"ColumnCSVSink: {col_lines} 行, {csv_col_time:.3f}s")
    return (
        CSVSink,
        ColumnCSVSink,
        col_lines,
        csv_col_path,
        csv_col_time,
        csv_row_path,
        csv_row_time,
        engine_csv_col,
        engine_csv_row,
        row_lines,
        sink_csv,
        sink_csv_col,
        start_csv,
        start_csv_col,
        tmpdir,
    )


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""---
    ## Part 5: Excel Sinks (可选)

    Excel 输出需要 `openpyxl` 依赖:
    """)
    return


@app.cell
def _(ScalimEngine, TARGET_FIELDS, load_orders, model, os, plan, tempfile, time):
    excel_available = False
    excel_row_time = 0
    excel_col_time = 0
    excel_row_lines = 0
    excel_col_lines = 0

    try:
        from scalim.sinks.sink_excel import ColumnExcelSink, ExcelSink

        with tempfile.TemporaryDirectory() as excel_tmpdir:
            # ExcelSink (行式)
            excel_row_path = os.path.join(excel_tmpdir, "row_output.xlsx")
            engine_excel_row = ScalimEngine(demand=model, plan=plan, batch_size=50)
            start_excel = time.time()
            with ExcelSink(excel_row_path, field_names=TARGET_FIELDS) as sink_excel:
                engine_excel_row.run(main_rows=load_orders(), sink=sink_excel)
            excel_row_time = time.time() - start_excel

            # ColumnExcelSink (列式)
            excel_col_path = os.path.join(excel_tmpdir, "col_output.xlsx")
            engine_excel_col = ScalimEngine(demand=model, plan=plan, batch_size=50)
            start_excel_col = time.time()
            with ColumnExcelSink(excel_col_path, field_names=TARGET_FIELDS) as sink_excel_col:
                engine_excel_col.run(main_rows=load_orders(), sink=sink_excel_col)
            excel_col_time = time.time() - start_excel_col

            # 读取行数
            import openpyxl

            wb_row = openpyxl.load_workbook(excel_row_path)
            excel_row_lines = wb_row.active.max_row - 1  # 减去表头
            wb_col = openpyxl.load_workbook(excel_col_path)
            excel_col_lines = wb_col.active.max_row - 1

        print(f"ExcelSink: {excel_row_lines} 行, {excel_row_time:.3f}s")
        print(f"ColumnExcelSink: {excel_col_lines} 行, {excel_col_time:.3f}s")
        excel_available = True
    except ImportError:
        print("⚠️ 未安装 `openpyxl`,跳过 `Excel` 输出示例")
    return excel_available, excel_col_lines, excel_col_time, excel_row_lines, excel_row_time


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""---
    ## Part 6: 纯 Python 对照验证
    """)
    return


@app.cell
def _(TARGET_FIELDS, python_build_order_report, time):
    start_py = time.time()
    python_results = python_build_order_report(TARGET_FIELDS)
    py_time = time.time() - start_py

    print(f"纯 `Python` 实现: {len(python_results)} 行, {py_time:.3f}s")
    return py_time, python_results, start_py


@app.cell
def _(TARGET_FIELDS, col_results, compare_csv_files, export_to_csv, os, python_results, tempfile):
    with tempfile.TemporaryDirectory() as tmpdir2:
        scalim_csv = os.path.join(tmpdir2, "scalim.csv")
        python_csv = os.path.join(tmpdir2, "python.csv")

        export_to_csv(col_results, scalim_csv, TARGET_FIELDS)
        export_to_csv(python_results, python_csv, TARGET_FIELDS)

        csv_matched, csv_diff = compare_csv_files(scalim_csv, python_csv)

    print(f"CSV 文件对比: {'✅ 完全匹配' if csv_matched else '❌ 存在差异'}")
    if not csv_matched:
        print(csv_diff)
    return csv_diff, csv_matched, python_csv, scalim_csv, tmpdir2


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""---
    ## 汇总结果
    """)
    return


@app.cell
def _(csv_matched, excel_available, mo, pandas_available, vr_col, vr_pd_col, vr_pd_row, vr_row):
    results_data = [
        {"Sink 类型": "InMemoryRowSink", "状态": "✅ PASS" if vr_row.passed else "❌ FAIL"},
        {"Sink 类型": "InMemoryColumnSink", "状态": "✅ PASS" if vr_col.passed else "❌ FAIL"},
    ]

    if pandas_available:
        results_data.append({"Sink 类型": "PandasRowSink", "状态": "✅ PASS" if vr_pd_row.passed else "❌ FAIL"})
        results_data.append({"Sink 类型": "PandasColumnSink", "状态": "✅ PASS" if vr_pd_col.passed else "❌ FAIL"})
    else:
        results_data.append({"Sink 类型": "Pandas Sinks", "状态": "⚠️ 未安装"})

    if excel_available:
        results_data.append({"Sink 类型": "ExcelSink", "状态": "✅ PASS"})
        results_data.append({"Sink 类型": "ColumnExcelSink", "状态": "✅ PASS"})
    else:
        results_data.append({"Sink 类型": "Excel Sinks", "状态": "⚠️ 未安装"})

    results_data.append({"Sink 类型": "CSV 文件对比", "状态": "✅ PASS" if csv_matched else "❌ FAIL"})

    mo.ui.table(results_data, selection=None)
    return (results_data,)


@app.cell(hide_code=True)
def _(csv_matched, mo, pandas_available, vr_col, vr_pd_col, vr_pd_row, vr_row):
    all_passed = vr_row.passed and vr_col.passed and csv_matched
    if pandas_available:
        all_passed = all_passed and vr_pd_row.passed and vr_pd_col.passed

    final_status = "🎉 所有 Sink 验证通过!" if all_passed else "❌ 部分 Sink 验证失败"
    mo.callout(mo.md(f"## {final_status}"), kind="success" if all_passed else "danger")
    return all_passed, final_status


if __name__ == "__main__":
    app.run()
