import marimo

import csv
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from scalim.dsl.yaml_dsl import RunOptions, run_workflow
from scalim.execution import versioned_outputs
from scalim_misc.demo_big_data_report.cases import build_test_config_small
from scalim_misc.demo_big_data_report.loaders import ECommerceConfig, get_config, set_config
from scalim_misc.examples._types import EXAMPLE_KIND_ORACLE, ExampleResult

__generated_with = "0.20.2"
app = marimo.App(width="full")

_EXAMPLE_ID = "demo_big_data_report/workflow_shared_workbooks"


def _read_csv_header_and_count_rows(path: Path) -> Tuple[List[str], int]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader, [])
        data_rows = sum(1 for _ in reader)
    return [str(x) for x in header], int(data_rows)


def _count_sheet_rows_and_header(workbook: Any, sheet_name: str) -> Tuple[Optional[List[str]], int]:
    if sheet_name not in workbook.sheetnames:
        return None, 0
    ws = workbook[sheet_name]
    it = ws.iter_rows(values_only=True)
    first = next(it, None)
    if first is None:
        return [], 0
    header = ["" if v is None else str(v) for v in first]
    rest_count = sum(1 for _ in it)
    return header, 1 + int(rest_count)


def _resolve_latest_book_artifact(out_root: Path, *, book_id: str) -> Tuple[Optional[Path], Dict[str, Any]]:
    """通过 `manifest/latest.json` 定位 `<out_root>/versions/<version_id>/books/<book_id>.xlsx`."""
    try:
        latest = versioned_outputs.read_latest(out_root)
        version_id = str(latest.get("version_id") or "")
        if not version_id:
            return None, {"latest": latest}
        manifest = versioned_outputs.read_version_manifest(out_root, version_id=version_id)
        books = manifest.get("books") if isinstance(manifest, dict) else None
        rel = books.get(book_id) if isinstance(books, dict) else None
        if not rel:
            rel = versioned_outputs.book_output_relpath(book_id=str(book_id))
        return out_root / "versions" / version_id / str(rel), {"latest": latest, "manifest": manifest}
    except Exception as exc:  # noqa: BLE001
        return None, {"exc_type": type(exc).__name__, "message": str(exc)}


def run_workflow_shared_workbooks(
    cfg: Optional[ECommerceConfig] = None,
    *,
    workflow_yaml_path: Optional[Path] = None,
) -> ExampleResult:
    if cfg is None:
        cfg = build_test_config_small()
    if workflow_yaml_path is None:
        demo_dir = Path(__file__).resolve().parents[1]
        workflow_yaml_path = demo_dir / "chapters_of_yaml_dsl" / "declared_yaml_dsl" / "workflow_demo_shared_workbooks.yaml"

    prev = get_config()
    set_config(cfg)
    try:
        allowed_modules = frozenset(["scalim_misc.demo_big_data_report.loaders"])
        repo_root = Path(__file__).resolve().parents[4]

        with tempfile.TemporaryDirectory(prefix="scalim-wf-wb-") as temp_dir:
            out_dir = Path(temp_dir).resolve()
            wf_copy = out_dir / "workflow.yaml"
            wf_copy.write_text(workflow_yaml_path.read_text(encoding="utf-8"), encoding="utf-8")
            demand_dir = workflow_yaml_path.parent
            (out_dir / "workflow_demo_shared_workbooks_demand.yaml").write_text(
                (demand_dir / "workflow_demo_shared_workbooks_demand.yaml").read_text(encoding="utf-8"),
                encoding="utf-8",
            )

            try:
                wf_result = run_workflow(
                    str(wf_copy),
                    options=RunOptions(
                        allowed_modules=allowed_modules,
                        batch_size=30,
                        init_vars={"order_ids": []},
                        allowed_yaml_roots=(str(repo_root),),
                    ),
                    path_aliases={"@": str(repo_root)},
                )
            except Exception as exc:  # noqa: BLE001
                summary = "workflow failed: {}: {}".format(type(exc).__name__, exc)
                return ExampleResult(
                    example_id=_EXAMPLE_ID,
                    passed=False,
                    kind=EXAMPLE_KIND_ORACLE,
                    summary=summary,
                    details={"exc_type": type(exc).__name__, "message": str(exc)},
                )

            errors = wf_result.errors()
            shared_report_sheet_root = out_dir / "out" / "shared_report_sheet"
            shared_report_append_root = out_dir / "out" / "shared_report_append"
            sheetbook_report_sheet_root = out_dir / "out" / "sheetbook_report_sheet"
            sheetbook_report_append_root = out_dir / "out" / "sheetbook_report_append"

            shared_report_sheet_xlsx, shared_sheet_meta = _resolve_latest_book_artifact(
                shared_report_sheet_root, book_id="shared_report_sheet"
            )
            shared_report_append_xlsx, shared_append_meta = _resolve_latest_book_artifact(
                shared_report_append_root, book_id="shared_report_append"
            )
            sheetbook_report_sheet_xlsx, sheetbook_sheet_meta = _resolve_latest_book_artifact(
                sheetbook_report_sheet_root, book_id="sheetbook_report_sheet"
            )
            sheetbook_report_append_xlsx, sheetbook_append_meta = _resolve_latest_book_artifact(
                sheetbook_report_append_root, book_id="sheetbook_report_append"
            )

            artifacts_ok = bool(
                shared_report_sheet_xlsx
                and shared_report_append_xlsx
                and sheetbook_report_sheet_xlsx
                and sheetbook_report_append_xlsx
                and shared_report_sheet_xlsx.exists()
                and shared_report_append_xlsx.exists()
                and sheetbook_report_sheet_xlsx.exists()
                and sheetbook_report_append_xlsx.exists()
            )

            wb_ok = False
            sb_ok = False
            wb_checks: Dict[str, Any] = {}
            sb_checks: Dict[str, Any] = {}

            if artifacts_ok:
                from openpyxl import load_workbook

                wb_sheet = load_workbook(shared_report_sheet_xlsx, read_only=True, data_only=True)  # type: ignore[arg-type]
                try:
                    wb_detail_header, wb_detail_rows = _count_sheet_rows_and_header(wb_sheet, "Detail")
                finally:
                    wb_sheet.close()

                wb_append = load_workbook(shared_report_append_xlsx, read_only=True, data_only=True)  # type: ignore[arg-type]
                try:
                    wb_append_header, wb_append_rows = _count_sheet_rows_and_header(wb_append, "DetailAppend")
                finally:
                    wb_append.close()

                single_run_rows = int(wb_detail_rows) - 1 if wb_detail_rows else 0
                expected_append_rows = 1 + (2 * single_run_rows) if single_run_rows >= 0 else 0
                wb_ok = bool(
                    wb_detail_header
                    and wb_append_header
                    and single_run_rows > 0
                    and wb_detail_header == wb_append_header
                    and wb_detail_rows == 1 + single_run_rows
                    and wb_append_rows == expected_append_rows
                )
                wb_checks = {
                    "header": wb_detail_header,
                    "detail": {"rows_total": wb_detail_rows, "expected": 1 + single_run_rows},
                    "detail_append": {"rows_total": wb_append_rows, "expected": expected_append_rows},
                }

                sb_sheet = load_workbook(sheetbook_report_sheet_xlsx, read_only=True, data_only=True)  # type: ignore[arg-type]
                try:
                    sb_detail_header, sb_detail_rows = _count_sheet_rows_and_header(sb_sheet, "Detail")
                finally:
                    sb_sheet.close()

                sb_append = load_workbook(sheetbook_report_append_xlsx, read_only=True, data_only=True)  # type: ignore[arg-type]
                try:
                    sb_append_header, sb_append_rows = _count_sheet_rows_and_header(sb_append, "DetailAppend")
                finally:
                    sb_append.close()

                sb_ok = bool(
                    wb_ok
                    and sb_detail_header == wb_detail_header
                    and sb_append_header == wb_detail_header
                    and sb_detail_rows == 1 + single_run_rows
                    and sb_append_rows == expected_append_rows
                )
                sb_checks = {
                    "header": sb_detail_header,
                    "detail": {"rows_total": sb_detail_rows, "expected": 1 + single_run_rows},
                    "detail_append": {"rows_total": sb_append_rows, "expected": expected_append_rows},
                }

            passed = bool(not errors and artifacts_ok and wb_ok and sb_ok)
            summary = "errors={} artifacts_ok={} wb_ok={} sheetbook_ok={}".format(
                len(errors),
                artifacts_ok,
                wb_ok,
                sb_ok,
            )
            if errors:
                summary = summary + "\nfirst_error: {} {}".format(errors[0].exc_type, errors[0].message)

            details: Dict[str, Any] = {
                "output_dir": str(out_dir),
                "workflow_yaml_path": str(workflow_yaml_path),
                "output_roots": {
                    "shared_report_sheet": str(shared_report_sheet_root),
                    "shared_report_append": str(shared_report_append_root),
                    "sheetbook_report_sheet": str(sheetbook_report_sheet_root),
                    "sheetbook_report_append": str(sheetbook_report_append_root),
                },
                "artifacts": {
                    "shared_report_sheet_xlsx": str(shared_report_sheet_xlsx) if shared_report_sheet_xlsx else None,
                    "shared_report_append_xlsx": str(shared_report_append_xlsx) if shared_report_append_xlsx else None,
                    "sheetbook_report_sheet_xlsx": str(sheetbook_report_sheet_xlsx) if sheetbook_report_sheet_xlsx else None,
                    "sheetbook_report_append_xlsx": str(sheetbook_report_append_xlsx) if sheetbook_report_append_xlsx else None,
                },
                "versioned_resolve": {
                    "shared_report_sheet": shared_sheet_meta,
                    "shared_report_append": shared_append_meta,
                    "sheetbook_report_sheet": sheetbook_sheet_meta,
                    "sheetbook_report_append": sheetbook_append_meta,
                },
                "workbook": wb_checks,
                "sheetbook": sb_checks,
                "errors": errors,
                "outcomes": wf_result.outcomes,
            }
            return ExampleResult(
                example_id=_EXAMPLE_ID,
                passed=passed,
                kind=EXAMPLE_KIND_ORACLE,
                summary=summary,
                details=details,
            )
    finally:
        set_config(prev)


def run_chapter() -> ExampleResult:
    return run_workflow_shared_workbooks()


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
        # demo_big_data_report / workflow_shared_workbooks

        ## 背景

        当团队开始把多份 demand YAML 组合成一个 workflow 后,很快会遇到“共享产物”的需求:

        - 多个节点把输出写入同一个 `xlsx`(不同 sheet)
        - 需要明确 append/overwrite 语义,并且可回归验证

        这就是 workflow YAML 的 `workflow.resources.books` + book-level 写策略(`write_defaults`) + demand outputs 绑定:

        - `workflow.resources.books`: 声明 workflow-scope 的共享 book 资源(`xlsx_file|xlsx_memory`)
        - `resources.books.*.write_defaults`: 声明该 book 的写入语义(sheet/append + 冲突策略等)
        - demand YAML 通过 `outputs[*].to.book/to.sheet` 绑定到共享 book; workflow 编译期推导 write nodes 并保证写入顺序确定性

        ## 需求方提问（自然语言）

        平台同学：我想把两个节点的明细输出追加到同一个报表里,能不能用 YAML 显式声明 append 语义并在 CI 里对拍？

        ## 对拍点（deterministic）

        - 两次相同的明细输出 append 到同一个 sheet,行数应为 2 倍
        - sheet mode(overwrite) 写入一次,行数应为 1 倍
        - workbook 与 sheetbook 导出的 header 与 CSV header 必须一致

        SSOT:
        - `notebooks/marimo/demo_big_data_report/chapters_of_yaml_dsl/ch090_workflow_shared_workbooks.py::run_workflow_shared_workbooks`
        """
    )
    return


@app.cell
def _():
    import marimo as mo

    return (mo,)


@app.cell
def _():
    from scalim_misc.notebook_support.pathing import ensure_repo_root_on_sys_path

    _ = ensure_repo_root_on_sys_path(__file__)
    demo_dir = Path(__file__).resolve().parents[1]
    workflow_yaml_path = demo_dir / "chapters_of_yaml_dsl" / "declared_yaml_dsl" / "workflow_demo_shared_workbooks.yaml"
    return demo_dir, workflow_yaml_path


@app.cell(hide_code=True)
def _(mo, workflow_yaml_path):
    from scalim_misc.notebook_support.yaml_excerpt import excerpt_head

    mo.md("## Workflow YAML (head)")
    mo.md("```yaml\n{}\n```".format(excerpt_head(workflow_yaml_path, max_lines=180)))
    return (excerpt_head,)


@app.cell
def _(workflow_yaml_path):
    cfg = build_test_config_small()
    result = run_workflow_shared_workbooks(cfg, workflow_yaml_path=workflow_yaml_path)
    return cfg, result


@app.cell(hide_code=True)
def _(mo, result):
    mo.callout(mo.md("## {}".format("PASS" if result.passed else "FAIL")), kind="success" if result.passed else "danger")
    mo.md("```\n{}\n```".format(result.summary))
    return


@app.cell(hide_code=True)
def _(mo, result):
    from scalim_misc.notebook_support.results_view import details_to_rows

    rows = details_to_rows(result.details)
    mo.ui.table(rows, selection=None)
    return (rows,)


if __name__ == "__main__":
    app.run()
