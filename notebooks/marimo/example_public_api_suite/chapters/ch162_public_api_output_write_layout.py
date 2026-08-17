import marimo

import importlib
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook

from scalim import execution as api
from scalim.events import Event, EventType
from scalim.ob.observer import Observer
from scalim.planning import PlanBuilder
from scalim.sinks import ColumnExcelSink, ExcelSink, StreamingColumnExcelSink
from scalim_misc.examples._types import EXAMPLE_KIND_ORACLE, ExampleResult
from scalim_misc.examples.public_api._fixtures import build_minimal_public_api_ir, build_minimal_public_api_runtime_bindings

__generated_with = "0.22.0"
app = marimo.App(width="full")
_EXAMPLE_ID = "example_public_api_suite/ch162_public_api_output_write_layout"

_run_ir_mod = importlib.import_module("scalim.execution.run_ir")


class _WriteTraceObserver(Observer):
    def __init__(self) -> None:
        self.event_types: Optional[Set[EventType]] = {
            EventType.ROW_WRITE,
            EventType.COLUMN_WRITE,
            EventType.PIPELINE_END,
        }
        self.row_field_counts: List[int] = []
        self.column_keys: List[str] = []
        self.column_row_counts: List[int] = []

    def on_event(self, event: Event) -> None:
        if event.event_type is EventType.ROW_WRITE:
            self.row_field_counts.append(int(getattr(event.payload, "field_count", 0)))
            return
        if event.event_type is EventType.COLUMN_WRITE:
            self.column_keys.append(str(getattr(event.payload, "field_key", "")))
            self.column_row_counts.append(int(getattr(event.payload, "row_count", 0)))


def _xlsx_rows(path: Path) -> List[Tuple[Any, ...]]:
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    try:
        return [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _run_layout(
    *,
    demand_ir: Any,
    layout: Any,
    runtime_bindings: Any,
    path: Path,
    output_write_layout: api.OutputWriteLayout,
    streaming: bool,
) -> Tuple[Any, _WriteTraceObserver]:
    if path.exists():
        path.unlink()
    observer = _WriteTraceObserver()
    result = api.run_ir(
        demand_ir,
        api.ExecutionRequest(
            export_layout=layout,
            output=api.OutputSpec(format="excel", path=str(path), streaming=streaming, include_header=True),
            runtime_bindings=runtime_bindings,
            batch_size=10,
            output_write_layout=output_write_layout,
            components=[observer],
        ),
    )
    return result, observer


def run_public_api_output_write_layout(tmp_dir: Optional[Path] = None) -> ExampleResult:
    """演示 `OutputWriteLayout`：工厂选型 + 事件取向 + 业务格子对拍。

    按数据形状由调用方显式调优；默认不设 `layout` 时行为与历史一致。
    `YAML` `books` / `composition` 不能设列布局。
    """
    base = tmp_dir if tmp_dir is not None else Path(".tmp/examples/ch162_output_write_layout")
    base.mkdir(parents=True, exist_ok=True)

    demand_ir = build_minimal_public_api_ir()
    runtime_bindings = build_minimal_public_api_runtime_bindings()
    plan = PlanBuilder(demand_ir).build()
    layout = api.export_layout_from_demand_ir(demand_ir, plan.target_fields)

    buffered_probe = base / "probe_buffered.xlsx"
    chunked_probe = base / "probe_chunked.xlsx"
    row_probe = base / "probe_row.xlsx"
    buffered_sink = _run_ir_mod._create_file_sink(
        api.OutputSpec(format="excel", path=str(buffered_probe), streaming=False),
        layout,
        output_write_layout=api.OutputWriteLayout.COLUMN_BUFFERED,
    )
    chunked_sink = _run_ir_mod._create_file_sink(
        api.OutputSpec(format="excel", path=str(chunked_probe), streaming=False),
        layout,
        output_write_layout=api.OutputWriteLayout.COLUMN_CHUNKED,
    )
    row_sink = _run_ir_mod._create_file_sink(
        api.OutputSpec(format="excel", path=str(row_probe), streaming=True),
        layout,
        output_write_layout=api.OutputWriteLayout.ROW_STREAM,
    )
    factory_ok = (
        isinstance(buffered_sink, ColumnExcelSink)
        and isinstance(chunked_sink, StreamingColumnExcelSink)
        and isinstance(row_sink, ExcelSink)
    )
    buffered_sink.close()
    row_sink.close()
    # 未写入的 `chunked sink`：不调用 `close`（`close` 需要先 `set_row_ids`）

    buffered_out = base / "run_buffered.xlsx"
    chunked_out = base / "run_chunked.xlsx"
    row_out = base / "run_row.xlsx"
    buffered_result, buffered_obs = _run_layout(
        demand_ir=demand_ir,
        layout=layout,
        runtime_bindings=runtime_bindings,
        path=buffered_out,
        output_write_layout=api.OutputWriteLayout.COLUMN_BUFFERED,
        streaming=False,
    )
    chunked_result, chunked_obs = _run_layout(
        demand_ir=demand_ir,
        layout=layout,
        runtime_bindings=runtime_bindings,
        path=chunked_out,
        output_write_layout=api.OutputWriteLayout.COLUMN_CHUNKED,
        streaming=False,
    )
    row_result, row_obs = _run_layout(
        demand_ir=demand_ir,
        layout=layout,
        runtime_bindings=runtime_bindings,
        path=row_out,
        output_write_layout=api.OutputWriteLayout.ROW_STREAM,
        streaming=True,
    )

    derived = api.resolve_output_write_layout(
        output_write_layout=None,
        streaming=False,
        output_format="excel",
        excel_column_residency=api.ExcelColumnResidency.CHUNKED,
        has_output_composition=False,
    )

    buffered_rows = _xlsx_rows(buffered_out)
    chunked_rows = _xlsx_rows(chunked_out)
    expected_fields = set(layout.field_ids)
    column_events_ok = (
        not buffered_obs.row_field_counts
        and not chunked_obs.row_field_counts
        and set(buffered_obs.column_keys) == expected_fields
        and set(chunked_obs.column_keys) == expected_fields
        and buffered_obs.column_row_counts == chunked_obs.column_row_counts
        and bool(buffered_obs.column_row_counts)
        and buffered_obs.column_row_counts[0] == 3
    )
    row_events_ok = (
        len(row_obs.row_field_counts) == 3
        and not row_obs.column_keys
        and all(count == len(expected_fields) for count in row_obs.row_field_counts)
    )

    passed = bool(
        factory_ok
        and buffered_result.total_rows == chunked_result.total_rows == row_result.total_rows == 3
        and buffered_out.exists()
        and chunked_out.exists()
        and row_out.exists()
        and buffered_rows == chunked_rows
        and derived is api.OutputWriteLayout.COLUMN_CHUNKED
        and column_events_ok
        and row_events_ok
        and not hasattr(api.OutputWriteLayout, "COLUMN_HOLD")
        and not hasattr(api.OutputWriteLayout, "COLUMN_WINDOW")
    )
    details: Dict[str, Any] = {
        "factory_buffered": type(buffered_sink).__name__,
        "factory_chunked": type(chunked_sink).__name__,
        "factory_row": type(row_sink).__name__,
        "buffered_rows": buffered_result.total_rows,
        "chunked_rows": chunked_result.total_rows,
        "row_stream_rows": row_result.total_rows,
        "xlsx_cells_equal": buffered_rows == chunked_rows,
        "buffered_column_writes": list(buffered_obs.column_keys),
        "chunked_column_writes": list(chunked_obs.column_keys),
        "row_write_count": len(row_obs.row_field_counts),
        "derived_from_residency_window": derived.value,
        "syntax_hint": "DemandRunRuntimeOptions(output_write_layout=OutputWriteLayout.COLUMN_CHUNKED)",
    }
    return ExampleResult(
        example_id=_EXAMPLE_ID,
        passed=passed,
        kind=EXAMPLE_KIND_ORACLE,
        summary="buffered={} chunked={} row={} factory_ok={} cells_eq={} col_ev={} row_ev={}".format(
            buffered_result.total_rows,
            chunked_result.total_rows,
            row_result.total_rows,
            factory_ok,
            buffered_rows == chunked_rows,
            column_events_ok,
            row_events_ok,
        ),
        details=details,
    )


def run_chapter() -> ExampleResult:
    return run_public_api_output_write_layout()


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
        # example_public_api_suite / ch162_public_api_output_write_layout

        本章目标:
        - 演示 Python 调优旋钮 `OutputWriteLayout`（`row_stream` / `column_buffered` / `column_chunked`）
        - 同 IR 下显式 buffered vs chunked：工厂选型不同，业务格子一致
        - 用 `COLUMN_WRITE` / `ROW_WRITE` 事件核对写出取向
        - **不是**自动切换；YAML books / composition 不能设列布局

        推荐写法:
        ```python
        from scalim.dsl.yaml_dsl import DemandRunRuntimeOptions, OutputWriteLayout
        DemandRunRuntimeOptions(output_write_layout=OutputWriteLayout.COLUMN_CHUNKED)
        ```
        迁移窗仍可用 `excel_column_residency=ExcelColumnResidency.CHUNKED`（未设 layout 时等价推导）。

        SSOT:
        - `notebooks/marimo/example_public_api_suite/chapters/ch162_public_api_output_write_layout.py::run_public_api_output_write_layout`
        - 人类文档: `docs/doc/getting-started/excel-column-residency.md`
        - agent: `agentdev/skills/scalim-yaml-dsl/references/streaming-column-excel-guidance.md`

        Gate:
        - `just examples`
        """
    )
    return


@app.cell
def _():
    import marimo as mo

    return (mo,)


@app.cell
def _():
    from scalim_misc.notebook_support.pathing import ensure_repo_root_on_sys_path

    _ = ensure_repo_root_on_sys_path(__file__)
    return


@app.cell
def _():
    result = run_public_api_output_write_layout()
    chapter_result = {
        "passed": result.passed,
        "summary": result.summary,
        "details": result.details if result.details is not None else {},
    }
    return (chapter_result,)


@app.cell(hide_code=True)
def _(mo, chapter_result):
    mo.callout(
        mo.md("## {}".format("PASS" if chapter_result["passed"] else "FAIL")),
        kind="success" if chapter_result["passed"] else "danger",
    )
    mo.md("```\n{}\n```".format(chapter_result["summary"]))
    return


@app.cell(hide_code=True)
def _(mo, chapter_result):
    from scalim_misc.notebook_support.results_view import details_to_rows

    rows = details_to_rows(chapter_result["details"])
    mo.ui.table(rows, selection=None)
    return (rows,)


if __name__ == "__main__":
    app.run()
